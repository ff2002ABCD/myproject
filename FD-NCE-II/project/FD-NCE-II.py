import numpy as np
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from scipy.integrate import odeint
import matplotlib.pyplot as plt
import os
import sys
from PIL import Image, ImageTk
import pandas as pd
from scipy import interpolate
import openpyxl  # 添加这行
from openpyxl.drawing.image import Image as XLImage  # 添加这行

def get_resource_path(relative_path):
    """获取资源的绝对路径，支持打包后的环境"""
    try:
        # 打包后的临时文件夹路径
        base_path = sys._MEIPASS
    except Exception:
        # 开发环境的路径
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, relative_path)

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'WenQuanYi Micro Hei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

class ChaosCircuit:
    """非线性电路混沌模拟器 - 基于蔡氏电路"""
    
    def __init__(self):
        # 电路参数
        self.C1 = 10e-9
        self.C2 = 100e-9
        self.L = 18e-3
        self.m0 = -0.5e-3
        self.m1 = -0.8e-3
        self.Bp = 1.0
        
        # 初始条件
        self.init_state = [0.1, 0.0, 0.0]
        
        # 仿真参数
        self.t_max = 0.1
        self.t_steps = 50000
        self.t = np.linspace(0, self.t_max, self.t_steps)
        
        # 舍弃瞬态的比例
        self.transient_ratio = 0.5
        
    def nonlinear_resistor(self, V):
        """非线性电阻的伏安特性"""
        if V <= -self.Bp:
            return self.m0 * V + (self.m0 - self.m1) * self.Bp
        elif V >= self.Bp:
            return self.m0 * V + (self.m1 - self.m0) * self.Bp
        else:
            return self.m1 * V
    
    def chua_circuit(self, state, t, R):
        """蔡氏电路微分方程"""
        V1, V2, I = state
        G = 1.0 / R
        IR = self.nonlinear_resistor(V1)
        
        dV1 = (G * (V2 - V1) - IR) / self.C1
        dV2 = (G * (V1 - V2) + I) / self.C2
        dI = -V2 / self.L
        
        return [dV1, dV2, dI]
    
    def simulate(self, R_value):
        """仿真特定电阻值的电路行为"""
        trajectory = odeint(self.chua_circuit, self.init_state, self.t, args=(R_value,))
        return trajectory
    
    def get_steady_state(self, trajectory):
        """获取稳态数据"""
        start_idx = int(len(trajectory) * self.transient_ratio)
        return trajectory[start_idx:]
    
    def get_voltage_current_characteristic(self, V_range=(-5, 5), points=1000):
        """获取非线性电阻的伏安特性数据"""
        V = np.linspace(V_range[0], V_range[1], points)
        I = np.array([self.nonlinear_resistor(v) for v in V])
        return V, I


class NonlinearResistorAnalyzer:
    """有源非线性电阻伏安特性分析器"""
    
    def __init__(self):
        # 测量参数
        self.V_range = (-5, 5)
        self.points = 1000
        self.data_file = None
        self.interp_func = None
        self.load_data()
        
    def load_data(self):
        """从Excel文件加载伏安特性数据"""
        try:
            # 获取数据文件路径
            data_path = get_resource_path(os.path.join("data", "伏安特性.xlsx"))
            
            # 读取Excel文件
            df = pd.read_excel(data_path, header=None)
            
            # 提取电压和电流数据
            voltage = df.iloc[:, 0].values
            current = df.iloc[:, 1].values
            
            # 创建插值函数
            self.interp_func = interpolate.interp1d(
                voltage, current, 
                kind='linear', 
                bounds_error=False, 
                fill_value=(current[0], current[-1])
            )
            
            print(f"成功加载数据文件：{data_path}")
            return True
            
        except Exception as e:
            print(f"加载数据文件失败：{e}")
            # 如果没有数据文件，使用理论曲线作为后备
            circuit = ChaosCircuit()
            V, I = circuit.get_voltage_current_characteristic(self.V_range, self.points)
            self.interp_func = interpolate.interp1d(
                V, I, 
                kind='linear', 
                bounds_error=False, 
                fill_value=(I[0], I[-1])
            )
            return False
    
    def get_current_at_voltage(self, voltage):
        """根据电压查询对应的电流（使用线性插值）"""
        if self.interp_func is not None:
            return float(self.interp_func(voltage))
        else:
            # 如果没有数据文件，使用理论计算
            circuit = ChaosCircuit()
            return circuit.nonlinear_resistor(voltage)
    
    def get_characteristic(self):
        """获取伏安特性数据（用于绘制曲线）"""
        if self.interp_func is not None:
            # 使用插值函数生成平滑曲线
            V = np.linspace(self.V_range[0], self.V_range[1], self.points)
            I = self.interp_func(V)
            return V, I
        else:
            # 使用理论计算
            circuit = ChaosCircuit()
            return circuit.get_voltage_current_characteristic(self.V_range, self.points)
    
    def get_measurement_points(self, num_points=20):
        """获取模拟测量点（用于表格显示）"""
        V, I = self.get_characteristic()
        
        # 均匀采样测量点
        indices = np.linspace(0, len(V)-1, num_points, dtype=int)
        measurement_V = V[indices]
        measurement_I = I[indices]
        
        return measurement_V, measurement_I

class RepeatButton:
    """支持长按重复触发的按钮"""
    
    def __init__(self, parent, text, command, repeat_delay=300, repeat_interval=50, **kwargs):
        """
        初始化重复按钮
        
        Args:
            parent: 父容器
            text: 按钮文本
            command: 点击命令
            repeat_delay: 长按后开始重复的延迟（毫秒）
            repeat_interval: 重复间隔（毫秒）
            **kwargs: 其他按钮参数
        """
        self.command = command
        self.repeat_delay = repeat_delay
        self.repeat_interval = repeat_interval
        self.repeat_job = None
        self.is_repeating = False
        
        # 提取颜色参数
        bg_color = kwargs.pop('bg', None)
        active_bg = kwargs.pop('activebackground', None)
        
        # 创建按钮
        self.button = tk.Button(parent, text=text, **kwargs)
        
        # 单独设置颜色参数
        if bg_color:
            self.button.config(bg=bg_color)
        if active_bg:
            self.button.config(activebackground=active_bg)
        
        # 绑定事件
        self.button.bind('<ButtonPress-1>', self.on_press)
        self.button.bind('<ButtonRelease-1>', self.on_release)
        self.button.bind('<Leave>', self.on_leave)
        
    def on_press(self, event):
        """鼠标按下事件"""
        # 立即执行一次命令
        self.command()
        
        # 延迟后开始重复
        self.repeat_job = self.button.after(self.repeat_delay, self.start_repeat)
        
    def on_release(self, event):
        """鼠标释放事件"""
        self.stop_repeat()
        
    def on_leave(self, event):
        """鼠标离开按钮事件"""
        # 如果鼠标离开按钮，停止重复
        if self.is_repeating:
            self.stop_repeat()
        
    def start_repeat(self):
        """开始重复执行"""
        self.is_repeating = True
        self.repeat()
        
    def repeat(self):
        """重复执行命令"""
        if self.is_repeating:
            self.command()
            self.repeat_job = self.button.after(self.repeat_interval, self.repeat)
        
    def stop_repeat(self):
        """停止重复"""
        self.is_repeating = False
        if self.repeat_job:
            self.button.after_cancel(self.repeat_job)
            self.repeat_job = None
            
    def pack(self, **kwargs):
        """包装按钮的pack方法"""
        self.button.pack(**kwargs)
        
    def grid(self, **kwargs):
        """包装按钮的grid方法"""
        self.button.grid(**kwargs)
        
    def config(self, **kwargs):
        """配置按钮属性"""
        self.button.config(**kwargs)


class ChaosVisualizerTk:
    """混沌现象可视化器 - Tkinter版本"""
    
    def __init__(self, circuit):
        self.circuit = circuit
        self.current_R = 1.79
        
        # 典型电阻值
        self.typical_values = {
            '单环': 1.790,
            '双环': 1.785,
            '四环': 1.7825,
            '三环': 1.7801,
            '混沌': 1.775
        }
        
        # 创建主窗口
        self.root = tk.Tk()
        self.root.title("FD-NCE-II 非线性电路混沌实验仪")
        self.root.geometry("1600x800")
        
        # 创建顶部选项卡框架
        self.top_frame = ttk.Frame(self.root)
        self.top_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        
        # 创建选项卡变量
        self.tab_var = tk.StringVar(value="chaos")  # 默认选择混沌实验
        
        # 创建两个选项卡按钮
        self.chaos_radio = ttk.Radiobutton(
            self.top_frame, 
            text="倍周期分岔和混沌现象的观测及相图描绘", 
            variable=self.tab_var, 
            value="chaos",
            command=self.switch_tab
        )
        self.chaos_radio.pack(side=tk.LEFT, padx=10)
        
        self.iv_radio = ttk.Radiobutton(
            self.top_frame, 
            text="有源非线性电阻的伏安特性测量", 
            variable=self.tab_var, 
            value="iv",
            command=self.switch_tab
        )
        self.iv_radio.pack(side=tk.LEFT, padx=10)
        
        # 创建内容框架（用于切换界面）
        self.content_frame = ttk.Frame(self.root)
        self.content_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建两个界面
        self.chaos_frame = ChaosInterface(self.content_frame, self)
        self.iv_frame = IVInterface(self.content_frame, self)
        
        # 默认显示混沌界面，隐藏伏安特性界面
        self.iv_frame.pack_forget()
        self.chaos_frame.pack(fill=tk.BOTH, expand=True)
        
        # 绑定窗口关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def switch_tab(self):
        """切换选项卡"""
        if self.tab_var.get() == "chaos":
            # 显示混沌界面，隐藏伏安特性界面
            self.iv_frame.pack_forget()
            self.chaos_frame.pack(fill=tk.BOTH, expand=True)
        else:  # iv
            # 显示伏安特性界面，隐藏混沌界面
            self.chaos_frame.pack_forget()
            self.iv_frame.pack(fill=tk.BOTH, expand=True)
    
    def on_closing(self):
        """窗口关闭事件"""
        plt.close('all')
        self.root.quit()
        self.root.destroy()
    
    def run(self):
        """运行应用"""
        self.root.mainloop()


class ChaosInterface(ttk.Frame):
    """混沌现象实验界面"""
    
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.circuit = app.circuit
        self.current_R = app.current_R
        
        # 拆分为W1（粗调）和W2（细调）
        self.W1 = 1.79  # 粗调部分
        self.W2 = 0.00  # 细调部分
        self.total_R = self.W1 + self.W2  # 总电阻
        
        self.typical_values = app.typical_values
        
        self.setup_interface()
        
        # 初始隐藏图形（因为电感初始是取下状态）
        self.after(100, self.hide_plots)
    
    def setup_interface(self):
        """设置界面"""
        # 创建左右两个主框架
        self.left_frame = ttk.Frame(self, padding="5")
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.right_frame = ttk.Frame(self, padding="5", width=650)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=False)
        self.right_frame.pack_propagate(False)  # 固定宽度
        
        # 设置右侧框架
        self.setup_right_frame()  # 修正为正确的方法名
        
        # 设置左侧图形界面
        self.setup_left_frame()
    
    def setup_left_frame(self):
        """设置左侧图形界面"""
        # 创建matplotlib图形
        self.fig = Figure(figsize=(10, 8), dpi=100)
        
        # 创建子图
        self.ax_phase = self.fig.add_subplot(2, 1, 1)
        self.ax_wave = self.fig.add_subplot(2, 1, 2)
        
        # 设置相图
        self.ax_phase.set_title('CH1-CH2 相图')
        self.ax_phase.set_xlabel('V1 (V)')
        self.ax_phase.set_ylabel('V2 (V)')
        self.ax_phase.grid(True, alpha=0.3)
        self.ax_phase.set_xlim([-10, 10])
        self.ax_phase.set_ylim([-3, 3])
        self.ax_phase.set_aspect('equal')
        
        # 设置波形图
        self.ax_wave.set_title('CH1-CH2 时域波形')
        self.ax_wave.set_xlabel('时间 (ms)')
        self.ax_wave.set_ylabel('电压 (V)')
        self.ax_wave.grid(True, alpha=0.3)
        self.ax_wave.set_xlim([50, 55])
        self.ax_wave.set_ylim([-10, 10])
        
        # 初始仿真
        traj = self.circuit.simulate(self.current_R * 1000)
        steady_traj = self.circuit.get_steady_state(traj)
        t_ms = self.circuit.t * 1000
        steady_t_ms = t_ms[int(len(t_ms) * self.circuit.transient_ratio):]
        
        # 绘制相图
        self.phase_line, = self.ax_phase.plot(steady_traj[:, 0], steady_traj[:, 1], 
                                              'b-', linewidth=0.8, alpha=0.8)
        
        # 绘制波形图
        self.ch1_line, = self.ax_wave.plot(steady_t_ms, steady_traj[:, 0], 
                                           'r-', label='CH1', linewidth=0.8)
        self.ch2_line, = self.ax_wave.plot(steady_t_ms, steady_traj[:, 1], 
                                           'b-', label='CH2', linewidth=0.8)
        self.ax_wave.legend(loc='upper right', fontsize=9)
        
        # # 电阻值显示
        # self.r_text = self.ax_phase.text(0.02, 0.98, f'R = {self.current_R:.4f} kΩ', 
        #                                 transform=self.ax_phase.transAxes, fontsize=11,
        #                                 verticalalignment='top', 
        #                                 bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
        
        # 将图形嵌入tkinter
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.left_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        
        # 添加matplotlib工具栏
        # self.toolbar = NavigationToolbar2Tk(self.canvas, self.left_frame)
        # self.toolbar.update()
        # self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    
    def setup_right_frame(self):
        """设置右侧控制面板"""
        # 添加图片
        self.setup_image()
        
        # 添加电感控制框架
        inductor_frame = ttk.Frame(self.right_frame)
        inductor_frame.pack(fill=tk.X, pady=5)
        
        # 电感状态变量
        self.inductor_state = tk.BooleanVar(value=False)
        
        # 创建按钮框架
        inductor_btn_frame = ttk.Frame(inductor_frame)
        inductor_btn_frame.pack(fill=tk.X, pady=5)
        
        # 摆放电感按钮
        self.inductor_on_btn = ttk.Button(inductor_btn_frame, text="摆放电感", 
                                        command=self.inductor_on,
                                        state=tk.NORMAL)
        self.inductor_on_btn.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        # 取下电感按钮
        self.inductor_off_btn = ttk.Button(inductor_btn_frame, text="取下电感", 
                                        command=self.inductor_off,
                                        state=tk.DISABLED)
        self.inductor_off_btn.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        # 电阻控制框架
        control_frame = ttk.Frame(self.right_frame)
        control_frame.pack(fill=tk.X, pady=10)
        
        # W1（粗调）滑块
        w1_frame = ttk.LabelFrame(control_frame, text="W1 粗调", padding="5")
        w1_frame.pack(fill=tk.X, pady=5)
        
        # 创建水平布局框架，包含滑块和按钮
        w1_horizontal_frame = ttk.Frame(w1_frame)
        w1_horizontal_frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(w1_horizontal_frame, text="范围:").pack(side=tk.LEFT)
        
        # W1滑块
        self.w1_slider = tk.Scale(w1_horizontal_frame, from_=1.55, to=1.82, 
                                resolution=0.001, orient=tk.HORIZONTAL,
                                length=350, command=self.on_w1_change, showvalue=False)
        self.w1_slider.set(self.W1)
        self.w1_slider.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # W1微调按钮框架
        w1_btn_frame = ttk.Frame(w1_horizontal_frame)
        w1_btn_frame.pack(side=tk.LEFT, padx=2)
        
        # W1- 按钮
        self.btn_w1_minus = RepeatButton(w1_btn_frame, text="W1-", 
                                        command=lambda: self.adjust_w1(-0.001),
                                        repeat_delay=300, repeat_interval=50,
                                        bg='lightcoral', activebackground='red',
                                        width=4, height=1)
        self.btn_w1_minus.pack(side=tk.LEFT, padx=1)
        
        # W1+ 按钮
        self.btn_w1_plus = RepeatButton(w1_btn_frame, text="W1+", 
                                    command=lambda: self.adjust_w1(0.001),
                                    repeat_delay=300, repeat_interval=50,
                                    bg='lightgreen', activebackground='green',
                                    width=4, height=1)
        self.btn_w1_plus.pack(side=tk.LEFT, padx=1)
        
        # W2（细调）滑块
        w2_frame = ttk.LabelFrame(control_frame, text="W2 细调", padding="5")
        w2_frame.pack(fill=tk.X, pady=5)
        
        # 创建水平布局框架，包含滑块和按钮
        w2_horizontal_frame = ttk.Frame(w2_frame)
        w2_horizontal_frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(w2_horizontal_frame, text="范围:").pack(side=tk.LEFT)
        
        # W2滑块
        self.w2_slider = tk.Scale(w2_horizontal_frame, from_=-0.02, to=0.02, 
                                resolution=0.0001, orient=tk.HORIZONTAL,
                                length=350, command=self.on_w2_change, showvalue=False)
        self.w2_slider.set(self.W2)
        self.w2_slider.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # W2微调按钮框架
        w2_btn_frame = ttk.Frame(w2_horizontal_frame)
        w2_btn_frame.pack(side=tk.LEFT, padx=2)
        
        # W2- 按钮
        self.btn_w2_minus = RepeatButton(w2_btn_frame, text="W2-", 
                                        command=lambda: self.adjust_w2(-0.0001),
                                        repeat_delay=300, repeat_interval=50,
                                        bg='lightblue', activebackground='blue',
                                        width=4, height=1)
        self.btn_w2_minus.pack(side=tk.LEFT, padx=1)
        
        # W2+ 按钮
        self.btn_w2_plus = RepeatButton(w2_btn_frame, text="W2+", 
                                    command=lambda: self.adjust_w2(0.0001),
                                    repeat_delay=300, repeat_interval=50,
                                    bg='lightyellow', activebackground='yellow',
                                    width=4, height=1)
        self.btn_w2_plus.pack(side=tk.LEFT, padx=1)
        
        # # W2复位按钮
        # self.btn_w2_reset = ttk.Button(w2_btn_frame, text="复位", 
        #                             command=self.reset_w2, width=4)
        # self.btn_w2_reset.pack(side=tk.LEFT, padx=5)
        
        # # 典型值按钮框架
        # typical_frame = ttk.Frame(self.right_frame, padding="10")
        # typical_frame.pack(fill=tk.X, pady=10)
        
        # # 创建典型值按钮
        # button_frame = ttk.Frame(typical_frame)
        # button_frame.pack(fill=tk.X)
        
        # colors = ['lightblue', 'lightgreen', 'lightcoral', 'plum', 'lightgray']
        # for i, (key, value) in enumerate(self.typical_values.items()):
        #     btn = tk.Button(button_frame, text=key, bg=colors[i % len(colors)],
        #                 command=lambda v=value: self.set_typical_value(v),
        #                 relief=tk.RAISED, borderwidth=2)
        #     btn.pack(side=tk.LEFT, padx=2, pady=2, fill=tk.X, expand=True)
    
    def setup_image(self):
        """设置右侧图片"""
        try:
            # 图片路径
            img_path = get_resource_path(os.path.join("background", "接示波器.jpg"))
            
            # 打开并调整图片大小
            pil_image = Image.open(img_path)
            
            # 计算调整后的大小
            basewidth = 530
            wpercent = (basewidth / float(pil_image.size[0]))
            hsize = int((float(pil_image.size[1]) * float(wpercent)))
            pil_image = pil_image.resize((basewidth, hsize), Image.Resampling.LANCZOS)
            
            # 创建图片显示框架（用于叠加）
            self.image_frame = ttk.Frame(self.right_frame)
            self.image_frame.pack(pady=10)
            
            # 创建画布用于显示图片和叠加电感图片
            self.image_canvas = tk.Canvas(self.image_frame, width=basewidth, height=hsize, highlightthickness=0)
            self.image_canvas.pack()
            
            # 转换为PhotoImage并保存引用
            self.bg_photo = ImageTk.PhotoImage(pil_image)
            
            # 在画布上显示背景图片
            self.image_canvas.create_image(0, 0, anchor=tk.NW, image=self.bg_photo, tags="bg")
            
            # 加载电感图片（但不显示）
            try:
                inductor_img_path = get_resource_path(os.path.join("background", "电感.jpg"))
                inductor_pil = Image.open(inductor_img_path)
                
                # ========== 修改这里：调整电感图片大小 ==========
                # 设置电感图片的宽度（根据需要调整）
                inductor_width = 40  # 可以改成你想要的大小，比如 150, 200, 250 等
                # 计算高度保持宽高比
                wpercent = (inductor_width / float(inductor_pil.size[0]))
                inductor_height = int((float(inductor_pil.size[1]) * float(wpercent)))
                inductor_pil = inductor_pil.resize((inductor_width, inductor_height), Image.Resampling.LANCZOS)
                # ==============================================
                
                self.inductor_photo = ImageTk.PhotoImage(inductor_pil)
                
                # ========== 修改这里：调整电感图片位置 ==========
                # 设置电感图片的位置坐标
                # 画布左上角是 (0,0)，右下角是 (basewidth, hsize)
                inductor_x = basewidth // 2 -218  # 水平居中
                inductor_y = hsize // 3 +138     # 垂直方向在顶部1/3处
                # 其他可选位置：
                # inductor_x = basewidth // 2      # 水平居中
                # inductor_y = hsize // 2           # 垂直居中
                # inductor_x = basewidth - 150      # 靠右
                # inductor_y = 50                    # 靠上
                # inductor_x = 150                   # 靠左
                # inductor_y = hsize - 150           # 靠下
                # ==============================================
                
                # 在画布上创建电感图片（初始隐藏）
                self.inductor_canvas_item = self.image_canvas.create_image(
                    inductor_x, inductor_y,  # 使用自定义位置
                    image=self.inductor_photo,
                    tags="inductor",
                    state="hidden"  # 初始隐藏
                )
            except Exception as e:
                print(f"电感图片加载失败: {e}")
                self.inductor_photo = None
                self.inductor_canvas_item = None
            
        except Exception as e:
            # 如果图片加载失败，显示错误信息
            error_label = ttk.Label(self.right_frame, 
                                text=f"图片加载失败\n{str(e)}\n请确保background文件夹中存在'接示波器.jpg'",
                                foreground='red', justify=tk.CENTER)
            error_label.pack(pady=10)
            print(f"图片加载错误: {e}")

    def inductor_on(self):
        """放上电感"""
        self.inductor_state.set(True)
        
        # 更新按钮状态
        self.inductor_on_btn.config(state=tk.DISABLED)
        self.inductor_off_btn.config(state=tk.NORMAL)
        
        # 显示电感图片
        if hasattr(self, 'inductor_canvas_item') and self.inductor_canvas_item:
            self.image_canvas.itemconfig(self.inductor_canvas_item, state="normal")
        
        # 显示左侧波形和相图
        self.show_plots()

    def inductor_off(self):
        """取下电感"""
        self.inductor_state.set(False)
        
        # 更新按钮状态
        self.inductor_on_btn.config(state=tk.NORMAL)
        self.inductor_off_btn.config(state=tk.DISABLED)
        
        # 隐藏电感图片
        if hasattr(self, 'inductor_canvas_item') and self.inductor_canvas_item:
            self.image_canvas.itemconfig(self.inductor_canvas_item, state="hidden")
        
        # 隐藏左侧波形和相图
        self.hide_plots()

    def show_plots(self):
        """显示波形和相图"""
        # 显示图形（通过更新数据实现）
        self.update_plot()

    def hide_plots(self):
        """隐藏波形和相图"""
        # 清空图形数据
        self.phase_line.set_data([], [])
        self.ch1_line.set_data([], [])
        self.ch2_line.set_data([], [])
        self.canvas.draw_idle()
    
    def on_slider_change(self, value):
        """滑块变化事件"""
        self.current_R = float(value)
        # self.r_var.set(f"{self.current_R:.4f} kΩ")
        self.update_plot()
    
    def adjust_resistance(self, delta):
        """调整电阻值"""
        new_R = self.current_R + delta
        if 1.55 <= new_R <= 1.82:
            self.slider.set(new_R)
    
    def on_w1_change(self, value):
        """W1滑块变化事件"""
        self.W1 = float(value)
        # self.w1_var.set(f"{self.W1:.3f} kΩ")
        self.update_total_resistance()

    def on_w2_change(self, value):
        """W2滑块变化事件"""
        self.W2 = float(value)
        # self.w2_var.set(f"{self.W2:+.4f} kΩ")
        self.update_total_resistance()

    def adjust_w1(self, delta):
        """调整W1值"""
        new_w1 = self.W1 + delta
        if 1.55 <= new_w1 <= 1.82:
            self.w1_slider.set(new_w1)

    def adjust_w2(self, delta):
        """调整W2值"""
        new_w2 = self.W2 + delta
        if -0.02 <= new_w2 <= 0.02:
            self.w2_slider.set(new_w2)

    def reset_w2(self):
        """复位W2为0"""
        self.w2_slider.set(0.0)

    def update_total_resistance(self):
        """更新总电阻"""
        self.total_R = self.W1 + self.W2
        # self.total_r_var.set(f"{self.total_R:.4f} kΩ")
        
        # 如果电感已放上，更新图形
        if self.inductor_state.get():
            self.update_plot()

    def set_typical_value(self, value):
        """设置典型值"""
        # 典型值设置到W1，W2复位到0
        self.w1_slider.set(value)
        self.reset_w2()

    def set_typical_value(self, value):
        """设置典型值"""
        self.slider.set(value)
    
    def update_plot(self):
        """更新图形"""
        # 如果电感没放上，不更新图形
        if not self.inductor_state.get():
            return
        
        # 使用总电阻
        R_kohm = self.total_R
        R_ohm = R_kohm * 1000
        
        traj = self.circuit.simulate(R_ohm)
        steady_traj = self.circuit.get_steady_state(traj)
        
        t_ms = self.circuit.t * 1000
        steady_t_ms = t_ms[int(len(t_ms) * self.circuit.transient_ratio):]
        
        self.phase_line.set_data(steady_traj[:, 0], steady_traj[:, 1])
        self.ch1_line.set_data(steady_t_ms, steady_traj[:, 0])
        self.ch2_line.set_data(steady_t_ms, steady_traj[:, 1])
        
        # 自动调整坐标轴范围以更好地显示波形
        self.ax_phase.relim()
        self.ax_phase.autoscale_view()
        self.ax_wave.relim()
        self.ax_wave.autoscale_view()
        
        self.canvas.draw_idle()


class IVInterface(ttk.Frame):
    """有源非线性电阻伏安特性测量界面"""
    
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.analyzer = NonlinearResistorAnalyzer()
        
        # 初始化数据存储
        self.measure_V = []
        self.measure_I = []

        # 添加拟合直线存储
        self.fit_lines = []  # 存储三条拟合直线的信息
        self.fit_line_plots = []  # 存储三条拟合直线的图形对象
        self.fit_texts = []  # 存储三条拟合直线的公式文本
        
        self.setup_interface()
    
    def setup_interface(self):
        """设置界面"""
        # 创建左右两个主框架
        self.left_frame = ttk.Frame(self, padding="5")
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.right_frame = ttk.Frame(self, padding="5", width=650)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=False)
        self.right_frame.pack_propagate(False)
        
        # 不再预先创建电压滑块，将在电阻箱控制中动态计算
        self.voltage_slider = None  # 保持属性存在，但设为None
        
        # 设置右侧图片（现在可以使用 voltage_slider）
        self.setup_right_image()
        
        # 设置电压控制（放在图片下面）
        self.setup_voltage_control()
        
        # 设置左侧内容（表格和曲线图）
        self.setup_left_content()
    
    def setup_right_image(self):
        """设置右侧图片"""
        try:
            # 图片路径
            img_path = get_resource_path(os.path.join("background", "接电压表.jpg"))
            
            # 打开并调整图片大小
            pil_image = Image.open(img_path)
            
            # 计算调整后的大小（宽度430像素，保持宽高比）
            basewidth = 530
            wpercent = (basewidth / float(pil_image.size[0]))
            hsize = int((float(pil_image.size[1]) * float(wpercent)))
            pil_image = pil_image.resize((basewidth, hsize), Image.Resampling.LANCZOS)
            
            # 转换为PhotoImage
            self.photo = ImageTk.PhotoImage(pil_image)
            
            # 创建图片显示框架（用于叠加文字）
            image_frame = ttk.Frame(self.right_frame)
            image_frame.pack(pady=10)
            
            # 创建画布用于显示图片和文字
            self.image_canvas = tk.Canvas(image_frame, width=basewidth, height=hsize, highlightthickness=0)
            self.image_canvas.pack()
            
            # 在画布上显示图片
            self.image_canvas.create_image(0, 0, anchor=tk.NW, image=self.photo)
            
            # 获取当前电压（从电阻计算）
            min_resistance = 0.0
            max_resistance = 99999.9
            min_voltage = -12.6  # 最小电阻对应最负的电压
            max_voltage = -0.01  # 最大电阻对应最接近0的电压

            if hasattr(self, 'resistance_value'):
                if max_resistance - min_resistance != 0:
                    current_voltage = min_voltage + (self.resistance_value - min_resistance) * (max_voltage - min_voltage) / (max_resistance - min_resistance)
                else:
                    current_voltage = min_voltage
            else:
                current_voltage = -6.3  # 默认值
            
            # 在图片上添加电压显示（白色背景，黑色文字）
            self.voltage_display = self.image_canvas.create_text(
                basewidth // 2-159, 80,  # 位于图片上方中间位置
                text=f"{current_voltage:.3f} V",
                fill='black',
                font=('Arial', 16, 'bold'),
                tags="voltage_text"
            )
            
            # 添加白色背景框
            bbox = self.image_canvas.bbox(self.voltage_display)
            if bbox:
                self.image_canvas.create_rectangle(
                    bbox[0]-10, bbox[1]-5, bbox[2]+10, bbox[3]+5,
                    fill='white', outline='black', width=2,
                    tags="voltage_bg"
                )
                # 将文字置前
                self.image_canvas.tag_raise(self.voltage_display)
            
            
        except Exception as e:
            # 如果图片加载失败，显示错误信息
            error_label = ttk.Label(self.right_frame, 
                                text=f"图片加载失败\n{str(e)}\n请确保background文件夹中存在'接电压表.jpg'",
                                foreground='red', justify=tk.CENTER)
            error_label.pack(pady=10)
            print(f"图片加载错误: {e}")
    
    def setup_voltage_control(self):
        """设置电压控制界面（在图片下方）- 6位电阻箱"""
        # 电压控制框架
        voltage_frame = ttk.LabelFrame(self.right_frame, text="电阻箱控制", padding="10")
        voltage_frame.pack(fill=tk.X, pady=10)
        
        # 电阻值显示
        # self.resistance_var = tk.StringVar(value="00000.0 Ω")
        # resistance_label = ttk.Label(voltage_frame, textvariable=self.resistance_var, 
        #                             font=('Courier', 16, 'bold'), foreground='blue')
        # resistance_label.pack(pady=5)
        
        # 电压对应显示
        # voltage_label = ttk.Label(voltage_frame, text="对应电压: -6.300 V", 
        #                         font=('Arial', 10), foreground='gray')
        # voltage_label.pack(pady=2)
        
        # 创建6位电阻的框架
        digits_frame = ttk.Frame(voltage_frame)
        digits_frame.pack(fill=tk.X, pady=10)
        
        # 电阻位数配置：5位整数 + 1位小数
        # 每一位的范围：0-9
        self.digit_vars = []  # 存储每一位的变量
        self.digit_frames = []  # 存储每一位的框架
        
        # 整数部分5位
        for i in range(5):
            self.create_digit_control(digits_frame, i, is_integer=True)
        
        # 小数点
        dot_label = ttk.Label(digits_frame, text=".", font=('Arial', 16, 'bold'))
        dot_label.pack(side=tk.LEFT, padx=2)
        
        # 小数部分1位
        self.create_digit_control(digits_frame, 5, is_integer=False)
        
        # 单位标签
        unit_label = ttk.Label(digits_frame, text="Ω", font=('Arial', 16, 'bold'))
        unit_label.pack(side=tk.LEFT, padx=5)
        
        # 初始化电阻值为0
        self.resistance_value = 0.0
        self.update_resistance_display()

    def create_digit_control(self, parent, position, is_integer=True):
        """创建每一位的控制组件"""
        digit_frame = ttk.Frame(parent)
        digit_frame.pack(side=tk.LEFT, padx=2)
        
        # 位名称
        if is_integer:
            if position == 0:
                name = "X10000"
            elif position == 1:
                name = "X1000"
            elif position == 2:
                name = "X100"
            elif position == 3:
                name = "X10"
            else:  # position == 4
                name = "X1"
        else:
            name = "X0.1"
        
        name_label = ttk.Label(digit_frame, text=name, font=('Arial', 8))
        name_label.pack()
        
        # 当前位数值显示
        digit_var = tk.StringVar(value="0")
        digit_label = ttk.Label(digit_frame, textvariable=digit_var, 
                            font=('Courier', 14, 'bold'), width=2)
        digit_label.pack()
        
        self.digit_vars.append({
            'var': digit_var,
            'position': position,
            'is_integer': is_integer,
            'value': 0
        })
        
        # 按钮框架
        btn_frame = ttk.Frame(digit_frame)
        btn_frame.pack()
        
        # + 按钮
        plus_btn = RepeatButton(btn_frame, text="+", 
                            command=lambda p=position: self.adjust_digit(p, 1),
                            repeat_delay=300, repeat_interval=100,
                            width=2, height=1)
        plus_btn.pack(side=tk.LEFT, padx=1)
        
        # - 按钮
        minus_btn = RepeatButton(btn_frame, text="-", 
                                command=lambda p=position: self.adjust_digit(p, -1),
                                repeat_delay=300, repeat_interval=100,
                                width=2, height=1)
        minus_btn.pack(side=tk.LEFT, padx=1)
        
        return digit_frame

    def adjust_digit(self, position, delta):
        """调整某一位的数字，支持循环（9+1=0，0-1=9）"""
        # 获取当前位的数值
        current_value = int(self.digit_vars[position]['var'].get())
        
        # 计算新值（带循环）
        if delta > 0:  # 增加
            new_value = (current_value + 1) % 10
        else:  # 减少
            new_value = (current_value - 1) % 10
        
        # 更新显示
        self.digit_vars[position]['var'].set(str(new_value))
        self.digit_vars[position]['value'] = new_value
        
        # 重新计算总电阻值
        self.calculate_total_resistance()

    def calculate_total_resistance(self):
        """根据各位数值计算总电阻"""
        # 整数部分：位置0-4
        integer_value = 0
        for i in range(5):
            integer_value = integer_value * 10 + self.digit_vars[i]['value']
        
        # 小数部分：位置5
        decimal_value = self.digit_vars[5]['value']
        
        # 总电阻 = 整数部分 + 小数部分/10
        self.resistance_value = integer_value + decimal_value / 10.0
        
        # 更新显示
        self.update_resistance_display()

    def update_resistance_display(self):
        """更新电阻显示和对应的电压"""
        # 计算对应电压：使用指数衰减函数（与measure_data中一致）
        min_voltage = -12.6   # 最小电阻时的电压（最负）
        max_voltage = -0.01   # 最大电阻时的电压（最接近0）
        
        # 计算衰减常数 tau，使得在 R=32000 时接近饱和
        tau = 32000 / 3  # 当 R=3*tau 时，达到 exp(-3)=0.05，即达到95%的饱和
        
        # 指数衰减公式
        voltage = max_voltage + (min_voltage - max_voltage) * np.exp(-self.resistance_value / tau)
        
        # 确保电压在合理范围内
        voltage = max(min_voltage, min(max_voltage, voltage))
        
        # 更新图片上的电压显示
        if hasattr(self, 'image_canvas') and hasattr(self, 'voltage_display'):
            try:
                self.image_canvas.itemconfig(self.voltage_display, text=f"{voltage:.3f}")
                
                # 更新背景框
                bbox = self.image_canvas.bbox(self.voltage_display)
                if bbox:
                    self.image_canvas.delete("voltage_bg")
                    self.image_canvas.create_rectangle(
                        bbox[0]-10, bbox[1]-5, bbox[2]+10, bbox[3]+5,
                        fill='white', outline='black', width=2,
                        tags="voltage_bg"
                    )
                    self.image_canvas.tag_raise(self.voltage_display)
            except:
                pass
    
    def setup_left_content(self):
        """设置左侧内容（表格和曲线图）"""
        # 创建上半部分：表格
        table_frame = ttk.LabelFrame(self.left_frame, text="测量数据", padding="5")
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        
        # 创建表格框架（带滚动条）
        table_container = ttk.Frame(table_frame)
        table_container.pack(fill=tk.BOTH, expand=True)
        
        # 创建垂直滚动条
        v_scrollbar = ttk.Scrollbar(table_container)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建水平滚动条
        h_scrollbar = ttk.Scrollbar(table_container, orient=tk.HORIZONTAL)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 创建表格（Treeview）- 交换电流和电阻列的位置
        columns = ('序号', '电压 (V)', '电阻 (Ω)', '电流 (mA)')
        self.table = ttk.Treeview(
            table_container, 
            columns=columns,
            show='headings',
            yscrollcommand=v_scrollbar.set,
            xscrollcommand=h_scrollbar.set,
            height=12
        )

        # 设置列标题 - 对应新的列顺序
        self.table.heading('序号', text='序号')
        self.table.heading('电压 (V)', text='电压 (V)')
        self.table.heading('电阻 (Ω)', text='电阻 (Ω)')
        self.table.heading('电流 (mA)', text='电流 (mA)')

        # 设置列宽 - 对应新的列顺序
        self.table.column('序号', width=60, anchor='center')
        self.table.column('电压 (V)', width=100, anchor='center')
        self.table.column('电阻 (Ω)', width=100, anchor='center')
        self.table.column('电流 (mA)', width=100, anchor='center')
        
        self.table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        v_scrollbar.config(command=self.table.yview)
        h_scrollbar.config(command=self.table.xview)
        
        # 创建按钮框架
        button_frame = ttk.Frame(table_frame)
        button_frame.pack(fill=tk.X, pady=5)
        
        # 测量按钮
        measure_btn = ttk.Button(button_frame, text="测量数据", command=self.measure_data)
        measure_btn.pack(side=tk.LEFT, padx=2)
        
        # 拟合直线1按钮
        fit1_btn = ttk.Button(button_frame, text="拟合直线1", command=self.fit_line1)
        fit1_btn.pack(side=tk.LEFT, padx=2)

        # 拟合直线2按钮
        fit2_btn = ttk.Button(button_frame, text="拟合直线2", command=self.fit_line2)
        fit2_btn.pack(side=tk.LEFT, padx=2)

        # 拟合直线3按钮
        fit3_btn = ttk.Button(button_frame, text="拟合直线3", command=self.fit_line3)
        fit3_btn.pack(side=tk.LEFT, padx=2)
        
        # 添加删除选中行按钮
        delete_btn = ttk.Button(button_frame, text="删除选中行", command=self.delete_selected_row)
        delete_btn.pack(side=tk.LEFT, padx=2)
        
       

        # 清除按钮
        clear_btn = ttk.Button(button_frame, text="清空数据", command=self.clear_data)
        clear_btn.pack(side=tk.LEFT, padx=2)
        
         # 导出按钮
        export_btn = ttk.Button(button_frame, text="导出数据", command=self.export_data)
        export_btn.pack(side=tk.LEFT, padx=2)
        
        # 添加导入数据按钮
        import_btn = ttk.Button(button_frame, text="导入数据", command=self.import_data)
        import_btn.pack(side=tk.LEFT, padx=2)

        # 创建下半部分：曲线图
        plot_frame = ttk.LabelFrame(self.left_frame, text="伏安特性曲线", padding="5")
        plot_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        
        # 创建matplotlib图形
        self.fig = Figure(figsize=(8, 5), dpi=100)
        self.ax = self.fig.add_subplot(111)
        
        # 设置图形
        self.ax.set_title('有源非线性电阻伏安特性')
        self.ax.set_xlabel('电压 (V)')
        self.ax.set_ylabel('电流 (mA)')
        self.ax.grid(True, alpha=0.3)
        
        # 设置横坐标范围与电压控制范围一致
        self.ax.set_xlim([-13, 1])  # 从-13V到1V，覆盖-12.6V到-0.01V的范围
        
        # 设置纵坐标范围 0-6 mA
        self.ax.set_ylim([0, 6])
        
        # 绘制测量点（初始为空）
        self.measure_scatter = self.ax.scatter([], [], c='red', s=30, marker='o', label='测量点', zorder=5)
        # 添加连线（初始为空）
        self.measure_line, = self.ax.plot([], [], 'r-', linewidth=1.5, label='测量曲线', alpha=0.7, zorder=4)

        # 添加三条拟合直线（初始隐藏）- 使用color参数指定颜色，格式字符串只指定线型
        self.fit_line1_plot, = self.ax.plot([], [], '--', linewidth=2, color='green', label='拟合直线1', alpha=0.8, visible=False)
        self.fit_line2_plot, = self.ax.plot([], [], '--', linewidth=2, color='blue', label='拟合直线2', alpha=0.8, visible=False)
        self.fit_line3_plot, = self.ax.plot([], [], '--', linewidth=2, color='magenta', label='拟合直线3', alpha=0.8, visible=False)

        # 将三条拟合直线添加到列表中
        self.fit_line_plots = [self.fit_line1_plot, self.fit_line2_plot, self.fit_line3_plot]

        # 初始化三个公式文本（初始为空）
        self.fit_texts = []
        colors = ['green', 'blue', 'magenta']
        # 为每个公式设置不同的 x 坐标，避免重叠
        # 直线1在左边 (0.02)，直线2在中间 (0.34)，直线3在右边 (0.66)
        x_positions = [0.02, 0.34, 0.66]  # 三个不同的水平位置
        for i in range(3):
            text = self.ax.text(x_positions[i], 0.98, '', 
                            transform=self.ax.transAxes,
                            verticalalignment='top',
                            horizontalalignment='left',
                            color=colors[i],
                            fontsize=9,
                            bbox=dict(boxstyle='round', facecolor='white', alpha=0.7, edgecolor=colors[i]),
                            visible=False)
            self.fit_texts.append(text)

        # 创建图例 - 确保所有线条都有正确的颜色
        # 先清除旧的图例
        if self.ax.legend_:
            self.ax.legend_.remove()

        # 创建新的图例，只包含可见的线条
        handles = [self.measure_scatter, self.measure_line]
        # 只有当拟合直线可见时才添加到图例
        for i, plot in enumerate(self.fit_line_plots):
            if plot.get_visible():
                handles.append(plot)

        self.ax.legend(handles=handles, loc='best')
        
        # 将图形嵌入tkinter
        self.canvas = FigureCanvasTkAgg(self.fig, master=plot_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        
        # # 添加matplotlib工具栏
        # self.toolbar = NavigationToolbar2Tk(self.canvas, plot_frame)
        # self.toolbar.update()
        # self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    
    def measure_data(self):
        """测量数据 - 根据当前电阻计算电压，再查询电流"""
        # 使用指数衰减函数：电压 = V_max + (V_min - V_max) * exp(-R / tau)
        # 其中 tau 是时间常数，控制衰减速度
        
        max_resistance = 99999.9
        min_voltage = -12.6   # 最小电阻时的电压（最负）
        max_voltage = -0.01   # 最大电阻时的电压（最接近0）
        
        # 计算衰减常数 tau，使得在 R=32000 时接近饱和
        # 饱和标准：达到最大电压的 95%
        tau = 32000 / 3  # 当 R=3*tau 时，达到 exp(-3)=0.05，即达到95%的饱和
        
        # 指数衰减公式
        voltage = max_voltage + (min_voltage - max_voltage) * np.exp(-self.resistance_value / tau)
        
        # 确保电压在合理范围内
        voltage = max(min_voltage, min(max_voltage, voltage))
        
        # 查询电流（mA）
        current_mA = self.analyzer.get_current_at_voltage(voltage)
        
        # 计算电阻（Ω）- 使用欧姆定律 R = V/I
        if abs(current_mA) > 1e-10:  # 避免除零
            resistance_ohm = abs(voltage / current_mA * 1000)  # 直接得到Ω
        else:
            resistance_ohm = float('inf')
        
        # ========== 新增：检查是否有重复电压点 ==========
        # 设置容差，判断两个电压是否相等（避免浮点数精度问题）
        tolerance = 1e-6
        
        # 检查当前电压是否已经存在于测量数据中
        is_duplicate = False
        for existing_voltage in self.measure_V:
            if abs(existing_voltage - voltage) < tolerance:
                is_duplicate = True
                break
        
        if is_duplicate:
            # 如果有重复点，不记录，提示用户
            messagebox.showwarning("重复数据", 
                                f"电压值 {voltage:.3f}V 已存在，不记录重复点。\n"
                                f"请调整电阻值后再测量。")
            return
        # ==============================================
        
        # 存储测量数据（用于绘图）
        self.measure_V.append(voltage)
        self.measure_I.append(current_mA)
        
        # 按电压从小到大排序数据
        self.sort_measure_data()
        
        # 重新绘制表格
        self.refresh_table()
        
        # 更新散点图
        self.update_scatter()
        
        # 可选：打印当前值用于调试
        print(f"R={self.resistance_value:.1f}Ω -> V={voltage:.3f}V")

    def sort_measure_data(self):
        """将测量数据按电压从小到大排序"""
        if self.measure_V and self.measure_I and len(self.measure_V) == len(self.measure_I):
            # 将电压和电流配对后排序
            paired_data = list(zip(self.measure_V, self.measure_I))
            paired_data.sort(key=lambda x: x[0])  # 按电压排序
            
            # 解包排序后的数据
            self.measure_V, self.measure_I = zip(*paired_data)
            self.measure_V = list(self.measure_V)
            self.measure_I = list(self.measure_I)

    def refresh_table(self):
        """刷新表格显示"""
        # 清空表格
        self.clear_table()
        
        # 确保数据已排序
        if len(self.measure_V) > 0 and len(self.measure_V) == len(self.measure_I):
            # 重新插入所有数据
            for i, (v, i_mA) in enumerate(zip(self.measure_V, self.measure_I), 1):
                # 计算电阻（欧姆）
                if abs(i_mA) > 1e-10:  # 避免除零
                    resistance_ohm = abs(v / i_mA * 1000)  # v是电压(V)，i_mA是电流(mA)
                    resistance_str = f'{resistance_ohm:.1f}'
                else:
                    resistance_str = '∞'
                
                # 插入数据
                self.table.insert('', 'end', values=(
                    i, 
                    f'{v:.3f}', 
                    resistance_str,
                    f'{i_mA:.3f}'
                ))
        else:
            print(f"警告：数据长度不一致 V={len(self.measure_V)}, I={len(self.measure_I)}")
        
    def calculate_curve(self):
        """计算并绘制伏安特性曲线（将散点连成直线）"""
        if not self.measure_V or not self.measure_I:
            messagebox.showinfo("提示", "没有测量数据，请先测量数据")
            return
        
        if len(self.measure_V) < 2:
            messagebox.showinfo("提示", "至少需要2个测量点才能连线")
            return
        
        
        messagebox.showinfo("完成", "曲线已更新")
    
    def delete_selected_row(self):
        """删除选中的行"""
        selected = self.table.selection()
        if not selected:
            messagebox.showinfo("提示", "请先选择要删除的行")
            return
        
        print(f"删除前 - 数据点数: {len(self.measure_V)}")
        print(f"表格选中行数: {len(selected)}")
        
        # 获取所有选中的行在表格中的索引位置
        all_items = self.table.get_children()
        selected_indices = [all_items.index(item) for item in selected]
        # 从大到小排序，这样删除时不会影响前面的索引
        selected_indices.sort(reverse=True)
        
        print(f"选中的行索引: {selected_indices}")
        
        # 根据索引删除数据点
        deleted_count = 0
        for idx in selected_indices:
            if idx < len(self.measure_V):  # 确保索引有效
                # 记录删除的数据用于调试
                deleted_v = self.measure_V.pop(idx)
                deleted_i = self.measure_I.pop(idx)
                print(f"删除索引 {idx}: V={deleted_v:.3f}, I={deleted_i:.3f}")
                deleted_count += 1
        
        print(f"实际删除点数: {deleted_count}")
        print(f"删除后 - 数据点数: {len(self.measure_V)}")
        
        if deleted_count == 0:
            messagebox.showwarning("警告", "没有删除任何数据点")
            return
        
        # 重新排序数据（确保数据有序）
        self.sort_measure_data()
        
        # 刷新表格显示
        self.refresh_table()
        
        # 更新散点图
        self.update_scatter()
        
        # 清空连线（数据点变化后应该重新连线）
        self.measure_line.set_data([], [])
        
        # # 如果有至少2个点，重新连线
        # if len(self.measure_V) >= 2:
        #     self.measure_line.set_data(self.measure_V, self.measure_I)
        
        # # 重新计算拟合直线（如果存在）
        # if self.fit_lines:
        #     self.recalculate_all_fits_with_selected_points()
        
        # 强制更新画布
        self.canvas.draw_idle()
        
        messagebox.showinfo("完成", f"已删除 {deleted_count} 个数据点")
    
    def recalculate_all_fits_with_selected_points(self):
        """重新计算所有拟合直线"""
        if not self.fit_lines:
            return
        
        # 清空现有拟合直线显示
        for plot in self.fit_line_plots:
            plot.set_visible(False)
            plot.set_data([], [])
        
        # 重新计算每个拟合
        new_fit_lines = []
        
        for i, fit_info in enumerate(self.fit_lines):
            if fit_info and len(self.measure_V) >= 2:
                # 使用所有数据点重新拟合
                coefficients = np.polyfit(self.measure_V, self.measure_I, 1)
                slope = coefficients[0]
                intercept = coefficients[1]
                
                # 计算新的 R^2
                I_mean = np.mean(self.measure_I)
                I_pred = np.polyval(coefficients, self.measure_V)
                ss_res = np.sum((np.array(self.measure_I) - I_pred) ** 2)
                ss_tot = np.sum((np.array(self.measure_I) - I_mean) ** 2)
                r_squared = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0
                
                # 生成新的拟合直线
                V_fit = np.linspace(min(self.measure_V), max(self.measure_V), 100)
                I_fit = np.polyval(coefficients, V_fit)
                
                # 更新直线
                colors = ['green', 'blue', 'magenta']
                self.fit_line_plots[i].set_data(V_fit, I_fit)
                self.fit_line_plots[i].set_color(colors[i])
                self.fit_line_plots[i].set_visible(True)
                
                # 保存新的拟合信息
                new_fit_lines.append({
                    'slope': slope,
                    'intercept': intercept,
                    'r_squared': r_squared,
                    'V_min': min(self.measure_V),
                    'V_max': max(self.measure_V),
                    'name': fit_info['name']
                })
            else:
                new_fit_lines.append(None)
        
        # 更新拟合信息
        self.fit_lines = new_fit_lines
        
        # 更新公式显示
        self.update_all_formulas()
        
        # 更新图例
        self.update_legend()

    def renumber_table(self):
        """重新编号表格中的序号"""
        items = self.table.get_children()
        for i, item in enumerate(items, 1):
            values = list(self.table.item(item, 'values'))
            values[0] = str(i)
            self.table.item(item, values=values)
    
    def clear_table(self):
        """清空表格"""
        for item in self.table.get_children():
            self.table.delete(item)
    
    def clear_data(self):
        """清除所有数据"""
        # 弹出确认对话框
        result = messagebox.askyesno("确认清除", "确定要清除所有测量数据吗？")
        if result:  # 如果用户点击"是"
            self.clear_table()
            self.measure_V = []
            self.measure_I = []
            
            # 清除拟合直线
            for plot in self.fit_line_plots:
                plot.set_visible(False)
                plot.set_data([], [])
            
            # 清除拟合信息
            self.fit_lines = []
            
            # 清除拟合文本
            for text in self.fit_texts:
                text.set_visible(False)
                text.set_text('')
            
            # 更新图例
            self.update_legend()
            
            # 更新散点图和连线图
            self.update_scatter()
            messagebox.showinfo("完成", "数据已清除")
            
    def update_scatter(self):
        """更新散点图"""
        if self.measure_V and self.measure_I:
            # 确保数据长度一致
            if len(self.measure_V) == len(self.measure_I):
                points = np.column_stack([self.measure_V, self.measure_I])
                self.measure_scatter.set_offsets(points)
            else:
                print(f"警告：数据长度不一致 V={len(self.measure_V)}, I={len(self.measure_I)}")
                self.measure_scatter.set_offsets(np.empty((0, 2)))
        else:
            self.measure_scatter.set_offsets(np.empty((0, 2)))
        self.canvas.draw_idle()

    def export_data(self):
        """导出数据到文件，并将曲线图截图保存到第二页"""
        if not self.measure_V:
            messagebox.showinfo("提示", "没有数据可导出")
            return
        
        # 选择保存位置
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")],
            title="保存测量数据"
        )
        
        if not filename:
            return
        
        try:
            file_ext = os.path.splitext(filename)[1].lower()
            
            if file_ext == '.csv':
                # 如果是CSV，只保存数据
                with open(filename, 'w', encoding='gbk') as f:
                    f.write("序号,电压(V),电阻(Ω),电流(mA)\n")
                    items = self.table.get_children()
                    for item in items:
                        values = self.table.item(item, 'values')
                        f.write(f"{values[0]},{values[1]},{values[2]},{values[3]}\n")
                messagebox.showinfo("成功", f"数据已导出到：{filename}")
                
            else:
                # 如果是Excel，先创建数据DataFrame
                data = []
                items = self.table.get_children()
                for item in items:
                    values = self.table.item(item, 'values')
                    data.append([values[0], values[1], values[2], values[3]])
                
                df = pd.DataFrame(data, columns=['序号', '电压(V)', '电阻(Ω)', '电流(mA)'])
                
                # 临时保存图形为图片
                temp_img_path = os.path.join(os.path.dirname(filename), '_temp_plot.png')
                self.fig.savefig(temp_img_path, dpi=150, bbox_inches='tight')
                
                # 使用openpyxl直接创建Excel文件
                from openpyxl import Workbook
                from openpyxl.drawing.image import Image as XLImage
                from openpyxl.styles import Font
                
                # 创建新的工作簿
                wb = Workbook()
                
                # 第一页：测量数据
                ws1 = wb.active
                ws1.title = '测量数据'
                
                # 写入标题
                headers = ['序号', '电压(V)', '电阻(Ω)', '电流(mA)']
                for col, header in enumerate(headers, 1):
                    cell = ws1.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                
                # 写入数据
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws1.cell(row=row_idx, column=col_idx, value=float(value) if value != '∞' else value)
                
                # 自动调整列宽
                for col in ws1.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws1.column_dimensions[column].width = adjusted_width
                
                # 第二页：伏安特性曲线
                ws2 = wb.create_sheet(title='伏安特性曲线')
                
                # 添加标题
                title_cell = ws2.cell(row=1, column=1, value='伏安特性曲线')
                title_cell.font = Font(size=14, bold=True)
                
                # 插入图片
                img = XLImage(temp_img_path)
                img.width = 600
                img.height = 400
                ws2.add_image(img, 'A3')
                
                # 添加说明
                ws2.cell(row=30, column=1, value=f'生成时间：{pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")}')
                ws2.cell(row=31, column=1, value=f'数据点数：{len(self.measure_V)}')
                
                # 保存Excel文件
                wb.save(filename)
                
                # 删除临时图片文件
                try:
                    os.remove(temp_img_path)
                except:
                    pass
                
                messagebox.showinfo("成功", f"数据已导出到：{filename}\n第一页：测量数据\n第二页：伏安特性曲线")
                
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")
            import traceback
            traceback.print_exc()

    def import_data(self):
        """从Excel文件导入数据"""
        # 选择文件
        filename = filedialog.askopenfilename(
            title="选择要导入的数据文件",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ]
        )
        
        if not filename:
            return
        
        try:
            # 根据文件扩展名选择读取方式
            if filename.endswith('.csv'):
                df = pd.read_csv(filename, encoding='gbk')
            else:
                df = pd.read_excel(filename)
            
            # 检查必要的列是否存在
            required_columns = ['电压(V)', '电流(mA)']
            # 尝试匹配可能的列名
            voltage_col = None
            current_col = None
            
            for col in df.columns:
                col_str = str(col)
                if '电压' in col_str or 'V' in col_str:
                    voltage_col = col
                if '电流' in col_str or 'mA' in col_str or 'I' in col_str:
                    current_col = col
            
            if voltage_col is None or current_col is None:
                # 如果找不到列名，假设第一列是电压，第二列是电流
                if len(df.columns) >= 2:
                    voltage_col = df.columns[0]
                    current_col = df.columns[1]
                else:
                    messagebox.showerror("错误", "无法识别数据格式，请确保文件包含电压和电流数据")
                    return
            
            # 询问是否清除现有数据
            if self.measure_V:
                result = messagebox.askyesno("确认", "是否清除现有数据后导入？\n选择'否'则追加数据")
                if result:
                    # 清除现有数据
                    self.clear_table()
                    self.measure_V = []
                    self.measure_I = []
                    self.fit_lines = []
                    for plot in self.fit_line_plots:
                        plot.set_visible(False)
                        plot.set_data([], [])
                    for text in self.fit_texts:
                        text.set_visible(False)
                        text.set_text('')
            else:
                result_append = True
            
            # 读取数据
            voltages = []
            currents = []
            
            for idx, row in df.iterrows():
                try:
                    v = float(row[voltage_col])
                    i = float(row[current_col])
                    voltages.append(v)
                    currents.append(i)
                except (ValueError, TypeError):
                    continue  # 跳过无法转换的行
            
            if not voltages:
                messagebox.showerror("错误", "没有有效的数值数据")
                return
            
            # 添加到现有数据
            self.measure_V.extend(voltages)
            self.measure_I.extend(currents)
            
            # 按电压排序
            self.sort_measure_data()
            
            # 刷新表格显示
            self.refresh_table()
            
            # 更新散点图
            self.update_scatter()
            
            # 更新图例
            self.update_legend()
            
            messagebox.showinfo("成功", f"成功导入 {len(voltages)} 条数据")
            
        except Exception as e:
            messagebox.showerror("错误", f"导入失败：{str(e)}")
            import traceback
            traceback.print_exc()

    def fit_line1(self):
        """拟合直线1"""
        self.fit_line(0, "拟合直线1")

    def fit_line2(self):
        """拟合直线2"""
        self.fit_line(1, "拟合直线2")

    def fit_line3(self):
        """拟合直线3"""
        self.fit_line(2, "拟合直线3")

    def fit_line(self, line_index, line_name):
        """拟合直线通用方法"""
        selected = self.table.selection()
        if not selected:
            messagebox.showinfo("提示", f"请先选择要用于{line_name}的数据点")
            return
        
        if len(selected) < 2:
            messagebox.showinfo("提示", f"{line_name}至少需要2个数据点")
            return
        
        # 获取选中的数据点
        V_points = []
        I_points = []
        
        for item in selected:
            values = self.table.item(item, 'values')
            if values:
                V_points.append(float(values[1]))  # 电压
                I_points.append(float(values[3]))  # 电流
        
        # 数据已经全局排序，但选中的点可能不连续，仍然需要排序
        sorted_pairs = sorted(zip(V_points, I_points))
        V_sorted, I_sorted = zip(*sorted_pairs)
        
        # 线性拟合
        coefficients = np.polyfit(V_sorted, I_sorted, 1)
        slope = coefficients[0]  # 斜率
        intercept = coefficients[1]  # 截距
        
        # 计算相关系数 R^2
        I_mean = np.mean(I_sorted)
        I_pred = np.polyval(coefficients, V_sorted)
        ss_res = np.sum((np.array(I_sorted) - I_pred) ** 2)
        ss_tot = np.sum((np.array(I_sorted) - I_mean) ** 2)
        r_squared = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0
        
        # 生成拟合直线的点（用于绘图）
        V_min = min(V_sorted)
        V_max = max(V_sorted)
        # 稍微扩展一点范围，让直线看起来更长
        V_range = V_max - V_min
        V_fit = np.linspace(V_min - V_range*0.1, V_max + V_range*0.1, 100)
        I_fit = np.polyval(coefficients, V_fit)
        
        # 更新对应的拟合直线 - 确保颜色正确
        self.fit_line_plots[line_index].set_data(V_fit, I_fit)
        self.fit_line_plots[line_index].set_visible(True)
        
        # 存储拟合信息 - 确保 fit_lines 列表长度足够
        while len(self.fit_lines) <= line_index:
            self.fit_lines.append(None)
        
        # 更新已有的拟合信息
        self.fit_lines[line_index] = {
            'slope': slope,
            'intercept': intercept,
            'r_squared': r_squared,
            'V_min': V_min,
            'V_max': V_max,
            'name': line_name
        }
        
        # 更新所有公式显示
        self.update_all_formulas()
        
        # 更新图例
        self.update_legend()
        
        self.canvas.draw_idle()
        
        # 显示消息框
        formula = f"{line_name}: I = {slope:.4f}V + {intercept:.4f}\nR^2 = {r_squared:.4f}"
        messagebox.showinfo("拟合完成", formula)

    def update_all_formulas(self):
        """更新所有拟合直线的公式显示"""
        colors = ['green', 'blue', 'magenta']
        
        # 遍历三条直线（索引0、1、2分别对应直线1、2、3）
        for i in range(3):
            # 检查是否有该直线的拟合信息
            if i < len(self.fit_lines) and self.fit_lines[i]:
                fit_info = self.fit_lines[i]
                # 生成公式文本
                formula = (f"{fit_info['name']}:\n"
                        f"I = {fit_info['slope']:.4f}V + {fit_info['intercept']:.4f}\n"
                        f"R^2 = {fit_info['r_squared']:.4f}")
                
                # 更新对应的文本
                self.fit_texts[i].set_text(formula)
                self.fit_texts[i].set_visible(True)
                
                # 更新文本框颜色
                self.fit_texts[i].set_color(colors[i])
                self.fit_texts[i].get_bbox_patch().set_edgecolor(colors[i])
                self.fit_texts[i].get_bbox_patch().set_alpha(0.7)
            else:
                # 如果没有拟合信息，隐藏文本
                self.fit_texts[i].set_visible(False)
                self.fit_texts[i].set_text('')
        
        # 强制更新画布
        self.canvas.draw_idle()

    def update_legend(self):
        """更新图例，只显示可见的线条"""
        # 清除旧的图例
        if self.ax.legend_:
            self.ax.legend_.remove()
        
        # 收集所有可见的线条
        handles = [self.measure_scatter, self.measure_line]
        
        # 添加可见的拟合直线
        colors = ['green', 'blue', 'magenta']
        labels = ['拟合直线1', '拟合直线2', '拟合直线3']
        for i, plot in enumerate(self.fit_line_plots):
            if plot.get_visible():
                # 确保线条有正确的颜色和标签
                plot.set_color(colors[i])
                plot.set_label(labels[i])
                handles.append(plot)
        
        # 创建新的图例
        if len(handles) > 2:  # 至少有一个拟合直线可见
            self.ax.legend(handles=handles, loc='best')

    def recalculate_all_fits(self):
        """重新计算所有拟合直线"""
        if not self.fit_lines:
            return
        
        # 清空现有拟合
        for plot in self.fit_line_plots:
            plot.set_visible(False)
            plot.set_data([], [])
        
        # 重新计算每个拟合
        new_fit_lines = []
        
        for i, fit_info in enumerate(self.fit_lines):
            if fit_info:
                # 获取该拟合原来使用的数据点（这里简化处理，使用所有数据点）
                # 实际应用中可能需要保存每个拟合使用的数据点索引
                if len(self.measure_V) >= 2:
                    # 使用所有数据点重新拟合
                    coefficients = np.polyfit(self.measure_V, self.measure_I, 1)
                    slope = coefficients[0]
                    intercept = coefficients[1]
                    
                    # 计算新的 R^2
                    I_mean = np.mean(self.measure_I)
                    I_pred = np.polyval(coefficients, self.measure_V)
                    ss_res = np.sum((np.array(self.measure_I) - I_pred) ** 2)
                    ss_tot = np.sum((np.array(self.measure_I) - I_mean) ** 2)
                    r_squared = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0
                    
                    # 生成新的拟合直线
                    V_fit = np.linspace(min(self.measure_V), max(self.measure_V), 100)
                    I_fit = np.polyval(coefficients, V_fit)
                    
                    # 更新直线 - 确保颜色正确
                    colors = ['green', 'blue', 'magenta']
                    self.fit_line_plots[i].set_data(V_fit, I_fit)
                    self.fit_line_plots[i].set_color(colors[i])
                    self.fit_line_plots[i].set_visible(True)
                    
                    # 保存新的拟合信息
                    new_fit_lines.append({
                        'slope': slope,
                        'intercept': intercept,
                        'r_squared': r_squared,
                        'V_min': min(self.measure_V),
                        'V_max': max(self.measure_V),
                        'name': fit_info['name']
                    })
                else:
                    new_fit_lines.append(None)
            else:
                new_fit_lines.append(None)
        
        # 更新拟合信息
        self.fit_lines = new_fit_lines
        
        # 更新公式显示
        self.update_all_formulas()
        
        # 更新图例
        self.update_legend()
        
        self.canvas.draw_idle()
        
        messagebox.showinfo("完成", "所有拟合直线已重新计算")

def main():
    circuit = ChaosCircuit()
    app = ChaosVisualizerTk(circuit)
    app.run()

if __name__ == "__main__":
    main()