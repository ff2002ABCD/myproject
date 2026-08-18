import sys
import os

import numpy as np
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from scipy import stats
from PIL import Image, ImageTk  # 添加PIL库支持
import matplotlib.pyplot as plt
import openpyxl

# 设置matplotlib中文字体
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'DejaVu Sans']  # 用来正常显示中文标签
plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号


# ---------------------------- 资源路径处理 ----------------------------
def get_resource_path(relative_path):
    """获取资源的绝对路径，支持打包后的环境"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# ---------------------------- 主应用程序 ----------------------------
class TempSensorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("温度传感器特性测量实验")
        self.root.geometry("1280x800")
        self.root.configure(bg='#f0f0f0')

        # 长按相关变量
        self.long_press_id = None
        self.long_press_delay = 300  # 首次延迟300ms
        self.long_press_interval = 100  # 之后每100ms触发一次
    

        # 全局数据存储 - 为每个实验独立存储
        self.table_data = {}
        self.init_table_data()
        
        # 曲线拟合相关变量 - 每个实验独立
        self.slope = {}
        self.r_value = {}
        self.intercept = {}
        
        # 当前选项卡
        self.current_tab = "Pt100恒电流法"
        
        # 温度渐变相关变量 - 每个实验独立
        self.current_temp_value = {}
        self.target_temp_value = {}
        self.temp_animation_id = {}
        
        # 拟合显示状态 - 每个实验独立
        self.fit_show_state = {}
        
        # 电阻箱值 - 每个实验独立
        self.resistance_box = {}
        self.resistance_box_digits = {}

        # 初始化各实验的数据
        self.init_experiment_data()
        
        # 创建界面各部分
        self.create_top_tab_bar()
        self.create_left_top_frame()
        self.create_left_bottom_frame()
        self.create_right_frame()
        
        # 默认显示Pt100恒电流法界面
        self.update_for_pt100_current()
        
        # 初始绘制散点图
        self.update_plot(show_fit=False)
    
    def init_experiment_data(self):
        """初始化所有实验的独立数据"""
        experiments = ["Pt100恒电流法", "Pt100直流电桥法", "NTC热敏电阻恒电流法", 
                    "NTC热敏电阻直流电桥法", "PN结温度传感器", "电压型集成温度传感器",
                    "电流型集成温度传感器"]
        
        for exp in experiments:
            # 初始化表格数据 - 9行
            self.table_data[exp] = []
            if "恒电流法" in exp and ("Pt100" in exp or "NTC热敏电阻" in exp):
                # 恒电流法（Pt100和NTC）：4列数据（序号、温度、U/mV、Rt/Ω）
                for i in range(1, 10):
                    self.table_data[exp].append([i, "", "", ""])
            elif "电流型" in exp:
                # 电流型：4列数据（序号、温度、U/V、I/uA）
                for i in range(1, 10):
                    self.table_data[exp].append([i, "", "", ""])
            else:
                # 其他（包括电桥法）：3列数据（序号、温度、值）
                for i in range(1, 10):
                    self.table_data[exp].append([i, "", ""])

            # 初始化电阻箱值（仅电桥法实验）
            if "电桥法" in exp:
                self.resistance_box[exp] = 1000.0  # 默认1000Ω
                self.resistance_box_digits[exp] = [0, 1, 0, 0, 0]  # 万=0, 千=1, 百=0, 十=0, 个=0
            
            # 初始化拟合变量
            self.slope[exp] = 0.0
            self.r_value[exp] = 0.0
            self.intercept[exp] = 0.0
            
            # 初始化温度变量
            self.current_temp_value[exp] = 30.0
            self.target_temp_value[exp] = 30.0
            self.temp_animation_id[exp] = None
            
            # 初始化拟合显示状态
            self.fit_show_state[exp] = False

    def update_resistance_digits(self, exp_name):
        """根据电阻值更新各位数字"""
        value = self.resistance_box.get(exp_name, 1000.0)
        # 分解为各位数字 (万, 千, 百, 十, 个)
        # 值范围 0-9999.9
        int_part = int(value)
        decimal_part = int(round((value - int_part) * 10))
        
        # 万位 (0-9)
        digit_wan = int_part // 10000
        # 千位
        digit_qian = (int_part // 1000) % 10
        # 百位
        digit_bai = (int_part // 100) % 10
        # 十位
        digit_shi = (int_part // 10) % 10
        # 个位
        digit_ge = int_part % 10
        
        self.resistance_box_digits[exp_name] = [digit_wan, digit_qian, digit_bai, digit_shi, digit_ge]

    

    def init_table_data(self):
        pass
    
    def create_top_tab_bar(self):
        top_frame = tk.Frame(self.root, bg='#d9d9d9', height=50)
        top_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        
        tabs = [
            "Pt100恒电流法", "Pt100直流电桥法", "NTC热敏电阻恒电流法",
            "NTC热敏电阻直流电桥法", "PN结温度传感器", "电流型集成温度传感器",
            "电压型集成温度传感器"
        ]
        
        self.tab_buttons = []
        for tab_name in tabs:
            btn = tk.Button(top_frame, text=tab_name, width=22, height=1,
                            command=lambda t=tab_name: self.on_tab_selected(t))
            btn.pack(side=tk.LEFT, padx=2, pady=5)
            self.tab_buttons.append(btn)
        
        self.highlight_tab(tabs[0])
    
    def on_tab_selected(self, tab_name):
        # 先保存当前实验的数据
        self.save_current_data()
        
        self.current_tab = tab_name
        self.highlight_tab(tab_name)
        
        if tab_name == "Pt100恒电流法":
            self.update_for_pt100_current()
        elif tab_name == "Pt100直流电桥法":
            self.update_for_pt100_bridge()
        elif tab_name == "NTC热敏电阻恒电流法":
            self.update_for_ntc_current()
        elif tab_name == "NTC热敏电阻直流电桥法":
            self.update_for_ntc_bridge()
        elif tab_name == "PN结温度传感器":
            self.update_for_pn_junction()
        elif tab_name == "电压型集成温度传感器":
            self.update_for_voltage_output()
        elif tab_name == "电流型集成温度传感器":
            self.update_for_current_output()
        else:
            self.show_other_tab_message(tab_name)
        
        # 显示或隐藏电阻箱（仅电桥法实验显示）
        if "电桥法" in tab_name:
            self.show_resistance_box(True)
        else:
            self.show_resistance_box(False)
        
        # 加载对应实验的数据并刷新显示
        self.load_experiment_data(tab_name)
        
        # 强制刷新表格显示（会清空并重新加载数据）
        self.refresh_table_display()

    def update_for_current_output(self):
        """电流型集成温度传感器时的左上区域布局"""
        self.update_device_display("current_output")

    def update_for_voltage_output(self):
        """电压型集成温度传感器时的左上区域布局"""
        self.update_device_display("voltage_output")

    def update_for_pn_junction(self):
        """PN结温度传感器时的左上区域布局"""
        self.update_device_display("pn_junction")
        
    def update_for_ntc_current(self):
        """NTC热敏电阻恒电流法时的左上区域布局"""
        self.update_device_display("ntc_current")

    def update_for_ntc_bridge(self):
        """NTC热敏电阻直流电桥法时的左上区域布局"""
        self.update_device_display("ntc_bridge")

    def save_current_data(self):
        """保存当前实验的数据"""
        exp_name = self.current_tab
        if exp_name in self.table_data:
            # 直接从表格控件读取
            table_rows = []
            is_current = "电流型集成温度传感器" in exp_name
            is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
            
            for row in self.table.get_children():
                values = self.table.item(row, 'values')
                if values:
                    if is_current:
                        # 电流型：4列数据
                        temp_val = values[1] if values[1] else ""
                        voltage_val = values[2] if values[2] else ""
                        current_val = values[3] if values[3] else ""
                        table_rows.append([int(values[0]), temp_val, voltage_val, current_val])
                    elif is_constant_current:
                        # 恒电流法：4列数据（序号、温度、U/mV、Rt/Ω）
                        temp_val = values[1] if values[1] else ""
                        u_val = values[2] if values[2] else ""
                        rt_val = values[3] if values[3] else ""
                        table_rows.append([int(values[0]), temp_val, u_val, rt_val])
                    else:
                        # 其他：3列数据
                        temp_val = values[1] if values[1] else ""
                        res_val = values[2] if values[2] else ""
                        table_rows.append([int(values[0]), temp_val, res_val])
            
            # 如果有数据且长度正确，更新存储
            expected_len = 4 if (is_current or is_constant_current) else 3
            if table_rows and len(table_rows) == len(self.table_data[exp_name]):
                self.table_data[exp_name] = table_rows
            elif table_rows:
                # 如果长度不一致，需要调整
                while len(table_rows) < 9:
                    if is_current or is_constant_current:
                        table_rows.append([len(table_rows) + 1, "", "", ""])
                    else:
                        table_rows.append([len(table_rows) + 1, "", ""])
                self.table_data[exp_name] = table_rows
            
            # 保存温度值
            try:
                self.current_temp_value[exp_name] = float(self.current_temp_var.get())
            except:
                pass

    def load_experiment_data(self, exp_name):
        """加载指定实验的数据"""
        if exp_name in self.table_data:
            # 确保数据完整 - 有9行
            is_current = "电流型集成温度传感器" in exp_name
            is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
            is_electrical_bridge = "电桥法" in exp_name and not is_constant_current
            
            # 根据实验类型确定列数
            if is_current or is_constant_current:
                cols = 4
            else:
                cols = 3
            
            # 确保数据行数正确，每行列数正确
            while len(self.table_data[exp_name]) < 9:
                if cols == 4:
                    self.table_data[exp_name].append([len(self.table_data[exp_name]) + 1, "", "", ""])
                else:
                    self.table_data[exp_name].append([len(self.table_data[exp_name]) + 1, "", ""])
            
            # 确保每行数据列数正确
            for i, row in enumerate(self.table_data[exp_name]):
                while len(row) < cols:
                    row.append("")
            
            # 刷新表格显示
            self.refresh_table_display()
            
            # 更新拟合结果显示
            slope_val = self.slope.get(exp_name, 0.0)
            intercept_val = self.intercept.get(exp_name, 0.0)
            r2_val = self.r_value.get(exp_name, 0.0)
            is_ntc = "NTC" in exp_name
            is_pn = "PN结" in exp_name
            is_voltage = "电压型集成温度传感器" in exp_name
            is_current = "电流型集成温度传感器" in exp_name
            
            if is_ntc:
                if intercept_val != 0.0:
                    formula_text = f"拟合公式: Rt = {intercept_val:.4f} × e^(-0.03×t)  (R² = {r2_val:.6f})"
                    self.formula_label.config(text=formula_text)
                    show_fit = self.fit_show_state.get(exp_name, False)
                    self.update_plot(show_fit=show_fit)
                else:
                    self.formula_label.config(text="拟合公式: 未计算")
                    self.fit_show_state[exp_name] = False
                    self.update_plot(show_fit=False)
            elif is_pn:
                if slope_val != 0.0:
                    formula_text = f"拟合公式: Ube = {slope_val:.4f} × t + {intercept_val:.4f}  (R² = {r2_val:.6f})"
                    self.formula_label.config(text=formula_text)
                    show_fit = self.fit_show_state.get(exp_name, False)
                    self.update_plot(show_fit=show_fit)
                else:
                    self.formula_label.config(text="拟合公式: 未计算")
                    self.fit_show_state[exp_name] = False
                    self.update_plot(show_fit=False)
            elif is_voltage:
                if slope_val != 0.0:
                    formula_text = f"拟合公式: U0 = {slope_val:.4f} × t + {intercept_val:.4f}  (R² = {r2_val:.6f})"
                    self.formula_label.config(text=formula_text)
                    show_fit = self.fit_show_state.get(exp_name, False)
                    self.update_plot(show_fit=show_fit)
                else:
                    self.formula_label.config(text="拟合公式: 未计算")
                    self.fit_show_state[exp_name] = False
                    self.update_plot(show_fit=False)
            elif is_current:
                if slope_val != 0.0:
                    formula_text = f"拟合公式: I = {slope_val:.4f} × t + {intercept_val:.4f}  (R² = {r2_val:.6f})"
                    self.formula_label.config(text=formula_text)
                    show_fit = self.fit_show_state.get(exp_name, False)
                    self.update_plot(show_fit=show_fit)
                else:
                    self.formula_label.config(text="拟合公式: 未计算")
                    self.fit_show_state[exp_name] = False
                    self.update_plot(show_fit=False)
            else:
                if slope_val != 0.0:
                    formula_text = f"拟合公式: Rt = {slope_val:.4f} × t + {intercept_val:.4f}  (R² = {r2_val:.6f})"
                    self.formula_label.config(text=formula_text)
                    show_fit = self.fit_show_state.get(exp_name, False)
                    self.update_plot(show_fit=show_fit)
                else:
                    self.formula_label.config(text="拟合公式: 未计算")
                    self.fit_show_state[exp_name] = False
                    self.update_plot(show_fit=False)
            
            # 更新温度显示
            if exp_name in self.current_temp_value:
                self.current_temp_var.set(f"{self.current_temp_value[exp_name]:.1f}")
                self.temp_progress.set(self.current_temp_value[exp_name])
                self.temp_value_label.config(text=f"{self.current_temp_value[exp_name]:.1f} ℃")
                self.set_temp_var.set(f"{self.current_temp_value[exp_name]:.1f}")
                self.update_voltage_display()

            # 如果是电桥法实验，恢复电阻箱值
            if "电桥法" in exp_name and exp_name in self.resistance_box:
                self.update_resistance_digits(exp_name)
                self.update_resistance_display(exp_name)
                self.update_voltage_display()

    def highlight_tab(self, selected_tab):
        for btn in self.tab_buttons:
            if btn.cget('text') == selected_tab:
                btn.config(bg='#4caf50', fg='white')
            else:
                btn.config(bg='#f0f0f0', fg='black')
    
    def show_other_tab_message(self, tab_name):
        for widget in self.left_top_inner_frame.winfo_children():
            widget.destroy()
        
        info_frame = tk.Frame(self.left_top_inner_frame, bg='#e0e0e0', relief=tk.RAISED, bd=2)
        info_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        label = tk.Label(info_frame, text=f"{tab_name}\n\n功能待定", 
                         font=("Arial", 16, "bold"), bg='#e0e0e0', fg='#666')
        label.pack(expand=True)
    
    def create_left_top_frame(self):
        self.left_top_frame = tk.LabelFrame(self.root, text="实验装置区域", font=("Arial", 12, "bold"),
                                            bg='#f0f0f0', padx=5, pady=5)
        self.left_top_frame.place(x=10, y=60, width=550, height=420)
        
        self.left_top_inner_frame = tk.Frame(self.left_top_frame, bg='#f0f0f0')
        self.left_top_inner_frame.pack(fill=tk.BOTH, expand=True)
    
    def update_for_pt100_current(self):
        """Pt100恒电流法时的左上区域布局：固定大小图片+三个显示框"""
        self.update_device_display("恒电流")

    def update_for_pt100_bridge(self):
        """Pt100直流电桥法时的左上区域布局"""
        self.update_device_display("电桥")
    
    def update_device_display(self, mode):
        """更新实验装置显示区域
        
        Args:
            mode: "恒电流" 或 "电桥" 或 "ntc_current" 或 "ntc_bridge" 或 "pn_junction" 或 "voltage_output" 或 "current_output"
        """
        for widget in self.left_top_inner_frame.winfo_children():
            widget.destroy()
        
        # 创建固定大小的画布
        self.device_canvas = tk.Canvas(self.left_top_inner_frame, highlightthickness=0, width=540, height=400, bg='#e0e0e0')
        self.device_canvas.pack(expand=True)
        
        # 根据模式选择背景图和参数
        if mode == "恒电流":
            img_name = "恒电流.jpg"
            self.r0 = 113.7
            self.temp_coeff = 0.398
            self.current_mode = "恒电流"
            self.sensor_type = "pt100"
        elif mode == "电桥":
            img_name = "电桥.jpg"
            self.r0 = 114.0
            self.temp_coeff = 0.391
            self.current_mode = "电桥"  # 标记为电桥模式
            self.sensor_type = "pt100"
        elif mode == "ntc_current":
            img_name = "恒电流.jpg"
            self.current_mode = "恒电流"
            self.sensor_type = "ntc"
        elif mode == "ntc_bridge":
            img_name = "电桥.jpg"
            self.current_mode = "电桥"  # 标记为电桥模式
            self.sensor_type = "ntc"
        elif mode == "pn_junction":
            img_name = "PN结.jpg"
            self.sensor_type = "pn"
            self.current_mode = "pn"
        elif mode == "voltage_output":
            img_name = "LM35.jpg"
            self.sensor_type = "voltage"
            self.current_mode = "voltage"
        else:  # current_output
            img_name = "AD590.jpg"
            self.sensor_type = "current"
            self.current_mode = "current"
        
        # 加载并固定大小图片
        img_path = get_resource_path(f"background/{img_name}")
        try:
            pil_image = Image.open(img_path)
            resized = pil_image.resize((540, 380), Image.Resampling.LANCZOS)
            self.bg_img = ImageTk.PhotoImage(resized)
            self.device_canvas.create_image(0, 0, anchor=tk.NW, image=self.bg_img, tags="bg_image")
        except Exception as e:
            print(f"图片加载失败: {e}")
            self.device_canvas.create_rectangle(0, 0, 540, 400, fill='#e0e0e0', tags="bg_image")
            self.device_canvas.create_text(270, 200, text=f"{mode}实验装置图", font=("Arial", 16), tags="bg_image")
        
        # 获取当前实验的数据
        exp_name = self.current_tab
        current_temp = self.current_temp_value.get(exp_name, 30.0)
        
        # 计算初始电压
        if self.sensor_type == "pt100":
            voltage = 0.114
        elif self.sensor_type == "ntc":
            resistance = 2060 * np.exp(-0.03 * current_temp)
            voltage = 0.001 * resistance
        elif self.sensor_type == "pn":
            # Ube = -2.157 * T + 669.1 (mV)
            voltage = -2.157 * current_temp + 669.1  # mV
            voltage = voltage / 1000  # 转换为V
        elif self.sensor_type == "voltage":
            # U0 = 9.92 * T + 29.85 (mV)
            voltage = 9.92 * current_temp + 29.85  # mV
            voltage = voltage / 1000  # 转换为V
        else:  # current
            # U = (1.031 * T + 271.0) / 1000 (V)
            voltage = (1.031 * current_temp + 271.0) / 1000
        
        # 电压显示（初始值显示4位小数）
        self.voltage_var = tk.StringVar(value=f"{voltage:.4f}")
        voltage_label = tk.Label(self.device_canvas, textvariable=self.voltage_var, width=9,  # 宽度加1
                                font=("Arial", 9), justify='center', bg='white', 
                                bd=2)
        self.device_canvas.create_window(87, 85, anchor=tk.NW, window=voltage_label)

        # 当前温度
        self.current_temp_var = tk.StringVar(value=f"{current_temp:.1f}")
        current_label = tk.Label(self.device_canvas, textvariable=self.current_temp_var, width=8,
                                font=("Arial", 9), justify='center', bg='white',
                                bd=2)
        self.device_canvas.create_window(367, 90, anchor=tk.NW, window=current_label)

        # 设定温度
        target_temp = self.target_temp_value.get(exp_name, 30.0)
        self.set_temp_var = tk.StringVar(value=f"{target_temp:.1f}")
        set_label = tk.Label(self.device_canvas, textvariable=self.set_temp_var, width=5,
                            font=("Arial", 9), justify='center', bg='white',
                            bd=2)
        self.device_canvas.create_window(368, 120, anchor=tk.NW, window=set_label)
        
        # 恢复电压表挡位指示灯位置（根据当前挡位）
        current_range = self.voltage_range_var.get()
        if current_range == "2V":
            self.indicator = self.device_canvas.create_oval(191+2, 100+1, 199+2, 108+1, fill='red', outline='red')
        else:
            self.indicator = self.device_canvas.create_oval(191+2, 84, 199+2, 92, fill='red', outline='red')
        
        # 更新电压显示
        self.update_voltage_display()

    def create_left_bottom_frame(self):
        self.left_bottom_frame = tk.LabelFrame(self.root, text="实验操作区域", font=("Arial", 12, "bold"),
                                            bg='#f0f0f0', padx=5, pady=5)
        self.left_bottom_frame.place(x=10, y=520, width=550, height=240)
        
        # 进度条 设定温度
        tk.Label(self.left_bottom_frame, text="设定温度调节:", font=("Arial", 10), bg='#f0f0f0').grid(row=0, column=0, sticky='w', pady=5, padx=5)
        exp_name = self.current_tab
        initial_temp = self.current_temp_value.get(exp_name, 30.0)
        self.temp_progress = ttk.Scale(self.left_bottom_frame, from_=30, to=100, orient=tk.HORIZONTAL, 
                                    length=150, value=initial_temp)
        self.temp_progress.grid(row=0, column=1, columnspan=2, pady=5, padx=5)
        self.temp_value_label = tk.Label(self.left_bottom_frame, text="30.0 ℃", width=8, bg='white', relief=tk.SUNKEN)
        self.temp_value_label.grid(row=0, column=3, padx=5)
        
        # 微调按钮 - 支持长按
        self.long_press_id = None
        
        def start_long_press(delta):
            adjust_temp(delta)
            self.long_press_id = self.root.after(300, lambda: repeat_long_press(delta))
        
        def repeat_long_press(delta):
            adjust_temp(delta)
            self.long_press_id = self.root.after(100, lambda: repeat_long_press(delta))
        
        def stop_long_press():
            if self.long_press_id is not None:
                self.root.after_cancel(self.long_press_id)
                self.long_press_id = None
        
        def adjust_temp(delta):
            val = self.temp_progress.get() + delta
            val = max(30, min(100, val))
            self.temp_progress.set(val)
            self.temp_value_label.config(text=f"{val:.1f} ℃")
            self.set_temp_var.set(f"{val:.1f}")
        
        btn_minus = tk.Button(self.left_bottom_frame, text="-", width=2)
        btn_minus.grid(row=0, column=4, padx=2)
        btn_minus.bind('<ButtonPress-1>', lambda e: start_long_press(-0.1))
        btn_minus.bind('<ButtonRelease-1>', lambda e: stop_long_press())
        btn_minus.bind('<Leave>', lambda e: stop_long_press())
        
        btn_plus = tk.Button(self.left_bottom_frame, text="+", width=2)
        btn_plus.grid(row=0, column=5, padx=2)
        btn_plus.bind('<ButtonPress-1>', lambda e: start_long_press(0.1))
        btn_plus.bind('<ButtonRelease-1>', lambda e: stop_long_press())
        btn_plus.bind('<Leave>', lambda e: stop_long_press())
        
        btn_confirm = tk.Button(self.left_bottom_frame, text="确定", command=self.confirm_set_temp, bg='#4caf50', fg='white', width=6)
        btn_confirm.grid(row=0, column=6, padx=5)
        
        def update_temp_label(*args):
            val = self.temp_progress.get()
            self.temp_value_label.config(text=f"{val:.1f} ℃")
            self.set_temp_var.set(f"{val:.1f}")
        self.temp_progress.configure(command=lambda x: update_temp_label())
        
        # === 电阻箱区域（仅电桥法实验显示）===
        self.resistance_frame = tk.Frame(self.left_bottom_frame, bg='#f0f0f0')
        self.resistance_frame.grid(row=1, column=0, columnspan=7, pady=5, sticky='w')
        
        tk.Label(self.resistance_frame, text="电阻箱:", font=("Arial", 10), bg='#f0f0f0').pack(side=tk.LEFT, padx=5)
        
        # 创建5位数字的+-按钮
        self.digit_buttons = []
        self.digit_labels = []
        digit_names = ["10000Ω", "1000Ω", "100Ω", "10Ω", "1Ω"]
        
        for i in range(5):
            digit_frame = tk.Frame(self.resistance_frame, bg='#f0f0f0')
            digit_frame.pack(side=tk.LEFT, padx=2)
            
            # 加按钮
            btn_up = tk.Button(digit_frame, text="+", width=2, height=1,
                            command=lambda idx=i: self.adjust_resistance_digit(idx, 1))
            btn_up.pack()
            
            # 数字显示
            lbl = tk.Label(digit_frame, text="0", width=3, height=1, 
                        font=("Arial", 12, "bold"), bg='white', relief=tk.RAISED)
            lbl.pack()
            self.digit_labels.append(lbl)
            
            # 减按钮
            btn_down = tk.Button(digit_frame, text="-", width=2, height=1,
                                command=lambda idx=i: self.adjust_resistance_digit(idx, -1))
            btn_down.pack()
            
            # 位数标签
            tk.Label(digit_frame, text=digit_names[i], font=("Arial", 8), bg='#f0f0f0').pack()
        
        # 小数点和小数位显示
        tk.Label(self.resistance_frame, text=".", font=("Arial", 14, "bold"), bg='#f0f0f0').pack(side=tk.LEFT, padx=1)
        
        # 小数位显示
        decimal_frame = tk.Frame(self.resistance_frame, bg='#f0f0f0')
        decimal_frame.pack(side=tk.LEFT, padx=2)
        
        btn_decimal_up = tk.Button(decimal_frame, text="+", width=2, height=1,
                                command=lambda: self.adjust_resistance_decimal(1))
        btn_decimal_up.pack()
        
        self.decimal_label = tk.Label(decimal_frame, text="0", width=3, height=1,
                                    font=("Arial", 12, "bold"), bg='white', relief=tk.RAISED)
        self.decimal_label.pack()
        
        btn_decimal_down = tk.Button(decimal_frame, text="-", width=2, height=1,
                                    command=lambda: self.adjust_resistance_decimal(-1))
        btn_decimal_down.pack()
        
        tk.Label(decimal_frame, text="0.1Ω", font=("Arial", 8), bg='#f0f0f0').pack()
        
        # 显示当前电阻值
        self.resistance_display_label = tk.Label(self.resistance_frame, text="1000.0 Ω", 
                                                font=("Arial", 10, "bold"), bg='#f0f0f0', fg='blue')
        self.resistance_display_label.pack(side=tk.LEFT, padx=10)
        
        # 初始状态：隐藏电阻箱（使用 grid_remove）
        self.resistance_frame.grid_remove()
        
        # 电压表挡位
        tk.Label(self.left_bottom_frame, text="电压表挡位:", font=("Arial", 10), bg='#f0f0f0').grid(row=2, column=0, sticky='w', pady=5, padx=5)
        self.voltage_range_var = tk.StringVar(value="2V")
        rb_2v = tk.Radiobutton(self.left_bottom_frame, text="2V", variable=self.voltage_range_var, value="2V", bg='#f0f0f0', command=self.on_voltage_range_change)
        rb_20v = tk.Radiobutton(self.left_bottom_frame, text="20V", variable=self.voltage_range_var, value="20V", bg='#f0f0f0', command=self.on_voltage_range_change)
        rb_2v.grid(row=2, column=1, sticky='w')
        rb_20v.grid(row=2, column=2, sticky='w')
        
        # 散热风扇开关
        tk.Label(self.left_bottom_frame, text="散热风扇开关:", font=("Arial", 10), bg='#f0f0f0').grid(row=3, column=0, sticky='w', pady=5, padx=5)
        self.fan_var = tk.StringVar(value="关")
        rb_fan_on = tk.Radiobutton(self.left_bottom_frame, text="开", variable=self.fan_var, value="开", bg='#f0f0f0', command=self.on_fan_switch)
        rb_fan_off = tk.Radiobutton(self.left_bottom_frame, text="关", variable=self.fan_var, value="关", bg='#f0f0f0', command=self.on_fan_switch)
        rb_fan_on.grid(row=3, column=1, sticky='w')
        rb_fan_off.grid(row=3, column=2, sticky='w')
    
    def adjust_resistance_digit(self, digit_index, delta):
        """调整电阻箱某一位数字（循环：0-9，9+变为0，0-变为9）"""
        exp_name = self.current_tab
        if exp_name not in self.resistance_box_digits:
            return
        
        digits = self.resistance_box_digits[exp_name]
        
        # 循环加减
        if delta > 0:
            # 加：9变为0
            if digits[digit_index] == 9:
                digits[digit_index] = 0
            else:
                digits[digit_index] += 1
        else:
            # 减：0变为9
            if digits[digit_index] == 0:
                digits[digit_index] = 9
            else:
                digits[digit_index] -= 1
        
        # 获取当前小数位
        try:
            decimal_part = int(self.decimal_label.cget('text'))
        except:
            decimal_part = 0
        
        # 计算完整电阻值
        int_value = digits[0] * 10000 + digits[1] * 1000 + digits[2] * 100 + digits[3] * 10 + digits[4]
        self.resistance_box[exp_name] = float(int_value) + decimal_part / 10.0
        
        self.update_resistance_display(exp_name)
        self.update_voltage_display()

    def adjust_resistance_decimal(self, delta):
        """调整电阻箱小数位（循环：0-9，9+变为0，0-变为9）"""
        exp_name = self.current_tab
        if exp_name not in self.resistance_box:
            return
        
        # 获取当前小数位
        try:
            current_decimal = int(self.decimal_label.cget('text'))
        except:
            current_decimal = 0
        
        # 循环加减
        if delta > 0:
            # 加：9变为0
            if current_decimal == 9:
                new_decimal = 0
            else:
                new_decimal = current_decimal + 1
        else:
            # 减：0变为9
            if current_decimal == 0:
                new_decimal = 9
            else:
                new_decimal = current_decimal - 1
        
        # 更新小数位显示
        self.decimal_label.config(text=str(new_decimal))
        
        # 获取当前整数部分
        digits = self.resistance_box_digits.get(exp_name, [0, 0, 0, 0, 0])
        int_value = digits[0] * 10000 + digits[1] * 1000 + digits[2] * 100 + digits[3] * 10 + digits[4]
        
        # 更新电阻值
        self.resistance_box[exp_name] = float(int_value) + new_decimal / 10.0
        
        # 更新显示
        self.update_resistance_display(exp_name)
        self.update_voltage_display()

    def update_resistance_display(self, exp_name):
        """更新电阻箱显示"""
        value = self.resistance_box.get(exp_name, 1000.0)
        digits = self.resistance_box_digits.get(exp_name, [0, 0, 0, 0, 0])
        
        # 更新各位数字显示
        for i, lbl in enumerate(self.digit_labels):
            lbl.config(text=str(digits[i]))
        
        # 更新小数位（从电阻值计算）
        decimal_part = int(round((value - int(value)) * 10))
        self.decimal_label.config(text=str(decimal_part))
        
        # 更新总显示
        self.resistance_display_label.config(text=f"{value:.1f} Ω")

    def show_resistance_box(self, show):
        """显示或隐藏电阻箱"""
        if show:
            self.resistance_frame.grid(row=1, column=0, columnspan=7, pady=5, sticky='w')
            # 初始化显示
            exp_name = self.current_tab
            # 确保小数位显示正确
            value = self.resistance_box.get(exp_name, 1000.0)
            decimal_part = int(round((value - int(value)) * 10))
            self.decimal_label.config(text=str(decimal_part))
            self.update_resistance_display(exp_name)
        else:
            self.resistance_frame.grid_remove()

    def confirm_set_temp(self):
        """确定设定温度，启动温度渐变动画"""
        exp_name = self.current_tab
        
        # 取消之前的动画
        if exp_name in self.temp_animation_id and self.temp_animation_id[exp_name] is not None:
            self.root.after_cancel(self.temp_animation_id[exp_name])
            self.temp_animation_id[exp_name] = None
        
        # 获取目标温度
        self.target_temp_value[exp_name] = self.temp_progress.get()
        
        # 获取当前温度
        try:
            self.current_temp_value[exp_name] = float(self.current_temp_var.get())
        except:
            self.current_temp_value[exp_name] = 30.0
        
        # 如果当前温度已经等于目标温度，不需要动画
        if abs(self.current_temp_value[exp_name] - self.target_temp_value[exp_name]) < 0.01:
            return
        
        # # 检查散热风扇是否阻止升温
        # fan_status = self.fan_var.get()
        # if fan_status == "开" and self.current_temp_value[exp_name] < self.target_temp_value[exp_name]:
        #     messagebox.showwarning("警告", 
        #         "散热风扇已开启，当前温度低于设定温度\n" +
        #         "散热风扇会阻止温度上升，请关闭散热风扇后再试")
        #     return
        
        # 开始温度渐变
        self.animate_temperature()
        
    def animate_temperature(self):
        """温度渐变动画，每0.2秒变化一次"""
        exp_name = self.current_tab
        
        # 获取散热风扇状态
        fan_status = self.fan_var.get()
        
        # 计算当前温度与目标温度的差值
        temp_diff = self.current_temp_value[exp_name] - self.target_temp_value[exp_name]
        
        if fan_status == "开":
            # 散热风扇打开时
            if temp_diff < 0:
                # 当前温度低于设定温度，升温慢（步长0.1）
                self.current_temp_value[exp_name] += 0.1
                if self.current_temp_value[exp_name] > self.target_temp_value[exp_name]:
                    self.current_temp_value[exp_name] = self.target_temp_value[exp_name]
            else:
                # 当前温度高于设定温度，正常降温
                if abs(temp_diff) < 2.0:
                    step = 0.25
                else:
                    step = 0.5
                self.current_temp_value[exp_name] -= step
                if self.current_temp_value[exp_name] < self.target_temp_value[exp_name]:
                    self.current_temp_value[exp_name] = self.target_temp_value[exp_name]
        else:
            # 散热风扇关闭时
            if temp_diff < 0:
                # 当前温度低于设定温度，正常升温
                if abs(temp_diff) < 2.0:
                    step = 0.25
                else:
                    step = 0.5
                self.current_temp_value[exp_name] += step
                if self.current_temp_value[exp_name] > self.target_temp_value[exp_name]:
                    self.current_temp_value[exp_name] = self.target_temp_value[exp_name]
            else:
                # 当前温度高于设定温度，降温慢（步长0.1）
                self.current_temp_value[exp_name] -= 0.1
                if self.current_temp_value[exp_name] < self.target_temp_value[exp_name]:
                    self.current_temp_value[exp_name] = self.target_temp_value[exp_name]
        
        # 更新显示
        self.current_temp_var.set(f"{self.current_temp_value[exp_name]:.1f}")
        self.update_voltage_display()
        
        # 检查是否达到目标温度
        if abs(self.current_temp_value[exp_name] - self.target_temp_value[exp_name]) < 0.01:
            self.temp_animation_id[exp_name] = None
            return
        
        # 继续动画（每0.2秒执行一次）
        self.temp_animation_id[exp_name] = self.root.after(200, self.animate_temperature)
        
    def on_voltage_range_change(self):
        self.update_voltage_display()
        self.update_indicator_position()  # 添加这一行
    
    def update_indicator_position(self):
        """更新电压表挡位指示灯位置"""
        try:
            if hasattr(self, 'device_canvas'):
                # 删除旧指示灯
                self.device_canvas.delete(self.indicator)
                # 根据挡位选择位置
                if self.voltage_range_var.get() == "2V":
                    # 2V位置 (195, 104) 圆点大小为8x8，中心在(195,104)
                    self.indicator = self.device_canvas.create_oval(191+2, 100+1, 199+2, 108+1, fill='red', outline='red')
                else:
                    # 20V位置 (195, 88)
                    self.indicator = self.device_canvas.create_oval(191+2, 84, 199+2, 92, fill='red', outline='red')
        except Exception as e:
            print(f"更新指示灯位置失败: {e}")

    def update_voltage_display(self):
        try:
            curr_temp = float(self.current_temp_var.get())
        except:
            curr_temp = 30.0
        
        # 根据传感器类型计算电压
        if hasattr(self, 'sensor_type'):
            if self.sensor_type == "pt100":
                if hasattr(self, 'current_mode'):
                    if self.current_mode == "恒电流":
                        resistance = 113.7 + 0.398 * (curr_temp - 30)
                        voltage = 0.001 * resistance
                    else:  # 电桥法
                        R1 = 1000.0
                        R2 = 1000.0
                        U0 = 2.0
                        R3 = 114.0 + 0.391 * (curr_temp - 30)
                        exp_name = self.current_tab
                        Rt = self.resistance_box.get(exp_name, 1000.0)
                        voltage = (R3 / (R1 + R3) - Rt / (R2 + Rt)) * U0
                else:
                    resistance = 113.7 + 0.398 * (curr_temp - 30)
                    voltage = 0.001 * resistance
            elif self.sensor_type == "ntc":
                if hasattr(self, 'current_mode') and self.current_mode == "电桥":
                    R1 = 1000.0
                    R2 = 1000.0
                    U0 = 2.0
                    R3 = 2041 * np.exp(-0.03 * curr_temp)
                    exp_name = self.current_tab
                    Rt = self.resistance_box.get(exp_name, 1000.0)
                    voltage = (R3 / (R1 + R3) - Rt / (R2 + Rt)) * U0
                else:
                    resistance = 2060 * np.exp(-0.03 * curr_temp)
                    voltage = 0.001 * resistance
            elif self.sensor_type == "pn":
                voltage_mv = -2.157 * curr_temp + 669.1
                voltage = voltage_mv / 1000
            elif self.sensor_type == "voltage":
                voltage_mv = 9.92 * curr_temp + 29.85
                voltage = voltage_mv / 1000
            else:  # current
                voltage = (1.031 * curr_temp + 271.0) / 1000
        else:
            resistance = 113.7 + 0.398 * (curr_temp - 30)
            voltage = 0.001 * resistance
        
        # 根据挡位显示（使用更精确的格式化，保留更多有效数字）
        if self.voltage_range_var.get() == "2V":
            self.voltage_var.set(f"{voltage:.4f}")  # 6位小数，提高精度
        else:
            self.voltage_var.set(f"{voltage:.3f}")  # 4位小数

    def on_fan_switch(self):
        status = self.fan_var.get()
        print(f"散热风扇已{status}")
    
    def update_table_headers(self):
        """根据当前实验类型更新表格列标题"""
        exp_name = self.current_tab
        is_pn = "PN结" in exp_name
        is_voltage = "电压型集成温度传感器" in exp_name
        is_current = "电流型集成温度传感器" in exp_name
        is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
        is_electrical_bridge = "电桥法" in exp_name and not is_constant_current
        
        if is_current:
            # 电流型：4列数据
            self.column_widths = {"序号": 40, "t/℃": 100, "U/V": 100, "I/uA": 100}
            self.table["columns"] = ("序号", "t/℃", "U/V", "I/uA")
            self.table.heading("序号", text="序号")
            self.table.heading("t/℃", text="t/℃")
            self.table.heading("U/V", text="U/V")
            self.table.heading("I/uA", text="I/uA")
            for col, width in self.column_widths.items():
                self.table.column(col, width=width, anchor='center')
        elif is_constant_current:
            # 恒电流法（Pt100和NTC）：4列数据：序号、温度、U/mV、Rt/Ω
            self.column_widths = {"序号": 40, "t/℃": 100, "U/mV": 100, "Rt/Ω": 100}
            self.table["columns"] = ("序号", "t/℃", "U/mV", "Rt/Ω")
            self.table.heading("序号", text="序号")
            self.table.heading("t/℃", text="t/℃")
            self.table.heading("U/mV", text="U/mV")
            self.table.heading("Rt/Ω", text="Rt/Ω")
            for col, width in self.column_widths.items():
                self.table.column(col, width=width, anchor='center')
        else:
            # 其他（包括电桥法）：3列数据
            if is_pn:
                label = "Ube/mV"
            elif is_voltage:
                label = "U0/mV"
            else:
                label = "Rt/Ω"
            
            self.column_widths = {"序号": 40, "t/℃": 140, label: 140}
            self.table["columns"] = ("序号", "t/℃", label)
            self.table.heading("序号", text="序号")
            self.table.heading("t/℃", text="t/℃")
            self.table.heading(label, text=label)
            for col, width in self.column_widths.items():
                self.table.column(col, width=width, anchor='center')

    def delete_selected_row(self):
        """删除表格中选中的行（支持多选）"""
        exp_name = self.current_tab
        
        # 获取所有选中的行
        selected = self.table.selection()
        if not selected:
            messagebox.showwarning("警告", "请先选中要删除的行（按住Ctrl或Shift可多选）")
            return
        
        # 获取所有选中行的序号
        selected_indices = []
        selected_data = []
        for item in selected:
            values = self.table.item(item, 'values')
            if values:
                row_index = int(values[0]) - 1
                selected_indices.append(row_index)
                selected_data.append(f"第{values[0]}行 (温度:{values[1]}, 值:{values[2]})")
        
        # 按索引从大到小排序，以便从后往前删除
        selected_indices.sort(reverse=True)
        
        # 确认删除
        msg = f"确定要删除以下 {len(selected)} 行数据吗？\n\n"
        msg += "\n".join(selected_data[:5])  # 最多显示5行
        if len(selected_data) > 5:
            msg += f"\n... 还有 {len(selected_data) - 5} 行"
        
        if not messagebox.askyesno("确认删除", msg):
            return
        
        # 检查数据是否存在
        if exp_name not in self.table_data:
            return
        
        exp_data = self.table_data[exp_name]
        
        # 判断实验类型，确定列数
        is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
        is_current = "电流型集成温度传感器" in exp_name
        cols = 4 if (is_constant_current or is_current) else 3
        
        # 从后往前删除（避免索引变化问题）
        for idx in selected_indices:
            if 0 <= idx < len(exp_data):
                del exp_data[idx]
        
        # 重新编号（从1开始）
        for i, row in enumerate(exp_data):
            row[0] = i + 1
        
        # 如果删除后少于9行，补全空行
        while len(exp_data) < 9:
            if cols == 4:
                exp_data.append([len(exp_data) + 1, "", "", ""])
            else:
                exp_data.append([len(exp_data) + 1, "", ""])
        
        # 保存数据
        self.table_data[exp_name] = exp_data
        
        # 清空拟合结果
        if exp_name in self.slope:
            self.slope[exp_name] = 0.0
            self.r_value[exp_name] = 0.0
            self.intercept[exp_name] = 0.0
            self.fit_show_state[exp_name] = False
        
        # 刷新表格显示
        self.refresh_table_display()
        self.update_plot(show_fit=False)
        self.formula_label.config(text="拟合公式: 未计算")
        
        messagebox.showinfo("删除成功", f"已删除 {len(selected)} 行数据")

    def create_right_frame(self):
        self.right_frame = tk.LabelFrame(self.root, text="数据记录区域", font=("Arial", 12, "bold"),
                                        bg='#f0f0f0', padx=5, pady=5)
        self.right_frame.place(x=570, y=60, width=700, height=700)
        
        # 表格 - 列标题根据实验类型动态设置
        columns = ("序号", "t/℃", "Rt/Ω")
        self.table = ttk.Treeview(self.right_frame, columns=columns, show='headings', height=10)
        self.table.heading("序号", text="序号")
        self.table.heading("t/℃", text="t/℃")
        self.table.heading("Rt/Ω", text="Rt/Ω")
        # 保存列宽配置
        self.column_widths = {"序号": 40, "t/℃": 140, "Rt/Ω": 140}
        self.table.column("序号", width=self.column_widths["序号"], anchor='center')
        self.table.column("t/℃", width=self.column_widths["t/℃"], anchor='center')
        self.table.column("Rt/Ω", width=self.column_widths["Rt/Ω"], anchor='center')
        
        self.table.bind("<Double-1>", self.on_cell_double_click)
        self.table.pack(pady=5, padx=5, fill=tk.X)
        
        self.refresh_table_display()
        
        # 曲线图
        self.fig = Figure(figsize=(6, 3.5), dpi=80)
        self.ax = self.fig.add_subplot(111)
        self.plot_canvas = FigureCanvasTkAgg(self.fig, master=self.right_frame)
        self.plot_canvas.get_tk_widget().pack(pady=5, fill=tk.BOTH, expand=True)
        
        info_frame = tk.Frame(self.right_frame, bg='#f0f0f0')
        info_frame.pack(pady=5, fill=tk.X)
        
        btn_frame = tk.Frame(self.right_frame, bg='#f0f0f0')
        btn_frame.pack(pady=10)
        
        # 记录数据按钮
        btn_record = tk.Button(btn_frame, text="记录数据", command=self.record_data, width=10)
        btn_record.pack(side=tk.LEFT, padx=3)
        
        
        
        # 计算按钮
        btn_calc = tk.Button(btn_frame, text="计算", command=self.calc_and_fit, width=10)
        btn_calc.pack(side=tk.LEFT, padx=3)

        # 删除行按钮
        btn_delete = tk.Button(btn_frame, text="删除行", command=self.delete_selected_row, width=10)
        btn_delete.pack(side=tk.LEFT, padx=3)
        
        # 清空数据按钮
        btn_clear = tk.Button(btn_frame, text="清空数据", command=self.clear_data, width=10)
        btn_clear.pack(side=tk.LEFT, padx=3)
        
        # 导出数据按钮
        btn_export = tk.Button(btn_frame, text="导出数据", command=self.export_data, width=10)
        btn_export.pack(side=tk.LEFT, padx=3)
        
        # 导入数据按钮
        btn_import = tk.Button(btn_frame, text="导入数据", command=self.import_data, width=10)
        btn_import.pack(side=tk.LEFT, padx=3)
        
        self.formula_label = tk.Label(self.right_frame, text="拟合公式: 未计算", font=("Arial",9), bg='#f0f0f0')
        self.formula_label.pack(pady=2)
        
        # 初始化时更新列标题
        self.update_table_headers()

    def record_data(self):
        """记录当前数据到表格"""
        exp_name = self.current_tab
        
        # 获取当前温度
        try:
            current_temp = float(self.current_temp_var.get())
        except:
            messagebox.showerror("错误", "无法获取当前温度")
            return
        
        # 判断实验类型
        is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
        is_current = "电流型集成温度传感器" in exp_name
        is_electrical_bridge = "电桥法" in exp_name and not is_constant_current
        is_pn = "PN结" in exp_name
        is_voltage = "电压型集成温度传感器" in exp_name
        
        # 获取当前电压/值
        if is_constant_current:
            # 恒电流法：获取电压值 U/mV
            try:
                # 从电压显示获取（单位为V，需要转换为mV）
                voltage_v = float(self.voltage_var.get())
                # 保留更多小数位，避免精度丢失
                value = round(voltage_v * 1000, 4)  # 转换为mV，保留4位小数
            except:
                messagebox.showerror("错误", "无法获取当前电压值")
                return
        elif is_electrical_bridge:
            # 电桥法：获取电阻箱值（已经是精确值）
            value = self.resistance_box.get(exp_name, 1000.0)
        elif is_current:
            # 电流型：获取电压值 U/V
            try:
                value = float(self.voltage_var.get())
            except:
                messagebox.showerror("错误", "无法获取当前电压值")
                return
        elif is_pn:
            # PN结：获取电压值 Ube/mV
            try:
                voltage_v = float(self.voltage_var.get())
                value = round(voltage_v * 1000, 4)  # 转换为mV，保留4位小数
            except:
                messagebox.showerror("错误", "无法获取当前电压值")
                return
        elif is_voltage:
            # 电压型：获取电压值 U0/mV
            try:
                voltage_v = float(self.voltage_var.get())
                value = round(voltage_v * 1000, 4)  # 转换为mV，保留4位小数
            except:
                messagebox.showerror("错误", "无法获取当前电压值")
                return
        else:
            messagebox.showerror("错误", "当前实验类型不支持记录数据")
            return
        
        # 检查是否有空行可填入
        exp_data = self.table_data.get(exp_name, [])
        target_row_index = -1
        
        # 查找第一个温度为空的行（从0开始）
        for i, row in enumerate(exp_data):
            if row[1] == "" or row[1] is None:
                target_row_index = i
                break
        
        # 如果没有空行，追加新行
        if target_row_index == -1:
            if len(exp_data) >= 9:
                messagebox.showwarning("警告", "数据已满（最多9行），请先清空数据")
                return
            # 追加新行
            if is_constant_current or is_current:
                new_row = [len(exp_data) + 1, "", "", ""]
            else:
                new_row = [len(exp_data) + 1, "", ""]
            exp_data.append(new_row)
            target_row_index = len(exp_data) - 1
        
        # 填入数据（保留足够精度）
        if is_constant_current or is_current:
            # 4列数据：序号、温度、值、自动计算列
            exp_data[target_row_index][1] = round(current_temp, 2)  # 温度保留2位小数
            exp_data[target_row_index][2] = value  # 值保留4位小数（已在上面处理）
            
            # 如果是恒电流法，自动计算Rt
            if is_constant_current:
                # Rt = U(mV) / 1(mA) = U(Ω)
                exp_data[target_row_index][3] = round(value, 4)  # 保留4位小数
            # 如果是电流型，自动计算电流
            if is_current:
                exp_data[target_row_index][3] = round(value * 1000, 4)  # I = U * 1000 (uA)
        else:
            # 3列数据：序号、温度、值
            exp_data[target_row_index][1] = round(current_temp, 2)  # 温度保留2位小数
            exp_data[target_row_index][2] = round(value, 4)  # 值保留4位小数
        
        # 保存数据
        self.table_data[exp_name] = exp_data
        
        # 刷新表格显示
        self.refresh_table_display()
        
        # 清空拟合结果
        if exp_name in self.slope:
            self.slope[exp_name] = 0.0
            self.r_value[exp_name] = 0.0
            self.intercept[exp_name] = 0.0
            self.fit_show_state[exp_name] = False
        
        self.formula_label.config(text="拟合公式: 未计算")
        self.update_plot(show_fit=False)
        
        # 根据实验类型显示不同的提示信息
        if is_constant_current:
            value_unit = "mV"
            auto_calc = f"Rt = {exp_data[target_row_index][3]:.4f} Ω"
        elif is_electrical_bridge:
            value_unit = "Ω"
            auto_calc = ""
        elif is_current:
            value_unit = "V"
            auto_calc = f"I = {exp_data[target_row_index][3]:.4f} uA"
        elif is_pn:
            value_unit = "mV"
            auto_calc = ""
        elif is_voltage:
            value_unit = "mV"
            auto_calc = ""
        else:
            value_unit = ""
            auto_calc = ""
        
        msg = f"已记录数据：\n温度 = {current_temp:.1f}℃\n值 = {value:.4f} {value_unit}"
        if auto_calc:
            msg += f"\n{auto_calc}"
        messagebox.showinfo("记录成功", msg)

    def refresh_table_display(self):
        """刷新表格显示"""
        # 先清空表格
        for row in self.table.get_children():
            self.table.delete(row)
        
        exp_name = self.current_tab
        is_current = "电流型集成温度传感器" in exp_name
        is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
        
        # 先更新列结构
        self.update_table_headers()
        
        if exp_name in self.table_data:
            # 确保有9行数据，并重新编号
            exp_data = self.table_data[exp_name]
            
            # 重新编号
            for i, row in enumerate(exp_data):
                row[0] = i + 1
            
            # 确保有9行
            cols = 4 if (is_current or is_constant_current) else 3
            while len(exp_data) < 9:
                if cols == 4:
                    exp_data.append([len(exp_data) + 1, "", "", ""])
                else:
                    exp_data.append([len(exp_data) + 1, "", ""])
            
            for row_data in exp_data:
                if is_current:
                    # 电流型：4列数据
                    temp_value = row_data[1] if len(row_data) > 1 and row_data[1] != "" else ""
                    voltage_value = row_data[2] if len(row_data) > 2 and row_data[2] != "" else ""
                    current_value = row_data[3] if len(row_data) > 3 and row_data[3] != "" else ""
                    self.table.insert("", tk.END, values=(row_data[0], temp_value, voltage_value, current_value))
                elif is_constant_current:
                    # 恒电流法：4列数据（序号、温度、U/mV、Rt/Ω）
                    temp_value = row_data[1] if len(row_data) > 1 and row_data[1] != "" else ""
                    u_value = row_data[2] if len(row_data) > 2 and row_data[2] != "" else ""
                    rt_value = row_data[3] if len(row_data) > 3 and row_data[3] != "" else ""
                    self.table.insert("", tk.END, values=(row_data[0], temp_value, u_value, rt_value))
                else:
                    # 其他：3列数据（序号、温度、值）
                    temp_value = row_data[1] if len(row_data) > 1 and row_data[1] != "" else ""
                    res_value = row_data[2] if len(row_data) > 2 and row_data[2] != "" else ""
                    self.table.insert("", tk.END, values=(row_data[0], temp_value, res_value))
        
        # 强制刷新列宽
        if hasattr(self, 'column_widths'):
            for col, width in self.column_widths.items():
                if col in self.table['columns']:
                    self.table.column(col, width=width, anchor='center')

    def on_cell_double_click(self, event):
        """双击表格单元格编辑"""
        region = self.table.identify_region(event.x, event.y)
        if region != "cell":
            return
        column = self.table.identify_column(event.x)
        row_id = self.table.identify_row(event.y)
        if not row_id:
            return
        col_index = int(column[1:]) - 1
        # 序号列不可编辑
        if col_index == 0:
            return
        
        # 获取当前值
        current_values = self.table.item(row_id, 'values')
        if not current_values:
            return
        
        exp_name = self.current_tab
        is_current = "电流型集成温度传感器" in exp_name
        is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
        
        # 电流型：电流列（第4列）不可编辑
        if is_current and col_index == 3:
            messagebox.showinfo("提示", "电流值由电压自动计算，不可手动编辑")
            return
        
        # 恒电流法：Rt列（第4列）不可编辑（由U自动计算）
        if is_constant_current and col_index == 3:
            messagebox.showinfo("提示", "电阻值由电压自动计算（R = U / 1mA）")
            return
        
        old_value = current_values[col_index] if current_values[col_index] != "" else ""
        
        # 获取单元格在表格中的位置
        x, y, w, h = self.table.bbox(row_id, column)
        
        # 创建输入框 - 直接放在表格上，坐标相对于表格
        entry = tk.Entry(self.table, justify='center')
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, old_value)
        entry.focus()
        
        def save_edit(event=None):
            new_val = entry.get().strip()
            entry.destroy()
            
            # 获取行索引
            row_idx = int(current_values[0]) - 1
            exp_name = self.current_tab
            is_current = "电流型集成温度传感器" in exp_name
            is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
            is_pn = "PN结" in exp_name
            is_voltage = "电压型集成温度传感器" in exp_name
            
            if exp_name in self.table_data and 0 <= row_idx < len(self.table_data[exp_name]):
                if is_current:
                    # 电流型：温度和电压可编辑，电流自动计算
                    if col_index == 1:  # 温度列
                        if new_val == "":
                            self.table_data[exp_name][row_idx][1] = ""
                            self.table_data[exp_name][row_idx][2] = ""
                            self.table_data[exp_name][row_idx][3] = ""
                        else:
                            try:
                                self.table_data[exp_name][row_idx][1] = float(new_val)
                            except ValueError:
                                messagebox.showerror("错误", "请输入有效的数字")
                                return
                    elif col_index == 2:  # 电压列
                        if new_val == "":
                            self.table_data[exp_name][row_idx][2] = ""
                            self.table_data[exp_name][row_idx][3] = ""
                        else:
                            try:
                                voltage = float(new_val)
                                self.table_data[exp_name][row_idx][2] = voltage
                                self.table_data[exp_name][row_idx][3] = voltage * 1000
                            except ValueError:
                                messagebox.showerror("错误", "请输入有效的数字")
                                return
                elif is_constant_current:
                    # 恒电流法：温度和U/mV可编辑，Rt自动计算（R = U / 1mA = U/1000 * 1000 = U）
                    # 注意：U单位是mV，I=1mA，所以 R = U/I = U(mV)/1(mA) = U(Ω)
                    if col_index == 1:  # 温度列
                        if new_val == "":
                            self.table_data[exp_name][row_idx][1] = ""
                        else:
                            try:
                                self.table_data[exp_name][row_idx][1] = float(new_val)
                            except ValueError:
                                messagebox.showerror("错误", "请输入有效的数字")
                                return
                    elif col_index == 2:  # U/mV列
                        if new_val == "":
                            self.table_data[exp_name][row_idx][2] = ""
                            self.table_data[exp_name][row_idx][3] = ""
                        else:
                            try:
                                u_mv = float(new_val)
                                self.table_data[exp_name][row_idx][2] = u_mv
                                # 自动计算电阻：R = U / I，I = 1mA，U单位mV，R单位Ω
                                # R = U(mV) / 1(mA) = U(Ω)
                                self.table_data[exp_name][row_idx][3] = u_mv
                            except ValueError:
                                messagebox.showerror("错误", "请输入有效的数字")
                                return
                else:
                    # 非电流型非恒电流法：温度列和值列都可编辑
                    if col_index == 1:  # 温度列
                        if new_val == "":
                            self.table_data[exp_name][row_idx][1] = ""
                        else:
                            try:
                                self.table_data[exp_name][row_idx][1] = float(new_val)
                            except ValueError:
                                messagebox.showerror("错误", "请输入有效的数字")
                                return
                    elif col_index == 2:  # 值列（Rt/Ω 或 Ube/mV 或 U0/mV）
                        if new_val == "":
                            self.table_data[exp_name][row_idx][2] = ""
                        else:
                            try:
                                self.table_data[exp_name][row_idx][2] = float(new_val)
                            except ValueError:
                                messagebox.showerror("错误", "请输入有效的数字")
                                return
                
                # 清空拟合结果
                if exp_name in self.slope:
                    self.slope[exp_name] = 0.0
                    self.r_value[exp_name] = 0.0
                    self.intercept[exp_name] = 0.0
                    self.fit_show_state[exp_name] = False
            
            # 刷新表格显示
            self.refresh_table_display()
            self.update_plot(show_fit=False)
            self.update_voltage_display()
            self.formula_label.config(text="拟合公式: 未计算")
            
            # 恢复列宽
            if hasattr(self, 'column_widths'):
                for col, width in self.column_widths.items():
                    if col in self.table['columns']:
                        self.table.column(col, width=width, anchor='center')
                        
        entry.bind('<Return>', save_edit)
        entry.bind('<FocusOut>', lambda e: save_edit())

    def update_plot(self, show_fit=False):
        """更新曲线图，过滤掉空值"""
        self.ax.clear()
        
        exp_name = self.current_tab
        is_pn = "PN结" in exp_name
        is_voltage = "电压型集成温度传感器" in exp_name
        is_current = "电流型集成温度传感器" in exp_name
        is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
        
        # 从当前实验的数据中获取
        temps = []
        values = []  # 电阻、电压或电流
        if exp_name in self.table_data:
            for row in self.table_data[exp_name]:
                if is_current:
                    # 电流型：使用电流值（第4列）作为纵坐标
                    if row[1] != "" and row[3] != "":
                        temps.append(float(row[1]))
                        values.append(float(row[3]))
                elif is_constant_current:
                    # 恒电流法：使用Rt值（第4列）作为纵坐标
                    if row[1] != "" and row[3] != "":
                        temps.append(float(row[1]))
                        values.append(float(row[3]))
                else:
                    if row[1] != "" and row[2] != "":
                        temps.append(float(row[1]))
                        values.append(float(row[2]))
        
        # 绘制散点图
        if temps:  # 只在有数据时绘制
            self.ax.scatter(temps, values, color='blue', s=50, label='实验数据')
        
        # 检查是否有拟合结果
        if show_fit and exp_name in self.slope and len(temps) >= 2:
            slope_val = self.slope.get(exp_name, 0.0)
            intercept_val = self.intercept.get(exp_name, 0.0)
            is_ntc = "NTC" in exp_name
            
            # NTC需要 intercept_val (k值) 不为0
            # 其他需要 slope_val 不为0
            if is_ntc:
                has_fit = intercept_val != 0.0
            else:
                has_fit = slope_val != 0.0
            
            if has_fit and max(temps) != min(temps):
                t_min, t_max = min(temps), max(temps)
                if t_min == t_max:
                    t_min = t_min - 1
                    t_max = t_max + 1
                t_line = np.linspace(t_min, t_max, 100)
                
                if is_ntc:
                    # NTC：指数曲线 R = k * e^(-0.03*T)
                    r_line = intercept_val * np.exp(-0.03 * t_line)
                    self.ax.plot(t_line, r_line, 'r-', linewidth=2, label=f'拟合曲线 (k={intercept_val:.4f})')
                else:
                    # 其他：线性直线
                    r_line = slope_val * t_line + intercept_val
                    self.ax.plot(t_line, r_line, 'r-', linewidth=2, label=f'拟合直线 (A={slope_val:.4f})')
        
        # 设置坐标轴标签
        self.ax.set_xlabel("t/℃", fontsize=11)
        if is_pn:
            self.ax.set_ylabel("Ube/mV", fontsize=11)
            self.ax.set_title("电压-温度特性曲线", fontsize=12)
        elif is_voltage:
            self.ax.set_ylabel("U0/mV", fontsize=11)
            self.ax.set_title("电压-温度特性曲线", fontsize=12)
        elif is_current:
            self.ax.set_ylabel("I/uA", fontsize=11)
            self.ax.set_title("电流-温度特性曲线", fontsize=12)
        else:
            self.ax.set_ylabel("Rt/Ω", fontsize=11)
            self.ax.set_title("电阻-温度特性曲线", fontsize=12)
        self.ax.grid(True, alpha=0.3)
        if temps:  # 只在有数据时显示图例
            self.ax.legend()
        self.plot_canvas.draw()
        
    def calc_and_fit(self):
        """计算拟合，支持线性和指数拟合"""
        exp_name = self.current_tab
        
        temps = []
        values = []
        is_current = "电流型集成温度传感器" in exp_name
        is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
        
        if exp_name in self.table_data:
            for row in self.table_data[exp_name]:
                if is_current:
                    # 电流型：使用电流值（第4列）进行拟合
                    if row[1] != "" and row[3] != "":
                        temps.append(float(row[1]))
                        values.append(float(row[3]))
                elif is_constant_current:
                    # 恒电流法：使用Rt值（第4列）进行拟合
                    if row[1] != "" and row[3] != "":
                        temps.append(float(row[1]))
                        values.append(float(row[3]))
                else:
                    if row[1] != "" and row[2] != "":
                        temps.append(float(row[1]))
                        values.append(float(row[2]))
        
        if len(temps) < 2:
            messagebox.showwarning("警告", "有效数据点不足，至少需要2组非空数据")
            return
        
        # 判断实验类型
        is_ntc = "NTC" in exp_name
        is_pn = "PN结" in exp_name
        is_voltage = "电压型集成温度传感器" in exp_name
        
        try:
            if is_ntc:
                # NTC热敏电阻：指数拟合 R = k * e^(-0.03*T)
                log_values = np.log(values)
                slope, intercept, r_value, p_value, std_err = stats.linregress(temps, log_values)
                
                k_value = np.exp(intercept)
                
                # 保存拟合结果
                self.slope[exp_name] = -0.03
                self.intercept[exp_name] = k_value
                self.r_value[exp_name] = r_value ** 2
                
                formula_text = f"拟合公式: Rt = {k_value:.4f} × e^(-0.03×t)  (R² = {r_value ** 2:.6f})"
                self.formula_label.config(text=formula_text)
                
                # 设置拟合显示状态为True
                self.fit_show_state[exp_name] = True
                self.update_plot(show_fit=True)
                
                messagebox.showinfo("计算完成", f"指数拟合已完成\n\n公式: Rt = {k_value:.4f} × e^(-0.03×t)\nR² = {r_value ** 2:.6f}")
            else:
                # 其他：线性拟合
                slope, intercept, r_value, p_value, std_err = stats.linregress(temps, values)
                
                self.slope[exp_name] = slope
                self.intercept[exp_name] = intercept
                self.r_value[exp_name] = r_value ** 2
                
                if is_pn:
                    formula_text = f"拟合公式: Ube = {slope:.4f} × t + {intercept:.4f}  (R² = {r_value ** 2:.6f})"
                elif is_voltage:
                    formula_text = f"拟合公式: U0 = {slope:.4f} × t + {intercept:.4f}  (R² = {r_value ** 2:.6f})"
                elif is_current:
                    formula_text = f"拟合公式: I = {slope:.4f} × t + {intercept:.4f}  (R² = {r_value ** 2:.6f})"
                else:
                    formula_text = f"拟合公式: Rt = {slope:.4f} × t + {intercept:.4f}  (R² = {r_value ** 2:.6f})"
                self.formula_label.config(text=formula_text)
                
                # 设置拟合显示状态为True
                self.fit_show_state[exp_name] = True
                self.update_plot(show_fit=True)
                
                if is_pn:
                    messagebox.showinfo("计算完成", f"线性拟合已完成\n\n公式: Ube = {slope:.4f} × t + {intercept:.4f}\nR² = {r_value ** 2:.6f}")
                elif is_voltage:
                    messagebox.showinfo("计算完成", f"线性拟合已完成\n\n公式: U0 = {slope:.4f} × t + {intercept:.4f}\nR² = {r_value ** 2:.6f}")
                elif is_current:
                    messagebox.showinfo("计算完成", f"线性拟合已完成\n\n公式: I = {slope:.4f} × t + {intercept:.4f}\nR² = {r_value ** 2:.6f}")
                else:
                    messagebox.showinfo("计算完成", f"线性拟合已完成\n\n公式: Rt = {slope:.4f} × t + {intercept:.4f}\nR² = {r_value ** 2:.6f}")
        except Exception as e:
            messagebox.showerror("错误", f"拟合失败：{str(e)}")

    def clear_data(self):
        """清空当前实验数据"""
        exp_name = self.current_tab
        is_current = "电流型集成温度传感器" in exp_name
        is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
        
        # 清空对应实验的数据 - 保留序号，清空数据列
        if exp_name in self.table_data:
            for i in range(len(self.table_data[exp_name])):
                if is_current or is_constant_current:
                    # 4列数据
                    self.table_data[exp_name][i][1] = ""
                    self.table_data[exp_name][i][2] = ""
                    self.table_data[exp_name][i][3] = ""
                else:
                    # 3列数据
                    self.table_data[exp_name][i][1] = ""
                    self.table_data[exp_name][i][2] = ""
        
        # 清空拟合结果
        if exp_name in self.slope:
            self.slope[exp_name] = 0.0
            self.r_value[exp_name] = 0.0
            self.intercept[exp_name] = 0.0
            self.fit_show_state[exp_name] = False
        
        # 刷新表格显示
        self.refresh_table_display()
        self.update_plot(show_fit=False)
        self.formula_label.config(text="拟合公式: 未计算")
        messagebox.showinfo("清空完成", f"{exp_name}的所有数据已清空")

    def export_data(self):
        """导出数据到Excel文件"""
        exp_name = self.current_tab
        self.save_current_data()
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"{exp_name}_数据.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("所有文件", "*.*")]
        )

        if file_path:
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment
                
                wb = openpyxl.Workbook()
                
                # 数据表格工作表
                ws_data = wb.active
                ws_data.title = "实验数据"
                
                # 设置表头
                is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
                is_current = "电流型集成温度传感器" in exp_name
                
                if is_constant_current:
                    headers = ["序号", "t/℃", "U/mV", "Rt/Ω"]
                elif is_current:
                    headers = ["序号", "t/℃", "U/V", "I/uA"]
                elif "PN结" in exp_name:
                    headers = ["序号", "t/℃", "Ube/mV"]
                elif "电压型集成温度传感器" in exp_name:
                    headers = ["序号", "t/℃", "U0/mV"]
                else:
                    headers = ["序号", "t/℃", "Rt/Ω"]
                
                for col, header in enumerate(headers, 1):
                    cell = ws_data.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # 写入数据 - 从当前实验获取
                if exp_name in self.table_data:
                    for row_idx, row_data in enumerate(self.table_data[exp_name], 2):
                        ws_data.cell(row=row_idx, column=1, value=row_data[0])
                        ws_data.cell(row=row_idx, column=2, value=row_data[1] if row_data[1] != "" else None)
                        if is_constant_current or is_current:
                            # 4列数据
                            ws_data.cell(row=row_idx, column=3, value=row_data[2] if row_data[2] != "" else None)
                            ws_data.cell(row=row_idx, column=4, value=row_data[3] if row_data[3] != "" else None)
                        else:
                            # 3列数据
                            ws_data.cell(row=row_idx, column=3, value=row_data[2] if row_data[2] != "" else None)
                
                # 拟合结果工作表
                ws_fit = wb.create_sheet("拟合结果")
                ws_fit.cell(row=1, column=1, value="参数").font = Font(bold=True)
                ws_fit.cell(row=1, column=2, value="数值").font = Font(bold=True)
                
                slope_val = self.slope.get(exp_name, 0.0)
                intercept_val = self.intercept.get(exp_name, 0.0)
                r2_val = self.r_value.get(exp_name, 0.0)
                is_ntc = "NTC" in exp_name
                is_pn = "PN结" in exp_name
                is_voltage = "电压型集成温度传感器" in exp_name
                is_current = "电流型集成温度传感器" in exp_name
                is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
                
                if is_ntc:
                    # NTC热敏电阻：指数拟合
                    ws_fit.cell(row=2, column=1, value="系数 k")
                    ws_fit.cell(row=2, column=2, value=intercept_val if intercept_val != 0 else "未计算")
                    ws_fit.cell(row=3, column=1, value="R^2")
                    ws_fit.cell(row=3, column=2, value=r2_val if r2_val != 0 else "未计算")
                    if intercept_val != 0:
                        formula = f"Rt = {intercept_val:.4f} × e^(-0.03×t)  (R^2 = {r2_val:.6f})"
                        ws_fit.cell(row=4, column=1, value="拟合公式")
                        ws_fit.cell(row=4, column=2, value=formula)
                elif is_pn:
                    # PN结温度传感器：线性拟合
                    ws_fit.cell(row=2, column=1, value="温度系数 A")
                    ws_fit.cell(row=2, column=2, value=slope_val if slope_val != 0 else "未计算")
                    ws_fit.cell(row=3, column=1, value="截距 intercept")
                    ws_fit.cell(row=3, column=2, value=intercept_val if intercept_val != 0 else "未计算")
                    ws_fit.cell(row=4, column=1, value="R^2")
                    ws_fit.cell(row=4, column=2, value=r2_val if r2_val != 0 else "未计算")
                    if slope_val != 0:
                        formula = f"Ube = {slope_val:.4f} × t + {intercept_val:.4f}  (R^2 = {r2_val:.6f})"
                        ws_fit.cell(row=5, column=1, value="拟合公式")
                        ws_fit.cell(row=5, column=2, value=formula)
                elif is_voltage:
                    # 电压型集成温度传感器：线性拟合
                    ws_fit.cell(row=2, column=1, value="温度系数 A")
                    ws_fit.cell(row=2, column=2, value=slope_val if slope_val != 0 else "未计算")
                    ws_fit.cell(row=3, column=1, value="截距 intercept")
                    ws_fit.cell(row=3, column=2, value=intercept_val if intercept_val != 0 else "未计算")
                    ws_fit.cell(row=4, column=1, value="R^2")
                    ws_fit.cell(row=4, column=2, value=r2_val if r2_val != 0 else "未计算")
                    if slope_val != 0:
                        formula = f"U0 = {slope_val:.4f} × t + {intercept_val:.4f}  (R^2 = {r2_val:.6f})"
                        ws_fit.cell(row=5, column=1, value="拟合公式")
                        ws_fit.cell(row=5, column=2, value=formula)
                elif is_current:
                    # 电流型集成温度传感器：线性拟合
                    ws_fit.cell(row=2, column=1, value="温度系数 A")
                    ws_fit.cell(row=2, column=2, value=slope_val if slope_val != 0 else "未计算")
                    ws_fit.cell(row=3, column=1, value="截距 intercept")
                    ws_fit.cell(row=3, column=2, value=intercept_val if intercept_val != 0 else "未计算")
                    ws_fit.cell(row=4, column=1, value="R^2")
                    ws_fit.cell(row=4, column=2, value=r2_val if r2_val != 0 else "未计算")
                    if slope_val != 0:
                        formula = f"I = {slope_val:.4f} × t + {intercept_val:.4f}  (R^2 = {r2_val:.6f})"
                        ws_fit.cell(row=5, column=1, value="拟合公式")
                        ws_fit.cell(row=5, column=2, value=formula)
                elif is_constant_current:
                    # 恒电流法（Pt100/NTC）：线性拟合
                    ws_fit.cell(row=2, column=1, value="温度系数 A")
                    ws_fit.cell(row=2, column=2, value=slope_val if slope_val != 0 else "未计算")
                    ws_fit.cell(row=3, column=1, value="截距 intercept")
                    ws_fit.cell(row=3, column=2, value=intercept_val if intercept_val != 0 else "未计算")
                    ws_fit.cell(row=4, column=1, value="R^2")
                    ws_fit.cell(row=4, column=2, value=r2_val if r2_val != 0 else "未计算")
                    if slope_val != 0:
                        formula = f"Rt = {slope_val:.4f} × t + {intercept_val:.4f}  (R^2 = {r2_val:.6f})"
                        ws_fit.cell(row=5, column=1, value="拟合公式")
                        ws_fit.cell(row=5, column=2, value=formula)
                else:
                    # Pt100电桥法：线性拟合
                    ws_fit.cell(row=2, column=1, value="温度系数 A")
                    ws_fit.cell(row=2, column=2, value=slope_val if slope_val != 0 else "未计算")
                    ws_fit.cell(row=3, column=1, value="截距 intercept")
                    ws_fit.cell(row=3, column=2, value=intercept_val if intercept_val != 0 else "未计算")
                    ws_fit.cell(row=4, column=1, value="R^2")
                    ws_fit.cell(row=4, column=2, value=r2_val if r2_val != 0 else "未计算")
                    if slope_val != 0:
                        formula = f"Rt = {slope_val:.4f} × t + {intercept_val:.4f}  (R^2 = {r2_val:.6f})"
                        ws_fit.cell(row=5, column=1, value="拟合公式")
                        ws_fit.cell(row=5, column=2, value=formula)
                
                # 保存文件
                wb.save(file_path)
                messagebox.showinfo("导出成功", f"数据已导出到：\n{file_path}")
                
            except ImportError:
                messagebox.showerror("错误", "请先安装openpyxl库：\npip install openpyxl")
            except Exception as e:
                messagebox.showerror("导出失败", f"导出时发生错误：{str(e)}")

    def import_data(self):
        """从Excel文件导入数据到当前实验"""
        exp_name = self.current_tab
        
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            try:
                import openpyxl
                
                wb = openpyxl.load_workbook(file_path, data_only=True)
                
                # 导入数据表格
                if "实验数据" in wb.sheetnames:
                    ws_data = wb["实验数据"]
                else:
                    ws_data = wb.active
                
                # 读取数据
                new_table_data = []
                is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
                is_current = "电流型集成温度传感器" in exp_name
                
                row = 2
                while row <= 10:  # 最多10行
                    seq_cell = ws_data.cell(row=row, column=1).value
                    if seq_cell is None:
                        break
                    
                    temp_val = ws_data.cell(row=row, column=2).value
                    
                    if is_constant_current or is_current:
                        # 4列数据
                        val_col3 = ws_data.cell(row=row, column=3).value
                        val_col4 = ws_data.cell(row=row, column=4).value
                        
                        temp = "" if temp_val is None else float(temp_val)
                        val3 = "" if val_col3 is None else float(val_col3)
                        val4 = "" if val_col4 is None else float(val_col4)
                        
                        new_table_data.append([int(seq_cell), temp, val3, val4])
                    else:
                        # 3列数据
                        val_col3 = ws_data.cell(row=row, column=3).value
                        
                        temp = "" if temp_val is None else float(temp_val)
                        val3 = "" if val_col3 is None else float(val_col3)
                        
                        new_table_data.append([int(seq_cell), temp, val3])
                    
                    row += 1
                
                # 如果读取到数据，更新当前实验的表格
                if new_table_data:
                    self.table_data[exp_name] = new_table_data
                else:
                    messagebox.showwarning("警告", "Excel文件中没有找到有效数据")
                    return
                
                # 导入拟合结果
                is_ntc = "NTC" in exp_name
                is_pn = "PN结" in exp_name
                is_voltage = "电压型集成温度传感器" in exp_name
                is_current = "电流型集成温度传感器" in exp_name
                is_constant_current = "恒电流法" in exp_name and ("Pt100" in exp_name or "NTC热敏电阻" in exp_name)
                
                if "拟合结果" in wb.sheetnames:
                    ws_fit = wb["拟合结果"]
                    
                    if is_ntc:
                        # NTC：读取k值和R^2
                        k_val = ws_fit.cell(row=2, column=2).value
                        r2_val = ws_fit.cell(row=3, column=2).value
                        
                        if k_val and k_val != "未计算":
                            self.slope[exp_name] = -0.03
                            self.intercept[exp_name] = float(k_val)
                            self.r_value[exp_name] = float(r2_val) if r2_val and r2_val != "未计算" else 0.0
                            self.fit_show_state[exp_name] = True
                            formula_text = f"拟合公式: Rt = {float(k_val):.4f} × e^(-0.03×t)  (R² = {self.r_value[exp_name]:.6f})"
                            self.formula_label.config(text=formula_text)
                            self.update_plot(show_fit=True)
                        else:
                            self.slope[exp_name] = 0.0
                            self.intercept[exp_name] = 0.0
                            self.r_value[exp_name] = 0.0
                            self.fit_show_state[exp_name] = False
                            self.formula_label.config(text="拟合公式: 未计算")
                            self.update_plot(show_fit=False)
                    elif is_pn:
                        # PN结：读取线性拟合参数
                        slope_val = ws_fit.cell(row=2, column=2).value
                        intercept_val = ws_fit.cell(row=3, column=2).value
                        r2_val = ws_fit.cell(row=4, column=2).value
                        
                        if slope_val and slope_val != "未计算":
                            self.slope[exp_name] = float(slope_val)
                            self.intercept[exp_name] = float(intercept_val) if intercept_val and intercept_val != "未计算" else 0.0
                            self.r_value[exp_name] = float(r2_val) if r2_val and r2_val != "未计算" else 0.0
                            self.fit_show_state[exp_name] = True
                            formula_text = f"拟合公式: Ube = {float(slope_val):.4f} × t + {self.intercept[exp_name]:.4f}  (R² = {self.r_value[exp_name]:.6f})"
                            self.formula_label.config(text=formula_text)
                            self.update_plot(show_fit=True)
                        else:
                            self.slope[exp_name] = 0.0
                            self.r_value[exp_name] = 0.0
                            self.intercept[exp_name] = 0.0
                            self.fit_show_state[exp_name] = False
                            self.formula_label.config(text="拟合公式: 未计算")
                            self.update_plot(show_fit=False)
                    elif is_voltage:
                        # 电压型集成温度传感器：读取线性拟合参数
                        slope_val = ws_fit.cell(row=2, column=2).value
                        intercept_val = ws_fit.cell(row=3, column=2).value
                        r2_val = ws_fit.cell(row=4, column=2).value
                        
                        if slope_val and slope_val != "未计算":
                            self.slope[exp_name] = float(slope_val)
                            self.intercept[exp_name] = float(intercept_val) if intercept_val and intercept_val != "未计算" else 0.0
                            self.r_value[exp_name] = float(r2_val) if r2_val and r2_val != "未计算" else 0.0
                            self.fit_show_state[exp_name] = True
                            formula_text = f"拟合公式: U0 = {float(slope_val):.4f} × t + {self.intercept[exp_name]:.4f}  (R² = {self.r_value[exp_name]:.6f})"
                            self.formula_label.config(text=formula_text)
                            self.update_plot(show_fit=True)
                        else:
                            self.slope[exp_name] = 0.0
                            self.r_value[exp_name] = 0.0
                            self.intercept[exp_name] = 0.0
                            self.fit_show_state[exp_name] = False
                            self.formula_label.config(text="拟合公式: 未计算")
                            self.update_plot(show_fit=False)
                    elif is_current:
                        # 电流型集成温度传感器：读取线性拟合参数
                        slope_val = ws_fit.cell(row=2, column=2).value
                        intercept_val = ws_fit.cell(row=3, column=2).value
                        r2_val = ws_fit.cell(row=4, column=2).value
                        
                        if slope_val and slope_val != "未计算":
                            self.slope[exp_name] = float(slope_val)
                            self.intercept[exp_name] = float(intercept_val) if intercept_val and intercept_val != "未计算" else 0.0
                            self.r_value[exp_name] = float(r2_val) if r2_val and r2_val != "未计算" else 0.0
                            self.fit_show_state[exp_name] = True
                            formula_text = f"拟合公式: I = {float(slope_val):.4f} × t + {self.intercept[exp_name]:.4f}  (R² = {self.r_value[exp_name]:.6f})"
                            self.formula_label.config(text=formula_text)
                            self.update_plot(show_fit=True)
                        else:
                            self.slope[exp_name] = 0.0
                            self.r_value[exp_name] = 0.0
                            self.intercept[exp_name] = 0.0
                            self.fit_show_state[exp_name] = False
                            self.formula_label.config(text="拟合公式: 未计算")
                            self.update_plot(show_fit=False)
                    elif is_constant_current:
                        # 恒电流法（Pt100/NTC）：读取线性拟合参数
                        slope_val = ws_fit.cell(row=2, column=2).value
                        intercept_val = ws_fit.cell(row=3, column=2).value
                        r2_val = ws_fit.cell(row=4, column=2).value
                        
                        if slope_val and slope_val != "未计算":
                            self.slope[exp_name] = float(slope_val)
                            self.intercept[exp_name] = float(intercept_val) if intercept_val and intercept_val != "未计算" else 0.0
                            self.r_value[exp_name] = float(r2_val) if r2_val and r2_val != "未计算" else 0.0
                            self.fit_show_state[exp_name] = True
                            formula_text = f"拟合公式: Rt = {float(slope_val):.4f} × t + {self.intercept[exp_name]:.4f}  (R² = {self.r_value[exp_name]:.6f})"
                            self.formula_label.config(text=formula_text)
                            self.update_plot(show_fit=True)
                        else:
                            self.slope[exp_name] = 0.0
                            self.r_value[exp_name] = 0.0
                            self.intercept[exp_name] = 0.0
                            self.fit_show_state[exp_name] = False
                            self.formula_label.config(text="拟合公式: 未计算")
                            self.update_plot(show_fit=False)
                    else:
                        # Pt100电桥法：读取线性拟合参数
                        slope_val = ws_fit.cell(row=2, column=2).value
                        intercept_val = ws_fit.cell(row=3, column=2).value
                        r2_val = ws_fit.cell(row=4, column=2).value
                        
                        if slope_val and slope_val != "未计算":
                            self.slope[exp_name] = float(slope_val)
                            self.intercept[exp_name] = float(intercept_val) if intercept_val and intercept_val != "未计算" else 0.0
                            self.r_value[exp_name] = float(r2_val) if r2_val and r2_val != "未计算" else 0.0
                            self.fit_show_state[exp_name] = True
                            formula_text = f"拟合公式: Rt = {float(slope_val):.4f} × t + {self.intercept[exp_name]:.4f}  (R² = {self.r_value[exp_name]:.6f})"
                            self.formula_label.config(text=formula_text)
                            self.update_plot(show_fit=True)
                        else:
                            self.slope[exp_name] = 0.0
                            self.r_value[exp_name] = 0.0
                            self.intercept[exp_name] = 0.0
                            self.fit_show_state[exp_name] = False
                            self.formula_label.config(text="拟合公式: 未计算")
                            self.update_plot(show_fit=False)
                else:
                    # 没有拟合结果工作表
                    self.slope[exp_name] = 0.0
                    self.r_value[exp_name] = 0.0
                    self.intercept[exp_name] = 0.0
                    self.fit_show_state[exp_name] = False
                    self.formula_label.config(text="拟合公式: 未计算")
                    self.update_plot(show_fit=False)
                
                self.refresh_table_display()
                messagebox.showinfo("导入成功", f"数据已从以下文件导入到 {exp_name}：\n{file_path}")
                
            except ImportError:
                messagebox.showerror("错误", "请先安装openpyxl库：\npip install openpyxl")
            except Exception as e:
                messagebox.showerror("导入失败", f"导入时发生错误：{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = TempSensorApp(root)
    root.mainloop()