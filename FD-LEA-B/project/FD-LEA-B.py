import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
import math
import os
import sys
import numpy as np  # 用于线性回归计算
import pandas as pd  # 用于数据导出导入（可选安装）

def get_resource_path(relative_path):
    """获取资源的绝对路径，支持打包后的环境"""
    try:
        # 打包后的临时文件夹路径
        base_path = sys._MEIPASS
    except Exception:
        # 开发环境的路径
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, relative_path)

class SimpleImageRotator:
    def __init__(self, root):
        self.root = root
        self.root.title("FD-LEA-B 线膨胀系数测试实验仪")
        self.root.geometry("1600x840")  # 增大窗口尺寸以适应四个区域
        
        # 初始化变量
        self.original_image = None
        self.preview_image = None
        self.tk_image = None
        self.angle = 0.0
        self.dragging = False
        self.last_angle = 0.0
        self.center_x = 0
        self.center_y = 0
        
        # 内圈图片相关变量
        self.inner_original_image = None
        self.inner_preview_image = None
        self.inner_tk_image = None
        
        # 图片路径
        self.background_path = get_resource_path("background")
        self.outer_image_path = os.path.join(self.background_path, "千分表外圈.jpg")
        self.inner_image_path = os.path.join(self.background_path, "千分表内圈.jpg")
        self.experiment_instrument_path = os.path.join(self.background_path, "实验仪.jpg")
        self.experiment_device_path = os.path.join(self.background_path, "实验装置.jpg")
        
        # 创建界面
        self.setup_ui()
        
        # 等待窗口初始化后加载图片
        self.root.after(100, self.initialize_app)
        
        # 绑定事件
        self.bind_events()

        # 新增温度相关变量
        self.room_temp = 25.0  # 室温
        self.current_temp = 25.0  # 当前温度
        self.target_temp = 0.0  # 目标温度
        self.temp_mode = "power_off"  # 模式: power_off, set, control
        self.display_mode = "current"  # 显示模式: current, target
        self.timer_running = False  # 定时器是否运行
        self.power_on = False  # 电源状态

        # 长按相关变量
        self.long_press_running = False
        self.long_press_command = None

        self.red_light_id = None  # 红灯的canvas对象ID
        self.red_light_on = False  # 红灯状态
        self.red_light_blinking = False  # 是否在闪烁
        
        self.waiting_for_temp_set = False  # 新增：等待温度设定状态

        # 千分表读数相关变量
        self.reading_value = 0.0  # 当前读数（mm）
        self.large_needle_angle = 0.0  # 大针角度（度）
        self.small_needle_angle = 0.0  # 小针角度（度）
        self.needle_radius = 170  # 指针长度
        self.small_needle_radius = 40  # 小针长度

        # 材料热膨胀系数（单位：1/°C）
        self.material_expansion_coefficient = {
            "铝棒": 23.1e-6,  # 铝的热膨胀系数
            "铜棒": 16.5e-6,  # 铜的热膨胀系数
            "铁棒": 11.8e-6   # 铁的热膨胀系数
        }

        # 样品尺寸
        self.sample_diameter = 8.0  # 直径8mm
        self.sample_length = 400.0  # 长度400mm

        # 新增：样品放置状态
        self.sample_placed = False  # 样品是否已放置
        self.selected_sample = None  # 当前选择的样品（放置后才生效）
        self.NO_SAMPLE_READING = -0.04  # 无样品时的基础读数 -40um = -0.04mm

        # 基础读数（不考虑温度影响的原始读数）
        self.base_reading = 0.0

        # 添加误差相关变量
        import random
        self.error_percent = random.uniform(2.0, 3.0)  # 误差幅度在2%-3%之间
        self.error_direction = random.choice([-1, 1])  # 随机选择正误差或负误差
        self.error_factor = 1 + (self.error_direction * self.error_percent / 100)
        print(f"系统初始化: 误差方向 = {'+' if self.error_direction > 0 else '-'}{self.error_percent:.2f}%, 误差因子 = {self.error_factor:.4f}")

    def calculate_thermal_expansion(self):
        """计算当前温度下的热膨胀量（使用固定误差）"""
        # 如果没有放置样品，热膨胀量为0
        if not self.sample_placed:
            return 0.0
        
        # 获取当前样品的热膨胀系数
        material = self.sample_var.get()
        alpha = self.material_expansion_coefficient.get(material, 23.1e-6)
        
        # 计算相对于室温的长度变化量
        # ΔL = L₀ * α * ΔT
        delta_temp = self.current_temp - self.room_temp
        theoretical_delta = self.sample_length * alpha * delta_temp  # 理论值
        
        # 应用固定误差（只在升温时添加误差）
        if delta_temp > 0:
            delta_length = theoretical_delta * self.error_factor
            error_sign = '+' if self.error_direction > 0 else '-'
            print(f"热膨胀计算: 材料={material}, α={alpha:.2e}, ΔT={delta_temp:.1f}°C")
            print(f"  理论值: {theoretical_delta:.6f}mm, 固定误差: {error_sign}{self.error_percent:.2f}%, 实际值: {delta_length:.6f}mm")
        else:
            delta_length = theoretical_delta
            print(f"热膨胀计算: 材料={material}, α={alpha:.2e}, ΔT={delta_temp:.1f}°C, 无误差（未升温）")
        
        return delta_length

    def initialize_app(self):
        """初始化应用程序，在窗口创建后执行"""
        # 计算初始中心点
        self.update_center()
        # 加载图片
        self.load_image()
        # 显示图片
        self.update_display_picture()
        # 加载实验仪和实验装置图片
        self.load_experiment_images()
        
        # 设置初始状态为无样品，固定-40um读数
        self.NO_SAMPLE_READING = -0.04  # -40um = -0.04mm
        self.base_reading = self.NO_SAMPLE_READING
        self.sample_placed = False
        self.selected_sample = None
        print(f"初始化: 无样品, 基础读数: {self.base_reading:.3f} mm (-40um)")
        
        # 更新样品标签
        self.sample_label.config(text="无样品")
        
        # 更新千分表显示
        self.update_reading_from_value()

    def setup_ui(self):
        # 设置背景色
        self.root.configure(bg="#f0f0f0")
        
        # 创建主框架，使用grid布局
        main_frame = tk.Frame(self.root, bg="#f0f0f0")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 配置grid权重
        main_frame.grid_columnconfigure(0, weight=1)  # 左列权重1
        main_frame.grid_columnconfigure(1, weight=1)  # 右列权重1
        main_frame.grid_rowconfigure(0, weight=1)     # 上行权重1
        main_frame.grid_rowconfigure(1, weight=1)     # 下行权重1
        
        # 左上区域：千分表区域
        self.top_left_frame = tk.Frame(main_frame, bg="white", relief=tk.RAISED, borderwidth=1)
        self.top_left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=(0, 5))
        
        # 创建画布
        self.canvas = tk.Canvas(
            self.top_left_frame,
            bg="white",
            highlightthickness=1,
            highlightbackground="#cccccc"
        )
        self.canvas.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 右上区域：实验仪区域
        self.top_right_frame = tk.Frame(main_frame, bg="white", relief=tk.RAISED, borderwidth=1)
        self.top_right_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=(0, 5))
        
        # 左下区域：实验装置区域
        self.bottom_left_frame = tk.Frame(main_frame, bg="white", relief=tk.RAISED, borderwidth=1)
        self.bottom_left_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 5), pady=(5, 0))
        
        # 右下区域：数据记录区域
        self.bottom_right_frame = tk.Frame(main_frame, bg="white", relief=tk.RAISED, borderwidth=1)
        self.bottom_right_frame.grid(row=1, column=1, sticky="nsew", padx=(5, 0), pady=(5, 0))
        
        # 设置各区域标题
        self.setup_region_titles()
        # 初始化数据记录区域
        self.setup_data_record_region()

        main_frame.grid_rowconfigure(0, weight=1)     # 上行权重1
        main_frame.grid_rowconfigure(1, weight=1)     # 下行权重1
        
        # 配置grid权重 - 确保1:1比例
        # for i in range(2):
        main_frame.grid_rowconfigure(0, weight=5, uniform="row")  # 添加uniform参数
        main_frame.grid_columnconfigure(0, weight=5)
        main_frame.grid_rowconfigure(1, weight=6, uniform="row")  # 添加uniform参数
        main_frame.grid_columnconfigure(1, weight=6)

        # 绑定画布尺寸变化事件
        self.canvas.bind("<Configure>", self.on_canvas_configure)

    def setup_region_titles(self):
        """设置各区域标题"""
        # 左上区域标题
        top_left_title = tk.Label(self.top_left_frame, text="千分表区域", 
                                 bg="white", font=("Arial", 12, "bold"))
        top_left_title.place(x=10, y=5)
        
        # 右上区域标题
        top_right_title = tk.Label(self.top_right_frame, text="实验仪区域", 
                                  bg="white", font=("Arial", 12, "bold"))
        top_right_title.place(x=10, y=5)
        
        # 左下区域标题
        bottom_left_title = tk.Label(self.bottom_left_frame, text="实验装置区域", 
                                    bg="white", font=("Arial", 12, "bold"))
        bottom_left_title.place(x=10, y=5)
        
        # 右下区域标题
        bottom_right_title = tk.Label(self.bottom_right_frame, text="数据记录区域", 
                                     bg="white", font=("Arial", 12, "bold"))
        bottom_right_title.place(x=10, y=5)

    def setup_data_record_region(self):
        """设置数据记录区域"""
        # 创建主框架，分为左右两部分
        data_main_frame = tk.Frame(self.bottom_right_frame, bg="white")
        data_main_frame.pack(fill="both", expand=False, padx=5, pady=5)
        
        # 左侧区域（实验参数和表格）- 固定比例
        left_data_frame = tk.Frame(data_main_frame, bg="white", relief=tk.RAISED, borderwidth=1)
        left_data_frame.pack(side="left", fill="both", expand=True, padx=(0, 2), pady=5)
        
        # 右侧区域（曲线、计算结果和按钮）- 固定比例
        right_data_frame = tk.Frame(data_main_frame, bg="white", relief=tk.RAISED, borderwidth=1)
        right_data_frame.pack(side="right", fill="both", expand=True, padx=(2, 0), pady=5)
        
        # 设置两侧区域的比例（左侧稍宽）
        data_main_frame.grid_columnconfigure(0, weight=5)  # 左侧权重5
        data_main_frame.grid_columnconfigure(1, weight=4)  # 右侧权重4
        
        # 左侧区域设置（实验参数和表格）
        self.setup_left_data_region(left_data_frame)
        
        # 右侧区域设置（曲线、计算结果和按钮）
        self.setup_right_data_region(right_data_frame)

    def setup_left_data_region(self, parent_frame):
        """设置左侧数据区域（实验参数和表格）"""
        # 移除滚动条，直接使用Frame
        self.left_content_frame = tk.Frame(parent_frame, bg="white")
        self.left_content_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 实验参数部分
        self.setup_experiment_params_in_left()
        
        # 数据表格部分
        self.setup_data_table_in_left()

    def setup_experiment_params_in_left(self):
        """设置实验参数显示（放在左侧区域）"""
        params_frame = tk.LabelFrame(self.left_content_frame, text="实验参数", 
                                    bg="white", font=("Arial", 11, "bold"))
        params_frame.pack(fill="x", padx=10, pady=(5, 5))  # 减小顶部间距
        
        # 使用更紧凑的网格布局
        # 第一行：实验样品
        row1 = tk.Frame(params_frame, bg="white")
        row1.pack(fill="x", padx=5, pady=3)
        tk.Label(row1, text="实验样品：", bg="white", 
                font=("Arial", 10)).pack(side="left", padx=2)
        self.sample_label = tk.Label(row1, text="铝棒", bg="white", 
                                    font=("Arial", 10), fg="blue")
        self.sample_label.pack(side="left", padx=2)
        
        # 第二行：加热前传感器温度
        row2 = tk.Frame(params_frame, bg="white")
        row2.pack(fill="x", padx=5, pady=3)
        tk.Label(row2, text="加热前传感器温度：", bg="white", 
                font=("Arial", 10)).pack(side="left", padx=2)
        tk.Label(row2, text="25.0°C", bg="white", 
                font=("Arial", 10), fg="gray").pack(side="left", padx=2)
        
        # 第三行：样品长度
        row3 = tk.Frame(params_frame, bg="white")
        row3.pack(fill="x", padx=5, pady=3)
        tk.Label(row3, text="样品长度 l =", bg="white", 
                font=("Arial", 10)).pack(side="left", padx=2)
        tk.Label(row3, text="0.400 m", bg="white", 
                font=("Arial", 10), fg="gray").pack(side="left", padx=2)
        
        # 第四行：样品直径
        row4 = tk.Frame(params_frame, bg="white")
        row4.pack(fill="x", padx=5, pady=3)
        tk.Label(row4, text="样品直径 d =", bg="white", 
                font=("Arial", 10)).pack(side="left", padx=2)
        tk.Label(row4, text="8.0 mm", bg="white", 
                font=("Arial", 10), fg="gray").pack(side="left", padx=2)

    def setup_data_table_in_left(self):
        """设置数据输入表格（放在左侧区域）"""
        table_frame = tk.LabelFrame(self.left_content_frame, text="实验数据记录", 
                                bg="white", font=("Arial", 11, "bold"))
        table_frame.pack(fill="both", expand=False, padx=10, pady=5)
        
        # 表格标题
        headers = ["序号", "温度θ/°C", "Δl/10^-6 m"]
        
        # 创建表头框架
        header_frame = tk.Frame(table_frame, bg="white")
        header_frame.pack(fill="x", padx=5, pady=(5, 2))
        
        # 序号列标题（宽度较小）
        tk.Label(header_frame, text=headers[0], bg="white", width=5,
                font=("Arial", 9, "bold")).pack(side="left", padx=2)
        
        # 温度列标题
        tk.Label(header_frame, text=headers[1], bg="white", width=12,
                font=("Arial", 9, "bold")).pack(side="left", padx=2)
        
        # 长度变化列标题·
        tk.Label(header_frame, text=headers[2], bg="white", width=12,
                font=("Arial", 9, "bold")).pack(side="left", padx=2)
        
        # 创建表格内容框架（用于放置所有行）
        table_content_frame = tk.Frame(table_frame, bg="white")
        table_content_frame.pack(fill="both", expand=True, padx=5, pady=2)
        
        # 创建8行数据
        self.data_entries = []
        for i in range(8):  # 8组数据
            row_frame = tk.Frame(table_content_frame, bg="white")
            row_frame.pack(fill="x", pady=1)
            
            # 序号标签
            tk.Label(row_frame, text=f"{i+1}", bg="white", width=5, 
                    font=("Arial", 9)).pack(side="left", padx=2)
            
            # 温度输入框
            temp_var = tk.StringVar(value="")
            temp_entry = tk.Entry(row_frame, textvariable=temp_var, width=12, 
                                font=("Arial", 9), justify="center",
                                relief="solid", bd=1)
            temp_entry.pack(side="left", padx=2, pady=1)
            
            # 长度变化输入框
            delta_var = tk.StringVar(value="")
            delta_entry = tk.Entry(row_frame, textvariable=delta_var, width=12, 
                                font=("Arial", 9), justify="center",
                                relief="solid", bd=1)
            delta_entry.pack(side="left", padx=2, pady=1)
            
            # 存储数据条目
            self.data_entries.append({
                "temp": temp_var,
                "delta": delta_var,
                "temp_entry": temp_entry,
                "delta_entry": delta_entry
            })
        
        # 添加空白区域以便表格居中
        spacer_frame = tk.Frame(table_content_frame, bg="white", height=10)
        spacer_frame.pack(fill="x")

    def setup_right_data_region(self, parent_frame):
        """设置右侧数据区域（曲线、计算结果和按钮）"""
        # 顶部：绘图区域
        plot_frame = tk.LabelFrame(parent_frame, text="Δl-θ关系曲线", 
                                bg="white", font=("Arial", 11, "bold"))
        plot_frame.pack(fill="both", expand=False, padx=10, pady=(10, 5))
        
        # 创建绘图画布 - 保持较小高度
        self.plot_canvas = tk.Canvas(plot_frame, bg="white", height=200,width=160)
        self.plot_canvas.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 绑定画布尺寸变化事件
        self.plot_canvas.bind("<Configure>", self.on_plot_canvas_configure)
        
        # 初始绘制坐标轴
        self.draw_plot_axes()
        
        # 中间：计算结果部分
        self.setup_calculation_results_in_right(parent_frame)
        
        # 底部：按钮区域
        self.setup_control_buttons_in_right(parent_frame)

    def setup_calculation_results_in_right(self, parent_frame):
        """设置计算结果区域（放在右侧区域）"""
        results_frame = tk.LabelFrame(parent_frame, text="计算结果", 
                                    bg="white", font=("Arial", 11, "bold"))
        results_frame.pack(fill="x", padx=10, pady=5)
        
        # 创建内部框架以更好地控制布局
        inner_frame = tk.Frame(results_frame, bg="white")
        inner_frame.pack(fill="x", padx=5, pady=5)
        
        # 线膨胀系数计算值
        calc_frame = tk.Frame(inner_frame, bg="white")
        calc_frame.pack(fill="x", pady=3)
        tk.Label(calc_frame, text="线膨胀系数计算值 α_cal =", bg="white", 
                font=("Arial", 10)).pack(side="left", padx=5)
        self.calc_coeff_label = tk.Label(calc_frame, text="---", bg="white", 
                                        font=("Arial", 10), fg="blue")
        self.calc_coeff_label.pack(side="left", padx=2)
        tk.Label(calc_frame, text="×10^-6/°C", bg="white", 
                font=("Arial", 10)).pack(side="left")
        
        # 线膨胀系数理论值
        theory_frame = tk.Frame(inner_frame, bg="white")
        theory_frame.pack(fill="x", pady=3)
        tk.Label(theory_frame, text="线膨胀系数理论值 α_theory =", bg="white", 
                font=("Arial", 10)).pack(side="left", padx=5)
        self.theory_coeff_label = tk.Label(theory_frame, text="23.1", bg="white", 
                                        font=("Arial", 10), fg="green")
        self.theory_coeff_label.pack(side="left", padx=2)
        tk.Label(theory_frame, text="×10^-6/°C", bg="white", 
                font=("Arial", 10)).pack(side="left")
        
        # 百分误差
        error_frame = tk.Frame(inner_frame, bg="white")
        error_frame.pack(fill="x", pady=3)
        tk.Label(error_frame, text="百分误差 =", bg="white", 
                font=("Arial", 10)).pack(side="left", padx=5)
        self.error_label = tk.Label(error_frame, text="---", bg="white", 
                                font=("Arial", 10), fg="red")
        self.error_label.pack(side="left", padx=2)
        tk.Label(error_frame, text="%", bg="white", 
                font=("Arial", 10)).pack(side="left")

    def setup_control_buttons_in_right(self, parent_frame):
        """设置控制按钮（放在右侧区域）"""
        button_frame = tk.Frame(parent_frame, bg="white")
        button_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        # 计算按钮
        calc_btn = tk.Button(button_frame, text="计算", 
                            bg="#4CAF50", fg="white",
                            font=("Arial", 11),
                            width=6,
                            command=self.calculate_fit)
        calc_btn.pack(side="left", padx=5, pady=5)
        
        # 清空数据按钮 - 新增
        clear_btn = tk.Button(button_frame, text="清空数据", 
                            bg="#FF5722", fg="white",  # 橙色按钮
                            font=("Arial", 11),
                            width=6,
                            command=self.clear_data)
        clear_btn.pack(side="left", padx=5, pady=5)

        # 导出数据按钮
        export_btn = tk.Button(button_frame, text="导出数据", 
                            bg="#2196F3", fg="white",
                            font=("Arial", 11),
                            width=6,
                            command=self.export_data)
        export_btn.pack(side="left", padx=5, pady=5)
        
        # 导入数据按钮
        import_btn = tk.Button(button_frame, text="导入数据", 
                            bg="#FF9800", fg="white",
                            font=("Arial", 11),
                            width=6,
                            command=self.import_data)
        import_btn.pack(side="left", padx=5, pady=5)

    def clear_data(self):
        """清空表格数据和曲线"""
        # 添加确认对话框
        import tkinter.messagebox as messagebox
        response = messagebox.askyesno(
            "确认清空",
            "确定要清空所有实验数据、计算结果和曲线吗？\n此操作不可撤销。"
        )
        
        if not response:
            return  # 用户取消
        
        # 清空所有数据输入框
        for entry in self.data_entries:
            entry["temp"].set("")  # 清空温度数据
            entry["delta"].set("")  # 清空长度变化数据
        
        # 清空计算结果
        self.calc_coeff_label.config(text="---")
        self.error_label.config(text="---")
        
        # 重绘坐标轴（清除拟合曲线）
        self.draw_plot_axes()
        
        # 输出提示信息
        print("已清空所有实验数据、计算结果和曲线")
        
        # 显示成功提示
        messagebox.showinfo("操作成功", "数据已成功清空")

    def on_plot_canvas_configure(self, event):
        """绘图画布尺寸变化时重绘坐标轴"""
        if event.width > 100 and event.height > 100:  # 确保有足够的绘制空间
            self.draw_plot_axes()

    def draw_plot_axes(self):
        """绘制坐标轴"""
        self.plot_canvas.delete("all")
        
        width = self.plot_canvas.winfo_width() or 250
        height = self.plot_canvas.winfo_height() or 120
        
        # 如果画布太小，暂时不绘制
        if width < 100 or height < 100:
            return
        
        # 绘制坐标轴
        margin = 50  # 增加边距以确保标签可见
        plot_width = width - 2 * margin
        plot_height = height - 2 * margin
        
        # 坐标轴
        self.plot_canvas.create_line(margin, height - margin, 
                                    width - margin, height - margin, width=2)  # X轴
        self.plot_canvas.create_line(margin, margin, 
                                    margin, height - margin, width=2)  # Y轴
        
        # 坐标轴标签 - 调整位置使其完全可见
        self.plot_canvas.create_text(
            width // 2, height - margin + 25,  # 向下移动一点
            text="温度 θ/°C", 
            font=("Arial", 10, "bold")
        )
        self.plot_canvas.create_text(
            margin - 35,  # 向左移动一点
            height // 2, 
            text="Δl/10⁻⁶ m",  # 使用上标符号
            font=("Arial", 10, "bold"), 
            angle=90
        )
        
        # 网格和刻度
        for i in range(0, 11, 2):
            # X轴刻度
            x = margin + i * plot_width / 10
            self.plot_canvas.create_line(x, height - margin - 5, 
                                        x, height - margin + 5, width=1)
            self.plot_canvas.create_text(x, height - margin + 15, 
                                        text=str(i*10), font=("Arial", 8))
            
            # Y轴刻度
            y = height - margin - i * plot_height / 10
            self.plot_canvas.create_line(margin - 5, y, margin + 5, y, width=1)
            self.plot_canvas.create_text(margin - 20, y,  # 向左移动一点
                                        text=str(i*100), font=("Arial", 8))
        
        # 添加标题
        # self.plot_canvas.create_text(
        #     width // 2, margin // 2,
        #     text="Δl-θ关系曲线",
        #     font=("Arial", 11, "bold"),
        #     fill="black"
        # )

    # 删除原有的setup_data_record_region函数中的重复内容
    def setup_experiment_params(self):
        """保留函数签名，但内容已被替换"""
        pass

    def setup_data_table(self):
        """保留函数签名，但内容已被替换"""
        pass

    def setup_plot_area(self):
        """保留函数签名，但内容已被替换"""
        pass

    def setup_calculation_results(self):
        """保留函数签名，但内容已被替换"""
        pass

    def setup_control_buttons(self):
        """保留函数签名，但内容已被替换"""
        pass

    def update_sample_label(self):
        """更新样品标签"""
        if self.sample_placed and self.selected_sample:
            self.sample_label.config(text=self.selected_sample)
            # 更新理论值
            coeff_dict = {
                "铝棒": "23.1",
                "铜棒": "16.5", 
                "铁棒": "11.8"
            }
            self.theory_coeff_label.config(text=coeff_dict.get(self.selected_sample, "23.1"))
        else:
            self.sample_label.config(text="无样品")
            self.theory_coeff_label.config(text="---")

    def calculate_fit(self):
        """计算拟合直线"""
        # 收集数据
        temps = []
        deltas = []
        
        for entry in self.data_entries:
            temp_str = entry["temp"].get().strip()
            delta_str = entry["delta"].get().strip()
            
            if temp_str and delta_str:
                try:
                    temp = float(temp_str)
                    delta = float(delta_str)
                    temps.append(temp)
                    deltas.append(delta)
                except ValueError:
                    continue
        
        if len(temps) < 2:
            print("数据点不足，无法计算拟合直线")
            return
        
        # 线性回归计算
        import numpy as np
        temps_arr = np.array(temps)
        deltas_arr = np.array(deltas)
        
        # 最小二乘法拟合
        A = np.vstack([temps_arr, np.ones(len(temps_arr))]).T
        slope, intercept = np.linalg.lstsq(A, deltas_arr, rcond=None)[0]
        
        # 计算相关系数
        correlation = np.corrcoef(temps_arr, deltas_arr)[0, 1]
        
        calc_coeff=slope/self.sample_length*1000
        # 更新计算结果
        self.calc_coeff_label.config(text=f"{calc_coeff:.2f}")
        
        # 计算误差
        try:
            theory_value = float(self.theory_coeff_label.cget("text"))
            error = abs((calc_coeff - theory_value) / theory_value) * 100
            self.error_label.config(text=f"{error:.2f}")
        except:
            self.error_label.config(text="---")
        
        # 绘制拟合曲线
        self.draw_fitted_line(temps_arr, deltas_arr, slope, intercept)
        
        # 显示拟合公式
        self.plot_canvas.create_text(
            self.plot_canvas.winfo_width() - 100, 30,
            text=f"Δl = {slope:.3f}θ + {intercept:.3f}",
            font=("Arial", 9, "bold"),
            fill="red",
            anchor="ne"
        )
        self.plot_canvas.create_text(
            self.plot_canvas.winfo_width() - 100, 50,
            text=f"R² = {correlation**2:.1f}",
            font=("Arial", 9),
            fill="red",
            anchor="ne"
        )
        
        print(f"拟合结果: 斜率={slope:.3f}, 截距={intercept:.3f}, R²={correlation**2:.1f}")

    def draw_fitted_line(self, temps, deltas, slope, intercept):
        """绘制拟合直线和数据点"""
        self.draw_plot_axes()  # 重绘坐标轴
        
        width = self.plot_canvas.winfo_width() or 400
        height = self.plot_canvas.winfo_height() or 250
        margin = 50
        plot_width = width - 2 * margin
        plot_height = height - 2 * margin
        
        # 绘制数据点
        for temp, delta in zip(temps, deltas):
            x = margin + (temp / 100) * plot_width
            y = height - margin - (delta / 1000) * plot_height
            
            self.plot_canvas.create_oval(x-3, y-3, x+3, y+3, fill="blue", outline="blue")
        
        # 绘制拟合直线
        if len(temps) > 0:
            x1 = margin
            y1 = height - margin - ((slope * 0 + intercept) / 1000) * plot_height
            
            x2 = width - margin
            y2 = height - margin - ((slope * 100 + intercept) / 1000) * plot_height
            
            self.plot_canvas.create_line(x1, y1, x2, y2, fill="red", width=2, dash=(4, 2))

    def export_data(self):
        """导出数据到Excel"""
        try:
            from tkinter import filedialog
            import pandas as pd
            import openpyxl
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as ExcelImage
            from PIL import ImageGrab
            import io
            
            # 选择保存路径
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            if not file_path:
                return
            
            # 1. 先导出数据
            data_list = []
            for i, entry in enumerate(self.data_entries, 1):
                temp = entry["temp"].get().strip() or ""
                delta = entry["delta"].get().strip() or ""
                data_list.append([i, temp, delta])
            
            # 创建Excel写入器
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 实验参数
                params_data = {
                    "参数": ["实验样品", "加热前传感器温度", "样品长度", "样品直径", 
                        "线膨胀系数计算值", "线膨胀系数理论值", "百分误差"],
                    "值": [
                        self.sample_label.cget("text"),
                        "25.0",
                        "0.400",
                        "8.0",
                        self.calc_coeff_label.cget("text"),
                        self.theory_coeff_label.cget("text"),
                        self.error_label.cget("text")
                    ],
                    "单位": [
                        "", "°C", "m", "mm", "×10^-6/°C", "×10^-6/°C", "%"
                    ]
                }
                params_df = pd.DataFrame(params_data)
                params_df.to_excel(writer, sheet_name='实验参数', index=False)
                
                # 表格数据
                data_df = pd.DataFrame(data_list, columns=["序号", "温度θ/°C", "Δl/10^-6 m"])
                data_df.to_excel(writer, sheet_name='实验数据', index=False)
                
                # 创建一个空的工作表用于图表
                pd.DataFrame().to_excel(writer, sheet_name='图表', index=False)
            
            # 2. 截取当前曲线图像并添加到Excel
            self.add_plot_screenshot_to_excel(file_path)
            
            print(f"数据和图像已导出到: {file_path}")
            
        except Exception as e:
            print(f"导出数据错误: {e}")

    def add_plot_screenshot_to_excel(self, excel_path):
        """截取曲线图像并添加到Excel"""
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as ExcelImage
            from PIL import ImageGrab
            import tempfile
            import os
            
            # 截取绘图区域的屏幕截图
            x = self.plot_canvas.winfo_rootx()
            y = self.plot_canvas.winfo_rooty()
            width = self.plot_canvas.winfo_width()
            height = self.plot_canvas.winfo_height()
            
            if width <= 10 or height <= 10:
                print("绘图区域太小，无法截取图像")
                return
            
            # 截取图像
            screenshot = ImageGrab.grab(bbox=(x, y, x + width, y + height))
            
            # 保存到临时文件
            temp_dir = tempfile.gettempdir()
            temp_image_path = os.path.join(temp_dir, "plot_screenshot.png")
            screenshot.save(temp_image_path, 'PNG')
            
            # 加载工作簿并添加图像
            wb = load_workbook(excel_path)
            
            # 获取或创建图表工作表
            if '图表' in wb.sheetnames:
                ws = wb['图表']
            else:
                ws = wb.create_sheet('图表')
            
            # 清除工作表中的所有内容
            ws.delete_rows(1, ws.max_row)
            ws.delete_cols(1, ws.max_column)
            
            # 添加图像
            img = ExcelImage(temp_image_path)
            
            # 调整图像大小
            scale_factor = 0.7  # 缩放因子
            img.width = int(img.width * scale_factor)
            img.height = int(img.height * scale_factor)
            
            # 将图像添加到A1单元格
            ws.add_image(img, 'A1')
            
            # 保存工作簿
            wb.save(excel_path)
            
            # 删除临时文件
            os.remove(temp_image_path)
            
            print("曲线图像已添加到Excel")
            
        except Exception as e:
            print(f"添加截图到Excel错误: {e}")

    def import_data(self):
        """从Excel导入数据"""
        try:
            from tkinter import filedialog
            import pandas as pd
            import numpy as np
            
            # 选择文件
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
            )
            
            if not file_path:
                return
            
            # 读取数据
            data_df = pd.read_excel(file_path, sheet_name='实验数据')
            
            # 填充表格，处理nan值
            for i, entry in enumerate(self.data_entries):
                if i < len(data_df):
                    # 处理温度数据
                    if "温度θ/°C" in data_df.columns:
                        temp_val = data_df.iloc[i]["温度θ/°C"]
                        # 检查是否为nan
                        if pd.isna(temp_val) or temp_val is np.nan or temp_val == "":
                            temp_val = ""
                        else:
                            # 转换为字符串，移除可能的尾随0
                            temp_val = str(float(temp_val)).rstrip('0').rstrip('.') if '.' in str(temp_val) else str(temp_val)
                    else:
                        temp_val = ""
                    
                    # 处理长度变化数据
                    if "Δl/10^-6 m" in data_df.columns:
                        delta_val = data_df.iloc[i]["Δl/10^-6 m"]
                        # 检查是否为nan
                        if pd.isna(delta_val) or delta_val is np.nan or delta_val == "":
                            delta_val = ""
                        else:
                            # 转换为字符串，移除可能的尾随0
                            delta_val = str(float(delta_val)).rstrip('0').rstrip('.') if '.' in str(delta_val) else str(delta_val)
                    else:
                        delta_val = ""
                    
                    # 设置到输入框
                    entry["temp"].set(temp_val)
                    entry["delta"].set(delta_val)
                else:
                    # 如果数据行数少于表格行数，清空多余行
                    entry["temp"].set("")
                    entry["delta"].set("")
            
            print(f"数据已从 {file_path} 导入")
            
            # 导入后自动重绘坐标轴（清除可能存在的旧曲线）
            self.draw_plot_axes()
            
            # # 可选：显示导入成功消息
            # self.show_import_success()
            
        except Exception as e:
            print(f"导入数据错误: {e}")
            # 显示错误消息
            import tkinter.messagebox as messagebox
            # messagebox.showerror("导入错误", f"导入数据时发生错误:\n{str(e)}")

    def show_import_success(self):
        """显示导入成功提示"""
        # 在绘图区域显示临时提示
        width = self.plot_canvas.winfo_width() or 250
        height = self.plot_canvas.winfo_height() or 150
        
        if width > 100 and height > 100:
            self.plot_canvas.create_text(
                width // 2, height // 2,
                text="数据导入成功",
                font=("Arial", 12, "bold"),
                fill="blue",
                tags="import_message"
            )
            
            # 2秒后移除提示
            self.root.after(2000, lambda: self.plot_canvas.delete("import_message"))

    def load_experiment_images(self):
        """加载实验仪和实验装置图片"""
        # 加载实验仪图片
        if os.path.exists(self.experiment_instrument_path):
            try:
                # 创建实验仪画布
                self.instrument_canvas = tk.Canvas(self.top_right_frame, bg="white", 
                                                  highlightthickness=0)
                self.instrument_canvas.pack(fill=tk.BOTH, expand=True)
                
                # 加载图片
                instrument_img = Image.open(self.experiment_instrument_path)
                instrument_img = instrument_img.resize((900, 380), Image.Resampling.LANCZOS)
                self.instrument_photo = ImageTk.PhotoImage(instrument_img)
                
                # 显示图片
                self.instrument_canvas.create_image(456, 191, image=self.instrument_photo)
                
                # 在图片上添加按钮和温度显示框
                self.setup_instrument_controls()
                
            except Exception as e:
                print(f"加载实验仪图片错误: {e}")
                self.create_default_instrument()
        else:
            self.create_default_instrument()
            
        # 加载实验装置图片
        if os.path.exists(self.experiment_device_path):
            try:
                # 创建实验装置画布
                self.device_canvas = tk.Canvas(self.bottom_left_frame, bg="white", 
                                              highlightthickness=0)
                self.device_canvas.pack(fill=tk.BOTH, expand=True)
                
                # 加载图片
                device_img = Image.open(self.experiment_device_path)
                device_img = device_img.resize((600, 220), Image.Resampling.LANCZOS)
                self.device_photo = ImageTk.PhotoImage(device_img)
                
                # 显示图片
                self.device_canvas.create_image(330, 155, image=self.device_photo)
                
                # 在图片右侧添加样品选择选项卡
                self.setup_sample_selection()
                
            except Exception as e:
                print(f"加载实验装置图片错误: {e}")
                self.create_default_device()
        else:
            self.create_default_device()

    def setup_instrument_controls(self):
        """设置实验仪控制按钮和温度显示"""
        # 直接使用画布添加控件，而不是用白色框架覆盖
        self.instrument_canvas.create_text(50, 20, text="实验仪控制", 
                                        font=("Arial", 12, "bold"), 
                                        fill="black", anchor="w")
        
        # 修改按钮配置
        buttons_config = [
            {"text": "复位", "x": 100-5+1, "y": 250-5-7, "size": 30, "is_round": True, "font_size": 9, "command": "reset"},
            {"text": "升温", "x": 100+45+2+1, "y": 250-5-7, "size": 30, "is_round": True, "font_size": 9, "command": "increase"},
            {"text": "降温", "x": 100+45+52+2+1, "y": 250-5-7, "size": 30, "is_round": True, "font_size": 9, "command": "decrease"},
            {"text": "确定", "x": 100+45+105+2+1, "y": 250-5-7, "size": 30, "is_round": True, "font_size": 9, "command": "confirm"},
            {"text": "I", "x": 801, "y": 238-7, "width": 3, "height": 1, "is_round": False, "font_size": 12, "command": "power_on"},  # I = 打开电源
            {"text": "O", "x": 801, "y": 273-7, "width": 3, "height": 1, "is_round": False, "font_size": 12, "command": "power_off"},  # O = 关闭电源
        ]

        # 在确定按钮右侧添加红灯
        red_light_x = 100+45+105+2 + 97+1  # 确定按钮x坐标 + 40像素
        red_light_y = 250-5 + 15-7  # 确定按钮y坐标 + 15像素（居中）
        
        # 创建红灯（初始为灰色，表示关闭）
        self.red_light_id = self.instrument_canvas.create_oval(
            red_light_x - 12, red_light_y - 12,
            red_light_x + 12, red_light_y + 12,
            fill="gray", outline="darkgray", width=1
        )
        
        # # 添加"红"字标签
        # self.instrument_canvas.create_text(
        #     red_light_x, red_light_y + 25,
        #     text="红",
        #     fill="black",
        #     font=("Arial", 8)
        # )

        self.instrument_buttons = {}
        for config in buttons_config:
            if config["is_round"]:
                # 创建圆形按钮
                btn = tk.Canvas(self.instrument_canvas, 
                            width=config["size"], 
                            height=config["size"],
                            highlightthickness=0,
                            bg='#D6D9DE')
                
                # 绘制圆形按钮
                btn.create_oval(2, 2, config["size"]-2, config["size"]-2, 
                            fill="#545454", outline="#545454", width=4)
                
                # 添加按钮文字
                btn.create_text(config["size"]//2, config["size"]//2,
                            text=config["text"],
                            fill="#545454",
                            font=("Arial", config["font_size"], "bold"))
                
                # 绑定点击事件
                command = config.get("command", "")
                if command:
                    def make_click_handler(cmd):
                        return lambda event: self.handle_button_command(cmd)
                    btn.bind("<Button-1>", make_click_handler(command))
                
                # 为升温降温按钮添加长按支持
                if command in ["increase", "decrease"]:
                    def make_press_handler(cmd):
                        return lambda event: self.start_long_press(cmd)
                    def make_release_handler():
                        return lambda event: self.stop_long_press()
                    btn.bind("<ButtonPress-1>", make_press_handler(command))
                    btn.bind("<ButtonRelease-1>", make_release_handler())
                
                # 悬停效果
                def on_enter(event, canvas):
                    canvas.itemconfig(1, fill="#545454")
                
                def on_leave(event, canvas):
                    canvas.itemconfig(1, fill="#545454")
                
                btn.bind("<Enter>", lambda e, c=btn: on_enter(e, c))
                btn.bind("<Leave>", lambda e, c=btn: on_leave(e, c))
                
                btn_window = self.instrument_canvas.create_window(
                    config["x"], config["y"], 
                    window=btn, 
                    anchor="nw"
                )
                
            else:
                # 创建矩形按钮（电源按钮）
                command = config.get("command", "")
                btn = tk.Button(self.instrument_canvas, text=config["text"], 
                    width=config["width"], height=config["height"],
                    bg="#DC6E88", fg="black",
                    font=("Arial", config["font_size"]),
                    relief="flat",
                    bd=0,
                    highlightthickness=0,
                    activebackground="#DC6E88",
                    activeforeground="black",
                    cursor="arrow")
                
                if command:
                    if command == "power_on":  # I 按钮
                        btn.config(command=lambda: self.handle_button_command("power_on"))
                    elif command == "power_off":  # O 按钮
                        btn.config(command=lambda: self.handle_button_command("power_off"))
                
                    btn_window = self.instrument_canvas.create_window(
                            config["x"], config["y"], 
                            window=btn, 
                            anchor="nw"
                )
            
            self.instrument_buttons[config["text"]] = btn
            
        # 温度显示框 - 也直接在画布上显示
        # # 先创建一个半透明的背景框
        # temp_bg = self.instrument_canvas.create_rectangle(
        #     180, 160, 420, 250, 
        #     fill="white", stipple="gray50",  # 使用点状图案实现半透明效果
        #     width=0
        # )
        
        # 创建温度显示框
        temp_frame = tk.Frame(self.instrument_canvas, bg="#DC6E88", borderwidth=2)
        temp_window = self.instrument_canvas.create_window(175, 142-7, window=temp_frame, anchor="center", 
                                                        width=150, height=40)
        
        # temp_label = tk.Label(temp_frame, text="温度显示", 
        #                     bg="#f8f8f8", font=("Arial", 12, "bold"))
        # temp_label.pack(pady=5)
        
        # 修改温度显示框
        self.temp_display = tk.Label(temp_frame, text="",  # 初始为空
                                    bg="#DC6E88", font=("Arial", 30, "bold"),
                                    fg="black")
        self.temp_display.pack()
        
        # 启动温度更新定时器
        self.start_temp_timer()

    def update_red_light(self):
        """更新红灯状态"""
        if not self.red_light_id:
            return
        
        if self.temp_mode == "control" and self.current_temp < self.target_temp:
            # 控温模式且当前温度小于目标温度时，红灯闪烁
            if not self.red_light_blinking:
                self.red_light_blinking = True
                self.start_red_light_blink()
        # else:
        #     # 其他情况下红灯关闭
        #     self.red_light_blinking = False
        #     self.instrument_canvas.itemconfig(self.red_light_id, fill="gray")

    def start_red_light_blink(self):
        """开始红灯闪烁"""
        if not self.red_light_blinking or not self.red_light_id:
            return
        
        # 切换红灯状态
        if self.red_light_on:
            self.instrument_canvas.itemconfig(self.red_light_id, fill="gray")
            self.red_light_on = False
        else:
            self.instrument_canvas.itemconfig(self.red_light_id, fill="red")
            self.red_light_on = True
        
        # 如果还在闪烁状态，继续闪烁（0.5秒间隔）
        if self.red_light_blinking:
            self.root.after(500, self.start_red_light_blink)

    def stop_red_light_blink(self):
        """停止红灯闪烁"""
        self.red_light_blinking = False
        if self.red_light_id:
            self.instrument_canvas.itemconfig(self.red_light_id, fill="gray")

    def handle_button_command(self, command):
        """处理按钮命令"""
        if command == "power_on":
            if self.power_on:
                return  # 电源已经打开，无需重复操作
            # 打开电源
            self.power_on = True
            print("电源已打开")
            self.reset_temperature()  # 打开电源自动触发复位
        elif command == "power_off":
            # 关闭电源
            self.power_off()
            print("电源已关闭")
        elif not self.power_on and command not in ["power_on", "reset"]:
            # 电源关闭时，只有打开电源和复位按钮可用
            print("电源关闭，请先打开电源")
            return
        
        # 其他按钮处理逻辑
        if command == "reset":
            if self.power_on:  # 只有在电源打开时才能复位
                # 如果是控温模式，先停止温度更新
                if self.temp_mode == "control":
                    self.timer_running = True
                self.reset_temperature()
                print("已执行复位操作")
            else:
                print("电源关闭，无法复位")
        elif command == "increase":
            if self.temp_mode == "set":  # 在设定模式下有效
                self.target_temp = min(self.target_temp + 1, 80)
                self.update_display()
                print(f"升温: 目标温度={self.target_temp}")
            elif self.waiting_for_temp_set:  # 新增：在等待状态下点击升温
                self.target_temp = 1.0  # 设置初始值
                self.temp_mode = "set"  # 进入温度设定模式
                self.waiting_for_temp_set = False  # 退出等待状态
                self.update_display()
                print(f"进入温度设定模式: 目标温度={self.target_temp}")
            else:
                print("升温按钮仅在温度设定模式或等待状态下有效")

        elif command == "decrease":
            if self.temp_mode == "set":  # 在设定模式下有效
                self.target_temp = max(self.target_temp - 1, 0)
                self.update_display()
                print(f"降温: 目标温度={self.target_temp}")
            elif self.waiting_for_temp_set:  # 新增：在等待状态下点击降温
                self.target_temp = 0.0  # 设置初始值
                self.temp_mode = "set"  # 进入温度设定模式
                self.waiting_for_temp_set = False  # 退出等待状态
                self.update_display()
                print(f"进入温度设定模式: 目标温度={self.target_temp}")
            else:
                print("降温按钮仅在温度设定模式或等待状态下有效")
        elif command == "confirm":
            if self.temp_mode == "set":
                # 从设定模式切换到控温模式
                self.temp_mode = "control"
                self.display_mode = "current"
                print(f"进入控温模式: 目标温度={self.target_temp}, 当前温度={self.current_temp}")
                self.update_display()
                # 检查是否需要闪烁红灯
                self.update_red_light()
            elif self.temp_mode == "control":
                # 在控温模式下切换显示
                self.display_mode = "target" if self.display_mode == "current" else "current"
                self.update_display()
                print(f"切换显示模式: {self.display_mode}")

    def reset_temperature(self):
        """复位温度"""
        # 先停止温度更新定时器
        self.timer_running = True  # 设置为True以阻止正在进行的温度更新
        
        # 重置温度相关变量
        self.target_temp = 0.0  # 重置目标温度为0
        self.display_mode = "target"  # 显示目标温度
        # self.current_temp = self.room_temp  # 当前温度重置为室温
        self.waiting_for_temp_set = False  # 退出等待状态
        self.temp_mode = "power_off"  # 重置为电源关闭模式
        self.long_press_running = False  # 停止长按
        
        # 重置基础读数（新增）
        # import random
        # self.base_reading = random.uniform(-0.1, 1.2)
        # print(f"复位: 重置基础读数为 {self.base_reading:.3f} mm")
        
        # 停止红灯闪烁
        self.stop_red_light_blink()
        
        # 清除之前的显示流程定时器
        if hasattr(self, '_reset_timer_id'):
            self.root.after_cancel(self._reset_timer_id)
        
        # 显示FdHC
        self.temp_display.config(text="FdHC")
        
        # 更新千分表显示（新增）
        self.update_reading_from_value()
        
        # 延迟1秒后显示 A25.0
        self._reset_timer_id = self.root.after(1000, self.show_temp_setting)

    def show_temp_setting(self):
        """显示温度设定界面"""
        if not self.power_on:
            return
        # 显示 "A"+当前温度
        self.temp_display.config(text=f"A{self.current_temp:.1f}")
        # 延迟1秒后显示 b==.=
        self._reset_timer_id = self.root.after(1000, self.show_b_setting)

    def show_b_setting(self):
        if not self.power_on:
            return
        self.temp_display.config(text="b0.0")
        # 延迟1秒后显示 b==.=
        self._reset_timer_id = self.root.after(1000, self.show_b_equal)

    def show_b_equal(self):
        """显示b==.=""" 
        if not self.power_on:
            return
        self.temp_display.config(text="b==.=")
        # 延迟后进入等待温度设定状态
        self._reset_timer_id = self.root.after(1000, self.enter_waiting_for_temp_set)
    
    def enter_waiting_for_temp_set(self):
        """进入等待温度设定状态"""
        if not self.power_on:
            return
        self.waiting_for_temp_set = True
        self.temp_mode = "waiting"  # 新增模式
        self.timer_running = False  # 重新启用温度更新
        print("已进入等待温度设定状态，请点击升温或降温按钮开始设定")
        # 在等待状态下显示 b==.=
        self.temp_display.config(text="b==.=")

    def finish_temp_setup(self):
        """完成温度设置流程"""
        pass

    def update_display(self):
        """更新温度显示"""
        if not self.power_on:
            self.temp_display.config(text="")
            return
        
        # 如果在复位流程中（temp_mode为power_off），不更新显示
        if self.temp_mode == "power_off":
            return
        
        if self.temp_mode == "waiting":
            # 等待状态下显示 b==.=
            self.temp_display.config(text="b==.=")
        elif self.temp_mode == "set":
            # 设定模式下显示目标温度
            self.temp_display.config(text=f"b{self.target_temp:.1f}")
        elif self.temp_mode == "control":
            # 控温模式下根据显示模式显示
            if self.display_mode == "current":
                self.temp_display.config(text=f"A{self.current_temp:.1f}")
            else:
                self.temp_display.config(text=f"b{self.target_temp:.1f}")
        else:
            # 其他模式
            self.temp_display.config(text="")

    def power_off(self):
        """关闭电源"""
        self.power_on = False
        self.temp_mode = "power_off"
        self.temp_display.config(text="")
        self.stop_red_light_blink()  # 停止红灯闪烁
        print("电源已关闭")

    def reset_temperature(self):
        """复位温度"""
        # 先停止温度更新定时器
        self.timer_running = True  # 设置为True以阻止正在进行的温度更新
        
        # 重置温度相关变量
        self.target_temp = 0.0  # 重置目标温度为0
        self.display_mode = "target"  # 显示目标温度
        # self.current_temp = self.room_temp  # 注释掉，不让温度立刻重置
        self.waiting_for_temp_set = False  # 退出等待状态
        self.temp_mode = "power_off"  # 重置为电源关闭模式
        self.long_press_running = False  # 停止长按
        
        # 复位时停止红灯闪烁
        self.stop_red_light_blink()
        
        # 清除之前的显示流程定时器
        if hasattr(self, '_reset_timer_id'):
            self.root.after_cancel(self._reset_timer_id)
        
        # 显示FdHC
        self.temp_display.config(text="FdHC")
        
        # 更新千分表显示
        self.update_reading_from_value()
        
        # 延迟1秒后显示 A当前温度
        self._reset_timer_id = self.root.after(1000, self.show_temp_setting)

    def update_temperature(self):
        """更新温度逻辑"""
        if not self.power_on:
            # 电源关闭时，重置定时器并返回
            # self.timer_running = False
            self.stop_red_light_blink()  # 停止红灯闪烁
            # self.root.after(1000, self.update_temperature)
            # return
        
        # 如果正在复位流程中（temp_mode为power_off），执行降温逻辑
        if self.temp_mode == "power_off":
            # 在复位模式下，让温度缓慢下降到室温
            if self.current_temp > self.room_temp:
                self.current_temp = max(self.current_temp - 0.1, self.room_temp)  # 每次降0.1°C
                print(f"复位降温中: 当前温度={self.current_temp:.1f}°C")
                self.update_display()  # 更新显示
                self.update_reading_from_value()  # 更新千分表读数
            elif self.current_temp < self.room_temp:
                self.current_temp = min(self.current_temp + 0.1, self.room_temp)  # 每次升0.1°C
                print(f"复位升温中: 当前温度={self.current_temp:.1f}°C")
                self.update_display()
                self.update_reading_from_value()
            
            # 停止红灯闪烁
            self.stop_red_light_blink()
            
            self.timer_running = False
            self.root.after(1000, self.update_temperature)
            return
        
        if self.timer_running:
            # 避免重复执行
            return
        
        self.timer_running = True
        
        print(f"更新温度: 模式={self.temp_mode}, 当前={self.current_temp:.1f}, 目标={self.target_temp:.1f}")
        
        if self.temp_mode == "control":
            # 控温模式 - 根据温度差动态调整变化速度
            temp_diff = abs(self.target_temp - self.current_temp)
            
            if self.current_temp < self.target_temp:
                # 升温逻辑
                # 根据温度差计算升温步长
                if temp_diff >= 5.0:
                    step = 0.5  # 温差大时快速升温
                elif temp_diff >= 2.0:
                    step = 0.2  # 温差中等时中速升温
                elif temp_diff >= 1.0:
                    step = 0.1  # 温差较小时慢速升温
                elif temp_diff >= 0.5:
                    step = 0.05  # 温差很小时很慢升温
                else:
                    step = 0.02  # 温差极小时极慢升温
                
                self.current_temp = min(self.current_temp + step, self.target_temp)
                print(f"升温中: 当前温度={self.current_temp:.2f}°C, 温差={temp_diff:.2f}°C, 步长={step:.3f}°C")
                self.update_display()
                # 更新红灯状态（如果还没到目标温度，保持闪烁）
                self.update_red_light()
                
            elif self.current_temp > self.target_temp:
                # 降温逻辑，但不能低于室温
                # 根据温度差计算降温步长
                if temp_diff >= 5.0:
                    step = 0.5  # 温差大时快速降温
                elif temp_diff >= 2.0:
                    step = 0.2  # 温差中等时中速降温
                elif temp_diff >= 1.0:
                    step = 0.1  # 温差较小时慢速降温
                elif temp_diff >= 0.5:
                    step = 0.05  # 温差很小时很慢降温
                else:
                    step = 0.02  # 温差极小时极慢降温
                
                new_temp = max(self.current_temp - step, max(self.target_temp, self.room_temp))
                self.current_temp = new_temp
                print(f"降温中: 当前温度={self.current_temp:.2f}°C, 温差={temp_diff:.2f}°C, 步长={step:.3f}°C")
                self.update_display()
                # 温度下降时，红灯保持闪烁（直到复位）
                # 注意：这里不调用 stop_red_light_blink()
            else:
                print(f"已达到目标温度: {self.current_temp:.1f}°C")
                # 达到目标温度后，红灯保持闪烁（直到复位）
                # 注意：这里不调用 stop_red_light_blink()
            
            # 在控温模式下，温度变化时更新千分表读数
            self.update_reading_from_value()
            
        elif self.temp_mode == "set" or self.temp_mode == "waiting":
            # 在设定或等待模式时，温度调整到室温
            if self.current_temp > self.room_temp:
                self.current_temp = max(self.current_temp - 0.1, self.room_temp)
                print(f"降温至室温: 当前温度={self.current_temp:.1f}°C")
                self.update_display()
            elif self.current_temp < self.room_temp:
                self.current_temp = min(self.current_temp + 0.1, self.room_temp)
                print(f"升温至室温: 当前温度={self.current_temp:.1f}°C")
                self.update_display()
            
            # 更新千分表读数
            self.update_reading_from_value()
            
            # 停止红灯闪烁
            self.stop_red_light_blink()
        
        self.timer_running = False
        # 每秒更新一次
        self.root.after(1000, self.update_temperature)

    def start_temp_timer(self):
        """启动温度更新定时器"""
        self.root.after(1000, self.update_temperature)

    def start_long_press(self, command):
        """开始长按"""
        if self.temp_mode == "set" or self.waiting_for_temp_set:
            self.long_press_running = True
            self.long_press_command = command
            self.do_long_press()

    def do_long_press(self):
        """执行长按操作"""
        if not self.long_press_running or (self.temp_mode != "set" and not self.waiting_for_temp_set):
            return
        
        # 如果在等待状态，先进入设定模式
        if self.waiting_for_temp_set:
            self.temp_mode = "set"
            self.waiting_for_temp_set = False
            print(f"通过长按进入温度设定模式: temp_mode={self.temp_mode}")
        
        if self.long_press_command == "increase":
            self.target_temp = min(self.target_temp + 1, 80)
        elif self.long_press_command == "decrease":
            self.target_temp = max(self.target_temp - 1, 0)
        
        self.update_display()
        # 每200ms执行一次
        if self.long_press_running:
            self.root.after(200, self.do_long_press)

    def stop_long_press(self):
        """停止长按"""
        self.long_press_running = False

    def setup_sample_selection(self):
        """设置样品选择按钮组（放在图片下方）"""
        # 创建样品选择框架（放在图片下方）
        sample_frame = tk.Frame(self.bottom_left_frame, bg="white")
        sample_frame.place(relx=0, rely=0.7, relwidth=1, relheight=0.3)  # 增加高度
        
        # 样品选择标签
        sample_label = tk.Label(sample_frame, text="选择样品:", 
                            bg="white", font=("Arial", 12, "bold"))
        sample_label.pack(pady=5)
        
        # 创建按钮组框架
        button_frame = tk.Frame(sample_frame, bg="white")
        button_frame.pack()
        
        # 创建样品选择变量
        self.sample_var = tk.StringVar(value="")
        
        # 创建三个样品选择按钮（三选一）
        samples = ["铝棒", "铜棒", "铁棒"]
        self.sample_buttons = {}
        
        for sample in samples:
            btn = tk.Radiobutton(button_frame, 
                            text=sample,
                            variable=self.sample_var,
                            value=sample,
                            bg="white",
                            font=("Arial", 11),
                            indicatoron=0,
                            selectcolor="#4a90e2",
                            width=4,
                            height=1,
                            state="normal",  # 初始为可用状态
                            command=lambda s=sample: self.on_sample_selected(s))
            btn.pack(side=tk.LEFT, padx=5)
            self.sample_buttons[sample] = btn
        
        # 新增：样品放置/取下按钮
        self.place_button = tk.Button(button_frame, text="放上样品",
                                    bg="#4CAF50", fg="white",
                                    font=("Arial", 11),
                                    width=6,
                                    command=self.toggle_sample_placement)
        self.place_button.pack(side=tk.LEFT, padx=10)
        
        # 量杆操作按钮
        rod_frame = tk.Frame(sample_frame, bg="white")
        rod_frame.pack(pady=10)
        
        # 压紧量杆按钮
        self.tighten_btn = tk.Button(rod_frame, text="压紧量杆", 
                            bg="#4CAF50", fg="white",
                            font=("Arial", 11),
                            state="disabled",  # 初始为禁用状态
                            command=self.tighten_rod)
        self.tighten_btn.pack(side=tk.LEFT, padx=5)
        
        # 放松量杆按钮
        self.loosen_btn = tk.Button(rod_frame, text="放松量杆", 
                            bg="#f44336", fg="white",
                            font=("Arial", 11),
                            state="disabled",  # 初始为禁用状态
                            command=self.loosen_rod)
        self.loosen_btn.pack(side=tk.LEFT, padx=5)
        
        # 初始设置为无样品状态
        self.base_reading = self.NO_SAMPLE_READING
        self.sample_label.config(text="无样品")
        self.update_reading_from_value()


    def tighten_rod(self):
        """压紧量杆 - 只修改基础读数"""
        if not self.sample_placed:
            print("无样品时无法操作量杆")
            return
        import random
        increase = random.uniform(0.05, 0.1)
        old_value = self.base_reading
        self.base_reading = self.clamp_reading(self.base_reading + increase)
        
        if self.base_reading == 1.2:
            print(f"压紧量杆: 基础读数增加 {increase:.3f} mm, 已达到上限 1.200 mm")
        else:
            print(f"压紧量杆: 基础读数增加 {increase:.3f} mm, 当前基础读数: {self.base_reading:.3f} mm")
        
        self.update_reading_from_value()

    def loosen_rod(self):
        """放松量杆 - 只修改基础读数"""
        if not self.sample_placed:
            print("无样品时无法操作量杆")
            return
        import random
        decrease = random.uniform(0.05, 0.1)
        old_value = self.base_reading
        self.base_reading = self.clamp_reading(self.base_reading - decrease)
        
        if self.base_reading == -0.1:
            print(f"放松量杆: 基础读数减少 {decrease:.3f} mm, 已达到下限 -0.100 mm")
        else:
            print(f"放松量杆: 基础读数减少 {decrease:.3f} mm, 当前基础读数: {self.base_reading:3f} mm")
        
        self.update_reading_from_value()

    def on_sample_selected(self, sample=None):
        """样品选择事件处理 - 只记录选择，不立即生效"""
        if sample is None:
            sample = self.sample_var.get()
        
        print(f"已选择样品: {sample}（尚未放置）")

    def toggle_sample_placement(self):
        """切换样品放置/取下状态"""
        if not self.sample_placed:
            # 放上样品
            selected = self.sample_var.get()
            if not selected:
                print("请先选择一个样品")
                return
            
            self.sample_placed = True
            self.selected_sample = selected
            self.place_button.config(text="取下样品", bg="#f44336")
            
            # 禁用样品选择按钮
            for btn in self.sample_buttons.values():
                btn.config(state="disabled")
            
            # 启用量杆操作按钮
            self.tighten_btn.config(state="normal")
            self.loosen_btn.config(state="normal")
            
            # 随机生成基础读数（-0.1到1.2mm）
            import random
            self.base_reading = random.uniform(-0.1, 1.2)
            
            # 更新样品标签和理论值
            self.sample_var.set(selected)  # 确保变量值正确
            self.update_sample_label()
            
            print(f"已放上样品: {selected}, 基础读数: {self.base_reading:.3f} mm")
        else:
            # 取下样品
            self.sample_placed = False
            self.selected_sample = None
            self.place_button.config(text="放上样品", bg="#4CAF50")
            
            # 启用样品选择按钮
            for btn in self.sample_buttons.values():
                btn.config(state="normal")
            
            # 禁用量杆操作按钮
            self.tighten_btn.config(state="disabled")
            self.loosen_btn.config(state="disabled")
            
            # 清空样品选择
            self.sample_var.set("")
            
            # 设置无样品基础读数（固定-40um）
            self.base_reading = self.NO_SAMPLE_READING
            
            # 更新样品标签为无样品状态
            self.sample_label.config(text="无样品")
            self.theory_coeff_label.config(text="---")  # 清空理论值
            
            print(f"已取下样品，基础读数恢复为: {self.base_reading:.3f} mm (-40um)")
        
        # 更新千分表显示
        self.update_reading_from_value()

    def create_default_instrument(self):
        """创建默认实验仪界面"""
        default_frame = tk.Frame(self.top_right_frame, bg="white")
        default_frame.pack(fill=tk.BOTH, expand=True)
        
        label = tk.Label(default_frame, text="实验仪区域\n(实验仪.jpg未找到)", 
                        bg="white", font=("Arial", 14), fg="#666666")
        label.pack(expand=True)
        
        # 仍然添加按钮和温度显示
        self.setup_instrument_controls()

    def create_default_device(self):
        """创建默认实验装置界面"""
        default_frame = tk.Frame(self.bottom_left_frame, bg="white")
        default_frame.pack(fill=tk.BOTH, expand=True)
        
        label = tk.Label(default_frame, text="实验装置区域\n(实验装置.jpg未找到)", 
                        bg="white", font=("Arial", 14), fg="#666666")
        label.pack(expand=True)
        
        # 仍然添加样品选择
        self.setup_sample_selection()

    def on_canvas_configure(self, event):
        """画布尺寸变化时更新中心点"""
        self.center_x = event.width // 2
        self.center_y = event.height // 2
        
        # 重新显示图片
        if self.preview_image:
            self.rotate_and_update()

    def update_center(self):
        """更新中心点坐标"""
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        
        if canvas_width <= 1 or canvas_height <= 1:
            # 如果画布还未初始化，使用预设值
            self.center_x = 200  # 调整为中心位置
            self.center_y = 200
        else:
            self.center_x = canvas_width // 2
            self.center_y = canvas_height // 2

    def load_image(self):
        """加载指定图片"""
        try:
            # 加载外圈图片（可旋转）
            if os.path.exists(self.outer_image_path):
                img = Image.open(self.outer_image_path)
                if img.mode != 'RGBA':
                    img = img.convert('RGBA')
                
                # 调整外圈图片透明度为0.5（128/255） - 新增部分
                datas = img.getdata()
                new_data = []
                for item in datas:
                    if item[3] > 50:  # 如果alpha值大于50
                        new_data.append((item[0], item[1], item[2], 150))  # 设置为半透明(128 = 0.5透明度)
                    else:
                        new_data.append(item)  # 保持透明
                
                img.putdata(new_data)
                self.original_image = img
                self.create_preview()
            else:
                self.create_default_outer_image()
            
            # 加载内圈图片（固定位置）
            if os.path.exists(self.inner_image_path):
                inner_img = Image.open(self.inner_image_path)
                if inner_img.mode != 'RGBA':
                    inner_img = inner_img.convert('RGBA')
                
                # 内圈图片保持不透明（或者也可以设置为不同的透明度）
                self.inner_original_image = inner_img
                self.create_inner_preview()
            else:
                self.create_default_inner_image()
                
        except Exception as e:
            self.create_default_outer_image()
            self.create_default_inner_image()

    def create_preview(self):
        """创建外圈图片预览图"""
        if self.original_image is None:
            return
        
        width, height = self.original_image.size
        max_size = 350  # 调整大小以适应新布局
        
        # 如果图片太大，进行缩放
        if width > max_size or height > max_size:
            # 计算缩放比例，保持宽高比
            scale = min(max_size / width, max_size / height)
            new_width = int(width * scale)
            new_height = int(height * scale)
            
            # 缩放图片
            self.preview_image = self.original_image.resize(
                (new_width, new_height), 
                Image.Resampling.LANCZOS
            )
            print(f"外圈图片缩放: {width}x{height} -> {new_width}x{new_height}")
        else:
            # 如果图片已经很小，直接使用
            self.preview_image = self.original_image.copy()
            print(f"外圈图片保持原尺寸: {width}x{height}")

    def create_inner_preview(self):
        """创建内圈图片预览图"""
        if self.inner_original_image is None:
            return
        
        width, height = self.inner_original_image.size
        
        # 内圈图片应该比外圈小
        if self.preview_image:
            outer_width, outer_height = self.preview_image.size
            target_size = min(outer_width, outer_height) * 0.64
        else:
            target_size = 120  # 默认大小
        
        # 计算缩放比例，保持宽高比
        scale = min(target_size / width, target_size / height)
        new_width = int(width * scale)
        new_height = int(height * scale)
        
        self.inner_preview_image = self.inner_original_image.resize(
            (new_width, new_height), 
            Image.Resampling.LANCZOS
        )
        print(f"内圈图片缩放: {width}x{height} -> {new_width}x{new_height}")

    def create_default_outer_image(self):
        """创建默认外圈图片"""
        width, height = 350, 350  # 增大尺寸
        from PIL import ImageDraw
        
        img = Image.new("RGBA", (width, height), (255, 255, 255, 255))
        draw = ImageDraw.Draw(img)
        
        center_x, center_y = width // 2, height // 2
        radius = 150  # 外圈半径
        
        # 绘制外圈
        draw.ellipse(
            [center_x - radius, center_y - radius, 
            center_x + radius, center_y + radius],
            outline="#4a90e2",
            width=3
        )
        
        # 绘制刻度（每36°一个长刻度，每3.6°一个短刻度）
        for i in range(100):  # 100个小刻度
            angle = math.radians(i * 3.6)
            # 长刻度（每10个小刻度）
            if i % 10 == 0:
                start_radius = radius - 20
                end_radius = radius - 10
                # 添加数字
                num = i // 10
                num_radius = radius - 30
                num_x = center_x + num_radius * math.sin(angle)
                num_y = center_y - num_radius * math.cos(angle)
                draw.text((num_x-5, num_y-5), str(num), fill="#333333")
            else:
                start_radius = radius - 15
                end_radius = radius - 10
            
            x1 = center_x + start_radius * math.sin(angle)
            y1 = center_y - start_radius * math.cos(angle)
            x2 = center_x + end_radius * math.sin(angle)
            y2 = center_y - end_radius * math.cos(angle)
            draw.line([x1, y1, x2, y2], fill="#333333", width=1)
        
        self.original_image = img
        self.preview_image = img.copy()

    def clamp_reading(self, value):
        """限制读数在有效范围内"""
        return max(-0.1, min(1.2, value))

    def draw_needles(self):
        """在画布上绘制大小指针"""
        self.canvas.delete("needle")  # 删除之前的指针
        
        center_x, center_y = self.center_x, self.center_y
        
        # 绘制大针（中心，顺时针方向）
        # 注意：画布的0°是向上，顺时针为正
        large_end_x = center_x + self.needle_radius * math.sin(math.radians(self.large_needle_angle))
        large_end_y = center_y - self.needle_radius * math.cos(math.radians(self.large_needle_angle))
        self.canvas.create_line(
            center_x, center_y, large_end_x, large_end_y,
            fill="red", width=3, arrow="last", tags="needle"
        )
        
        # 绘制小针（右上方）
        small_center_x = center_x + 62  # 小针中心偏右
        small_center_y = center_y - 20  # 小针中心偏上
        
        # 小针的角度已经是计算好的最终角度，直接使用
        small_end_x = small_center_x + self.small_needle_radius * math.sin(math.radians(self.small_needle_angle))
        small_end_y = small_center_y - self.small_needle_radius * math.cos(math.radians(self.small_needle_angle))
        
        self.canvas.create_line(
            small_center_x, small_center_y, small_end_x, small_end_y,
            fill="blue", width=2, arrow="last", tags="needle"
        )
        
        # 标记小针中心点
        self.canvas.create_oval(
            small_center_x - 3, small_center_y - 3,
            small_center_x + 3, small_center_y + 3,
            fill="blue", outline="blue", tags="needle"
        )

    def update_reading_from_value(self):
        """根据当前读数更新指针角度"""
        # 计算热膨胀量（如果当前在控温模式）
        thermal_expansion = 0.0
        thermal_expansion = self.calculate_thermal_expansion()
        
        # 总读数 = 基础读数 + 热膨胀量
        total_reading = self.base_reading + thermal_expansion
        
        # 限制总读数在范围内
        total_reading = self.clamp_reading(total_reading)
        
        # 大针：顺时针每360°对应0.2mm增加
        self.large_needle_angle = (total_reading * 1800) % 360  # 每mm对应1800°
        
        # 小针：逆时针每36°对应0.2mm增加
        self.small_needle_angle = 90 - (total_reading * 180) % 360
        self.small_needle_angle = self.small_needle_angle % 360
        
        # 绘制指针
        self.draw_needles()
        
        # 更新显示读数（在千分表区域显示）
        self.canvas.delete("reading_text")
        
        # 显示详细信息
        reading_text = f"读数: {total_reading:.3f} mm"
        if thermal_expansion != 0:
            reading_text += f"\n基础: {self.base_reading:.3f} mm"
            reading_text += f"\n热膨胀: {thermal_expansion:+.3f} mm"
        
        if total_reading <= -0.099:  # 接近下限
            reading_text += " (下限)"
        elif total_reading >= 1.199:  # 接近上限
            reading_text += " (上限)"
        
        # self.canvas.create_text(
        #     self.center_x, self.center_y + 60,  # 下移一点以容纳更多行
        #     text=reading_text,
        #     fill="black", font=("Arial", 10, "bold"),
        #     tags="reading_text"
        # )
        
        print(f"指针角度 - 大针: {self.large_needle_angle:.1f}°, 小针: {self.small_needle_angle:.1f}°")


    def create_default_inner_image(self):
        """创建默认内圈图片"""
        width, height = 120, 120  # 调整大小
        from PIL import ImageDraw
        
        img = Image.new("RGBA", (width, height), (255, 255, 255, 0))  # 透明背景
        draw = ImageDraw.Draw(img)
        
        # 绘制一个简单的内圈
        draw.ellipse(
            [10, 10, width-10, height-10],
            outline="#27ae60",  # 绿色
            width=3
        )
        
        # 绘制中心点
        draw.ellipse(
            [width//2 - 5, height//2 - 5, width//2 + 5, height//2 + 5],
            fill="#e74c3c",  # 红色
            outline="white"
        )
        
        self.inner_original_image = img
        self.inner_preview_image = img.copy()
        print("创建默认内圈图片")

    def bind_events(self):
        # 绑定鼠标事件
        self.canvas.bind("<ButtonPress-1>", self.start_drag)
        self.canvas.bind("<B1-Motion>", self.drag)
        self.canvas.bind("<ButtonRelease-1>", self.end_drag)
        
        # 绑定鼠标滚轮事件
        self.canvas.bind("<MouseWheel>", self.zoom_image)

    def start_drag(self, event):
        """开始拖动"""
        if self.preview_image is None:
            return
        
        self.dragging = True
        
        # 确保中心点是最新的
        self.update_center()
        
        # 计算初始角度
        dx = event.x - self.center_x
        dy = event.y - self.center_y
        self.last_angle = math.degrees(math.atan2(dx, -dy))
        
        if self.last_angle < 0:
            self.last_angle += 360

    def drag(self, event):
        """拖动旋转"""
        if not self.dragging or self.preview_image is None:
            return
        
        # 计算当前角度
        dx = event.x - self.center_x
        dy = event.y - self.center_y
        current_angle = math.degrees(math.atan2(dx, -dy))
        
        if current_angle < 0:
            current_angle += 360
        
        # 计算角度变化
        angle_change = current_angle - self.last_angle
        
        # 处理角度跳变
        if angle_change > 180:
            angle_change -= 360
        elif angle_change < -180:
            angle_change += 360
        
        # 更新角度
        self.angle += angle_change
        self.angle %= 360
        
        # 旋转并更新显示
        self.rotate_and_update()
        
        self.last_angle = current_angle

    def end_drag(self, event):
        """结束拖动"""
        self.dragging = False

    def zoom_image(self, event):
        """缩放图片"""
        if self.preview_image is None:
            return
        
        # 计算缩放比例
        scale_factor = 1.1 if event.delta > 0 else 0.9
        
        # 更新显示
        self.rotate_and_update()

    def rotate_and_update(self):
        """旋转图片并更新显示"""
        if self.preview_image is None:
            return
        
        try:
            # 旋转外圈图片
            if self.angle != 0:
                rotated_img = self.preview_image.rotate(
                    -self.angle, 
                    expand=True, 
                    resample=Image.BILINEAR
                )
            else:
                rotated_img = self.preview_image.copy()
            
            # 更新显示
            self.tk_image = ImageTk.PhotoImage(rotated_img)
            # 确保内圈图片也被加载
            if not self.inner_tk_image and self.inner_preview_image:
                self.inner_tk_image = ImageTk.PhotoImage(self.inner_preview_image)
            
            self.update_display_with_image(rotated_img)
            
        except Exception as e:
            print(f"旋转错误: {e}")

    def update_display_picture(self):
        """更新显示（不旋转）"""
        if self.preview_image is None:
            return
        
        try:
            # 更新外圈图片
            self.tk_image = ImageTk.PhotoImage(self.preview_image)
            # 更新内圈图片
            if self.inner_preview_image:
                self.inner_tk_image = ImageTk.PhotoImage(self.inner_preview_image)
            
            self.update_display_with_image(self.preview_image)
            
        except Exception as e:
            print(f"更新显示错误: {e}")

    def update_display_with_image(self, image):
        """在画布上显示图片（叠加内圈和外圈）"""
        self.canvas.delete("all")
        
        # 先显示内圈图片（底层）
        if self.inner_tk_image:
            self.canvas.create_image(
                self.center_x,
                self.center_y,
                image=self.inner_tk_image,
                anchor=tk.CENTER
            )
        
        # 再显示外圈图片（上层，可旋转）
        if self.tk_image:
            self.canvas.create_image(
                self.center_x,
                self.center_y,
                image=self.tk_image,
                anchor=tk.CENTER
            )
        
        # 绘制中心点
        marker_radius = 4
        self.canvas.create_oval(
            self.center_x - marker_radius,
            self.center_y - marker_radius,
            self.center_x + marker_radius,
            self.center_y + marker_radius,
            fill="red",
            outline="white",
            width=1
        )
        
        # 绘制指针（新增）
        self.draw_needles()
        
        # 显示读数（新增）
        thermal_expansion = 0.0
        if self.temp_mode == "control":
            thermal_expansion = self.calculate_thermal_expansion()
        
        total_reading = self.base_reading + thermal_expansion
        total_reading = self.clamp_reading(total_reading)
        
        reading_text = f"读数: {total_reading:.3f} mm"
        if thermal_expansion != 0:
            reading_text += f"\n基础: {self.base_reading:.3f} mm"
            reading_text += f"\n热膨胀: {thermal_expansion:+.3f} mm"
        
        # self.canvas.create_text(
        #     self.center_x, self.center_y + 60,
        #     text=reading_text,
        #     fill="black", font=("Arial", 10, "bold"),
        #     tags="reading_text"
        # )

def main():
    root = tk.Tk()
    app = SimpleImageRotator(root)
    root.mainloop()

if __name__ == "__main__":
    main()