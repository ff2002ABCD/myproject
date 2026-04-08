import tkinter as tk
from tkinter import ttk, font
from PIL import Image, ImageTk
import os
import sys
import time
import random
import math
from tkinter import messagebox, filedialog
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import pandas as pd  # 用于Excel操作
import numpy as np   # 用于数值计算

def get_resource_path(relative_path):
    """获取资源的绝对路径，支持打包后的环境"""
    try:
        # 打包后的临时文件夹路径
        base_path = sys._MEIPASS
    except Exception:
        # 开发环境的路径
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, relative_path)

class ExperimentApp:
    def __init__(self, root):
        self.root = root
        self.root.title("实验系统")
        self.root.geometry("1600x800")
        self.root.resizable(False, False)
        # 当前界面状态
        self.current_interface = "平衡法"  # 平衡法/动态法/参数设置
        # 新增：电源状态控制
        self.power_on = False  # 电源状态，初始为关闭
        self.selected_index = 0  # 当前选中项的索引
        # 新增：存储基本电荷分析结果
        self.balance_e_experimental = None  # 平衡法测得的基本电荷
        self.balance_e_error = None         # 平衡法百分误差
        self.dynamic_e_experimental = None  # 动态法测得的基本电荷
        self.dynamic_e_error = None         # 动态法百分误差
        # 长按相关变量
        self.auto_adjust_interval = None  # 自动调整定时器
        self.is_auto_adjusting = False    # 是否正在自动调整
        self.auto_adjust_direction = None # 自动调整方向 ('increase' 或 'decrease')
        self.auto_adjust_speed = 100      # 初始速度（毫秒）
        self.min_adjust_speed = 30        # 最快调整速度（毫秒）

        # 计时相关变量
        self.is_timing = False  # 是否正在计时
        self.start_time = 0  # 开始时间
        self.timer_id = None  # 定时器ID
        self.elapsed_time = 0.00  # 经过的时间（秒）
    
        # 电压控制参数
        self.lift_voltage = tk.IntVar(value=250)  # 提升电压，初始250V
        self.balance_voltage_slider = tk.IntVar(value=50)  # 平衡电压滑块，初始250V
        
        # 油滴相关变量
        self.oil_droplets = []  # 存储油滴对象
        self.is_oil_simulating = False  # 是否正在模拟油滴运动
        self.oil_timer_id = None  # 油滴模拟定时器ID
        self.focus_scale = tk.DoubleVar(value=2.0)  # 对焦进度条，控制油滴大小 (1-3倍)
        
        # 物理常数
        self.electron_charge = 1.602e-19  # 电子电荷 (C)
        self.plate_distance = 5.00e-3  # 极板间距 (m)
        self.oil_density = 981  # 油滴密度 (kg/m³)
        self.gravity = 9.794  # 重力加速度 (m/s²)
        
        # 数据表格相关
        self.data_table = None  # 存储表格引用
        self.data_tree = None  # Treeview表格对象
        self.chart_canvas = None  # 图表画布
        self.method_data = {"平衡法": [], "动态法": []}  # 存储两种方法的数据
        self.current_table_data = []  # 当前显示的数据

        # 数据存储
        self.balance_params = {
            "平衡电压": "0.0",
            "下落时间": "0.00",
        }
        
        self.dynamic_params = {
            "平衡电压": "0.0",
            "下落时间": "0.00",
            "上升电压": "0.0",
            "上升时间": "0.00",
        }
        
        self.settings_params = {
            "运动距离": "2.0",
            "油滴密度": "981",
            "室温": "24",
            "大气压强": "101325",
            "重力加速度": "9.794",
            "修正常数": "0.00823",
        }
        
        # 电荷显示框数据 - 平衡法和动态法共用，不清除
        self.charge_display_data = ["0.0", "0.0", "0.0", "0.0", "0.0", "0.0", "0.0"]
        self.current_charge_index = 0  # 当前要更新的电荷显示框索引

        # 存储图片引用
        self.images = {}
        self.photo_refs = {}  # 防止图片被垃圾回收
        
        # 数据记录区域数据存储（新增）
        self.balance_table_data = []  # 存储平衡法表格数据
        self.dynamic_table_data = []  # 存储动态法表格数据
        
        # 参数存储（新增）
        self.balance_params_storage = {
            "l": "2.0",
            "T": "297.15",
            "P": "101325",
            "g": "9.794",
            "rho1": "981",
            "rho2": "1.188",
            "eta": "1.832×10⁻⁵"
        }
        
        self.dynamic_params_storage = {
            "l": "2.0",
            "T": "297.15",
            "P": "101325",
            "g": "9.794",
            "rho1": "981",
            "rho2": "1.188",
            "eta": "1.832×10⁻⁵"
        }

        # 创建主框架
        self.create_frames()
        
        # 加载图片
        self.load_all_images()
        
        # 初始化显示屏区域
        self.init_display_area()
        
        # 创建实验仪区域按钮
        self.create_instrument_buttons()
        
        # 创建数据记录区域
        self.create_data_record_area()
        
        self.create_charge_display_boxes()
        # 显示实验装置区域图片
        self.show_experiment_device()
        # 禁用其他控制按钮
        if hasattr(self, 'spray_button'):
            self.spray_button.config(state='disabled')
        if hasattr(self, 'clear_button'):
            self.clear_button.config(state='disabled')
        
        if hasattr(self, 'increase_button'):
            self.increase_button.config(state='disabled')
        if hasattr(self, 'decrease_button'):
            self.decrease_button.config(state='disabled')
        # self.compare_radius_calculation()
        self.verify_calculation_consistency()
        
    def create_charge_display_boxes(self):
        """创建七个电荷显示框 - 叠加在显示屏区域的背景图上"""
        # 注意：这里使用 self.frame_top_left 作为父容器
        # 电荷显示框容器
        self.charge_frame = tk.Frame(self.frame_top_left, bg='white')
        self.charge_frame.place(x=10+440+20+25, y=320-200, width=70, height=180)
        
        # 初始状态：电源关闭，不显示电荷框
        self.charge_frame.place_forget()
        
        # 创建七个显示框
        self.charge_labels = []
        positions = [(0, 0), (0, 25), (0, 50), (0, 75),
                    (0, 100), (0, 125), (0, 150)]
        
        for i in range(7):
            frame = tk.Frame(self.charge_frame, bd=1, relief=tk.SUNKEN, bg='white')
            frame.place(x=positions[i][0], y=positions[i][1], width=70, height=25)
            
            # 显示框编号
            tk.Label(frame, text=f"{i+1}", font=('Arial', 8), 
                    bg='lightgray', width=3).place(x=0, y=0)
            
            # 电荷值标签
            label = tk.Label(frame, text=self.charge_display_data[i], 
                        font=('Arial', 8, 'bold'), bg='white')
            label.place(x=30, y=0)
            self.charge_labels.append(label)

    def update_charge_indicator(self):
        """更新电荷显示框的当前索引指示器"""
        if not hasattr(self, 'charge_indicator'):
            return
        
        # 指示器应该放在每个显示框下方
        positions = [(45, 70), (145, 70), (245, 70), (345, 70),
                    (95, 120), (195, 120), (295, 120)]
        
        if 0 <= self.current_charge_index < len(positions):
            self.charge_indicator.place(x=positions[self.current_charge_index][0], 
                                    y=positions[self.current_charge_index][1])

    def update_charge_display(self, charge_value):
        """更新电荷显示框的值"""
        # 检查电源状态
        if not self.power_on:
            print("电源关闭，无法更新电荷显示")
            return
        
        # 如果电荷显示框还没创建，先创建
        if not hasattr(self, 'charge_labels'):
            self.create_charge_display_boxes()
        
        # 确保电荷框可见
        if hasattr(self, 'charge_frame') and self.charge_frame.winfo_ismapped() == 0:
            self.charge_frame.place(x=10+440+20+25, y=320-200, width=70, height=180)
        
        # 将电荷值转换为科学计数法显示（×10⁻¹⁹ C）
        if charge_value != 0:
            charge_display = f"{charge_value/1e-19:.3f}"
        else:
            charge_display = "0.00"
        
        # 更新当前显示框
        self.charge_display_data[self.current_charge_index] = charge_display
        
        # 确保列表存在且有足够元素
        if hasattr(self, 'charge_labels') and self.current_charge_index < len(self.charge_labels):
            self.charge_labels[self.current_charge_index].config(text=charge_display)
        
        # 移动到下一个显示框（循环）
        self.current_charge_index = (self.current_charge_index + 1) % 7
            
    def create_frames(self):
        """创建四个区域"""
        # 左上区域 - 显示屏区域
        self.frame_top_left = tk.Frame(self.root, bd=2, relief=tk.SUNKEN, bg='white')
        self.frame_top_left.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        
        # 右上区域 - 实验仪区域
        self.frame_top_right = tk.Frame(self.root, bd=2, relief=tk.SUNKEN, bg='white')
        self.frame_top_right.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        
        # 左下区域 - 实验装置区域
        self.frame_bottom_left = tk.Frame(self.root, bd=2, relief=tk.SUNKEN, bg='white')
        self.frame_bottom_left.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        
        # 右下区域 - 数据记录区域
        self.frame_bottom_right = tk.Frame(self.root, bd=2, relief=tk.RAISED)
        self.frame_bottom_right.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        
        # 配置网格权重
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)
        
        self.root.grid_columnconfigure(0, weight=1)  # 左列
        self.root.grid_columnconfigure(1, weight=2)  # 右列

        # 区域标签
        tk.Label(self.frame_top_left, text="显示屏区域", font=('Arial', 12, 'bold')).pack(pady=5)
        tk.Label(self.frame_bottom_left, text="实验装置区域", font=('Arial', 12, 'bold')).pack(pady=5)
        tk.Label(self.frame_top_right, text="实验仪区域", font=('Arial', 12, 'bold')).pack(pady=5)
        tk.Label(self.frame_bottom_right, text="数据记录区域", font=('Arial', 12, 'bold')).pack(pady=5)
    
    def load_all_images(self):
        """加载所有需要的图片"""
        try:
            # 获取程序目录
            current_dir = os.path.dirname(os.path.abspath(__file__))
            background_dir = os.path.join(current_dir, 'background')
            
            # 检查background目录是否存在
            if not os.path.exists(background_dir):
                print(f"警告: background目录不存在: {background_dir}")
                os.makedirs(background_dir, exist_ok=True)
            
            # 显示屏区域图片
            balance_img_path = os.path.join(background_dir, "平衡法.jpg")
            dynamic_img_path = os.path.join(background_dir, "动态法.jpg")
            
            # 实验装置区域图片
            device_img_path = os.path.join(background_dir, "实验装置.jpg")
            
            # 实验仪区域图片
            instrument_img_path = os.path.join(background_dir, "FD-MLG-A.jpg")
            
            # 加载图片
            if os.path.exists(balance_img_path):
                self.images['平衡法'] = Image.open(balance_img_path)
                print("加载平衡法图片成功")
            else:
                print(f"平衡法图片不存在: {balance_img_path}")
                self.images['平衡法'] = Image.new('RGB', (400, 300), color='lightblue')
            
            if os.path.exists(dynamic_img_path):
                self.images['动态法'] = Image.open(dynamic_img_path)
                print("加载动态法图片成功")
            else:
                print(f"动态法图片不存在: {dynamic_img_path}")
                self.images['动态法'] = Image.new('RGB', (400, 300), color='lightgreen')
            
            if os.path.exists(device_img_path):
                self.images['实验装置'] = Image.open(device_img_path)
                print("加载实验装置图片成功")
            else:
                print(f"实验装置图片不存在: {device_img_path}")
                self.images['实验装置'] = Image.new('RGB', (100, 300), color='lightgray')
            
            if os.path.exists(instrument_img_path):
                self.images['实验仪'] = Image.open(instrument_img_path)
                print("加载实验仪图片成功")
            else:
                print(f"实验仪图片不存在: {instrument_img_path}")
                self.images['实验仪'] = Image.new('RGB', (800, 300), color='gray')
                
        except Exception as e:
            print(f"加载图片时出错: {e}")
            # 创建空白图片
            for name in ['平衡法', '动态法', '参数设置', '实验装置', '实验仪']:
                self.images[name] = Image.new('RGB', (400, 300), color='gray')
    
    def show_experiment_device(self):
        """显示实验装置区域的图片和电压调节滑块"""
        # 创建主容器Frame，水平排列图片和滑块
        main_container = tk.Frame(self.frame_bottom_left, bg='white')
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 左侧：实验装置图片
        left_frame = tk.Frame(main_container, bg='white')
        left_frame.pack(side=tk.LEFT, padx=(0, 20))
        
        try:
            # 获取实验装置图片
            if '实验装置' in self.images:
                img = self.images['实验装置']
                # 调整图片大小以适应区域
                img_resized = img.resize((200, 300), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img_resized)
                
                # 创建标签显示图片
                label = tk.Label(left_frame, image=photo, bg='white')
                label.image = photo  # 保持引用
                label.pack()
        except Exception as e:
            print(f"显示实验装置图片失败: {e}")
            tk.Label(left_frame, text="实验装置图片", 
                    font=('Arial', 12), bg='white', width=20, height=10).pack()
        
        # 右侧：电压调节滑块和油滴控制按钮
        right_frame = tk.Frame(main_container, bg='white')
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 创建滑块
        self.create_voltage_sliders(right_frame)

    def create_voltage_sliders(self, parent_frame):
        """在指定父容器中创建电压调节滑块和油滴控制"""
        # 创建滑块框架
        slider_frame = tk.Frame(parent_frame, bg='white')
        slider_frame.pack(pady=20, padx=10, fill=tk.BOTH, expand=True)
        
        # ====== 提升电压滑块 ======
        lift_frame = tk.Frame(slider_frame, bg='white')
        lift_frame.pack(fill=tk.X, pady=(0, 25))
        
        tk.Label(lift_frame, text="提升电压:", 
                font=('Arial', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        
        # 创建滑块和数值显示
        lift_slider_frame = tk.Frame(lift_frame, bg='white')
        lift_slider_frame.pack(fill=tk.X)
        
        # 数值标签
        self.lift_voltage_label = tk.Label(lift_slider_frame, 
                                        text=f"{self.lift_voltage.get()} V", 
                                        font=('Arial', 10, 'bold'), 
                                        bg='white', 
                                        width=6,
                                        fg='blue')
        self.lift_voltage_label.pack(side=tk.LEFT, padx=(0, 15))
        
        # 滑块
        self.lift_slider = tk.Scale(lift_slider_frame, 
                                from_=200, to=300, 
                                orient=tk.HORIZONTAL,
                                variable=self.lift_voltage,
                                length=100,
                                showvalue=0,
                                command=self.on_lift_voltage_changed,
                                bg='white',
                                highlightbackground='white',
                                troughcolor='lightgray',
                                sliderrelief=tk.RAISED,
                                sliderlength=20)
        self.lift_slider.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # 范围标签
        range_label = tk.Label(lift_slider_frame, text="(200-300V)", 
                            font=('Arial', 9), bg='white', fg='gray')
        range_label.pack(side=tk.LEFT, padx=(15, 0))
        
        # ====== 平衡电压滑块和微调按钮 ======
        balance_frame = tk.Frame(slider_frame, bg='white')
        balance_frame.pack(fill=tk.X, pady=(0, 25))
        
        tk.Label(balance_frame, text="平衡电压:", 
                font=('Arial', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        
        # 创建滑块和数值显示
        balance_slider_frame = tk.Frame(balance_frame, bg='white')
        balance_slider_frame.pack(fill=tk.X)
        
        # 数值标签
        self.balance_voltage_label = tk.Label(balance_slider_frame, 
                                            text=f"{self.balance_voltage_slider.get()} V", 
                                            font=('Arial', 10, 'bold'), 
                                            bg='white', 
                                            width=6,
                                            fg='green')
        self.balance_voltage_label.pack(side=tk.LEFT, padx=(0, 15))
        
        # 创建微调按钮容器
        adjustment_frame = tk.Frame(balance_slider_frame, bg='white')
        adjustment_frame.pack(side=tk.LEFT, padx=(0, 10))
        
        # 减小按钮 (-1V)
        self.decrease_button = tk.Button(adjustment_frame, text="-", 
                                    font=('Arial', 8, 'bold'),
                                    width=2, height=1,
                                    bg='lightblue',
                                    relief=tk.RAISED,
                                    activebackground='skyblue')
        self.decrease_button.pack(side=tk.LEFT, padx=(0, 2))
        
        # 绑定鼠标事件
        self.decrease_button.bind("<ButtonPress-1>", 
                                lambda e: self.start_auto_adjust('decrease'))
        self.decrease_button.bind("<ButtonRelease-1>", 
                                lambda e: self.stop_auto_adjust())
        
        # 滑块
        self.balance_slider = tk.Scale(balance_slider_frame, 
                                    from_=0, to=500, 
                                    orient=tk.HORIZONTAL,
                                    variable=self.balance_voltage_slider,
                                    length=80,  # 稍微缩短滑块长度
                                    showvalue=0,
                                    command=self.on_balance_voltage_changed,
                                    bg='white',
                                    highlightbackground='white',
                                    troughcolor='lightgray',
                                    sliderrelief=tk.RAISED,
                                    sliderlength=20)
        self.balance_slider.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # 增加按钮 (+1V)
        self.increase_button = tk.Button(adjustment_frame, text="+", 
                                    font=('Arial', 8, 'bold'),
                                    width=2, height=1,
                                    bg='lightcoral',
                                    relief=tk.RAISED,
                                    activebackground='salmon')
        self.increase_button.pack(side=tk.LEFT, padx=(2, 0))
        
        # 绑定鼠标事件
        self.increase_button.bind("<ButtonPress-1>", 
                                lambda e: self.start_auto_adjust('increase'))
        self.increase_button.bind("<ButtonRelease-1>", 
                                lambda e: self.stop_auto_adjust())
        
        # # 保留原有的点击事件
        # self.decrease_button.config(command=self.decrease_balance_voltage)
        # self.increase_button.config(command=self.increase_balance_voltage)
        
        # 快速调整按钮容器
        quick_adjust_frame = tk.Frame(balance_slider_frame, bg='white')
        quick_adjust_frame.pack(side=tk.LEFT, padx=(10, 0))
        
        # # -10V 快速调整按钮
        # tk.Button(quick_adjust_frame, text="-10", 
        #         font=('Arial', 7, 'bold'),
        #         width=3, height=1,
        #         command=lambda: self.adjust_balance_voltage(-10),
        #         bg='lightblue',
        #         relief=tk.RAISED,
        #         activebackground='skyblue').pack(side=tk.LEFT, padx=(0, 2))
        
        # # +10V 快速调整按钮
        # tk.Button(quick_adjust_frame, text="+10", 
        #         font=('Arial', 7, 'bold'),
        #         width=3, height=1,
        #         command=lambda: self.adjust_balance_voltage(10),
        #         bg='lightcoral',
        #         relief=tk.RAISED,
        #         activebackground='salmon').pack(side=tk.LEFT, padx=(0, 0))
        
        # 范围标签
        range_label2 = tk.Label(balance_slider_frame, text="(0-500V)", 
                            font=('Arial', 9), bg='white', fg='gray')
        range_label2.pack(side=tk.LEFT, padx=(15, 0))
        
        # ====== 油滴控制按钮组 ======
        oil_control_frame = tk.Frame(slider_frame, bg='white')
        oil_control_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 喷入油雾按钮
        self.spray_button = tk.Button(oil_control_frame, text="喷入油雾", 
                                    font=('Arial', 10, 'bold'),
                                    bg='lightblue',
                                    command=self.spray_oil_droplets,
                                    relief=tk.RAISED,
                                    width=12)
        self.spray_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # 清空油滴按钮
        self.clear_button = tk.Button(oil_control_frame, text="清空油滴", 
                                    font=('Arial', 10, 'bold'),
                                    bg='lightcoral',
                                    command=self.clear_oil_droplets,
                                    relief=tk.RAISED,
                                    width=12)
        self.clear_button.pack(side=tk.LEFT)
        
        # ====== 对焦进度条 ======
        focus_frame = tk.Frame(slider_frame, bg='white')
        focus_frame.pack(fill=tk.X, pady=(10, 0))
        
        tk.Label(focus_frame, text="对焦调节:", 
                font=('Arial', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        
        # 创建滑块和数值显示
        focus_slider_frame = tk.Frame(focus_frame, bg='white')
        focus_slider_frame.pack(fill=tk.X)
        
        # 数值标签
        self.focus_label = tk.Label(focus_slider_frame, 
                                text=f"{self.focus_scale.get():.1f}倍", 
                                font=('Arial', 10, 'bold'), 
                                bg='white', 
                                width=6,
                                fg='purple')
        self.focus_label.pack(side=tk.LEFT, padx=(0, 15))
        
        # 滑块
        self.focus_slider = tk.Scale(focus_slider_frame, 
                                from_=1, to=3, 
                                orient=tk.HORIZONTAL,
                                variable=self.focus_scale,
                                length=100,
                                resolution=0.1,
                                showvalue=0,
                                command=self.on_focus_changed,
                                bg='white',
                                highlightbackground='white',
                                troughcolor='lightgray',
                                sliderrelief=tk.RAISED,
                                sliderlength=20)
        self.focus_slider.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # 范围标签
        range_label3 = tk.Label(focus_slider_frame, text="(1-3倍)", 
                            font=('Arial', 9), bg='white', fg='gray')
        range_label3.pack(side=tk.LEFT, padx=(15, 0))

    def start_auto_adjust(self, direction):
        """开始自动调整平衡电压（长按功能）"""
        # 停止之前的自动调整
        if not self.power_on:
            return
        if self.auto_adjust_interval:
            self.root.after_cancel(self.auto_adjust_interval)
            self.auto_adjust_interval = None
        
        # 设置调整方向
        self.is_auto_adjusting = True
        self.auto_adjust_direction = direction
        self.auto_adjust_speed = 200  # 初始速度：200ms/次
        
        # 立即调整一次
        if direction == 'increase':
            self.increase_balance_voltage()
        else:  # 'decrease'
            self.decrease_balance_voltage()
        
        # 开始自动调整循环
        self.schedule_next_auto_adjust()

    def schedule_next_auto_adjust(self):
        """安排下一次自动调整"""
        if not self.is_auto_adjusting:
            return
        
        # 加速调整（长按时间越长，调整越快）
        self.auto_adjust_speed = max(self.min_adjust_speed, 
                                    int(self.auto_adjust_speed * 0.9))
        
        # 安排下一次调整
        self.auto_adjust_interval = self.root.after(
            self.auto_adjust_speed, 
            self.perform_auto_adjust
        )

    def perform_auto_adjust(self):
        """执行自动调整"""
        if not self.is_auto_adjusting:
            return
        
        # 根据方向调整电压
        if self.auto_adjust_direction == 'increase':
            self.increase_balance_voltage()
        else:  # 'decrease'
            self.decrease_balance_voltage()
        
        # 安排下一次调整
        self.schedule_next_auto_adjust()

    def stop_auto_adjust(self):
        """停止自动调整"""
        self.is_auto_adjusting = False
        
        if self.auto_adjust_interval:
            self.root.after_cancel(self.auto_adjust_interval)
            self.auto_adjust_interval = None
        
        # 重置调整速度
        self.auto_adjust_speed = 100

    def decrease_balance_voltage(self):
        """减小平衡电压 1V"""
        current_value = self.balance_voltage_slider.get()
        if current_value > 0:
            new_value = current_value - 1
            self.balance_voltage_slider.set(new_value)
            self.on_balance_voltage_changed(new_value)
            # 更新按钮状态（如果是长按则不会调用）
            if not self.is_auto_adjusting:
                print(f"平衡电压减小: {current_value}V → {new_value}V")

    def increase_balance_voltage(self):
        """增加平衡电压 1V"""
        current_value = self.balance_voltage_slider.get()
        if current_value < 500:
            new_value = current_value + 1
            self.balance_voltage_slider.set(new_value)
            self.on_balance_voltage_changed(new_value)
            # 更新按钮状态（如果是长按则不会调用）
            if not self.is_auto_adjusting:
                print(f"平衡电压增加: {current_value}V → {new_value}V")

    def adjust_balance_voltage(self, delta):
        """按指定值调整平衡电压"""
        current_value = self.balance_voltage_slider.get()
        new_value = current_value + delta
        
        # 限制在有效范围内
        new_value = max(0, min(500, new_value))
        
        if new_value != current_value:
            self.balance_voltage_slider.set(new_value)
            self.on_balance_voltage_changed(new_value)
            
            direction = "增加" if delta > 0 else "减小"
            print(f"平衡电压{direction}{abs(delta)}V: {current_value}V → {new_value}V")

    def on_focus_changed(self, value):
        """对焦滑块变化时的回调"""
        self.focus_label.config(text=f"{float(value):.1f}倍")
        # 更新所有油滴的大小
        self.update_oil_droplet_sizes()

    def update_oil_droplet_sizes(self):
        """更新所有油滴的大小（根据对焦倍数）"""
        for droplet in self.oil_droplets:
            # 获取基础显示半径
            base_radius = droplet['display_radius_base']
            
            # 计算新的显示半径（基于对焦倍数）
            new_radius = int(base_radius * self.focus_scale.get())
            
            # 更新Canvas上的油滴大小
            if 'id' in droplet:
                self.display_canvas.coords(droplet['id'],
                                        droplet['x'] - new_radius,
                                        droplet['y'] - new_radius,
                                        droplet['x'] + new_radius,
                                        droplet['y'] + new_radius)
            
            # 更新存储的半径
            droplet['radius'] = new_radius

    def spray_oil_droplets(self):
        """喷入油雾 - 在显示屏区域生成随机油滴"""
        print("喷入油雾...")
        if not self.power_on:
            print("电源关闭，无法喷入油雾")
            return
        # 清空旧的油滴
        self.clear_oil_droplets()
        
        # 随机生成3-15个油滴
        num_droplets = random.randint(3, 15)
        
        # 获取显示屏区域的尺寸
        canvas_width = 400
        canvas_height = 300
        
        for i in range(num_droplets):
            # 随机油滴类型和大小
            size_type = random.choices(['small', 'medium', 'large'], weights=[0.45, 0.45, 0.1])[0]
            
            # 定义不同大小油滴的实际半径范围（微米）
            if size_type == 'small':
                actual_radius = random.uniform(0.5e-6, 1.0e-6)  # 0.5-1.0微米
                display_radius_base = 2  # 小油滴：基础显示半径2像素
            elif size_type == 'medium':
                actual_radius = random.uniform(1.0e-6, 1.5e-6)  # 1.0-1.5微米
                display_radius_base = 3  # 中油滴：基础显示半径3像素
            else:  # large
                actual_radius = random.uniform(1.5e-6, 2.0e-6)  # 1.5-2.0微米
                display_radius_base = 4  # 大油滴：基础显示半径4像素
            
            # 根据对焦倍数计算显示半径
            display_radius = int(display_radius_base * self.focus_scale.get())
            
            # 随机位置（根据油滴大小留出不同的边界）
            x = random.randint(display_radius*2, canvas_width - display_radius*2)
            y = random.randint(display_radius*2, canvas_height - display_radius*2)
            
            # 随机电荷数（1-8个电子）
            electron_count = random.randint(0, 8)
            charge = electron_count * self.electron_charge
            
            # 计算油滴质量
            volume = (4/3) * math.pi * (actual_radius ** 3)
            mass = self.oil_density * volume
            
            # 计算该油滴的平衡电压
            balanced_voltage = (mass * self.gravity * self.plate_distance) / charge if charge > 0 else 250
            
            # 限制平衡电压在合理范围内
            balanced_voltage = max(50, min(450, balanced_voltage))
            
            # 为油滴添加轻微的水平速度（随机扰动）
            velocity_x = random.uniform(-0.0001, 0.0001)
            
            # 在Canvas上绘制油滴（不同大小的红点）
            droplet_id = self.display_canvas.create_oval(
                x - display_radius, y - display_radius,
                x + display_radius, y + display_radius,
                fill='red', outline='darkred', width=1
            )
            
            # 存储油滴信息
            droplet = {
                'id': droplet_id,
                'x': x,
                'y': y,
                'radius': display_radius,
                'display_radius_base': display_radius_base,  # 存储基础显示半径
                'actual_radius': actual_radius,
                'charge': charge,
                'electron_count': electron_count,
                'mass': mass,
                'velocity_y': 0,
                'velocity_x': velocity_x,
                'balanced_voltage': balanced_voltage,
                'is_balanced': False,
                'last_update_time': time.time(),
                'color': 'red',
                'base_color': 'red',
                'size_type': size_type,
            }
            
            self.oil_droplets.append(droplet)
            print(f"油滴{i+1}: 大小={size_type}, 实际半径={actual_radius*1e6:.2f}μm, 显示半径={display_radius}px, 电荷={electron_count}e, 平衡电压={balanced_voltage:.0f}V")
        
        print(f"生成了 {num_droplets} 个大小不同的油滴")
        
        # 开始油滴运动模拟
        if not self.is_oil_simulating:
            self.start_oil_simulation()

    def clear_oil_droplets(self):
        """清空所有油滴"""
        print("清空油滴...")
        
        # 停止油滴模拟
        if self.is_oil_simulating and self.oil_timer_id:
            self.root.after_cancel(self.oil_timer_id)
            self.oil_timer_id = None
            self.is_oil_simulating = False
        
        # 从Canvas上删除所有油滴和标签
        for droplet in self.oil_droplets:
            self.display_canvas.delete(droplet['id'])
            # if 'label_id' in droplet:
            #     self.display_canvas.delete(droplet['label_id'])
        
        # 清空油滴列表
        self.oil_droplets.clear()

    def start_oil_simulation(self):
        """开始油滴运动模拟"""
        self.is_oil_simulating = True
        self.update_oil_droplet_motion()

    # 在 update_oil_droplet_motion 方法中进行以下修改：

    def update_oil_droplet_motion(self):
        """根据电荷公式反推速度更新油滴运动"""
        if not self.is_oil_simulating or not self.oil_droplets:
            self.is_oil_simulating = False
            return
        
        mode = self.voltage_mode.get()
        current_time = time.time()
        
        # 获取实验参数
        l = float(self.settings_params["运动距离"]) / 1000  # 运动距离 m
        rho = float(self.settings_params["油滴密度"])        # 油滴密度 kg/m³
        g = float(self.settings_params["重力加速度"])        # 重力加速度 m/s²
        d = self.plate_distance                             # 极板间距 m
        b = float(self.settings_params["修正常数"])          # 修正常数 N/m
        p = float(self.settings_params["大气压强"])          # 大气压强 Pa
        T = float(self.settings_params["室温"]) + 273.15     # 温度 K
        
        # ========== 统一使用Sutherland公式计算空气粘滞系数 ==========
        T0 = 273.15  # 参考温度 0°C
        eta0 = 1.716e-5  # 参考温度下的空气粘滞系数
        S = 110.4  # Sutherland常数
        
        eta = eta0 * (T/T0)**1.5 * (T0 + S) / (T + S)  # Pa·s
        # ========================================================
        
        # 速度放大系数
        SPEED_MULTIPLIER = 1
        
        for i, droplet in enumerate(self.oil_droplets):
            dt = current_time - droplet['last_update_time']
            if dt <= 0:
                dt = 0.05
            
            # 计算油滴实际半径和质量
            a = droplet['actual_radius']
            mass = droplet['mass']
            
            # 统一使用修正因子
            correction = 1 + b / (p * a)
            
            # 计算自由下落终端速度（基础速度，所有模式共用）
            v_free_fall = (2/9) * (rho * g * a**2) / eta
            v_free_fall = v_free_fall / correction  # 修正
            v_free_fall *= SPEED_MULTIPLIER  # 放大
            
            if mode == "平衡":
                # 平衡模式
                U = self.balance_voltage_slider.get()
                
                if U == 0:
                    # 电压为0时，就是自由下落
                    droplet['velocity_y'] = v_free_fall
                    droplet['is_balanced'] = False
                else:
                    # 计算电场力
                    E = U / d if d > 0 else 0  # 电场强度 V/m
                    F_e = droplet['charge'] * E  # 电场力
                    
                    # 计算重力
                    F_g = mass * g
                    
                    # 计算净力
                    net_force = F_e - F_g
                    
                    if abs(net_force) > 1e-20:
                        # 从力平衡计算速度
                        v_net = net_force / (6 * math.pi * eta * a * correction)
                        v_net *= SPEED_MULTIPLIER
                        
                        # 方向：电场力大于重力时向上，否则向下
                        if net_force > 0:
                            droplet['velocity_y'] = -abs(v_net)  # 向上为负
                        else:
                            droplet['velocity_y'] = abs(v_net)   # 向下为正
                    else:
                        droplet['velocity_y'] = 0  # 平衡状态
                
                # 判断平衡状态 (在平衡电压±1V范围内)
                if U > 0:  # 只有电压>0时才可能平衡
                    voltage_diff = abs(U - droplet['balanced_voltage'])
                    if voltage_diff < 1.0:
                        droplet['is_balanced'] = True
                        droplet['velocity_y'] = 0
                        if droplet['color'] != 'green':
                            self.display_canvas.itemconfig(droplet['id'], fill='green')
                            droplet['color'] = 'green'
                        droplet['last_update_time'] = current_time
                        continue
                    else:
                        droplet['is_balanced'] = False
                        if droplet['color'] != droplet['base_color']:
                            self.display_canvas.itemconfig(droplet['id'], fill=droplet['base_color'])
                            droplet['color'] = droplet['base_color']
                else:
                    droplet['is_balanced'] = False
                    if droplet['color'] != droplet['base_color']:
                        self.display_canvas.itemconfig(droplet['id'], fill=droplet['base_color'])
                        droplet['color'] = droplet['base_color']
                        
            elif mode == "提升":
                # 使用公式反推计算上升速度
                v_e = self.calculate_rise_speed_from_formulas(droplet)
                
                if v_e is not None and v_e > 0:
                    # 成功计算上升速度
                    droplet['velocity_y'] = -v_e  # 向上为负
                else:
                    # 如果计算失败，油滴无法上升，开始下落
                    v_g = self.calculate_fall_speed_from_balance_formula(droplet)
                    if v_g is not None:
                        droplet['velocity_y'] = v_g
                    else:
                        droplet['velocity_y'] = v_free_fall  # 使用自由下落速度
                
            else:  # 下落模式
                # 下落模式：使用平衡法公式反推速度
                v_g = self.calculate_fall_speed_from_balance_formula(droplet)
                if v_g is not None:
                    droplet['velocity_y'] = v_g
                else:
                    droplet['velocity_y'] = v_free_fall  # 使用自由下落速度
                
                # 标记为非平衡状态
                droplet['is_balanced'] = False
                if droplet['color'] != droplet['base_color']:
                    self.display_canvas.itemconfig(droplet['id'], fill=droplet['base_color'])
                    droplet['color'] = droplet['base_color']
            
            droplet['velocity_y'] *= 1  # 速度因子，保持原样
            
            # 转换速度：实际速度 -> 像素速度
            pixel_per_meter = 115000
            velocity_y_pixels = droplet['velocity_y'] * pixel_per_meter * dt
            
            # 更新位置
            new_x = droplet['x']
            new_y = droplet['y'] + velocity_y_pixels
            
            # 边界检查
            canvas_height = 300
            canvas_width = 400
            
            # 左右边界反弹
            if droplet['velocity_x'] != 0:
                if new_x < droplet['radius']:
                    new_x = droplet['radius']
                    droplet['velocity_x'] = -droplet['velocity_x'] * 0.8
                elif new_x > canvas_width - droplet['radius']:
                    new_x = canvas_width - droplet['radius']
                    droplet['velocity_x'] = -droplet['velocity_x'] * 0.8
            
            # 上下边界检查
            if new_y < droplet['radius']:
                new_y = droplet['radius']
                droplet['velocity_y'] = 0
            elif new_y > canvas_height - droplet['radius']:
                new_y = canvas_height - droplet['radius']
                droplet['velocity_y'] = 0
            
            # 更新Canvas上的位置
            dx = new_x - droplet['x']
            dy = new_y - droplet['y']
            
            if abs(dx) > 0.01 or abs(dy) > 0.01:
                self.display_canvas.move(droplet['id'], dx, dy)
            
            # 更新存储的位置和最后更新时间
            droplet['x'] = new_x
            droplet['y'] = new_y
            droplet['last_update_time'] = current_time
        
        if not self.oil_droplets:
            self.is_oil_simulating = False
            return
        
        # 增加刷新频率
        self.oil_timer_id = self.root.after(30, self.update_oil_droplet_motion)

    def calculate_fall_speed_from_balance_formula(self, droplet):
        """用平衡法公式计算下落速度"""
        # 获取参数
        l = float(self.settings_params["运动距离"]) / 1000
        rho = float(self.settings_params["油滴密度"])
        g = float(self.settings_params["重力加速度"])
        d = self.plate_distance
        b = float(self.settings_params["修正常数"])
        p = float(self.settings_params["大气压强"])
        
        # ========== 统一使用Sutherland公式计算空气粘滞系数 ==========
        T = float(self.settings_params["室温"]) + 273.15
        T0 = 273.15
        eta0 = 1.716e-5
        S = 110.4
        eta = eta0 * (T/T0)**1.5 * (T0 + S) / (T + S)
        # ========================================================
        
        # 油滴参数
        a = droplet['actual_radius']
        q = droplet['charge']
        U_balance = droplet['balanced_voltage']
        
        if q <= 0 or U_balance <= 0:
            return None
        
        try:
            correction = 1 + b / (p * a)
            
            # 计算 t_g
            K = q * U_balance * math.sqrt(2 * rho * g) / (18 * math.pi * d)
            K_pow = K ** (2/3)
            t_g = (eta * l) / (K_pow * correction)
            
            # 计算 v_g = l / t_g
            v_g = l / t_g if t_g > 0 else 0
            return v_g
            
        except:
            return None
    
    def calculate_rise_speed_force_balance(self, droplet):
        """力学平衡法计算上升速度（备用）"""
        # 获取参数
        d = self.plate_distance
        g = float(self.settings_params["重力加速度"])
        b = float(self.settings_params["修正常数"])
        p = float(self.settings_params["大气压强"])
        
        # 空气粘滞系数
        T = float(self.settings_params["室温"]) + 273.15
        eta = 1.818e-5 * (T/293.15)**0.735
        
        # 油滴参数
        a = droplet['actual_radius']
        q = droplet['charge']
        mass = droplet['mass']
        U_rise = self.lift_voltage.get() + self.balance_voltage_slider.get()
        
        if U_rise <= 0:
            return 0
        
        # 力学平衡计算
        correction = 1 + b / (p * a)
        E_rise = U_rise / d
        F_e_up = q * E_rise
        F_g = mass * g
        net_force = F_e_up - F_g
        
        if net_force > 0:
            v_rise = net_force / (6 * math.pi * eta * a * correction)
            return v_rise
        else:
            return 0  # 无法上升
    
    def calculate_rise_speed_from_formulas(self, droplet):
        """从平衡法和动态法公式反推上升速度"""
        # 获取实验参数
        l = float(self.settings_params["运动距离"]) / 1000  # m
        rho = float(self.settings_params["油滴密度"])        # kg/m³
        g = float(self.settings_params["重力加速度"])        # m/s²
        d = self.plate_distance                             # 极板间距 m
        b = float(self.settings_params["修正常数"])          # 修正常数 N/m
        p = float(self.settings_params["大气压强"])          # 大气压强 Pa
        T = float(self.settings_params["室温"]) + 273.15     # K
        
        # ========== 统一使用Sutherland公式计算空气粘滞系数 ==========
        T0 = 273.15
        eta0 = 1.716e-5
        S = 110.4
        eta = eta0 * (T/T0)**1.5 * (T0 + S) / (T + S)  # Pa·s
        # ========================================================
        
        # 油滴参数
        a = droplet['actual_radius']          # 油滴实际半径 (m)
        q = droplet['charge']                 # 油滴电荷 (C)
        U_balance = droplet['balanced_voltage']  # 平衡电压 (V)
        U_rise = self.lift_voltage.get() + self.balance_voltage_slider.get()  # 上升电压 (V)
        
        # 检查输入有效性
        if q <= 0 or U_balance <= 0 or U_rise <= 0 or a <= 0:
            print(f"无效参数: q={q}, U_balance={U_balance}, U_rise={U_rise}, a={a}")
            return None
        
        try:
            # ========== 步骤1：用平衡法公式计算下落时间 t_g ==========
            correction = 1 + b / (p * a)  # 修正因子
            
            # 计算中间项 K = q * U_balance * sqrt(2ρg) / (18πd)
            K = q * U_balance * math.sqrt(2 * rho * g) / (18 * math.pi * d)
            
            if K <= 0:
                print(f"K值无效: {K}")
                return None
            
            # 计算 (K)^(2/3)
            K_pow = K ** (2/3)
            
            # 计算下落时间 t_g
            t_g = (eta * l) / (K_pow * correction)
            
            if t_g <= 0 or t_g > 100:  # 合理范围检查
                print(f"下落时间异常: t_g={t_g}")
                return None
            
            print(f"计算得到下落时间: t_g = {t_g:.4f} s")
            
            # ========== 步骤2：用动态法公式计算上升时间 t_e ==========
            # 计算动态法公式中的分母项
            # C = (18π/√(2ρg)) * d * (ηl/correction)^(3/2) * √(1/t_g)
            C1 = 18 * math.pi / math.sqrt(2 * rho * g)  # 第一部分
            C2 = math.pow((eta * l) / correction, 1.5)  # 第二部分
            C3 = math.sqrt(1 / t_g)                     # 第三部分
            
            denominator = C1 * d * C2 * C3  # 总分母
            
            if denominator <= 0:
                print(f"分母无效: {denominator}")
                return None
            
            # 计算 1/t_e
            inv_t_e = (q * U_rise) / denominator - (1 / t_g)
            
            if inv_t_e <= 0:
                # 这意味着电场力不足以克服重力
                print(f"无法上升: 1/t_e = {inv_t_e}")
                return None
            
            # 计算上升时间 t_e
            t_e = 1 / inv_t_e
            
            if t_e <= 0 or t_e > 100:  # 合理范围检查
                print(f"上升时间异常: t_e={t_e}")
                return None
            
            print(f"计算得到上升时间: t_e = {t_e:.4f} s")
            
            # ========== 步骤3：计算上升速度 ==========
            v_e = l / t_e
            
             # 添加上升速度修正因子（如果计算结果偏大）
            correction_factor = 0.94  # 根据实际情况调整，0.5表示减小一半
            
            v_e = v_e * correction_factor

            print(f"计算得到上升速度: v_e = {v_e:.2e} m/s")
            
            return v_e
            
        except (ValueError, ZeroDivisionError) as e:
            print(f"计算上升速度时出错: {e}")
            return None
    
    def update_balance_status(self):
        """更新平衡状态显示"""
        pass
        # if not self.oil_droplets:
        #     # 如果没有油滴，显示提示
        #     if hasattr(self, 'balance_status_text'):
        #         self.display_canvas.delete(self.balance_status_text)
            
        #     current_voltage = self.balance_voltage_slider.get()
        #     self.balance_status_text = self.display_canvas.create_text(
        #         200, 20, text=f"电压:{current_voltage}V | 无油滴 (点击'喷入油雾')", 
        #         font=('Arial', 10, 'bold'), fill='gray'
        #     )
        #     return
        
        # balanced_count = sum(1 for d in self.oil_droplets if d['is_balanced'])
        # total_count = len(self.oil_droplets)
        
        # current_voltage = self.balance_voltage_slider.get()
        
        # # 在屏幕上显示平衡状态
        # if hasattr(self, 'balance_status_text'):
        #     self.display_canvas.delete(self.balance_status_text)
        
        # # 显示平衡的油滴编号
        # balanced_droplets = []
        # for i, droplet in enumerate(self.oil_droplets):
        #     voltage_diff = abs(current_voltage - droplet['balanced_voltage'])
        #     if voltage_diff < 1.0:
        #         balanced_droplets.append(str(i+1))
        
        # if balanced_droplets:
        #     status_text = f"电压:{current_voltage}V | 平衡油滴: #{','.join(balanced_droplets)}"
        #     color = "green"
        # else:
        #     # 找到最接近平衡的油滴
        #     closest_droplet = min(
        #         self.oil_droplets, 
        #         key=lambda d: abs(d['balanced_voltage'] - current_voltage)
        #     )
        #     closest_diff = closest_droplet['balanced_voltage'] - current_voltage
        #     closest_index = self.oil_droplets.index(closest_droplet) + 1
            
        #     status_text = f"电压:{current_voltage}V | 最近:油滴#{closest_index}(差{closest_diff:+.1f}V)"
        #     color = "blue" if abs(closest_diff) < 5 else "black"
        
        # self.balance_status_text = self.display_canvas.create_text(
        #     200, 20, text=status_text, 
        #     font=('Arial', 10, 'bold'), fill=color
        # )

    def on_lift_voltage_changed(self, value):
        """提升电压滑块变化时的回调"""
        self.lift_voltage_label.config(text=f"{value} V")
        self.update_voltage_display()
        
    def on_balance_voltage_changed(self, value):
        """平衡电压滑块变化时的回调"""
        voltage = int(value)
        self.balance_voltage_label.config(text=f"{voltage} V")
        
        # 更新参数
        self.balance_params["平衡电压"] = f"{voltage:.0f}"
        self.dynamic_params["平衡电压"] = f"{voltage:.0f}"
        
        # # 如果油滴存在，显示平衡状态
        # if self.oil_droplets:
        #     balanced_droplets = []
        #     for i, droplet in enumerate(self.oil_droplets):
        #         voltage_diff = abs(voltage - droplet['balanced_voltage'])
        #         if voltage_diff < 1.0:
        #             balanced_droplets.append(i+1)
            
        #     if balanced_droplets:
        #         print(f"🎯 电压{voltage}V: 油滴#{balanced_droplets}已达到平衡！")
        #     else:
        #         # 找到最接近平衡的油滴
        #         closest_droplet = min(
        #             self.oil_droplets, 
        #             key=lambda d: abs(d['balanced_voltage'] - voltage)
        #         )
        #         closest_index = self.oil_droplets.index(closest_droplet) + 1
        #         diff = closest_droplet['balanced_voltage'] - voltage
                
        #         if abs(diff) < 2:
        #             print(f"🔍 接近平衡: 油滴#{closest_index} 差{diff:+.1f}V")
        #         elif abs(diff) < 5:
        #             direction = "调高" if diff > 0 else "调低"
        #             print(f"📏 调整建议: {direction}电压{diff:.1f}V可使油滴#{closest_index}平衡")
        
        self.update_voltage_display()
        
    def update_voltage_display(self):
        """根据电压模式和滑块值更新显示屏上的电压显示"""
        mode = self.voltage_mode.get()
        balance_value = self.balance_voltage_slider.get()
        lift_value = self.lift_voltage.get()
        
        if mode == "提升":
            # 上升电压 = 提升电压 + 平衡电压
            total_voltage = lift_value + balance_value
            self.dynamic_params["上升电压"] = f"{total_voltage:.0f}"
            
            # 更新动态法界面的上升电压显示
            if hasattr(self, 'rise_voltage_text'):
                self.display_canvas.itemconfig(self.rise_voltage_text, 
                                            text=f"{total_voltage:.0f} V")
            
            print(f"提升模式: 上升电压 = {lift_value} + {balance_value} = {total_voltage}V")
            
        elif mode == "平衡":
            # 平衡电压 = 平衡电压滑块值
            self.balance_params["平衡电压"] = f"{balance_value:.0f}"

            self.dynamic_params["平衡电压"] = f"{balance_value:.0f}"
            
            # 更新平衡法界面的平衡电压显示
            if hasattr(self, 'balance_voltage_text'):
                self.display_canvas.itemconfig(self.balance_voltage_text, 
                                            text=f"{balance_value:.0f} V")
            
            if hasattr(self, 'dynamic_balance_text'):
                self.display_canvas.itemconfig(self.dynamic_balance_text, 
                                            text=f"{balance_value:.0f} V")
            print(f"平衡模式: 平衡电压 = {balance_value}V")
            
        elif mode == "下落":
            print(f"下落模式: 平衡电压 = {balance_value}V")
            # 在下落模式下，可以添加其他逻辑
        #   
    def init_display_area(self):
        """初始化显示屏区域"""
        # # 创建画布用于显示图片和参数
        # self.display_canvas = tk.Canvas(self.frame_top_left, width=400, height=300)
        # self.display_canvas.pack(padx=10, pady=10)
        
        # # 不需要Frame容器了
        # # 直接使用Canvas存储界面元素
        # self.display_items = {}  # 存储Canvas上的元素引用
        
        # # 初始化显示
        # self.switch_display_interface("平衡法")
        pass

    def create_instrument_buttons(self):
        """在实验仪区域创建控制按钮"""
        # 先显示背景图片
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            img_path3 = get_resource_path(os.path.join('background', "FD-MLG-A.jpg"))
            img3 = Image.open(img_path3)
            img3 = img3.resize((800, 300), Image.Resampling.LANCZOS)
            self.photo3 = ImageTk.PhotoImage(img3)
            bg_label = tk.Label(self.frame_top_right, image=self.photo3)
            bg_label.place(x=0, y=0, relwidth=1, relheight=1)
        except Exception as e:
            print(f"无法加载实验仪图片: {e}")
        
        # 不使用Frame，直接在区域上放置按钮
        self.power_on_button = tk.Button(
            self.frame_top_right,
            text="I",
            command=self.turn_on_power,
            font=('Arial', 9),
            bg='#DC6E89',
            fg='black',
            width=3,
            height=1,
            relief=tk.FLAT,
            # activebackground='lightgreen'
        )
        # 调整位置到你想要放置的位置
        self.power_on_button.place(x=718, y=185)  # 可自定义位置
        
        # 关闭电源按钮 - 位置 (x=500, y=130)
        self.power_off_button = tk.Button(
            self.frame_top_right,
            text="O",
            command=self.turn_off_power,
            font=('Arial', 9),
            bg='#DC6E89',
            fg='black',
            width=3,
            height=1,
            relief=tk.FLAT,
            # activebackground='lightcoral'
        )
        self.power_off_button.place(x=718, y=130+82)  # 可自定义位置
        # ====== 电压控制按钮组 ======
        # 这三个按钮是互斥的单选按钮，只能选择一个
        
        # 当前选择的电压模式
        self.voltage_mode = tk.StringVar(value="平衡")  # 初始选择平衡模式
        
        # 提升按钮 - 可以自定义位置 (x=150, y=100)
        self.lift_button = tk.Radiobutton(
            self.frame_top_right, 
            text="", 
            variable=self.voltage_mode, 
            value="提升",
            command=self.on_voltage_mode_changed,
            font=('Arial', 8),
            bg='white',
            fg='white',
            # selectcolor='#333333',
            activebackground='#545454',
            activeforeground='white',
            indicatoron=0,  # 不使用小圆圈指示器，显示为普通按钮
            width=6,
            height=1,
            relief=tk.FLAT
        )
        self.lift_button.place(x=150-50-2-12, y=100+60-5)  # 可自定义位置
        
        # 平衡按钮 - 可以自定义位置 (x=150, y=130)
        self.balance_button = tk.Radiobutton(
            self.frame_top_right, 
            text="", 
            variable=self.voltage_mode, 
            value="平衡",
            command=self.on_voltage_mode_changed,
            font=('Arial', 8),
            bg='white',
            fg='white',
            # selectcolor='#333333',
            activebackground='#545454',
            activeforeground='white',
            indicatoron=0,
            width=6,
            height=1,
            relief=tk.FLAT
        )
        self.balance_button.place(x=100-2-12, y=195-5)  # 可自定义位置
        
        # 下落按钮 - 可以自定义位置 (x=150, y=160)
        self.fall_button = tk.Radiobutton(
            self.frame_top_right, 
            text="", 
            variable=self.voltage_mode, 
            value="下落",
            command=self.on_voltage_mode_changed,
            font=('Arial', 8),
            bg='white',
            fg='white',
            # selectcolor='#333333',
            activebackground='#545454',
            activeforeground='white',
            indicatoron=0,
            width=6,
            height=1,
            relief=tk.FLAT
        )
        self.fall_button.place(x=150-50-2-12, y=160+70-5)  # 可自定义位置
        
        # 初始处于平衡状态
        self.balance_button.select()  # 确保平衡按钮被选中
        self.on_balance_mode()  # 调用平衡模式初始化函数
        
        # ====== 原来的控制按钮 ======
        # 向上按钮 - 位置 (x=300, y=100)
        up_button = tk.Button(self.frame_top_right, text="", font=('Arial', 8), 
                            width=2, height=1, command=self.navigate_up, bg='#545454',relief=tk.FLAT)
        up_button.place(x=300-30+37-12, y=100+72-7)
        
        # 向下按钮 - 位置 (x=300, y=170)
        down_button = tk.Button(self.frame_top_right, text="", font=('Arial', 8),
                                width=2, height=1, command=self.navigate_down, bg='#545454',relief=tk.FLAT)
        down_button.place(x=300-30+37-12, y=170+74-7)
        
        # 确定按钮 - 位置 (x=380, y=135)
        ok_button = tk.Button(self.frame_top_right, text="", font=('Arial', 8),
                            width=2, height=1, command=self.select_item, bg='#545454',relief=tk.FLAT)
        ok_button.place(x=350+37-12, y=172-7)
        
        # 返回按钮 - 位置 (x=220, y=135)
        back_button = tk.Button(self.frame_top_right, text="", font=('Arial', 8),
                                width=2, height=1, command=self.go_back, bg='#545454',relief=tk.FLAT)
        back_button.place(x=350+37-12, y=244-7)

        # 初始状态：电源关闭，显示空白屏幕
        self.initialize_display()

    def turn_on_power(self):
        """打开电源"""
        if not self.power_on:
            self.power_on = True
            print("电源已打开")
            
            # 启用电压控制按钮
            self.lift_button.config(state='normal')
            self.balance_button.config(state='normal')
            self.fall_button.config(state='normal')
            
            # 启用其他控制按钮（如果需要）
            if hasattr(self, 'spray_button'):
                self.spray_button.config(state='normal')
            if hasattr(self, 'clear_button'):
                self.clear_button.config(state='normal')
            
            # 启用滑块
            if hasattr(self, 'lift_slider'):
                self.lift_slider.config(state='normal')
            if hasattr(self, 'balance_slider'):
                self.balance_slider.config(state='normal')
            if hasattr(self, 'focus_slider'):
                self.focus_slider.config(state='normal')

            if hasattr(self, 'increase_button'):
                self.increase_button.config(state='normal')
            if hasattr(self, 'decrease_button'):
                self.decrease_button.config(state='normal')
            # 显示电荷计算框
            if hasattr(self, 'charge_frame'):
                self.charge_frame.place(x=10+440+20+25, y=320-200, width=70, height=180)
            
            # 打开显示屏，显示平衡法界面
            self.show_display("平衡法")
            
            # 更新按钮状态（可选：改变按钮颜色或文本）
            # self.power_on_button.config(bg='lightgreen')
            # self.power_off_button.config(bg='white')

    def turn_off_power(self):
        """关闭电源"""
        if self.power_on:
            self.power_on = False
            print("电源已关闭")
            
            # 禁用所有控制按钮
            self.lift_button.config(state='disabled')
            self.balance_button.config(state='disabled')
            self.fall_button.config(state='disabled')
            
            # 禁用其他控制按钮
            if hasattr(self, 'spray_button'):
                self.spray_button.config(state='disabled')
            if hasattr(self, 'clear_button'):
                self.clear_button.config(state='disabled')
            
            if hasattr(self, 'increase_button'):
                self.increase_button.config(state='disabled')
            if hasattr(self, 'decrease_button'):
                self.decrease_button.config(state='disabled')

            # 禁用滑块
            if hasattr(self, 'lift_slider'):
                self.lift_slider.config(state='disabled')
            if hasattr(self, 'balance_slider'):
                self.balance_slider.config(state='disabled')
            if hasattr(self, 'focus_slider'):
                self.focus_slider.config(state='disabled')
            
            # 隐藏电荷计算框
            if hasattr(self, 'charge_frame'):
                self.charge_frame.place_forget()
            
            # 停止所有正在进行的操作
            self.stop_all_operations()
            
            # 关闭显示屏，显示空白
            self.hide_display()
            
            # 更新按钮状态
            # self.power_on_button.config(bg='white')
            # self.power_off_button.config(bg='lightcoral')

    def initialize_display(self):
        """初始化显示屏"""
        # 创建显示屏区域但不显示内容
        self.display_canvas = tk.Canvas(self.frame_top_left, width=400, height=300, bg='black')
        self.display_canvas.pack(padx=10, pady=10)
        
        # 初始状态：电源关闭，显示空白
        self.hide_display()

    def hide_display(self):
        """隐藏显示屏内容（电源关闭状态）"""
        self.display_canvas.delete("all")
        # 显示黑色背景
        self.display_canvas.config(bg='black')
        
        # 可选：显示"电源关闭"的提示文字
        self.display_canvas.create_text(
            200, 150, 
            text="电源关闭", 
            font=('Arial', 16, 'bold'), 
            fill='gray',
            tags="power_off_text"
        )
        
        # 重置界面状态
        self.current_interface = None
        self.selected_index = 0
        
        # 停止油滴模拟
        if self.is_oil_simulating:
            self.clear_oil_droplets()

        # 重置平衡法时间数据
        self.balance_params["下落时间"] = "0.00"
        
        # 重置动态法时间数据
        self.dynamic_params["下落时间"] = "0.00"
        self.dynamic_params["上升时间"] = "0.00"
        
        self.charge_display_data = ["0.0", "0.0", "0.0", "0.0", "0.0", "0.0", "0.0"]
        # 重置电荷显示框的显示值（只重置显示，不清除存储的数据）
        for i in range(7):
            if hasattr(self, 'charge_labels') and i < len(self.charge_labels):
                self.charge_labels[i].config(text="0.0")
        
        # 重置当前电荷索引
        self.current_charge_index = 0


    def show_display(self, interface_name="平衡法"):
        """显示指定界面（电源打开状态）"""
        # 清除当前显示
        self.display_canvas.delete("all")
        
        # 设置背景色
        self.display_canvas.config(bg='white')
        
        # 切换界面
        self.switch_display_interface(interface_name)

    def stop_all_operations(self):
        """停止所有正在进行的操作"""
        # 停止计时
        if self.is_timing:
            self.is_timing = False
            if self.timer_id:
                self.root.after_cancel(self.timer_id)
                self.timer_id = None
        
        # 停止油滴模拟
        if self.is_oil_simulating:
            self.is_oil_simulating = False
            if self.oil_timer_id:
                self.root.after_cancel(self.oil_timer_id)
                self.oil_timer_id = None
        
        # 停止自动调整
        if self.is_auto_adjusting:
            self.stop_auto_adjust()

    def on_voltage_mode_changed(self):
        """电压模式改变时的回调函数"""
        mode = self.voltage_mode.get()
        print(f"电压模式已切换为: {mode}")
        
        # 根据不同的电压模式执行不同的操作
        if mode == "提升":
            self.on_lift_mode()
        elif mode == "平衡":
            self.on_balance_mode()
        elif mode == "下落":
            self.on_fall_mode()
        
        # 更新电压显示
        self.update_voltage_display()

    def on_lift_mode(self):
        """提升模式下的操作"""
        print("切换到提升模式")
        # 在提升模式下，更新动态法界面的上升电压
        
    def on_balance_mode(self):
        """平衡模式下的操作"""
        print("切换到平衡模式")
        # 在平衡模式下，更新平衡法界面的平衡电压
        
    def on_fall_mode(self):
        """下落模式下的操作"""
        print("切换到下落模式：电场关闭，油滴只受重力作用自由下落")
        print("提示：在下落模式下，可以测量油滴的下落时间")
        
        # 更新电压显示
        self.update_voltage_display()
        
        # 如果有油滴，重置颜色为红色（因为不再平衡）
        for droplet in self.oil_droplets:
            if droplet['color'] != 'red':
                self.display_canvas.itemconfig(droplet['id'], fill='red')
                droplet['color'] = 'red'
            droplet['is_balanced'] = False
            # 重置最后更新时间，避免dt过大
            droplet['last_update_time'] = time.time()
        
        # 确保模拟继续运行
        if self.oil_droplets and not self.is_oil_simulating:
            self.start_oil_simulation()
    
    def show_balance_interface_canvas(self):
        """在Canvas上显示平衡法界面"""
        canvas = self.display_canvas
        
        # 存储所有可选择的项目（按照你想要的导航顺序）
        self.balance_items = []
        
        # 实验方法标题 - 位置 (x=20, y=20)
        method_text = canvas.create_text(20, 55, text="平衡法", 
                                        font=('Arial', 10, 'bold'), anchor='w', tags="balance_method")
        self.balance_items.append(method_text)  # 索引0：实验方法
        
        # 平衡电压 - 位置 (x=20, y=50) - 改为只读标签
        self.balance_voltage_text = canvas.create_text(30, 100, text=f"{self.balance_params['平衡电压']} V", 
                                                    font=('Arial', 10), anchor='w', fill='black')
        # 注意：平衡电压不加入导航循环
        
        # 下落时间 - 位置 (x=20, y=80) - 改为只读标签
        self.fall_time_text = canvas.create_text(30, 145, text=f"{self.balance_params['下落时间']} s", 
                                            font=('Arial', 10), anchor='w', fill='black')
        self.balance_items.append(self.fall_time_text)  # 索引1：下落时间
        
        # 计算并保存按钮 - 位置 (x=20, y=110)
        calc_button = canvas.create_text(10, 270, text="计算并保存", 
                                        font=('Arial', 10), anchor='w', tags="calc_save")
        self.balance_items.append(calc_button)  # 索引2：计算并保存
        
        # 参数设置按钮 - 位置 (x=20, y=280)
        settings_text = canvas.create_text(330, 270, text="参数设置", 
                                        font=('Arial', 10), anchor='w', tags="settings")
        self.balance_items.append(settings_text)  # 索引3：参数设置
        
        # 重新绘制油滴
        self.redraw_oil_droplets()


    def switch_display_interface(self, interface_name):
        """切换显示屏区域的界面"""
        # 如果正在计时，先停止计时
        if self.is_timing:
            if self.current_interface == "平衡法":
                self.stop_timing("平衡法", "下落时间")
            elif self.current_interface == "动态法":
                # 判断当前选中的是哪个时间
                if hasattr(self, 'dynamic_items') and self.selected_index < len(self.dynamic_items):
                    if self.selected_index == 1:  # 上升时间
                        self.stop_timing("动态法", "上升时间")
                    elif self.selected_index == 2:  # 下落时间
                        self.stop_timing("动态法", "下落时间")
        
        # 切换界面时重置所有时间数据
        if interface_name != self.current_interface:  # 只有真的切换界面时才重置
            print(f"从 {self.current_interface} 切换到 {interface_name}")
            
            # 只在切换到不同实验方法时重置时间数据，不清空油滴
            if interface_name in ["平衡法", "动态法"] and self.current_interface in ["平衡法", "动态法"]:
                # 从平衡法切换到动态法或反之，重置时间数据
                self.balance_params["下落时间"] = "0.0"
                self.dynamic_params["下落时间"] = "0.0"
                self.dynamic_params["上升时间"] = "0.0"
            elif interface_name == "参数设置":
                # 切换到参数设置时，保留时间数据，因为可能会返回
                print("切换到参数设置界面")
            elif self.current_interface == "参数设置" and interface_name in ["平衡法", "动态法"]:
                # 从参数设置返回时，不重置时间数据
                print(f"从参数设置返回到 {interface_name}")
            
            # 注意：不清空油滴！油滴在切换实验方法时保持存在
        
        self.current_interface = interface_name
        
        # 清除Canvas上的所有元素（除了背景）
        self.display_canvas.delete("all")
        
        # 只有在切换到参数设置时才清空油滴
        # if interface_name == "参数设置":
        #     print("切换到参数设置，清空油滴")
        #     self.clear_oil_droplets()
        # else:
        #     # 切换到平衡法或动态法时，不清空油滴
        #     print(f"切换到{interface_name}，保持现有{len(self.oil_droplets)}个油滴")
        
        # 更新背景图片
        if interface_name in self.images:
            img = self.images[interface_name].resize((400, 300), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self.photo_refs[interface_name] = photo  # 保存引用
            self.display_canvas.create_image(0, 0, image=photo, anchor='nw')
        
        # 重置选中索引为0（总是从第一个项目开始）
        self.selected_index = 0
        
        # 根据界面显示不同的参数
        if interface_name == "平衡法":
            self.show_balance_interface_canvas()
        elif interface_name == "动态法":
            self.show_dynamic_interface_canvas()
        elif interface_name == "参数设置":
            self.show_settings_interface_canvas()
        
        # 延迟高亮，确保所有控件都已创建
        self.root.after(100, self.highlight_selected_canvas)
        
        # 切换界面后更新电压显示
        self.root.after(150, self.update_voltage_display)
        
        # 如果还有油滴，重新开始模拟
        if interface_name in ["平衡法", "动态法"] and self.oil_droplets and not self.is_oil_simulating:
            print(f"切换到{interface_name}，重新开始油滴模拟")
            self.start_oil_simulation()
    
    def highlight_selected_canvas(self):
        """高亮Canvas上的当前选中项"""
        # 清除所有文本类型的高亮
        try:
            highlighted_items = self.display_canvas.find_withtag("highlight")
            for item_id in highlighted_items:
                try:
                    # 只对文本类型应用fill选项
                    item_type = self.display_canvas.type(item_id)
                    if item_type == "text":
                        # 检查这个项目是否是正在计时的时间显示
                        if self.is_timing:
                            if self.current_interface == "平衡法" and hasattr(self, 'fall_time_text') and item_id == self.fall_time_text:
                                continue  # 跳过正在计时的项目
                            elif self.current_interface == "动态法":
                                if (hasattr(self, 'rise_time_text') and item_id == self.rise_time_text) or \
                                (hasattr(self, 'dynamic_fall_text') and item_id == self.dynamic_fall_text):
                                    continue  # 跳过正在计时的项目
                        
                        self.display_canvas.itemconfig(item_id, fill="black", font=('Arial', 10))
                except:
                    pass
        except:
            pass
        
        # 根据当前界面获取项目列表
        if self.current_interface == "平衡法":
            items = self.balance_items
        elif self.current_interface == "动态法":
            items = self.dynamic_items
        elif self.current_interface == "参数设置":
            items = self.settings_items
        else:
            return
        
        # 高亮当前选中项
        if 0 <= self.selected_index < len(items):
            item_id = items[self.selected_index]
            
            try:
                # 检查项目类型
                item_type = self.display_canvas.type(item_id)
                if item_type == "text":
                    # 检查这个项目是否是正在计时的时间显示
                    if self.is_timing:
                        if self.current_interface == "平衡法" and hasattr(self, 'fall_time_text') and item_id == self.fall_time_text:
                            return  # 如果是正在计时的项目，不设置为蓝色
                        elif self.current_interface == "动态法":
                            if (hasattr(self, 'rise_time_text') and item_id == self.rise_time_text) or \
                            (hasattr(self, 'dynamic_fall_text') and item_id == self.dynamic_fall_text):
                                return  # 如果是正在计时的项目，不设置为蓝色
                    
                    # 设置为蓝色高亮
                    self.display_canvas.itemconfig(item_id, fill="blue", font=('Arial', 10, 'bold'))
            except Exception as e:
                print(f"高亮时出错: {e}")
            
            # 给选中项添加highlight标签
            self.display_canvas.addtag_withtag("highlight", item_id)
    
    def show_dynamic_interface_canvas(self):
        """在Canvas上显示动态法界面"""
        canvas = self.display_canvas
        
        # 存储所有可选择的项目（按照你想要的导航顺序）
        self.dynamic_items = []
        
        # 实验方法标题 - 位置 (x=20, y=20)
        method_text = canvas.create_text(20, 55, text="动态法", 
                                        font=('Arial', 10, 'bold'), anchor='w', tags="dynamic_method")
        self.dynamic_items.append(method_text)  # 索引0：实验方法
        
        # 上升时间 - 位置 (x=20, y=50) - 改为只读标签
        self.rise_time_text = canvas.create_text(30, 190, text=f"{self.dynamic_params['上升时间']} s", 
                                            font=('Arial', 10), anchor='w', fill='black')
        self.dynamic_items.append(self.rise_time_text)  # 索引1：上升时间
        
        # 下落时间 - 位置 (x=20, y=80) - 改为只读标签
        self.dynamic_fall_text = canvas.create_text(30, 235, text=f"{self.dynamic_params['下落时间']} s", 
                                                font=('Arial', 10), anchor='w', fill='black')
        self.dynamic_items.append(self.dynamic_fall_text)  # 索引2：下落时间
        
        # 上升电压 - 位置 (x=20, y=110) - 改为只读标签
        self.rise_voltage_text = canvas.create_text(30, 80+65, text=f"{self.dynamic_params['上升电压']} V", 
                                                font=('Arial', 10), anchor='w', fill='black')
        
        # 平衡电压 - 位置 (x=20, y=140) - 改为只读标签
        self.dynamic_balance_text = canvas.create_text(30, 100, text=f"{self.dynamic_params['平衡电压']} V", 
                                                    font=('Arial', 10), anchor='w', fill='black')
        
        # 计算并保存按钮 - 位置 (x=20, y=170)
        dynamic_calc_button = canvas.create_text(10, 270, text="计算并保存", 
                                                font=('Arial', 10), anchor='w')
        self.dynamic_items.append(dynamic_calc_button)  # 索引3：计算并保存
        
        # 参数设置按钮 - 位置 (x=20, y=200)
        dynamic_settings_button = canvas.create_text(330, 270, text="参数设置", 
                                                    font=('Arial', 10), anchor='w')
        self.dynamic_items.append(dynamic_settings_button)  # 索引4：参数设置
        
        # 重新绘制油滴
        self.redraw_oil_droplets()

    def redraw_oil_droplets(self):
        """重新绘制所有油滴到Canvas上"""
        if not self.oil_droplets:
            return
        
        print(f"重新绘制 {len(self.oil_droplets)} 个大小不同的油滴")
        
        # 重新创建所有油滴在Canvas上
        for droplet in self.oil_droplets:
            x = droplet['x']
            y = droplet['y']
            radius = droplet['radius']
            color = droplet['color']
            
            # 在Canvas上重新绘制油滴
            droplet['id'] = self.display_canvas.create_oval(
                x - radius, y - radius,
                x + radius, y + radius,
                fill=color, outline='darkred', width=1
            )

    def show_settings_interface_canvas(self):
        """在Canvas上显示参数设置界面"""
        canvas = self.display_canvas
        
        # 存储所有可选择的项目
        self.settings_items = []
        
        # 运动距离 - 位置 (x=20, y=20)
        canvas.create_text(20, 20, text="运动距离:", font=('Arial', 10), anchor='w')
        self.distance_var = tk.StringVar(value=self.settings_params["运动距离"])
        self.distance_entry = tk.Entry(canvas, textvariable=self.distance_var,
                                    width=10, font=('Arial', 10), bg='white')
        distance_window = canvas.create_window(100, 20, window=self.distance_entry, anchor='w')
        canvas.create_text(180, 20, text="mm", font=('Arial', 9), anchor='w')
    
        # 油滴密度 - 位置 (x=20, y=50)
        canvas.create_text(20, 50, text="油滴密度:", font=('Arial', 10), anchor='w')
        self.density_var = tk.StringVar(value=self.settings_params["油滴密度"])
        self.density_entry = tk.Entry(canvas, textvariable=self.density_var,
                                    width=10, font=('Arial', 10), bg='white')
        density_window = canvas.create_window(100, 50, window=self.density_entry, anchor='w')
        canvas.create_text(180, 50, text="kg/m³", font=('Arial', 9), anchor='w')
        
        # 室温 - 位置 (x=20, y=80)
        canvas.create_text(20, 80, text="室温:", font=('Arial', 10), anchor='w')
        self.temp_var = tk.StringVar(value=self.settings_params["室温"])
        self.temp_entry = tk.Entry(canvas, textvariable=self.temp_var,
                                width=10, font=('Arial', 10), bg='white')
        temp_window = canvas.create_window(100, 80, window=self.temp_entry, anchor='w')
        canvas.create_text(180, 80, text="℃", font=('Arial', 9), anchor='w')

        # 大气压强 - 位置 (x=20, y=110)
        canvas.create_text(20, 110, text="大气压强:", font=('Arial', 10), anchor='w')
        self.pressure_var = tk.StringVar(value=self.settings_params["大气压强"])
        self.pressure_entry = tk.Entry(canvas, textvariable=self.pressure_var,
                                    width=10, font=('Arial', 10), bg='white')
        pressure_window = canvas.create_window(100, 110, window=self.pressure_entry, anchor='w')
        canvas.create_text(180, 110, text="Pa", font=('Arial', 9), anchor='w')

        # 重力加速度 - 位置 (x=20, y=140)
        canvas.create_text(20, 140, text="重力加速度:", font=('Arial', 10), anchor='w')
        self.gravity_var = tk.StringVar(value=self.settings_params["重力加速度"])
        self.gravity_entry = tk.Entry(canvas, textvariable=self.gravity_var,
                                    width=10, font=('Arial', 10), bg='white')
        gravity_window = canvas.create_window(100, 140, window=self.gravity_entry, anchor='w')
        canvas.create_text(180, 140, text="m/s²", font=('Arial', 9), anchor='w')

        # 修正常数 - 位置 (x=20, y=170)
        canvas.create_text(20, 170, text="修正常数:", font=('Arial', 10), anchor='w')
        self.correction_var = tk.StringVar(value=self.settings_params["修正常数"])
        self.correction_entry = tk.Entry(canvas, textvariable=self.correction_var,
                                    width=10, font=('Arial', 10), bg='white')
        correction_window = canvas.create_window(100, 170, window=self.correction_entry, anchor='w')
        canvas.create_text(180, 170, text="N/m", font=('Arial', 9), anchor='w')

        # 保存并返回按钮 - 位置 (x=20, y=210)
        save_return_text = canvas.create_text(20, 210, text="保存并返回", 
                                            font=('Arial', 10, 'bold'), anchor='w')
        self.settings_items.append(save_return_text)
        
        # 重置为默认按钮 - 位置 (x=20, y=240)
        reset_text = canvas.create_text(20, 240, text="重置为默认", 
                                    font=('Arial', 10, 'bold'), anchor='w')
        self.settings_items.append(reset_text)
    

    def start_timing(self, method, time_type):
        """开始计时 - 从0开始"""
        self.is_timing = True
        self.elapsed_time = 0.0  # 重置为0
        self.start_time = time.time()  # 使用time模块获取当前时间（秒）
        
        # 清除之前的定时器
        if self.timer_id:
            self.root.after_cancel(self.timer_id)
            self.timer_id = None
        
        # 立即显示0.00s
        if method == "平衡法" and time_type == "下落时间" and hasattr(self, 'fall_time_text'):
            self.display_canvas.itemconfig(self.fall_time_text, 
                                        text="0.00 s")
            # 将时间显示变为红色，表示正在计时
            self.display_canvas.itemconfig(self.fall_time_text, fill="red")
            # 从highlight标签中移除，防止被重置颜色
            self.display_canvas.dtag(self.fall_time_text, "highlight")
        elif method == "动态法":
            if time_type == "上升时间" and hasattr(self, 'rise_time_text'):
                self.display_canvas.itemconfig(self.rise_time_text, 
                                            text="0.00 s")
                # 将时间显示变为红色，表示正在计时
                self.display_canvas.itemconfig(self.rise_time_text, fill="red")
                # 从highlight标签中移除，防止被重置颜色
                self.display_canvas.dtag(self.rise_time_text, "highlight")
            elif time_type == "下落时间" and hasattr(self, 'dynamic_fall_text'):
                self.display_canvas.itemconfig(self.dynamic_fall_text, 
                                            text="0.00 s")
                # 将时间显示变为红色，表示正在计时
                self.display_canvas.itemconfig(self.dynamic_fall_text, fill="red")
                # 从highlight标签中移除，防止被重置颜色
                self.display_canvas.dtag(self.dynamic_fall_text, "highlight")
        
        print(f"开始计时: {method} - {time_type}")
        print("提示：正在计时中，上下键已被禁用")
        
        # 开始定时器
        self.update_timer(method, time_type)

    def stop_timing(self, method, time_type):
        """停止计时"""
        if not self.is_timing:
            return
        
        self.is_timing = False
        
        # 取消定时器
        if self.timer_id:
            self.root.after_cancel(self.timer_id)
            self.timer_id = None
        
        # 更新时间数据
        final_time = self.elapsed_time  # 使用已经计算好的时间
        
        if method == "平衡法" and time_type == "下落时间":
            self.balance_params[time_type] = f"{final_time:.2f}"
            print(f"停止计时: {method} - {time_type} = {final_time:.2f}s")
            
            # 恢复时间显示颜色 - 直接调用高亮方法让系统处理颜色
            if hasattr(self, 'fall_time_text'):
                # 重新高亮当前选中的项目
                self.highlight_selected_canvas()
            
        elif method == "动态法":
            self.dynamic_params[time_type] = f"{final_time:.2f}"
            print(f"停止计时: {method} - {time_type} = {final_time:.2f}s")
            
            # 恢复时间显示颜色 - 直接调用高亮方法让系统处理颜色
            if hasattr(self, 'rise_time_text') or hasattr(self, 'dynamic_fall_text'):
                # 重新高亮当前选中的项目
                self.highlight_selected_canvas()
        
        print("提示：计时已停止，可以移动光标")

    def update_timer(self, method, time_type):
        """更新计时器显示 - 50毫秒刷新频率"""
        if not self.is_timing:
            return
        
        # 计算经过的时间（使用time模块）
        current_time = time.time()
        self.elapsed_time = current_time - self.start_time  # 直接计算经过的时间
        
        # 更新显示（保留2位小数）
        time_str = f"{self.elapsed_time:.2f} s"
        if method == "平衡法" and time_type == "下落时间" and hasattr(self, 'fall_time_text'):
            self.display_canvas.itemconfig(self.fall_time_text, 
                                        text=time_str)
        elif method == "动态法":
            if time_type == "上升时间" and hasattr(self, 'rise_time_text'):
                self.display_canvas.itemconfig(self.rise_time_text, 
                                            text=time_str)
            elif time_type == "下落时间" and hasattr(self, 'dynamic_fall_text'):
                self.display_canvas.itemconfig(self.dynamic_fall_text, 
                                            text=time_str)
        
        # 继续定时器，使用50毫秒的间隔（每秒20次刷新）
        self.timer_id = self.root.after(50, lambda: self.update_timer(method, time_type))

    # 修改 calculate_elapsed_time 方法
    def calculate_elapsed_time(self):
        """计算经过的时间（秒）- 直接从elapsed_time变量获取"""
        return self.elapsed_time

    # 修改 navigate_up 和 navigate_down 方法
    def navigate_up(self):
        """向上导航"""
        # 如果正在计时，不允许移动光标
        if self.is_timing:
            print("正在计时，请先停止计时再移动光标")
            return
        
        if self.current_interface == "平衡法":
            self.selected_index = (self.selected_index - 1) % len(self.balance_items)
        elif self.current_interface == "动态法":
            self.selected_index = (self.selected_index - 1) % len(self.dynamic_items)
        elif self.current_interface == "参数设置":
            self.selected_index = (self.selected_index - 1) % len(self.settings_items)
        
        # 调用Canvas版本的高亮方法
        self.highlight_selected_canvas()

    def navigate_down(self):
        """向下导航"""
        # 如果正在计时，不允许移动光标
        if self.is_timing:
            print("正在计时，请先停止计时再移动光标")
            return
        
        if self.current_interface == "平衡法":
            self.selected_index = (self.selected_index + 1) % len(self.balance_items)
        elif self.current_interface == "动态法":
            self.selected_index = (self.selected_index + 1) % len(self.dynamic_items)
        elif self.current_interface == "参数设置":
            self.selected_index = (self.selected_index + 1) % len(self.settings_items)
        
        # 调用Canvas版本的高亮方法
        self.highlight_selected_canvas()

    
    
    def select_item(self):
        """确定按钮的功能"""
        # 检查当前是否正在计时
        if self.is_timing:
            # 如果正在计时，停止计时
            if self.current_interface == "平衡法":
                self.stop_timing("平衡法", "下落时间")
            elif self.current_interface == "动态法":
                if self.selected_index == 1:  # 上升时间
                    self.stop_timing("动态法", "上升时间")
                elif self.selected_index == 2:  # 下落时间
                    self.stop_timing("动态法", "下落时间")
            return
        
        # 如果不是计时状态，正常处理选择
        if self.current_interface == "平衡法":
            self.handle_balance_selection()
        elif self.current_interface == "动态法":
            self.handle_dynamic_selection()
        elif self.current_interface == "参数设置":
            self.handle_settings_selection()
    
    def handle_balance_selection(self):
        """处理平衡法界面的选择"""
        if self.selected_index == 0:  # 实验方法
            # 切换到动态法界面
            self.switch_display_interface("动态法")
        elif self.selected_index == 1:  # 下落时间
            # 开始计时
            self.start_timing("平衡法", "下落时间")
        elif self.selected_index == 2:  # 计算并保存
            self.calculate_and_save("平衡法")
        elif self.selected_index == 3:  # 参数设置
            self.switch_display_interface("参数设置")

    def handle_dynamic_selection(self):
        """处理动态法界面的选择"""
        if self.selected_index == 0:  # 实验方法
            # 切换到平衡法界面
            self.switch_display_interface("平衡法")
        elif self.selected_index == 1:  # 上升时间
            # 开始计时
            self.start_timing("动态法", "上升时间")
        elif self.selected_index == 2:  # 下落时间
            # 开始计时
            self.start_timing("动态法", "下落时间")
        elif self.selected_index == 3:  # 计算并保存
            self.calculate_and_save("动态法")
        elif self.selected_index == 4:  # 参数设置
            self.switch_display_interface("参数设置")


    def update_balance_display(self):
        """更新平衡法界面的显示值"""
        # 如果当前是平衡模式，从滑块获取平衡电压
        if self.voltage_mode.get() == "平衡":
            balance_value = self.balance_voltage_slider.get()
            self.balance_params["平衡电压"] = f"{balance_value:.0f}"
        
        if hasattr(self, 'balance_voltage_text'):
            # 更新平衡电压显示
            self.display_canvas.itemconfig(self.balance_voltage_text, 
                                        text=f"{self.balance_params['平衡电压']} V")
        if hasattr(self, 'fall_time_text'):
            # 更新下落时间显示
            self.display_canvas.itemconfig(self.fall_time_text, 
                                        text=f"{self.balance_params['下落时间']} s")

    def update_dynamic_display(self):
        """更新动态法界面的显示值"""
        # 如果当前是提升模式，计算上升电压
        if self.voltage_mode.get() == "提升":
            total_voltage = self.lift_voltage.get() + self.balance_voltage_slider.get()
            self.dynamic_params["上升电压"] = f"{total_voltage:.0f}"
        
        if hasattr(self, 'rise_time_text'):
            # 更新上升时间显示
            self.display_canvas.itemconfig(self.rise_time_text, 
                                        text=f"{self.dynamic_params['上升时间']} s")
        if hasattr(self, 'dynamic_fall_text'):
            # 更新下落时间显示
            self.display_canvas.itemconfig(self.dynamic_fall_text, 
                                        text=f"{self.dynamic_params['下落时间']} s")
        if hasattr(self, 'rise_voltage_text'):
            # 更新上升电压显示
            self.display_canvas.itemconfig(self.rise_voltage_text, 
                                        text=f"{self.dynamic_params['上升电压']} V")
        if hasattr(self, 'dynamic_balance_text'):
            # 更新平衡电压显示
            self.display_canvas.itemconfig(self.dynamic_balance_text, 
                                        text=f"{self.dynamic_params['平衡电压']} V")
        
    def handle_settings_selection(self):
        """处理参数设置界面的选择"""
        # 保存数据
        self.settings_params["运动距离"] = self.distance_var.get()
        self.settings_params["油滴密度"] = self.density_var.get()
        self.settings_params["室温"] = self.temp_var.get()
        self.settings_params["大气压强"] = self.pressure_var.get()
        self.settings_params["重力加速度"] = self.gravity_var.get()
        self.settings_params["修正常数"] = self.correction_var.get()
        
        if self.selected_index == 0:  # 保存并返回
            # 保存设置
            print("保存设置")
            # 总是返回到平衡法界面
            self.switch_display_interface("平衡法")
        elif self.selected_index == 1:  # 重置为默认
            self.reset_to_default()
    
    def calculate_and_save(self, method):
        """计算并保存数据"""
        # 检查电源状态
        if not self.power_on:
            print("电源关闭，无法计算")
            return
        
        print(f"执行{method}计算并保存")
        
        # 根据实验方法计算油滴电荷
        if method == "平衡法":
            # 使用实际测量数据 - 电压取整数，时间保留2位小数
            balance_voltage = int(float(self.balance_params["平衡电压"]))  # 电压取整
            fall_time = round(float(self.balance_params["下落时间"]), 2)  # 时间保留2位小数
            
            print(f"计算参数: 电压={balance_voltage}V (整数), 时间={fall_time}s (2位小数)")
            
            if balance_voltage == 0 or fall_time == 0:
                print("错误：平衡电压或下落时间为零，无法计算")
                return
            
            # 计算油滴电荷
            charge = self.calculate_balance_method_charge(balance_voltage, fall_time)
            print(f"平衡法计算结果: 电荷 = {charge:.2e} C")
            
            # 更新显示框，显示4位小数
            self.update_charge_display(charge)
            
        else:  # 动态法
            # 使用实际测量数据 - 电压取整数，时间保留2位小数
            balance_voltage = int(float(self.dynamic_params["平衡电压"]))  # 电压取整
            fall_time = round(float(self.dynamic_params["下落时间"]), 2)  # 时间保留2位小数
            rise_voltage = int(float(self.dynamic_params["上升电压"]))  # 电压取整
            rise_time = round(float(self.dynamic_params["上升时间"]), 2)  # 时间保留2位小数
            
            print(f"动态法计算参数:")
            print(f"  平衡电压={balance_voltage}V, 下落时间={fall_time}s")
            print(f"  上升电压={rise_voltage}V, 上升时间={rise_time}s")
            
            if balance_voltage == 0 or fall_time == 0 or rise_voltage == 0 or rise_time == 0:
                print("错误：电压或时间为零，无法计算")
                return
            
            # 计算油滴电荷
            charge = self.calculate_dynamic_method_charge(balance_voltage, fall_time, rise_voltage, rise_time)
            print(f"动态法计算结果: 电荷 = {charge:.2e} C")
            
            # 更新显示框，显示4位小数
            self.update_charge_display(charge)

    
    def calculate_balance_method_charge(self, U, t_g):
        """计算平衡法油滴电荷 - 统一计算公式"""
        import math
        
        # 获取实验参数
        rho1 = float(self.settings_params["油滴密度"])          # 油滴密度 (kg/m³)
        g = float(self.settings_params["重力加速度"])          # 重力加速度 (m/s²)
        l = float(self.settings_params["运动距离"]) / 1000     # 运动距离 (m)
        b = float(self.settings_params["修正常数"])           # 修正常数 (N/m)
        p = float(self.settings_params["大气压强"])           # 大气压强 (Pa)
        d = 5.00e-3                                           # 极板间距 (m)
        
        # 计算空气密度 ρ2（使用理想气体状态方程）
        T = float(self.settings_params["室温"]) + 273.15      # 转换为开尔文
        R = 287.05  # 空气气体常数 J/(kg·K)
        rho2 = p / (R * T)  # 空气密度 (kg/m³)
        
        # ========== 统一使用Sutherland公式计算空气粘滞系数 ==========
        T0 = 273.15
        eta0 = 1.716e-5
        S = 110.4
        eta = eta0 * (T/T0)**1.5 * (T0 + S) / (T + S)
        # ========================================================
        
        # 计算下落速度
        v_g = l / t_g  # 下落速度 (m/s)
        
        # 计算油滴半径 a（考虑空气浮力）
        # 第一次近似，不考虑修正
        a_approx = math.sqrt((9 * eta * v_g) / (2 * (rho1 - rho2) * g))
        
        # 迭代修正（考虑空气阻力修正）
        a = a_approx
        for i in range(5):
            correction_factor = 1 + b / (p * a)
            new_a = math.sqrt((9 * eta * l) / (2 * (rho1 - rho2) * g * t_g * correction_factor))
            # 检查收敛性
            if abs(new_a - a) < 1e-10:
                break
            a = new_a
        
        # 最终修正因子
        correction_factor_final = 1 + b / (p * a)
        
        # 标准平衡法电荷公式（考虑空气浮力）
        part1 = 18 * math.pi / math.sqrt(2 * (rho1 - rho2) * g)
        part2 = math.pow((eta * l) / (t_g * correction_factor_final), 1.5)
        q = part1 * part2 * (d / U)
        
        # 添加详细的调试信息
        print("\n=== 统一公式后的显示屏区域计算 ===")
        print(f"输入: U={U}V, t_g={t_g}s")
        print(f"参数: rho1={rho1}kg/m³, rho2={rho2:.3f}kg/m³")
        print(f"      g={g}m/s², l={l}m, d={d}m")
        print(f"      b={b}N/m, p={p}Pa, T={T}K")
        print(f"      eta={eta:.6e}Pa·s (使用Sutherland公式)")
        print(f"计算: v_g={v_g:.6e}m/s")
        print(f"      a={a:.6e}m, a×10⁷={a*1e7:.3f}")
        print(f"      correction_factor={correction_factor_final:.6f}")
        print(f"      q={q:.6e}C, q×10¹⁹={q/1e-19:.3f}")
        
        return q

    def get_unified_params(self, method="balance"):
        """获取统一的参数，确保两边计算一致"""
        params = {}
        
        # 基本参数（从settings_params获取）
        params['l'] = float(self.settings_params["运动距离"]) / 1000
        params['rho1'] = float(self.settings_params["油滴密度"])
        params['g'] = float(self.settings_params["重力加速度"])
        params['b'] = float(self.settings_params["修正常数"])
        
        # 温度相关参数统一计算
        if method == "balance":
            T_val = self.balance_T_var.get()
        else:
            T_val = self.dynamic_T_var.get()
        
        T = float(T_val) if T_val else 24.0
        # 如果是摄氏度，转换为开尔文
        if T < 100:  # 假设是摄氏度
            T = T + 273.15
        
        params['T'] = T
        
        # 压强
        if method == "balance":
            P = float(self.balance_P_var.get())
        else:
            P = float(self.dynamic_P_var.get())
        params['P'] = P
        
        # 统一计算空气密度
        R = 287.05
        params['rho2'] = P / (R * T)
        
        # 统一计算空气粘滞系数（Sutherland公式）
        T0 = 273.15
        eta0 = 1.716e-5
        S = 110.4
        params['eta'] = eta0 * (T/T0)**1.5 * (T0 + S) / (T + S)
        
        # 固定参数
        params['d'] = 5.00e-3  # 极板间距
        
        return params

    def verify_calculation_consistency(self):
        """验证显示屏区域和数据记录区域计算一致性"""
        print("\n" + "="*60)
        print("计算一致性验证 - 统一η精度后")
        print("="*60)
        
        # 测试用例：使用相同的输入
        test_U = 250  # V
        test_tg = 15.0  # s
        
        print(f"\n【输入参数】")
        print(f"电压 U = {test_U} V")
        print(f"下落时间 t_g = {test_tg} s")
        
        # 1. 获取显示屏区域计算结果和参数
        print(f"\n【1. 显示屏区域计算】")
        print("-"*40)
        
        # 使用与显示屏区域相同的方法获取参数
        rho1_display = float(self.settings_params["油滴密度"])
        g_display = float(self.settings_params["重力加速度"])
        b_display = float(self.settings_params["修正常数"])
        p_display = float(self.settings_params["大气压强"])
        l_display = float(self.settings_params["运动距离"]) / 1000
        
        # 温度
        T_celsius_display = float(self.settings_params["室温"]) if "室温" in self.settings_params else 24.0
        T_display = T_celsius_display + 273.15
        
        # 使用统一的η计算函数
        eta_info_display = self.calculate_unified_eta(T_display)
        eta_display = eta_info_display['eta']
        eta_display_str = eta_info_display['eta_display']
        
        print(f"参数来源: settings_params")
        print(f"油滴密度 ρ1 = {rho1_display} kg/m³")
        print(f"重力加速度 g = {g_display} m/s²")
        print(f"修正常数 b = {b_display} N/m")
        print(f"大气压强 P = {p_display} Pa")
        print(f"运动距离 l = {l_display*1000:.2f} mm = {l_display:.6f} m")
        print(f"室温(℃) = {T_celsius_display}")
        print(f"室温(K) = {T_display:.2f}")
        
        # 空气密度计算
        R = 287.05
        rho2_display = p_display / (R * T_display)
        print(f"空气密度 ρ2 = {rho2_display:.6f} kg/m³ (计算: P/(R*T))")
        print(f"空气粘滞系数 η = {eta_display_str} = {eta_display:.6e} Pa·s (统一Sutherland公式)")
        
        d = 5.00e-3
        print(f"极板间距 d = {d*1000:.2f} mm = {d:.6f} m")
        
        # 显示屏区域计算过程
        v_g_display = l_display / test_tg
        
        # 计算油滴半径（迭代）
        a_approx = math.sqrt((9 * eta_display * v_g_display) / (2 * (rho1_display - rho2_display) * g_display))
        a_display = a_approx
        for i in range(5):
            correction_factor = 1 + b_display / (p_display * a_display)
            new_a = math.sqrt((9 * eta_display * l_display) / (2 * (rho1_display - rho2_display) * g_display * test_tg * correction_factor))
            if abs(new_a - a_display) < 1e-10:
                break
            a_display = new_a
        
        correction_factor_final = 1 + b_display / (p_display * a_display)
        
        # 显示屏区域电荷公式
        part1_display = 18 * math.pi / math.sqrt(2 * (rho1_display - rho2_display) * g_display)
        part2_display = math.pow((eta_display * l_display) / (test_tg * correction_factor_final), 1.5)
        display_charge = part1_display * part2_display * (d / test_U)
        
        print(f"\n显示屏区域电荷计算结果:")
        print(f"  q = {display_charge:.6e} C")
        print(f"  q×10¹⁹ = {display_charge/1e-19:.6f} ×10⁻¹⁹C")
        
        # 2. 获取数据记录区域计算结果和参数
        print(f"\n【2. 数据记录区域计算】")
        print("-"*40)
        
        # 获取数据记录区域的参数（使用统一的方法获取）
        l_table = l_display  # 假设相同
        
        # 从界面获取温度，但使用统一计算
        if hasattr(self, 'balance_T_var'):
            T_str = self.balance_T_var.get()
            T_table_input = float(T_str) if T_str else T_display
        else:
            T_table_input = T_display
        
        # 检查是否是摄氏度
        if T_table_input < 100:  # 假设是摄氏度
            T_table = T_table_input + 273.15
            print(f"数据记录区域输入温度: {T_table_input}°C，转换为: {T_table:.2f}K")
        else:
            T_table = T_table_input
        
        # 使用统一的η计算函数
        eta_info_table = self.calculate_unified_eta(T_table)
        eta_table = eta_info_table['eta']
        eta_table_str = eta_info_table['eta_display']
        
        # 获取其他参数
        rho1_table = rho1_display  # 假设相同
        g_table = g_display  # 假设相同
        P_table = p_display  # 假设相同
        
        # 计算空气密度
        rho2_table = P_table / (R * T_table)
        
        print(f"参数来源: 界面变量 + 统一计算")
        print(f"油滴密度 ρ1 = {rho1_table} kg/m³")
        print(f"重力加速度 g = {g_table} m/s²")
        print(f"修正常数 b = 0.00823 N/m (数据记录区域硬编码)")
        print(f"大气压强 P = {P_table} Pa")
        print(f"运动距离 l = {l_table*1000:.2f} mm = {l_table:.6f} m")
        print(f"温度(K) = {T_table:.2f}")
        print(f"空气密度 ρ2 = {rho2_table:.6f} kg/m³")
        print(f"空气粘滞系数 η = {eta_table_str} = {eta_table:.6e} Pa·s (统一Sutherland公式)")
        
        # 数据记录区域计算过程
        v_g_table = l_table / test_tg
        
        # 数据记录区域的修正常数
        b_table = 0.00823  # 硬编码值
        
        # 计算油滴半径（数据记录区域的方法，迭代3次）
        a_table_approx = math.sqrt((9 * eta_table * v_g_table) / (2 * (rho1_table - rho2_table) * g_table))
        a_table = a_table_approx
        for i in range(3):  # 数据记录区域通常迭代3次
            correction_table = 1 + b_table / (P_table * a_table)
            a_table = math.sqrt((9 * eta_table * l_table) / (2 * (rho1_table - rho2_table) * g_table * test_tg * correction_table))
        
        correction_table_final = 1 + b_table / (P_table * a_table)
        
        # 数据记录区域电荷公式
        part1_table = 18 * math.pi / math.sqrt(2 * (rho1_table - rho2_table) * g_table)
        part2_table = math.pow((eta_table * l_table) / (test_tg * correction_table_final), 1.5)
        table_charge = part1_table * part2_table * (d / test_U)
        
        print(f"\n数据记录区域电荷计算结果:")
        print(f"  q = {table_charge:.6e} C")
        print(f"  q×10¹⁹ = {table_charge/1e-19:.6f} ×10⁻¹⁹C")
        
        # 3. 关键参数对比（重点看η）
        print(f"\n【3. 关键参数对比】")
        print("-"*40)
        
        # η的详细对比
        print(f"空气粘滞系数 η 对比:")
        print(f"  显示屏区域: {eta_display_str}")
        print(f"  数据记录区域: {eta_table_str}")
        print(f"  η实际值:")
        print(f"    显示屏区域: {eta_display:.12e} Pa·s")
        print(f"    数据记录区域: {eta_table:.12e} Pa·s")
        print(f"  η差异: {abs(eta_table - eta_display):.2e} Pa·s")
        print(f"  η相对误差: {abs(eta_table - eta_display)/eta_display*100:.8f}%")
        
        # 温度对比
        print(f"\n温度对比:")
        print(f"  显示屏区域: T = {T_display:.6f} K")
        print(f"  数据记录区域: T = {T_table:.6f} K")
        print(f"  温度差异: {abs(T_table - T_display):.6f} K")
        print(f"  温度相对误差: {abs(T_table - T_display)/T_display*100:.6f}%")
        
        # 4. 参数差异分析
        print(f"\n【4. 参数差异分析】")
        print("-"*40)
        
        print(f"{'参数':<20} {'显示屏区域':<15} {'数据记录区域':<15} {'差异':<15} {'相对误差':<15}")
        print(f"{'-'*20:<20} {'-'*15:<15} {'-'*15:<15} {'-'*15:<15} {'-'*15:<15}")
        
        params_compare = [
            ("T (K)", f"{T_display:.3f}", f"{T_table:.3f}", 
            f"{abs(T_table-T_display):.3f}", f"{abs(T_table-T_display)/T_display*100:.6f}%"),
            ("η×10⁵", f"{eta_display/1e-5:.6f}", f"{eta_table/1e-5:.6f}",
            f"{abs(eta_table-eta_display)/1e-5:.6f}", f"{abs(eta_table-eta_display)/eta_display*100:.6f}%"),
            ("ρ2 (kg/m³)", f"{rho2_display:.6f}", f"{rho2_table:.6f}",
            f"{abs(rho2_table-rho2_display):.6f}", f"{abs(rho2_table-rho2_display)/rho2_display*100:.6f}%"),
            ("a (×10⁻⁷m)", f"{a_display*1e7:.3f}", f"{a_table*1e7:.3f}",
            f"{abs(a_table-a_display)*1e7:.3f}", f"{abs(a_table-a_display)/a_display*100:.6f}%"),
            ("修正因子", f"{correction_factor_final:.6f}", f"{correction_table_final:.6f}",
            f"{abs(correction_table_final-correction_factor_final):.6f}", f"{abs(correction_table_final-correction_factor_final)/correction_factor_final*100:.6f}%"),
        ]
        
        for param_name, display_val, table_val, diff_val, diff_percent in params_compare:
            print(f"{param_name:<20} {display_val:<15} {table_val:<15} {diff_val:<15} {diff_percent:<15}")
        
        # 5. 最终结果对比
        print(f"\n【5. 计算结果对比】")
        print("-"*40)
        
        print(f"显示屏区域电荷: {display_charge:.12e} C")
        print(f"数据记录区域电荷: {table_charge:.12e} C")
        print(f"绝对差异: {abs(display_charge - table_charge):.2e} C")
        print(f"相对误差: {abs(display_charge - table_charge)/display_charge*100:.8f}%")
        
        # 计算电子数差异
        e = 1.602e-19
        n_display = display_charge / e
        n_table = table_charge / e
        print(f"\n电子数对比:")
        print(f"显示屏区域: q/e = {display_charge:.12e}/{e:.6e} = {n_display:.8f}")
        print(f"数据记录区域: q/e = {table_charge:.12e}/{e:.6e} = {n_table:.8f}")
        print(f"电子数差异: {abs(n_display - n_table):.8f}")
        print(f"电子数相对误差: {abs(n_display - n_table)/n_display*100:.8f}%")
        
        # 6. η精度对结果的影响分析
        print(f"\n【6. η精度影响分析】")
        print("-"*40)
        
        # 假设η显示值（保留3位小数）与实际计算值的差异
        eta_display_rounded = round(eta_display / 1e-5, 3) * 1e-5
        eta_table_rounded = round(eta_table / 1e-5, 3) * 1e-5
        
        print(f"η实际计算值:")
        print(f"  显示屏区域: {eta_display:.12e} Pa·s")
        print(f"  数据记录区域: {eta_table:.12e} Pa·s")
        print(f"η保留3位小数后的值:")
        print(f"  显示屏区域: {eta_display_rounded:.12e} Pa·s")
        print(f"  数据记录区域: {eta_table_rounded:.12e} Pa·s")
        print(f"η舍入误差:")
        print(f"  显示屏区域: {abs(eta_display_rounded - eta_display):.2e} Pa·s")
        print(f"  数据记录区域: {abs(eta_table_rounded - eta_table):.2e} Pa·s")
        
        # 计算舍入误差对电荷的影响（q ∝ η^1.5）
        q_display_rounded = display_charge * (eta_display_rounded/eta_display)**1.5
        q_table_rounded = table_charge * (eta_table_rounded/eta_table)**1.5
        
        print(f"\nη舍入对电荷的影响:")
        print(f"  显示屏区域舍入后电荷: {q_display_rounded:.12e} C")
        print(f"  显示屏区域舍入影响: {(q_display_rounded/display_charge-1)*100:.8f}%")
        print(f"  数据记录区域舍入后电荷: {q_table_rounded:.12e} C")
        print(f"  数据记录区域舍入影响: {(q_table_rounded/table_charge-1)*100:.8f}%")
        
        # 7. 结论
        print(f"\n【7. 结论】")
        print("-"*40)
        
        error_percent = abs(display_charge - table_charge)/display_charge*100
        
        if error_percent < 0.001:
            print(f"✅ 计算结果完全一致！相对误差: {error_percent:.8f}%")
        elif error_percent < 0.01:
            print(f"✅ 计算结果基本一致！相对误差: {error_percent:.8f}% (<0.01%)")
        elif error_percent < 0.1:
            print(f"⚠️  计算结果有微小差异！相对误差: {error_percent:.8f}% (<0.1%)")
        else:
            print(f"❌ 计算结果有显著差异！相对误差: {error_percent:.8f}%")
        
        # 主要误差来源分析
        if abs(T_table - T_display) > 0.1:
            print(f"主要误差来源: 温度不一致 (ΔT={abs(T_table-T_display):.3f}K)")
        elif abs(eta_table - eta_display)/eta_display*100 > 0.01:
            print(f"主要误差来源: η计算不一致 (Δη/η={abs(eta_table-eta_display)/eta_display*100:.6f}%)")
        elif abs(b_table - b_display)/b_display*100 > 0.1:
            print(f"主要误差来源: 修正常数不一致 (Δb/b={abs(b_table-b_display)/b_display*100:.2f}%)")
        else:
            print(f"主要误差来源: 综合因素 (温度、η、迭代次数等)")
        
        print("\n" + "="*60)
        
        return {
            'display_charge': display_charge,
            'table_charge': table_charge,
            'error_percent': error_percent,
            'eta_display': eta_display,
            'eta_table': eta_table,
            'T_display': T_display,
            'T_table': T_table
        }

    def calculate_dynamic_method_charge(self, U_balance, t_fall, U_rise, t_rise):
        """计算动态法油滴电荷 - 使用标准密立根公式"""
        import math
        
        # 获取参数
        l = float(self.settings_params["运动距离"]) / 1000  # m
        rho = float(self.settings_params["油滴密度"])        # kg/m³
        g = float(self.settings_params["重力加速度"])        # m/s²
        d = 5.00e-3                                         # 极板间距 m
        b = float(self.settings_params["修正常数"])          # 修正常数 N/m
        p = float(self.settings_params["大气压强"])          # 大气压强 Pa
        
        # ========== 统一使用Sutherland公式计算空气粘滞系数 ==========
        T = float(self.settings_params["室温"]) + 273.15      # 转换为开尔文
        T0 = 273.15
        eta0 = 1.716e-5
        S = 110.4
        eta = eta0 * (T/T0)**1.5 * (T0 + S) / (T + S)  # Pa·s
        # ========================================================
        
        # 计算速度
        v_g = l / t_fall  # 下落速度 m/s
        v_e = l / t_rise  # 上升速度 m/s
        
        # 计算油滴半径 a（迭代计算，考虑修正因子）
        # 第一次近似（不考虑修正）
        a_approx = math.sqrt((9 * eta * v_g) / (2 * rho * g))
        
        # 迭代修正（通常2-3次迭代即可）
        for _ in range(3):
            correction = 1 + b / (p * a_approx)
            a_approx = math.sqrt((9 * eta * l) / (2 * rho * g * t_fall * correction))
        
        a = a_approx
        correction_final = 1 + b / (p * a)
        
        # 标准动态法电荷公式
        # q = (18π/√(2ρg)) * d * (ηl/(1+b/pa))^(3/2) * (1/t_e + 1/t_g) * √(1/t_g) / U
        # 其中 U 应该是上升电压 U_rise
        
        # 分步计算
        part1 = 18 * math.pi / math.sqrt(2 * rho * g)
        part2 = math.pow((eta * l) / (correction_final), 1.5)
        part3 = (1/t_rise + 1/t_fall) * math.sqrt(1/t_fall)
        
        q = part1 * d * part2 * part3 / U_rise
        
        return q

    def reset_to_default(self):
        """重置为默认参数"""
        default_settings = {
            "运动距离": "2.0",
            "油滴密度": "981",
            "室温": "24",
            "大气压强": "101325",
            "重力加速度": "9.7940",
            "修正常数": "0.00823",
        }
        
        self.settings_params.update(default_settings)
        
        # 重置时间数据（但不重置电荷显示框）
        self.balance_params["下落时间"] = "0.0"
        self.dynamic_params["下落时间"] = "0.0"
        self.dynamic_params["上升时间"] = "0.0"
        
        # 更新界面显示
        self.distance_var.set(default_settings["运动距离"])
        self.density_var.set(default_settings["油滴密度"])
        self.temp_var.set(default_settings["室温"])
        self.pressure_var.set(default_settings["大气压强"])
        self.gravity_var.set(default_settings["重力加速度"])
        self.correction_var.set(default_settings["修正常数"])
        
        print("已重置为默认参数")
    
    def go_back(self):
        """返回按钮的功能"""
        if self.current_interface == "参数设置":
            # 从参数设置返回时，总是回到平衡法界面
            self.switch_display_interface("平衡法")
        else:
            print("返回功能")
            # 可以添加其他返回逻辑
    
    def create_data_record_area(self):
        """创建数据记录区域"""
        # 创建选项卡框架
        tab_frame = tk.Frame(self.frame_bottom_right)
        tab_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 创建选项卡
        tk.Label(tab_frame, text="选择实验方法:", font=('Arial', 10)).pack(side=tk.LEFT, padx=5)
        
        # 创建选项变量
        self.method_var = tk.StringVar(value="平衡法")
        
        # 创建单选按钮
        balance_radio = tk.Radiobutton(tab_frame, text="平衡法", variable=self.method_var, 
                                    value="平衡法", command=self.update_content)
        balance_radio.pack(side=tk.LEFT, padx=10)
        
        dynamic_radio = tk.Radiobutton(tab_frame, text="动态法", variable=self.method_var, 
                                    value="动态法", command=self.update_content)
        dynamic_radio.pack(side=tk.LEFT, padx=10)
        
        # 创建内容显示区域
        self.content_frame = tk.Frame(self.frame_bottom_right, bd=1, relief=tk.SUNKEN)
        self.content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 初始显示平衡法内容
        self.update_content()
    
    def update_content(self):
        """根据选择更新内容"""
        # 清除当前内容
        # self.save_dynamic_table_data()
        # self.save_balance_table_data()
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        method = self.method_var.get()
        
        if method == "平衡法":
            self.show_balance_method()
        else:
            self.show_dynamic_method()
    
    def show_balance_method(self):
        """显示平衡法内容 - 从存储中恢复数据"""
        # 创建带滚动条的主容器
        main_container = tk.Frame(self.content_frame)
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # 创建垂直滚动条
        v_scrollbar_main = tk.Scrollbar(main_container)
        v_scrollbar_main.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建水平滚动条
        h_scrollbar_main = tk.Scrollbar(main_container, orient=tk.HORIZONTAL)
        h_scrollbar_main.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 创建Canvas作为滚动区域
        canvas = tk.Canvas(main_container, 
                        yscrollcommand=v_scrollbar_main.set,
                        xscrollcommand=h_scrollbar_main.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        v_scrollbar_main.config(command=canvas.yview)
        h_scrollbar_main.config(command=canvas.xview)
        
        # 在Canvas中创建Frame作为内容容器
        content_frame = tk.Frame(canvas)
        
        # 将Frame添加到Canvas
        canvas_window = canvas.create_window((0, 0), window=content_frame, anchor='nw')
        
        # 配置Canvas滚动
        def configure_canvas(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(canvas_window, width=canvas.winfo_width())
        
        canvas.bind('<Configure>', configure_canvas)
        content_frame.bind('<Configure>', configure_canvas)
        
        # 1. 参数显示区域 - 从存储中恢复
        param_frame = tk.LabelFrame(content_frame, text="实验参数", font=('Arial', 10, 'bold'))
        param_frame.pack(fill=tk.X, padx=5, pady=5)
        
        param_grid = tk.Frame(param_frame)
        param_grid.pack(padx=10, pady=10)
        
        # 第一行
        row = 0
        tk.Label(param_grid, text="极板间距 d =", font=('Arial', 9)).grid(row=row, column=0, sticky='w', padx=(0, 5))
        tk.Label(param_grid, text="5.00 mm", font=('Arial', 9, 'bold'), fg='blue').grid(row=row, column=1, sticky='w')
        tk.Label(param_grid, text="(不可修改)", font=('Arial', 8), fg='gray').grid(row=row, column=2, sticky='w', padx=(5, 20))
        
        tk.Label(param_grid, text="油滴下落距离 l =", font=('Arial', 9)).grid(row=row, column=3, sticky='w', padx=(20, 5))
        self.balance_l_var = tk.StringVar(value=self.balance_params_storage["l"])
        l_entry = tk.Entry(param_grid, textvariable=self.balance_l_var, width=8, font=('Arial', 9))
        l_entry.grid(row=row, column=4, sticky='w', padx=5)
        tk.Label(param_grid, text="mm", font=('Arial', 9)).grid(row=row, column=5, sticky='w')
        
        # 第二行
        row += 1
        tk.Label(param_grid, text="室温 T =", font=('Arial', 9)).grid(row=row, column=0, sticky='w', padx=(0, 5))
        self.balance_T_var = tk.StringVar(value=self.balance_params_storage["T"])
        T_entry = tk.Entry(param_grid, textvariable=self.balance_T_var, width=8, font=('Arial', 9))
        T_entry.grid(row=row, column=1, sticky='w', padx=5)
        tk.Label(param_grid, text="K", font=('Arial', 9)).grid(row=row, column=2, sticky='w')
        
        tk.Label(param_grid, text="大气压强 P =", font=('Arial', 9)).grid(row=row, column=3, sticky='w', padx=(20, 5))
        self.balance_P_var = tk.StringVar(value=self.balance_params_storage["P"])
        P_entry = tk.Entry(param_grid, textvariable=self.balance_P_var, width=8, font=('Arial', 9))
        P_entry.grid(row=row, column=4, sticky='w', padx=5)
        tk.Label(param_grid, text="Pa", font=('Arial', 9)).grid(row=row, column=5, sticky='w')
        
        # 第三行
        row += 1
        tk.Label(param_grid, text="重力加速度 g =", font=('Arial', 9)).grid(row=row, column=0, sticky='w', padx=(0, 5))
        self.balance_g_var = tk.StringVar(value=self.balance_params_storage["g"])
        g_entry = tk.Entry(param_grid, textvariable=self.balance_g_var, width=8, font=('Arial', 9))
        g_entry.grid(row=row, column=1, sticky='w', padx=5)
        tk.Label(param_grid, text="m/s²", font=('Arial', 9)).grid(row=row, column=2, sticky='w')
        
        tk.Label(param_grid, text="油滴密度 ρ1 =", font=('Arial', 9)).grid(row=row, column=3, sticky='w', padx=(20, 5))
        self.balance_rho1_var = tk.StringVar(value=self.balance_params_storage["rho1"])
        # 修改为输入框而不是标签
        rho1_entry = tk.Entry(param_grid, textvariable=self.balance_rho1_var, width=8, font=('Arial', 9))
        rho1_entry.grid(row=row, column=4, sticky='w', padx=5)
        tk.Label(param_grid, text="kg/m³", font=('Arial', 9)).grid(row=row, column=5, sticky='w')
        
        # 第四行
        row += 1
        tk.Label(param_grid, text="空气密度 ρ2 =", font=('Arial', 9)).grid(row=row, column=0, sticky='w', padx=(0, 5))
        self.balance_rho2_var = tk.StringVar(value=self.balance_params_storage["rho2"])
        tk.Label(param_grid, textvariable=self.balance_rho2_var, font=('Arial', 9, 'bold'), 
                fg='green', width=10).grid(row=row, column=1, sticky='w', padx=5)
        tk.Label(param_grid, text="kg/m³", font=('Arial', 9)).grid(row=row, column=2, sticky='w')
        
        tk.Label(param_grid, text="空气粘滞系数 η =", font=('Arial', 9)).grid(row=row, column=3, sticky='w', padx=(20, 5))
        self.balance_eta_var = tk.StringVar(value=self.balance_params_storage["eta"])
        tk.Label(param_grid, textvariable=self.balance_eta_var, font=('Arial', 9, 'bold'), 
                fg='green', width=12).grid(row=row, column=4, sticky='w', padx=5)
        tk.Label(param_grid, text="kg/(m·s)", font=('Arial', 9)).grid(row=row, column=5, sticky='w')
        
        # 让参数网格自适应
        for i in range(6):
            param_grid.columnconfigure(i, weight=1)
        
        # 2. 创建表格框架（固定高度，使用滚动条）
        table_frame = tk.LabelFrame(content_frame, text="数据记录表", font=('Arial', 10, 'bold'))
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 设置表格区域固定高度
        self.table_container = tk.Frame(table_frame, height=200)
        self.table_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.table_container.pack_propagate(False)
        
        # 创建表格
        columns = ("电压U/V", "下落时间第一次tg1/s", "下落时间第二次tg2/s", "下落时间第三次tg3/s",
                "平均下落时间tg/s", "下落速度vg/×10⁻⁵m/s", "油滴半径a/×10⁻⁷m", 
                "油滴带电量q/×10⁻¹⁹C", "带电子数n")
        
        # 创建Treeview
        self.balance_tree = ttk.Treeview(self.table_container, columns=columns, show="headings", height=6)
        
        # 设置列宽
        col_widths = [80+20, 100+20+10, 100+20+10, 100+20+10, 100+20+10, 100+20+10, 100+20, 100+20, 80+20]
        
        for i, col in enumerate(columns):
            self.balance_tree.heading(col, text=col)
            self.balance_tree.column(col, width=col_widths[i], minwidth=50,stretch=False)
    
        # 创建垂直滚动条
        tree_v_scrollbar = tk.Scrollbar(self.table_container, orient=tk.VERTICAL)
        tree_v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建水平滚动条
        tree_h_scrollbar = tk.Scrollbar(self.table_container, orient=tk.HORIZONTAL)
        tree_h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 配置Treeview
        self.balance_tree.configure(yscrollcommand=tree_v_scrollbar.set, 
                                xscrollcommand=tree_h_scrollbar.set)
        self.balance_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        tree_v_scrollbar.config(command=self.balance_tree.yview)
        tree_h_scrollbar.config(command=self.balance_tree.xview)
        
        # 从存储中恢复数据，如果没有数据则添加初始3行空行
        if self.balance_table_data:
            for row_data in self.balance_table_data:
                self.balance_tree.insert("", "end", values=row_data)
        else:
            for _ in range(3):
                self.balance_tree.insert("", "end", values=("", "", "", "", "", "", "", "", ""))
        
        # 绑定事件
        self.balance_tree.bind("<KeyRelease>", self.on_balance_table_keyrelease)
        self.balance_tree.bind("<Double-1>", self.on_tree_double_click_selective)
        
        # 3. 创建图表区域
        chart_frame = tk.LabelFrame(content_frame, text="电荷-电子数关系图", font=('Arial', 10, 'bold'))
        chart_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        chart_container = tk.Frame(chart_frame, height=300)
        chart_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        chart_container.pack_propagate(False)
        
        self.balance_chart_canvas = tk.Canvas(chart_container, bg='white', width=600, height=300)
        self.balance_chart_canvas.pack(fill=tk.BOTH, expand=True)
        
        # 初始绘制坐标轴
        self.draw_balance_chart_axes()
        
        # 4. 创建按钮区域
        button_frame = tk.Frame(content_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 按钮
        buttons = [
            ("计算", self.calculate_balance_data),
            ("删除选中行", self.delete_balance_selected_row),
            ("清空数据", self.clear_balance_data),
            ("导出数据", self.export_balance_data),
            ("导入数据", self.import_balance_data)
        ]
        
        for text, command in buttons:
            tk.Button(button_frame, text=text, command=command,
                    font=('Arial', 9), width=12).pack(side=tk.LEFT, padx=5)
        
        # 确保内容框架正确配置
        content_frame.update_idletasks()
        
        # 如果有数据，绘制图表
        if self.balance_table_data and any(any(cell for cell in row[7:9] if cell) for row in self.balance_table_data):
            self.root.after(100, self.plot_balance_chart)

        # self.calculate_balance_data()
        # # 计算完成后，确保图表被绘制
        # self.plot_balance_chart()
    
    def on_tree_double_click_selective(self, event):
        """选择性双击编辑 - 只允许编辑前4列"""
        region = self.balance_tree.identify("region", event.x, event.y)
        
        if region == "cell":
            # 获取单元格信息
            row = self.balance_tree.identify_row(event.y)
            column = self.balance_tree.identify_column(event.x)
            
            # 获取列索引（从1开始）
            col_index = int(column.replace('#', '')) - 1
            
            # 只允许编辑前4列（电压和三次下落时间）
            if col_index > 3:  # 第5列及以后不可编辑
                # 显示提示信息
                self.show_tooltip("此列数据为自动计算，不可手动修改", event.x_root, event.y_root)
                return
            
            # 获取单元格位置
            x, y, width, height = self.balance_tree.bbox(row, column)
            
            # 获取当前值
            item = self.balance_tree.item(row)
            values = item['values']
            
            if col_index < len(values):
                current = values[col_index] if values[col_index] is not None else ""
            else:
                current = ""
            
            # 创建输入框
            entry = tk.Entry(self.balance_tree, 
                            font=('Arial', 9),
                            borderwidth=0)
            
            # 放置在单元格上
            entry.place(x=x, y=y, width=width, height=height)
            entry.insert(0, current)
            entry.select_range(0, tk.END)
            entry.focus()
            
            def save_and_quit(event=None):
                new_value = entry.get()
                
                # 验证输入是否为数字
                if col_index in [0, 1, 2, 3]:  # 电压和时间需要是数字
                    try:
                        if new_value.strip():  # 如果不是空值
                            float(new_value)
                    except ValueError:
                        messagebox.showerror("错误", "请输入有效的数字")
                        entry.focus()
                        return
                
                # 更新数据
                new_values = list(values)
                if col_index < len(new_values):
                    new_values[col_index] = new_value
                else:
                    new_values.append(new_value)
                
                # 更新行数据
                self.balance_tree.item(row, values=new_values)
                entry.destroy()
                
                # 如果前4列都有数据，自动计算该行
                if all(new_values[:4]) and all(v.strip() for v in new_values[:4]):
                    self.calculate_single_row(row, new_values)
            
            def cancel_edit(event=None):
                entry.destroy()
            
            # 绑定事件
            entry.bind('<Return>', save_and_quit)
            entry.bind('<FocusOut>', save_and_quit)
            entry.bind('<Escape>', cancel_edit)

    def show_tooltip(self, message, x, y):
        """显示提示信息"""
        # 创建一个临时窗口显示提示
        tooltip = tk.Toplevel()
        tooltip.wm_overrideredirect(True)
        tooltip.wm_geometry(f"+{x}+{y}")
        
        label = tk.Label(tooltip, text=message, 
                        font=('Arial', 8), 
                        bg='yellow', 
                        relief='solid', 
                        borderwidth=1)
        label.pack()
        
        # 1秒后自动关闭
        tooltip.after(1000, tooltip.destroy)

    def calculate_single_row(self, item_id, values):
        """计算单行数据"""
        try:
            # 获取参数
            U = float(values[0])  # 电压
            tg1 = float(values[1])  # 下落时间1
            tg2 = float(values[2])  # 下落时间2
            tg3 = float(values[3])  # 下落时间3
            
            # 计算平均下落时间
            tg_avg = (tg1 + tg2 + tg3) / 3
            values[4] = f"{tg_avg:.2f}"
            
            # 计算下落速度 vg (l=2.00mm=0.002m)
            l = 0.002  # 2.00mm = 0.002m
            vg = l / tg_avg  # m/s
            vg_show = vg * 1e5  # 转换为×10⁻⁵m/s
            values[5] = f"{vg_show:.3f}"
            
            # 计算油滴半径 a (平衡法公式)
            # 从参数显示区域获取参数值
            try:
                rho1 = float(self.balance_rho1_var.get())
                rho2 = float(self.balance_rho2_var.get())
                eta = float(self.balance_eta_var.get().replace('×10⁻⁵', '')) * 1e-5
                g = float(self.balance_g_var.get())
            except:
                # 如果参数未设置，使用默认值
                rho1 = 977.0
                rho2 = 1.158
                eta = 1.855e-5
                g = 9.794
            
            # 使用修正的斯托克斯公式
            a = math.sqrt((9 * eta * vg) / (2 * (rho1 - rho2) * g))
            a_show = a * 1e7  # 转换为×10⁻⁷m
            values[6] = f"{a_show:.3f}"
            
            # 计算油滴带电量 q (平衡法公式)
            d = 5.00e-3  # 极板间距 5mm=0.005m
            q = (6 * math.pi * eta * a * d * vg) / U
            q_show = q / 1e-19  # 转换为×10⁻¹⁹
            values[7] = f"{q_show:.3f}"
            
            # 计算带电子数 n
            e = 1.602e-19  # 元电荷
            n = round(q / e)
            values[8] = f"{n}"
            
            # 更新行数据
            self.balance_tree.item(item_id, values=values)
            
        except (ValueError, ZeroDivisionError) as e:
            print(f"计算错误: {e}")
            # 如果计算错误，清空计算结果
            values[4:] = ["", "", "", ""]
            self.balance_tree.item(item_id, values=values)


    def on_balance_table_keyrelease(self, event):
        """平衡法表格键盘释放事件 - 检测最后一行是否有数据，并保存当前数据"""
        # 获取所有行
        children = self.balance_tree.get_children()
        if not children:
            return
        
        # 检查最后一行
        last_item = children[-1]
        values = self.balance_tree.item(last_item, 'values')
        
        # 如果最后一行有任何数据，添加新行
        if values and any(v.strip() for v in values[:4]):  # 检查前4列（手动输入列）
            self.balance_tree.insert("", "end", values=("", "", "", "", "", "", "", "", ""))
        
        # 保存当前所有数据到存储
        self.save_balance_table_data()

    def calculate_balance_data(self):
        """计算平衡法数据"""
        # 1. 更新自动计算的参数
        # 清空图表
        self.draw_balance_chart_axes()
        try:
            # 获取温度值
            T_str = self.balance_T_var.get()
            T = float(T_str)
            
            # 检查温度是否合理
            if T < 200 or T > 400:
                if 0 <= T <= 50:
                    T_celsius = T
                    T = T_celsius + 273.15
                    print(f"INFO: 检测到摄氏度输入 {T_celsius}°C，已转换为 {T}K")
                    self.balance_T_var.set(f"{T:.2f}")
            
            P = float(self.balance_P_var.get())
            g = float(self.balance_g_var.get())
            
            # 计算空气密度 ρ2
            R = 287.05
            rho2 = P / (R * T)
            self.balance_rho2_var.set(f"{rho2:.3f}")  # 保持三位小数
            
            # ========== 统一计算空气粘滞系数 η ==========
            T0 = 273.15
            eta0 = 1.716e-5
            S = 110.4
            
            # 使用Sutherland公式计算（与显示屏区域统一）
            ratio = T / T0
            power_term = ratio ** 1.5
            sutherland_term = (T0 + S) / (T + S)
            eta = eta0 * power_term * sutherland_term
            
            # 统一显示为 ×10⁻⁵ 单位，保留三位小数
            eta_display = eta / 1e-5
            self.balance_eta_var.set(f"{eta_display:.3f}×10⁻⁵")  # 统一三位小数
            
            print(f"DEBUG 数据记录区域: eta={eta:.6e} Pa·s (Sutherland公式)")
            print(f"DEBUG 数据记录区域: eta×10⁵={eta_display:.3f}×10⁻⁵ Pa·s (统一三位小数)")
            
            # 油滴密度 ρ1 现在从输入框获取，不再自动计算
            rho1_str = self.balance_rho1_var.get()
            if not rho1_str.strip():
                rho1 = 981.0  # 默认值
                self.balance_rho1_var.set(f"{rho1}")
            else:
                rho1 = float(rho1_str)
            
            print(f"DEBUG: 油滴密度 rho1 = {rho1} (手动输入)")
            
        except ValueError as e:
            print(f"ERROR: 数值转换错误: {e}")
            messagebox.showerror("错误", f"参数输入有误，请检查数字格式\n错误: {e}")
            return
        except Exception as e:
            print(f"ERROR: 计算过程中出错: {e}")
            messagebox.showerror("错误", f"计算过程中出错: {e}")
            return
        
        # 2. 计算表格中每一行的数据
        for item in self.balance_tree.get_children():
            values = list(self.balance_tree.item(item, 'values'))
            
            # 只计算有基本数据的行
            if len(values) >= 4 and all(values[:4]):
                try:
                    # 提取输入数据 - 电压取整数，时间保留2位小数
                    U = int(float(values[0]))  # 电压取整
                    tg1 = round(float(values[1]), 2)  # 时间保留2位小数
                    tg2 = round(float(values[2]), 2)
                    tg3 = round(float(values[3]), 2)
                    
                    print(f"\n=== 数据记录区域计算 ===")
                    print(f"输入: U={U}V (整数), tg1={tg1}s, tg2={tg2}s, tg3={tg3}s (2位小数)")
                    
                    # 计算平均下落时间（保留2位小数）
                    tg_avg = round((tg1 + tg2 + tg3) / 3, 2)  # 平均时间也保留2位小数
                    values[4] = f"{tg_avg:.2f}"
                    
                    # 计算下落速度 vg (l=2.00mm=0.002m)
                    l = float(self.balance_l_var.get()) / 1000  # mm转m
                    vg = l / tg_avg  # m/s
                    vg_show = vg * 1e5  # 转换为×10⁻⁵m/s
                    values[5] = f"{vg_show:.3f}"  # 保持3位小数
                    
                    # 计算油滴半径 a (平衡法公式)
                    rho1_val = float(self.balance_rho1_var.get())
                    
                    # ===== 关键修改：统一使用计算出的eta值 =====
                    # 从统一计算的eta变量获取，而不是从界面显示值转换
                    # eta_val = float(self.balance_eta_var.get().replace('×10⁻⁵', '')) * 1e-5
                    eta_val = eta  # 直接使用上面计算出的eta值
                    # ==========================================
                    
                    # 使用修正的斯托克斯公式
                    a = math.sqrt((9 * eta_val * vg) / (2 * (rho1_val - rho2) * g))
                    a_show = a * 1e7  # 转换为×10⁻⁷m
                    values[6] = f"{a_show:.3f}"  # 保持3位小数
                    
                    b = 0.00823  # 默认修正常数 N/m
                    p = float(self.balance_P_var.get())  # 大气压强 Pa
                    
                    # 计算修正因子
                    correction = 1 + b / (p * a)
                    
                    # 标准平衡法公式
                    part1 = 18 * math.pi / math.sqrt(2 * (rho1_val - rho2) * g)
                    part2 = math.pow((eta_val * l) / (tg_avg * correction), 1.5)
                    d = 5.00e-3
                    q = part1 * part2 * (d / U)
                    q_show = q / 1e-19  # 转换为×10⁻¹⁹
                    
                    # 电荷显示使用4位小数
                    values[7] = f"{q_show:.3f}"  # 4位小数
                    
                    # 计算带电子数 n
                    e = 1.602e-19  # 元电荷
                    n = round(q / e)
                    values[8] = f"{n}"
                    
                    print(f"计算过程:")
                    print(f"  平均时间: tg_avg={tg_avg:.2f}s")
                    print(f"  下落速度: vg={vg:.6e} m/s")
                    print(f"  油滴半径: a={a:.6e} m")
                    print(f"  修正因子: correction={correction:.6f}")
                    print(f"  计算结果: q={q:.6e} C, q_show={q_show:.3f} ×10⁻¹⁹")
                    
                    # 更新行数据
                    self.balance_tree.item(item, values=values)
                    
                except (ValueError, ZeroDivisionError) as e:
                    print(f"计算错误: {e}")
        
        # 在计算完成后保存数据
        self.save_balance_table_data()
        # 3. 绘制图表
        self.plot_balance_chart()
    
    def calculate_unified_eta(self, T, method="balance"):
        """统一计算空气粘滞系数"""
        T0 = 273.15
        eta0 = 1.716e-5
        S = 110.4
        
        # Sutherland公式
        eta = eta0 * (T/T0)**1.5 * (T0 + S) / (T + S)
        
        # 统一显示格式：×10⁻⁵ 单位，保留三位小数
        eta_display = eta / 1e-5
        
        return {
            'eta': eta,  # 实际值，用于计算
            'eta_display': f"{eta_display:.3f}×10⁻⁵",  # 显示值，统一三位小数
            'eta_value': eta_display  # 数值，用于比较
        }

    def plot_balance_chart(self):
        """绘制平衡法图表数据点"""
        canvas = self.balance_chart_canvas
        
        # 收集数据
        n_values = []
        q_values = []
        
        for item in self.balance_tree.get_children():
            values = self.balance_tree.item(item, 'values')
            if len(values) >= 9 and values[7] and values[8]:  # q列和n列
                try:
                    q = float(values[7])  # 已经是×10⁻¹⁹单位
                    n = int(values[8])
                    n_values.append(n)
                    q_values.append(q)
                except (ValueError, TypeError):
                    continue
        
        if not n_values:
            # 如果没有数据，显示提示
            self.draw_balance_chart_axes()
            canvas.create_text(canvas.winfo_width()//2, canvas.winfo_height()//2, 
                            text="暂无数据，请先计算数据", 
                            font=('Arial', 10), fill='gray')
            return
        
        # 计算数据范围
        n_min = min(n_values) if n_values else 0
        n_max = max(n_values) if n_values else 10
        q_min = min(q_values) if q_values else 0
        q_max = max(q_values) if q_values else 20
        
        # 添加边距
        n_range = max(1, n_max - n_min)
        q_range = max(1, q_max - q_min)
        
        n_min_display = max(0, n_min - n_range * 0.1)
        n_max_display = n_max + n_range * 0.1
        q_min_display = max(0, q_min - q_range * 0.1)
        q_max_display = q_max + q_range * 0.1
        
        # 确保最小值不为负数
        n_min_display = max(0, n_min_display)
        q_min_display = max(0, q_min_display)
        
        # 确保范围合适
        if n_max_display - n_min_display < 2:
            n_max_display = n_min_display + 2
        
        if q_max_display - q_min_display < 2:
            q_max_display = q_min_display + 2
        
        # 绘制坐标轴
        self.draw_balance_chart_axes(n_min_display, n_max_display, q_min_display, q_max_display)
        
        # 获取坐标信息
        if not hasattr(self, 'balance_chart_info'):
            return
        
        info = self.balance_chart_info
        margin_top = info['margin_top']
        margin_left = info['margin_left']
        graph_width = info['graph_width']
        graph_height = info['graph_height']
        
        # 绘制数据点
        for n, q in zip(n_values, q_values):
            # 计算坐标位置
            x = margin_left + ((n - n_min_display) / (n_max_display - n_min_display)) * graph_width
            y = margin_top + ((q_max_display - q) / (q_max_display - q_min_display)) * graph_height
            
            # 绘制数据点
            canvas.create_oval(x-4, y-4, x+4, y+4, 
                            fill='red', outline='darkred', width=2)
            
            # 添加标签显示具体值
            label = f"({n}, {q:.3f})"
            canvas.create_text(x, y-15, text=label, 
                            font=('Arial', 8, 'bold'), fill='blue')
        
        # 绘制拟合直线
        if len(n_values) >= 2:
            try:
                n_array = np.array(n_values)
                q_array = np.array(q_values)
                
                # 线性拟合
                slope, intercept = np.polyfit(n_array, q_array, 1)
                
                # 计算直线端点
                x1 = margin_left
                x2 = margin_left + graph_width
                
                y1 = margin_top + ((q_max_display - (slope * n_min_display + intercept)) / 
                                (q_max_display - q_min_display)) * graph_height
                y2 = margin_top + ((q_max_display - (slope * n_max_display + intercept)) / 
                                (q_max_display - q_min_display)) * graph_height
                
                # 绘制拟合直线
                canvas.create_line(x1, y1, x2, y2, 
                                fill='blue', width=2, dash=(4, 2))
                
                # 在图表左上角显示拟合公式和测量结果
                # 计算左上角起始位置
                text_start_x = margin_left + 10  # 左边距+10像素
                text_start_y = margin_top + 10   # 上边距+10像素
                
                # 显示拟合公式
                formula_text = f"q = {slope:.3f}n + {intercept:.3f}"
                canvas.create_text(text_start_x, text_start_y, 
                                text=formula_text, font=('Arial', 9, 'bold'), 
                                fill='darkblue', anchor='nw')
                
                # 计算元电荷测量值和误差
                e_standard = 1.602
                e_measured = abs(slope)
                
                if e_measured > 0:
                    error_percent = abs((e_measured - e_standard) / e_standard) * 100
                else:
                    error_percent = float('inf')
                
                # 显示标准值和测量值
                standard_text = f"e₀: {e_standard}"
                measure_text = f"e: {e_measured:.3f}"
                
                canvas.create_text(text_start_x, text_start_y + 15, 
                                text=standard_text, font=('Arial', 9), 
                                fill='darkgreen', anchor='nw')
                
                canvas.create_text(text_start_x, text_start_y + 30, 
                                text=measure_text, font=('Arial', 9, 'bold'), 
                                fill='blue', anchor='nw')
                
                # 显示误差（带颜色）
                error_color = 'red' if error_percent > 5 else 'orange' if error_percent > 2 else 'green'
                error_text = f"δ: {error_percent:.2f}%"
                canvas.create_text(text_start_x, text_start_y + 45, 
                                text=error_text, font=('Arial', 9, 'bold'), 
                                fill=error_color, anchor='nw')
                
                # 添加图例说明
                legend_x = text_start_x
                legend_y = text_start_y + 65
                canvas.create_text(legend_x, legend_y, 
                                text="(e₀:标准值, e:测量值, δ:误差)", 
                                font=('Arial', 8), 
                                fill='gray', anchor='nw')
                
                self.balance_e_experimental = e_measured
                self.balance_e_error = error_percent
                
            except Exception as e:
                print(f"绘制拟合直线时出错: {e}")
        
        

    def delete_balance_selected_row(self):
        """删除平衡法选中行"""
        selection = self.balance_tree.selection()
        if not selection:
            messagebox.showinfo("提示", "请先选择要删除的行")
            return
        
        if messagebox.askyesno("确认", "确定要删除选中的行吗？"):
            for item in selection:
                self.balance_tree.delete(item)
        
        # 保存数据
        self.save_balance_table_data()

    def clear_balance_data(self):
        """清空平衡法数据"""
        if messagebox.askyesno("确认", "确定要清空所有数据吗？"):
            # 清空表格数据
            for item in self.balance_tree.get_children():
                self.balance_tree.delete(item)
            
            # 添加初始的3行空行
            for _ in range(3):
                self.balance_tree.insert("", "end", values=("", "", "", "", "", "", "", "", ""))
            
            # 清空图表
            self.draw_balance_chart_axes()
            
            # 保存数据
            self.save_balance_table_data()

    def draw_balance_chart_axes(self, n_min=None, n_max=None, q_min=None, q_max=None):
        """绘制平衡法图表坐标轴 - 支持动态范围"""
        canvas = self.balance_chart_canvas
        canvas.delete("all")
        
        # 获取画布实际尺寸
        width = canvas.winfo_width()
        height = canvas.winfo_height()
        
        # 如果画布尺寸太小，使用更大的默认尺寸
        if width < 300 or height < 200:
            # 获取父容器的尺寸作为参考
            parent = canvas.master
            if parent:
                parent_width = parent.winfo_width()
                parent_height = parent.winfo_height()
                if parent_width > 100 and parent_height > 100:
                    width = parent_width - 40
                    height = parent_height - 40
                else:
                    width = 500
                    height = 300
            else:
                width = 500
                height = 300
            
            canvas.config(width=width, height=height)
        
        # 边距设置
        margin_top = 40
        margin_bottom = 70
        margin_left = 60
        margin_right = 20
        
        graph_width = width - margin_left - margin_right
        graph_height = height - margin_top - margin_bottom
        
        # 绘制坐标轴
        canvas.create_line(margin_left, margin_top, 
                        margin_left, margin_top + graph_height, 
                        width=3, fill='black')
        canvas.create_line(margin_left, margin_top + graph_height, 
                        margin_left + graph_width, margin_top + graph_height, 
                        width=3, fill='black')
        
        # 坐标轴标签
        font_size = 10
        
        # Y轴标签
        canvas.create_text(margin_left - 40, margin_top + graph_height/2, 
                        text="q/×10⁻¹⁹C", 
                        angle=90, 
                        font=('Arial', font_size, 'bold'),
                        fill='darkblue')
        
        # X轴标签
        canvas.create_text(margin_left + graph_width/2, margin_top + graph_height + 40, 
                        text="电子数 n", 
                        font=('Arial', font_size, 'bold'),
                        fill='darkblue')
        
        # 图表标题
        canvas.create_text(width//2, 20, 
                        text="电荷-电子数关系图", 
                        font=('Arial', 12, 'bold'),
                        fill='darkred')
        
        # 设置坐标轴范围
        n_min_display = n_min if n_min is not None else 0
        n_max_display = n_max if n_max is not None else 10
        q_min_display = q_min if q_min is not None else 0
        q_max_display = q_max if q_max is not None else 20
        
        if n_max_display <= n_min_display:
            n_max_display = n_min_display + 1
        if q_max_display <= q_min_display:
            q_max_display = q_min_display + 1
        
        # 存储坐标信息
        self.balance_chart_info = {
            'margin_top': margin_top,
            'margin_left': margin_left,
            'graph_width': graph_width,
            'graph_height': graph_height,
            'n_min': n_min_display,
            'n_max': n_max_display,
            'q_min': q_min_display,
            'q_max': q_max_display
        }
        
        # 绘制Y轴刻度 - 固定5个刻度，保留小数点后一位
        y_ticks = 5
        for i in range(y_ticks + 1):
            y = margin_top + (graph_height / y_ticks) * (y_ticks - i)
            # 刻度线
            canvas.create_line(margin_left - 8, y, margin_left, y, width=2, fill='black')
            
            # 刻度标签 - 保留小数点后一位
            value = q_min_display + (q_max_display - q_min_display) * (i / y_ticks)
            label = f"{value:.1f}"  # 保留一位小数
                
            canvas.create_text(margin_left - 12, y, 
                            text=label, 
                            anchor='e', 
                            font=('Arial', font_size-1),
                            fill='darkgreen')
        
        # 绘制X轴刻度 - 显示合适的整数刻度
        tick_range = n_max_display - n_min_display
        
        # 确定刻度间隔
        if tick_range <= 2:
            tick_interval = 1
        elif tick_range <= 5:
            tick_interval = 1
        elif tick_range <= 10:
            tick_interval = 2
        elif tick_range <= 15:
            tick_interval = 3
        elif tick_range <= 20:
            tick_interval = 4
        elif tick_range <= 30:
            tick_interval = 5
        else:
            tick_interval = 10
        
        # 计算第一个刻度位置（向上取整到最近的tick_interval的倍数）
        if n_min_display >= 0:
            first_tick = math.ceil(n_min_display / tick_interval) * tick_interval
        else:
            first_tick = math.floor(n_min_display / tick_interval) * tick_interval
        
        # 计算最后一个刻度位置
        last_tick = math.floor(n_max_display / tick_interval) * tick_interval
        
        # 生成刻度值
        tick_values = []
        current = first_tick
        while current <= last_tick:
            if n_min_display <= current <= n_max_display:
                tick_values.append(current)
            current += tick_interval
        
        # 如果刻度太少，确保至少有2个刻度
        if len(tick_values) < 2:
            if n_min_display >= 0:
                tick_values = [math.floor(n_min_display), math.ceil(n_max_display)]
            else:
                tick_values = [0, math.ceil(n_max_display)]
        
        # 绘制X轴刻度
        for value in tick_values:
            # 计算x位置
            x = margin_left + ((value - n_min_display) / (n_max_display - n_min_display)) * graph_width
            
            # 刻度线
            canvas.create_line(x, margin_top + graph_height, 
                            x, margin_top + graph_height + 8, 
                            width=2, fill='black')
            
            # 刻度标签 - 显示整数
            canvas.create_text(x, margin_top + graph_height + 15, 
                            text=f"{int(round(value))}", 
                            anchor='n', 
                            font=('Arial', font_size-1),
                            fill='darkgreen')
        
        # 添加网格线
        grid_color = '#E0E0E0'
        grid_dash = (3, 2)
        
        # Y轴网格线
        for i in range(1, y_ticks):
            y = margin_top + (graph_height / y_ticks) * (y_ticks - i)
            canvas.create_line(margin_left, y, margin_left + graph_width, y, 
                            fill=grid_color, width=1, dash=grid_dash)
        
        # X轴网格线
        for value in tick_values:
            x = margin_left + ((value - n_min_display) / (n_max_display - n_min_display)) * graph_width
            canvas.create_line(x, margin_top, x, margin_top + graph_height, 
                            fill=grid_color, width=1, dash=grid_dash)
        
        # 添加坐标轴箭头
        arrow_size = 10
        canvas.create_polygon(
            margin_left - arrow_size//2, margin_top,
            margin_left + arrow_size//2, margin_top,
            margin_left, margin_top - arrow_size,
            fill='black'
        )
        
        canvas.create_polygon(
            margin_left + graph_width, margin_top + graph_height - arrow_size//2,
            margin_left + graph_width, margin_top + graph_height + arrow_size//2,
            margin_left + graph_width + arrow_size, margin_top + graph_height,
            fill='black'
        )

    def export_balance_data(self):
        """导出平衡法数据到Excel（参数和表格分开sheet）"""
        try:
            # 弹出保存文件对话框
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                title="保存平衡法数据",
                initialfile="平衡法实验数据.xlsx"
            )
            
            if not file_path:
                return
            
            # 收集参数数据
            params_data = []
            
            # 参数表头
            params_data.append(["=== 平衡法实验参数 ==="])
            params_data.append(["参数名称", "参数值", "单位"])
            
            # 添加固定参数
            params_data.append(["实验方法", "平衡法", ""])
            params_data.append(["极板间距", "5.00", "mm"])
            
            # 添加可调参数
            if hasattr(self, 'balance_l_var'):
                l_val = self.balance_l_var.get()
                params_data.append(["运动距离", l_val, "mm"])
            
            if hasattr(self, 'balance_T_var'):
                T_val = self.balance_T_var.get()
                params_data.append(["室温", T_val, "K"])
            
            if hasattr(self, 'balance_P_var'):
                P_val = self.balance_P_var.get()
                params_data.append(["大气压强", P_val, "Pa"])
            
            if hasattr(self, 'balance_g_var'):
                g_val = self.balance_g_var.get()
                params_data.append(["重力加速度", g_val, "m/s²"])
            
            if hasattr(self, 'balance_rho1_var'):
                rho1_val = self.balance_rho1_var.get()
                params_data.append(["油滴密度", rho1_val, "kg/m³"])
            
            if hasattr(self, 'balance_rho2_var'):
                rho2_val = self.balance_rho2_var.get()
                params_data.append(["空气密度", rho2_val, "kg/m³"])
            
            if hasattr(self, 'balance_eta_var'):
                eta_val = self.balance_eta_var.get()
                params_data.append(["空气粘滞系数", eta_val, "kg/(m·s)"])
            
            # 分析结果
            params_data.append([])  # 空行
            params_data.append(["=== 基本电荷分析结果 ==="])
            params_data.append(["项目", "数值", "单位"])
            params_data.append(["基本电荷理论值", "1.602", "×10⁻¹⁹ C"])
            
            if self.balance_e_experimental is not None:
                params_data.append(["基本电荷实验值", f"{self.balance_e_experimental:.3f}", "×10⁻¹⁹ C"])
            else:
                params_data.append(["基本电荷实验值", "未计算", ""])
            
            if self.balance_e_error is not None:
                params_data.append(["百分误差", f"{self.balance_e_error:.2f}", "%"])
            else:
                params_data.append(["百分误差", "未计算", ""])
            
            # 创建参数DataFrame
            params_df = pd.DataFrame(params_data)
            
            # 收集表格数据
            table_data = []
            
            # 表格标题
            table_data.append(["=== 平衡法数据记录表 ==="])
            
            # 表头
            headers = [self.balance_tree.heading(col)['text'] for col in self.balance_tree['columns']]
            table_data.append(headers)
            
            # 表格数据
            has_data = False
            for item in self.balance_tree.get_children():
                values = self.balance_tree.item(item, 'values')
                row_values = []
                for v in values:
                    if v is None or str(v).strip() == "":
                        row_values.append("")
                    else:
                        row_values.append(str(v).strip())
                
                if any(row_values):  # 至少有一个非空值
                    table_data.append(row_values)
                    has_data = True
            
            if not has_data:
                table_data.append(["暂无数据"])
            
            # 创建表格DataFrame
            table_df = pd.DataFrame(table_data)
            
            # 保存到Excel（两个sheet）
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 保存参数到第一个sheet
                params_df.to_excel(writer, sheet_name='实验参数', index=False, header=False)
                
                # 调整参数sheet列宽
                worksheet_params = writer.sheets['实验参数']
                worksheet_params.column_dimensions['A'].width = 25
                worksheet_params.column_dimensions['B'].width = 20
                worksheet_params.column_dimensions['C'].width = 15
                
                # 保存表格到第二个sheet
                table_df.to_excel(writer, sheet_name='数据记录表', index=False, header=False)
                
                # 调整表格sheet列宽
                worksheet_table = writer.sheets['数据记录表']
                for i, col in enumerate(table_df.columns):
                    col_letter = get_column_letter(i + 1)
                    worksheet_table.column_dimensions[col_letter].width = 15
            
            # 显示导出成功消息
            e_exp_display = f"{self.balance_e_experimental:.3f}" if self.balance_e_experimental is not None else "未计算"
            e_err_display = f"{self.balance_e_error:.2f}%" if self.balance_e_error is not None else "未计算"
            
            messagebox.showinfo("导出成功", 
                            f"平衡法数据已成功导出到：\n{file_path}\n\n"
                            f"导出内容：\n"
                            f"1. 实验参数（参数和计算结果）\n"
                            f"2. 数据记录表\n\n"
                            f"基本电荷分析结果：\n"
                            f"- e理论值: 1.602 ×10⁻¹⁹ C\n"
                            f"- e实验值: {e_exp_display} ×10⁻¹⁹ C\n"
                            f"- 百分误差: {e_err_display}\n\n"
                            f"数据记录表: {len(table_data)-2} 行数据")
            
            print(f"平衡法数据已导出到: {file_path}")
            
        except Exception as e:
            messagebox.showerror("导出失败", f"导出数据时出错：\n{str(e)}")
            import traceback
            traceback.print_exc()

    def load_balance_params_from_excel(self, file_path):
        """从Excel加载平衡法参数"""
        try:
            # 检查文件是否存在
            if not os.path.exists(file_path):
                print(f"文件不存在: {file_path}")
                return False
            
            # 尝试读取Excel文件
            try:
                excel_file = pd.ExcelFile(file_path)
            except Exception as e:
                print(f"无法读取Excel文件 {file_path}: {e}")
                return False
            
            sheet_names = excel_file.sheet_names
            print(f"Excel文件中的sheet: {sheet_names}")
            
            param_data = []
            
            # 优先查找包含"平衡"的sheet，否则使用第一个sheet
            target_sheet = None
            for sheet in sheet_names:
                if '平衡' in sheet:
                    target_sheet = sheet
                    break
            
            if not target_sheet and sheet_names:
                target_sheet = sheet_names[0]  # 使用第一个sheet
            
            try:
                if target_sheet:
                    print(f"读取sheet: {target_sheet}")
                    # 使用header=None确保读取所有行
                    df = pd.read_excel(file_path, sheet_name=target_sheet, header=None)
                    param_data = df.values.tolist()
                    print(f"读取到 {len(param_data)} 行数据")
            except Exception as e:
                print(f"读取sheet失败: {e}")
                return False
            
            if not param_data:
                print("没有读取到数据")
                return False
            
            # 查找实验参数区域
            param_section_start = -1
            param_section_end = -1
            
            # 首先查找实验参数标题
            for i, row in enumerate(param_data):
                if i >= 100:  # 检查前100行
                    break
                if isinstance(row, list) and len(row) > 0:
                    # 将单元格转换为字符串
                    cell_str = str(row[0]) if row[0] is not None and not pd.isna(row[0]) else ""
                    # 检查是否是实验参数标题
                    if "实验参数" in cell_str or "参数" in cell_str:
                        param_section_start = i
                        print(f"找到参数标题在第 {i} 行: {cell_str}")
                        break
            
            if param_section_start == -1:
                # 如果没有找到明确的参数标题，尝试查找表格开始位置
                for i, row in enumerate(param_data):
                    if i >= 50:
                        break
                    if isinstance(row, list) and len(row) > 0:
                        # 检查是否是参数行（包含"="或":"）
                        cell_str = str(row[0]) if row[0] is not None and not pd.isna(row[0]) else ""
                        if "=" in cell_str or ":" in cell_str:
                            param_section_start = i
                            print(f"找到参数行在第 {i} 行: {cell_str}")
                            break
            
            if param_section_start == -1:
                print("没有找到参数区域")
                return False
            
            # 查找参数区域结束位置（通常是遇到空行或下一个标题）
            for i in range(param_section_start + 1, min(param_section_start + 30, len(param_data))):
                if i >= len(param_data):
                    break
                
                row = param_data[i]
                if not isinstance(row, list) or len(row) == 0:
                    continue
                
                # 检查是否是空行或下一个标题
                first_cell = str(row[0]) if row[0] is not None and not pd.isna(row[0]) else ""
                
                # 如果是空行或下一个标题，结束参数区域
                if not first_cell.strip() or "===" in first_cell or "---" in first_cell or "###" in first_cell:
                    param_section_end = i
                    break
            
            if param_section_end == -1:
                # 如果没有明确的结束，假设参数区域持续20行
                param_section_end = min(param_section_start + 20, len(param_data))
            
            print(f"参数区域: 第 {param_section_start} 行到第 {param_section_end} 行")
            
            # 提取参数
            param_dict = {}
            
            for i in range(param_section_start, param_section_end):
                if i >= len(param_data):
                    break
                
                row = param_data[i]
                if not isinstance(row, list) or len(row) < 2:
                    continue
                
                # 提取参数名和参数值
                param_name = ""
                param_value = ""
                
                # 尝试多种格式
                if row[0] is not None and not pd.isna(row[0]):
                    cell_str = str(row[0]).strip()
                    # 尝试解析"参数名 = 值"或"参数名:"格式
                    if "=" in cell_str:
                        parts = cell_str.split("=", 1)
                        param_name = parts[0].strip()
                        if len(row) > 1 and row[1] is not None and not pd.isna(row[1]):
                            param_value = str(row[1]).strip()
                        elif len(parts) > 1:
                            param_value = parts[1].strip()
                    elif ":" in cell_str:
                        parts = cell_str.split(":", 1)
                        param_name = parts[0].strip()
                        if len(row) > 1 and row[1] is not None and not pd.isna(row[1]):
                            param_value = str(row[1]).strip()
                        elif len(parts) > 1:
                            param_value = parts[1].strip()
                    else:
                        param_name = cell_str
                        if len(row) > 1 and row[1] is not None and not pd.isna(row[1]):
                            param_value = str(row[1]).strip()
                
                # 清理参数名
                if param_name:
                    # 移除常见的单位符号
                    for unit in ["=", ":", "(", ")", "[", "]", "{", "}", "单位", "Unit", "unit"]:
                        param_name = param_name.replace(unit, "")
                    param_name = param_name.strip()
                    
                    if param_name and param_value:
                        param_dict[param_name] = param_value
                        print(f"提取到参数: {param_name} = {param_value}")
            
            print(f"平衡法提取到的参数字典: {param_dict}")
            
            if param_dict:
                # 更新平衡法参数
                self.update_balance_params_from_dict_extended(param_dict)
                return True
            else:
                print("没有提取到任何参数")
                return False
            
        except Exception as e:
            print(f"加载平衡法参数时出错: {e}")
            import traceback
            traceback.print_exc()
            return False

    def load_dynamic_params_from_excel(self, file_path):
        """从Excel加载动态法参数 - 增强版"""
        try:
            if not os.path.exists(file_path):
                print(f"文件不存在: {file_path}")
                return False
            
            # 尝试读取Excel文件
            try:
                excel_file = pd.ExcelFile(file_path)
            except Exception as e:
                print(f"无法读取Excel文件 {file_path}: {e}")
                return False
            
            sheet_names = excel_file.sheet_names
            print(f"Excel文件中的sheet: {sheet_names}")
            
            param_data = []
            
            # 优先查找包含"动态"的sheet，否则尝试所有sheet
            target_sheets = []
            for sheet in sheet_names:
                if '动态' in sheet:
                    target_sheets.append(sheet)
                elif '数据' in sheet or '参数' in sheet:
                    target_sheets.append(sheet)
            
            if not target_sheets:
                target_sheets = sheet_names[:1]  # 使用第一个sheet
            
            for target_sheet in target_sheets:
                try:
                    print(f"尝试读取sheet: {target_sheet}")
                    # 使用header=None确保读取所有行
                    df = pd.read_excel(file_path, sheet_name=target_sheet, header=None)
                    param_data = df.values.tolist()
                    print(f"从sheet '{target_sheet}' 读取到 {len(param_data)} 行数据")
                    break  # 成功读取一个sheet就停止
                except Exception as e:
                    print(f"读取sheet '{target_sheet}' 失败: {e}")
                    continue
            
            if not param_data:
                print("没有读取到数据")
                return False
            
            # 扩展的参数映射字典
            param_mapping = {
                '运动距离': 'l',
                '运动距离 l': 'l',
                '距离': 'l',
                'l': 'l',
                '下落距离': 'l',
                
                '室温': 'T',
                '室温 T': 'T',
                '温度': 'T',
                'T': 'T',
                '室温(℃)': 'T',
                '室温(K)': 'T',
                
                '大气压强': 'P',
                '大气压强 P': 'P',
                '压强': 'P',
                'P': 'P',
                '大气压': 'P',
                
                '重力加速度': 'g',
                '重力加速度 g': 'g',
                'g': 'g',
                '重力': 'g',
                
                '油滴密度': 'rho1',
                '油滴密度 ρ1': 'rho1',
                'ρ1': 'rho1',
                '密度1': 'rho1',
                '油滴': 'rho1',
                
                '空气密度': 'rho2',
                '空气密度 ρ2': 'rho2',
                'ρ2': 'rho2',
                '密度2': 'rho2',
                '空气': 'rho2',
                
                '空气粘滞系数': 'eta',
                '空气粘滞系数 η': 'eta',
                'η': 'eta',
                '粘滞系数': 'eta',
                '粘滞': 'eta'
            }
            
            # 用于存储提取的参数
            param_dict = {}
            
            # 查找参数区域的策略
            param_sections_found = []
            
            # 策略1: 查找包含"参数"的标题行
            for i, row in enumerate(param_data):
                if i >= 100:  # 检查前100行
                    break
                if isinstance(row, list) and len(row) > 0:
                    first_cell = str(row[0]) if row[0] is not None and not pd.isna(row[0]) else ""
                    if "实验参数" in first_cell or "参数" in first_cell:
                        param_sections_found.append(i)
                        print(f"找到参数标题在第 {i} 行: {first_cell}")
            
            # 策略2: 如果没找到明确标题，查找包含参数名称的行
            if not param_sections_found:
                for i, row in enumerate(param_data):
                    if i >= 50:  # 检查前50行
                        break
                    if isinstance(row, list) and len(row) > 1:
                        first_cell = str(row[0]) if row[0] is not None and not pd.isna(row[0]) else ""
                        second_cell = str(row[1]) if len(row) > 1 and row[1] is not None and not pd.isna(row[1]) else ""
                        
                        # 检查是否是参数行（包含参数名和数值）
                        for key in param_mapping.keys():
                            if key in first_cell:
                                param_sections_found.append(i)
                                print(f"找到参数行在第 {i} 行: {first_cell} = {second_cell}")
                                break
            
            # 策略3: 查找任何包含等号或冒号的行
            if not param_sections_found:
                for i, row in enumerate(param_data):
                    if i >= 50:
                        break
                    if isinstance(row, list) and len(row) > 1:
                        first_cell = str(row[0]) if row[0] is not None and not pd.isna(row[0]) else ""
                        second_cell = str(row[1]) if len(row) > 1 and row[1] is not None and not pd.isna(row[1]) else ""
                        
                        if ('=' in first_cell or ':' in first_cell) and second_cell.strip():
                            param_sections_found.append(i)
                            print(f"找到参数行(格式匹配)在第 {i} 行: {first_cell}")
            
            if not param_sections_found:
                print("没有找到参数区域")
                return False
            
            # 提取所有找到的参数行
            for start_row in param_sections_found:
                # 从开始行向下检查最多20行
                for i in range(start_row, min(start_row + 20, len(param_data))):
                    if i >= len(param_data):
                        break
                    
                    row = param_data[i]
                    if not isinstance(row, list) or len(row) < 2:
                        continue
                    
                    # 提取参数名和值
                    param_name = ""
                    param_value = ""
                    
                    if row[0] is not None and not pd.isna(row[0]):
                        cell_str = str(row[0]).strip()
                        
                        # 多种格式解析
                        if '=' in cell_str:
                            parts = cell_str.split('=', 1)
                            param_name = parts[0].strip()
                            if len(parts) > 1:
                                param_value = parts[1].strip()
                            elif len(row) > 1 and row[1] is not None and not pd.isna(row[1]):
                                param_value = str(row[1]).strip()
                        elif ':' in cell_str:
                            parts = cell_str.split(':', 1)
                            param_name = parts[0].strip()
                            if len(parts) > 1:
                                param_value = parts[1].strip()
                            elif len(row) > 1 and row[1] is not None and not pd.isna(row[1]):
                                param_value = str(row[1]).strip()
                        else:
                            param_name = cell_str
                            if len(row) > 1 and row[1] is not None and not pd.isna(row[1]):
                                param_value = str(row[1]).strip()
                    
                    # 清理参数名
                    if param_name:
                        # 移除常见的单位符号和非字母数字字符
                        cleanup_chars = ['=', ':', '(', ')', '[', ']', '{', '}', 
                                    '单位', 'Unit', 'unit', '参数', 'Parameter']
                        for char in cleanup_chars:
                            param_name = param_name.replace(char, "")
                        param_name = param_name.strip()
                        
                        if param_name and param_value and param_value.strip():
                            param_dict[param_name] = param_value.strip()
                            print(f"提取到参数: '{param_name}' = '{param_value}'")
            
            print(f"动态法提取到的参数字典: {param_dict}")
            
            if param_dict:
                # 更新动态法参数
                self.update_dynamic_params_from_dict_extended(param_dict)
                return True
            else:
                print("没有提取到任何参数")
                return False
                
        except Exception as e:
            print(f"加载动态法参数时出错: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def update_balance_params_from_dict_extended(self, param_dict):
        """从字典更新平衡法参数到界面 - 简化版"""
        try:
            print(f"开始更新平衡法参数，参数数量: {len(param_dict)}")
            
            # 简单的参数映射
            param_mapping = {
                '运动距离': 'l',
                '室温': 'T',
                '大气压强': 'P',
                '重力加速度': 'g',
                '油滴密度': 'rho1',
                '空气密度': 'rho2',
                '空气粘滞系数': 'eta'
            }
            
            updated_params = []
            
            for param_name, param_value in param_dict.items():
                # 查找匹配的参数
                for map_name, map_key in param_mapping.items():
                    if map_name in param_name:
                        if param_value and param_value.strip():
                            # 更新存储
                            self.balance_params_storage[map_key] = param_value.strip()
                            
                            # 更新界面
                            if map_key == 'l' and hasattr(self, 'balance_l_var'):
                                self.balance_l_var.set(param_value.strip())
                                updated_params.append(f"运动距离: {param_value}")
                            elif map_key == 'T' and hasattr(self, 'balance_T_var'):
                                self.balance_T_var.set(param_value.strip())
                                updated_params.append(f"室温: {param_value}")
                            elif map_key == 'P' and hasattr(self, 'balance_P_var'):
                                self.balance_P_var.set(param_value.strip())
                                updated_params.append(f"大气压强: {param_value}")
                            elif map_key == 'g' and hasattr(self, 'balance_g_var'):
                                self.balance_g_var.set(param_value.strip())
                                updated_params.append(f"重力加速度: {param_value}")
                            elif map_key == 'rho1' and hasattr(self, 'balance_rho1_var'):
                                self.balance_rho1_var.set(param_value.strip())
                                updated_params.append(f"油滴密度: {param_value}")
                            elif map_key == 'rho2' and hasattr(self, 'balance_rho2_var'):
                                self.balance_rho2_var.set(param_value.strip())
                                updated_params.append(f"空气密度: {param_value}")
                            elif map_key == 'eta' and hasattr(self, 'balance_eta_var'):
                                # 统一处理η的显示格式
                                try:
                                    eta_val = float(param_value.replace('×10⁻⁵', ''))
                                    self.balance_eta_var.set(f"{eta_val:.3f}×10⁻⁵")  # 统一三位小数
                                except:
                                    self.balance_eta_var.set(param_value.strip())
                                updated_params.append(f"空气粘滞系数: {param_value}")
                        break
            
            if updated_params:
                print(f"平衡法参数更新完成，更新了 {len(updated_params)} 个参数")
            else:
                print("没有更新任何参数")
            
            print(f"最终参数存储: {self.balance_params_storage}")
                            
        except Exception as e:
            print(f"更新平衡法参数时出错: {e}")

    def import_balance_data(self):
        """从Excel导入平衡法数据（参数和表格分开sheet）"""
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
                title="选择要导入的平衡法数据文件"
            )
            
            if not file_path:
                return
            
            print(f"开始导入文件: {file_path}")
            
            try:
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names
                print(f"Excel文件中包含的sheet: {sheet_names}")
            except Exception as e:
                messagebox.showerror("错误", f"无法读取Excel文件：\n{str(e)}")
                return
            
            # 1. 导入参数（从"实验参数"sheet）
            params_loaded = False
            if '实验参数' in sheet_names:
                print("找到'实验参数'sheet，开始导入参数...")
                params_loaded = self.load_params_from_sheet(file_path, '实验参数', 'balance')
            else:
                print("没有找到'实验参数'sheet，跳过参数导入")
            
            # 2. 导入表格数据（从"数据记录表"sheet）
            data_loaded = False
            imported_rows = 0
            
            if '数据记录表' in sheet_names:
                print("找到'数据记录表'sheet，开始导入表格数据...")
                imported_rows = self.load_table_data_from_sheet(file_path, '数据记录表', 'balance')
                if imported_rows > 0:
                    data_loaded = True
            else:
                print("没有找到'数据记录表'sheet，跳过表格数据导入")
            
            # 3. 显示导入结果
            if data_loaded:
                # 添加一个空行用于继续输入
                self.balance_tree.insert("", "end", values=[""] * len(self.balance_tree['columns']))
                
                # 重新计算数据
                print("重新计算数据...")
                self.root.after(500, self.calculate_balance_data)
            
            # 显示结果
            message_text = f"平衡法数据导入完成：\n\n"
            message_text += f"参数导入: {'成功' if params_loaded else '失败'}\n"
            message_text += f"表格导入: {'成功' if data_loaded else '失败'}\n"
            if data_loaded:
                message_text += f"导入行数: {imported_rows}行\n"
            
            messagebox.showinfo("导入完成", message_text)
            
        except Exception as e:
            messagebox.showerror("导入失败", f"导入数据时出错：\n{str(e)}")
            import traceback
            traceback.print_exc()

    def extract_params_from_excel(self, df, method_type):
        """从Excel数据中提取参数 - 安全版本"""
        try:
            if not isinstance(df, pd.DataFrame):
                print("extract_params_from_excel: df不是DataFrame")
                return
            
            # 将DataFrame转换为列表以便处理
            data = df.values.tolist()
            
            param_section_start = -1
            for i, row in enumerate(data[:50]):  # 只检查前50行
                if isinstance(row, list) and len(row) > 0:
                    first_cell = str(row[0]) if row[0] is not None and not pd.isna(row[0]) else ""
                    if method_type == 'balance' and "平衡法实验参数" in first_cell:
                        param_section_start = i
                        break
                    elif method_type == 'dynamic' and "动态法实验参数" in first_cell:
                        param_section_start = i
                        break
            
            if param_section_start == -1:
                return
            
            # 提取参数
            param_dict = {}
            start_row = param_section_start + 2
            
            for i in range(start_row, min(start_row + 20, len(data))):
                if i >= len(data):
                    break
                
                row = data[i]
                if isinstance(row, list) and len(row) >= 2:
                    param_name = str(row[0]) if row[0] is not None and not pd.isna(row[0]) else ""
                    param_value = str(row[1]) if row[1] is not None and not pd.isna(row[1]) else ""
                    
                    if param_name and param_value and not any(marker in param_name for marker in ['===', '---', '###']):
                        param_dict[param_name.strip()] = param_value.strip()
            
            # 调用更新方法
            if method_type == 'balance':
                self.update_balance_params_from_dict_extended(param_dict)
            else:
                self.update_dynamic_params_from_dict_extended(param_dict)
                
        except Exception as e:
            print(f"extract_params_from_excel出错: {e}")

    def update_param_from_excel(self, param_name, param_value, method_type):
        """从Excel更新参数到界面"""
        try:
            # 提取纯数值部分
            value_num = ''.join(c for c in param_value if c.isdigit() or c == '.' or c == '-')
            
            if not value_num:
                return
            
            if method_type == 'balance':
                if "运动距离" in param_name:
                    self.balance_l_var.set(value_num)
                    self.balance_params_storage['l'] = value_num
                elif "室温" in param_name:
                    self.balance_T_var.set(value_num)
                    self.balance_params_storage['T'] = value_num
                elif "大气压强" in param_name:
                    self.balance_P_var.set(value_num)
                    self.balance_params_storage['P'] = value_num
                elif "重力加速度" in param_name:
                    self.balance_g_var.set(value_num)
                    self.balance_params_storage['g'] = value_num
                elif "油滴密度" in param_name:
                    self.balance_rho1_var.set(value_num)
                    self.balance_params_storage['rho1'] = value_num
                elif "空气密度" in param_name:
                    self.balance_rho2_var.set(value_num)
                    self.balance_params_storage['rho2'] = value_num
                elif "空气粘滞系数" in param_name:
                    self.balance_eta_var.set(param_value)  # 保留原始格式
                    self.balance_params_storage['eta'] = param_value
            
            elif method_type == 'dynamic':
                if "运动距离" in param_name:
                    self.dynamic_l_var.set(value_num)
                    self.dynamic_params_storage['l'] = value_num
                elif "室温" in param_name:
                    self.dynamic_T_var.set(value_num)
                    self.dynamic_params_storage['T'] = value_num
                elif "大气压强" in param_name:
                    self.dynamic_P_var.set(value_num)
                    self.dynamic_params_storage['P'] = value_num
                elif "重力加速度" in param_name:
                    self.dynamic_g_var.set(value_num)
                    self.dynamic_params_storage['g'] = value_num
                elif "油滴密度" in param_name:
                    self.dynamic_rho1_var.set(value_num)
                    self.dynamic_params_storage['rho1'] = value_num
                elif "空气密度" in param_name:
                    self.dynamic_rho2_var.set(value_num)
                    self.dynamic_params_storage['rho2'] = value_num
                elif "空气粘滞系数" in param_name:
                    self.dynamic_eta_var.set(param_value)
                    self.dynamic_params_storage['eta'] = param_value
                    
        except Exception as e:
            print(f"更新参数时出错: {e}")

    def show_dynamic_method(self):
        """显示动态法内容 - 从存储中恢复数据"""
        # 创建带滚动条的主容器
        main_container = tk.Frame(self.content_frame)
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # 创建垂直滚动条
        v_scrollbar_main = tk.Scrollbar(main_container)
        v_scrollbar_main.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建水平滚动条
        h_scrollbar_main = tk.Scrollbar(main_container, orient=tk.HORIZONTAL)
        h_scrollbar_main.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 创建Canvas作为滚动区域
        canvas = tk.Canvas(main_container, 
                        yscrollcommand=v_scrollbar_main.set,
                        xscrollcommand=h_scrollbar_main.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        v_scrollbar_main.config(command=canvas.yview)
        h_scrollbar_main.config(command=canvas.xview)
        
        # 在Canvas中创建Frame作为内容容器
        content_frame = tk.Frame(canvas)
        
        # 将Frame添加到Canvas
        canvas_window = canvas.create_window((0, 0), window=content_frame, anchor='nw')
        
        # 配置Canvas滚动
        def configure_canvas(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(canvas_window, width=canvas.winfo_width())
        
        canvas.bind('<Configure>', configure_canvas)
        content_frame.bind('<Configure>', configure_canvas)
        
        # ====== 实验参数区域 ======
        param_frame = tk.LabelFrame(content_frame, text="实验参数", font=('Arial', 10, 'bold'))
        param_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 创建参数网格
        param_grid = tk.Frame(param_frame)
        param_grid.pack(padx=10, pady=10)
        
        # 第一行
        row = 0
        tk.Label(param_grid, text="极板间距 d =", font=('Arial', 9)).grid(row=row, column=0, sticky='w', padx=(0, 5))
        tk.Label(param_grid, text="5.00 mm", font=('Arial', 9, 'bold'), fg='blue').grid(row=row, column=1, sticky='w')
        tk.Label(param_grid, text="(不可修改)", font=('Arial', 8), fg='gray').grid(row=row, column=2, sticky='w', padx=(5, 20))
        
        tk.Label(param_grid, text="运动距离 l =", font=('Arial', 9)).grid(row=row, column=3, sticky='w', padx=(20, 5))
        self.dynamic_l_var = tk.StringVar(value=self.dynamic_params_storage["l"])
        tk.Entry(param_grid, textvariable=self.dynamic_l_var, width=8, font=('Arial', 9)).grid(row=row, column=4, sticky='w', padx=5)
        tk.Label(param_grid, text="mm", font=('Arial', 9)).grid(row=row, column=5, sticky='w')
        
        # 第二行
        row += 1
        tk.Label(param_grid, text="室温 T =", font=('Arial', 9)).grid(row=row, column=0, sticky='w', padx=(0, 5))
        self.dynamic_T_var = tk.StringVar(value=self.dynamic_params_storage["T"])
        tk.Entry(param_grid, textvariable=self.dynamic_T_var, width=8, font=('Arial', 9)).grid(row=row, column=1, sticky='w', padx=5)
        tk.Label(param_grid, text="K", font=('Arial', 9)).grid(row=row, column=2, sticky='w')
        
        tk.Label(param_grid, text="大气压强 P =", font=('Arial', 9)).grid(row=row, column=3, sticky='w', padx=(20, 5))
        self.dynamic_P_var = tk.StringVar(value=self.dynamic_params_storage["P"])
        tk.Entry(param_grid, textvariable=self.dynamic_P_var, width=10, font=('Arial', 9)).grid(row=row, column=4, sticky='w', padx=5)
        tk.Label(param_grid, text="Pa", font=('Arial', 9)).grid(row=row, column=5, sticky='w')
        
        # 第三行
        row += 1
        tk.Label(param_grid, text="重力加速度 g =", font=('Arial', 9)).grid(row=row, column=0, sticky='w', padx=(0, 5))
        self.dynamic_g_var = tk.StringVar(value=self.dynamic_params_storage["g"])
        tk.Entry(param_grid, textvariable=self.dynamic_g_var, width=8, font=('Arial', 9)).grid(row=row, column=1, sticky='w', padx=5)
        tk.Label(param_grid, text="m/s²", font=('Arial', 9)).grid(row=row, column=2, sticky='w')
        
        tk.Label(param_grid, text="油滴密度 ρ1 =", font=('Arial', 9)).grid(row=row, column=3, sticky='w', padx=(20, 5))
        self.dynamic_rho1_var = tk.StringVar(value=self.dynamic_params_storage["rho1"])
        # 修改为输入框而不是标签
        tk.Entry(param_grid, textvariable=self.dynamic_rho1_var, width=8, font=('Arial', 9)).grid(row=row, column=4, sticky='w', padx=5)
        tk.Label(param_grid, text="kg/m³", font=('Arial', 9)).grid(row=row, column=5, sticky='w')
        
        # 第四行
        row += 1
        tk.Label(param_grid, text="空气密度 ρ2 =", font=('Arial', 9)).grid(row=row, column=0, sticky='w', padx=(0, 5))
        self.dynamic_rho2_var = tk.StringVar(value=self.dynamic_params_storage["rho2"])
        tk.Label(param_grid, textvariable=self.dynamic_rho2_var, font=('Arial', 9, 'bold'), 
                fg='green', width=10).grid(row=row, column=1, sticky='w', padx=5)
        tk.Label(param_grid, text="kg/m³", font=('Arial', 9)).grid(row=row, column=2, sticky='w')
        
        tk.Label(param_grid, text="空气粘滞系数 η =", font=('Arial', 9)).grid(row=row, column=3, sticky='w', padx=(20, 5))
        self.dynamic_eta_var = tk.StringVar(value=self.dynamic_params_storage["eta"])
        tk.Label(param_grid, textvariable=self.dynamic_eta_var, font=('Arial', 9, 'bold'), 
                fg='green', width=12).grid(row=row, column=4, sticky='w', padx=5)
        tk.Label(param_grid, text="kg/(m·s)", font=('Arial', 9)).grid(row=row, column=5, sticky='w')
        
        # 让参数网格自适应
        for i in range(6):
            param_grid.columnconfigure(i, weight=1)
        
        # ====== 数据记录表格区域 ======
        table_frame = tk.LabelFrame(content_frame, text="数据记录表", font=('Arial', 10, 'bold'))
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 设置表格容器固定高度
        self.dynamic_table_container = tk.Frame(table_frame, height=250)
        self.dynamic_table_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.dynamic_table_container.pack_propagate(False)
        
        # 创建表格（14列）
        columns = ("电压U/V", 
                "上升时间第一次te1/s", "上升时间第二次te2/s", "上升时间第三次te3/s", "平均上升时间te/s",
                "下落时间第一次tg1/s", "下落时间第二次tg2/s", "下落时间第三次tg3/s", "平均下落时间tg/s",
                "上升速度ve/×10⁻⁵m/s", "下落速度vg/×10⁻⁵m/s", 
                "油滴半径a/×10⁻⁷m", "油滴带电量q/×10⁻¹⁹C", "带电子数n")
        
        # 创建Treeview
        self.dynamic_tree = ttk.Treeview(self.dynamic_table_container, columns=columns, show="headings", height=6)
        
        # 设置列宽
        col_widths = [80, 100+40, 100+40, 100+40, 100+40, 
                    100+40, 100+40, 100+40, 100+40, 
                    100+40, 100+40, 100+40, 100+40, 80]
        
        for i, col in enumerate(columns):
            self.dynamic_tree.heading(col, text=col)
            self.dynamic_tree.column(col, width=col_widths[i], minwidth=50,stretch=False)
        
        # 创建垂直滚动条
        tree_v_scrollbar = tk.Scrollbar(self.dynamic_table_container, orient=tk.VERTICAL)
        tree_v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建水平滚动条
        tree_h_scrollbar = tk.Scrollbar(self.dynamic_table_container, orient=tk.HORIZONTAL)
        tree_h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 配置Treeview
        self.dynamic_tree.configure(yscrollcommand=tree_v_scrollbar.set, 
                                xscrollcommand=tree_h_scrollbar.set)
        self.dynamic_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        tree_v_scrollbar.config(command=self.dynamic_tree.yview)
        tree_h_scrollbar.config(command=self.dynamic_tree.xview)
        
        # 从存储中恢复数据，如果没有数据则添加初始3行空行
        if self.dynamic_table_data:
            for row_data in self.dynamic_table_data:
                self.dynamic_tree.insert("", "end", values=row_data)
        else:
            for _ in range(3):
                self.dynamic_tree.insert("", "end", values=[""] * len(columns))
        
        # 绑定键盘事件
        self.dynamic_tree.bind("<KeyRelease>", self.on_dynamic_table_keyrelease)
        # 绑定双击编辑事件
        self.dynamic_tree.bind("<Double-1>", self.on_dynamic_tree_double_click)
        
        # ====== 图表区域 ======
        chart_frame = tk.LabelFrame(content_frame, text="电荷-电子数关系图", font=('Arial', 10, 'bold'))
        chart_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        chart_container = tk.Frame(chart_frame, height=300)
        chart_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        chart_container.pack_propagate(False)
        
        self.dynamic_chart_canvas = tk.Canvas(chart_container, bg='white', width=600, height=300)
        self.dynamic_chart_canvas.pack(fill=tk.BOTH, expand=True)
        
        # 初始绘制坐标轴
        self.draw_dynamic_chart_axes()
        
        # ====== 按钮区域 ======
        button_frame = tk.Frame(content_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 按钮
        buttons = [
            ("计算", self.calculate_dynamic_data),
            ("删除选中行", self.delete_dynamic_selected_row),
            ("清空数据", self.clear_dynamic_data),
            ("导出数据", self.export_dynamic_data),
            ("导入数据", self.import_dynamic_data)
        ]
        
        for text, command in buttons:
            tk.Button(button_frame, text=text, command=command,
                    font=('Arial', 9), width=12).pack(side=tk.LEFT, padx=5)
        
        # 确保内容框架正确配置
        content_frame.update_idletasks()
        
        # 如果有数据，绘制图表
        if self.dynamic_table_data and any(any(cell for cell in row[12:14] if cell) for row in self.dynamic_table_data):
            self.root.after(100, self.plot_dynamic_chart)

    def on_dynamic_table_keyrelease(self, event):
        """动态法表格键盘释放事件 - 检测最后一行是否有数据，并保存当前数据"""
        # 获取所有行
        children = self.dynamic_tree.get_children()
        if not children:
            return
        
        # 检查最后一行
        last_item = children[-1]
        values = self.dynamic_tree.item(last_item, 'values')
        
        # 如果最后一行有手动输入列的数据（第0-8列），添加新行
        if values and any(v.strip() for v in values[:9]):
            # 添加新空行
            self.dynamic_tree.insert("", "end", values=[""] * len(self.dynamic_tree['columns']))
        
        # 保存当前所有数据到存储
        self.save_dynamic_table_data()

    def save_balance_table_data(self):
        """保存平衡法表格数据到存储"""
        self.balance_table_data = []
        for item in self.balance_tree.get_children():
            values = self.balance_tree.item(item, 'values')
            self.balance_table_data.append(values)
        
        # 保存参数
        self.balance_params_storage["l"] = self.balance_l_var.get()
        self.balance_params_storage["T"] = self.balance_T_var.get()
        self.balance_params_storage["P"] = self.balance_P_var.get()
        self.balance_params_storage["g"] = self.balance_g_var.get()
        self.balance_params_storage["rho1"] = self.balance_rho1_var.get()
        self.balance_params_storage["rho2"] = self.balance_rho2_var.get()
        self.balance_params_storage["eta"] = self.balance_eta_var.get()

    def save_dynamic_table_data(self):
        """保存动态法表格数据到存储"""
        self.dynamic_table_data = []
        for item in self.dynamic_tree.get_children():
            values = self.dynamic_tree.item(item, 'values')
            self.dynamic_table_data.append(values)
        
        # 保存参数
        self.dynamic_params_storage["l"] = self.dynamic_l_var.get()
        self.dynamic_params_storage["T"] = self.dynamic_T_var.get()
        self.dynamic_params_storage["P"] = self.dynamic_P_var.get()
        self.dynamic_params_storage["g"] = self.dynamic_g_var.get()
        self.dynamic_params_storage["rho1"] = self.dynamic_rho1_var.get()
        self.dynamic_params_storage["rho2"] = self.dynamic_rho2_var.get()
        self.dynamic_params_storage["eta"] = self.dynamic_eta_var.get()

    def on_dynamic_tree_double_click(self, event):
        """动态法表格双击编辑"""
        region = self.dynamic_tree.identify("region", event.x, event.y)
        
        if region == "cell":
            # 获取单元格信息
            row = self.dynamic_tree.identify_row(event.y)
            column = self.dynamic_tree.identify_column(event.x)
            
            # 获取列索引（从1开始）
            col_index = int(column.replace('#', '')) - 1
            
            # 只允许编辑前8列（手动输入列）：电压(0)、三次上升时间(1-3)、三次下落时间(5-7)
            # 第4列（平均上升时间）和第8列（平均下落时间）不可编辑
            non_editable_columns = [4, 8]  # 平均上升时间列、平均下落时间列
            
            if col_index in non_editable_columns:  # 平均时间列不可编辑
                self.show_tooltip("此列数据为自动计算，不可手动修改", event.x_root, event.y_root)
                return
            
            # 获取单元格位置
            x, y, width, height = self.dynamic_tree.bbox(row, column)
            
            # 获取当前值
            item = self.dynamic_tree.item(row)
            values = item['values']
            
            if col_index < len(values):
                current = values[col_index] if values[col_index] is not None else ""
            else:
                current = ""
            
            # 创建输入框
            entry = tk.Entry(self.dynamic_tree, 
                            font=('Arial', 9),
                            borderwidth=0)
            
            # 放置在单元格上
            entry.place(x=x, y=y, width=width, height=height)
            entry.insert(0, current)
            entry.select_range(0, tk.END)
            entry.focus()
            
            def save_and_quit(event=None):
                new_value = entry.get()
                
                # 验证输入是否为数字（电压和时间需要是数字）
                if col_index in [0, 1, 2, 3, 5, 6, 7]:  # 电压和时间列
                    try:
                        if new_value.strip():  # 如果不是空值
                            float(new_value)
                    except ValueError:
                        messagebox.showerror("错误", "请输入有效的数字")
                        entry.focus()
                        return
                
                # 更新数据
                new_values = list(values)
                if col_index < len(new_values):
                    new_values[col_index] = new_value
                else:
                    new_values.append(new_value)
                
                # 更新行数据
                self.dynamic_tree.item(row, values=new_values)
                entry.destroy()
            
            def cancel_edit(event=None):
                entry.destroy()
            
            # 绑定事件
            entry.bind('<Return>', save_and_quit)
            entry.bind('<FocusOut>', save_and_quit)
            entry.bind('<Escape>', cancel_edit)

    def get_balance_analysis_results(self):
        """获取平衡法分析结果"""
        return {
            '平衡法_e实验值': self.balance_e_experimental,
            '平衡法_百分误差': self.balance_e_error,
            '电子电荷标准值': 1.602
        }

    def get_dynamic_analysis_results(self):
        """获取动态法分析结果"""
        return {
            '动态法_e实验值': self.dynamic_e_experimental,
            '动态法_百分误差': self.dynamic_e_error,
            '电子电荷标准值': 1.602
        }

    def get_balance_params_dict(self):
        """获取平衡法所有参数"""
        params = {}
        
        try:
            # 基础参数
            params['实验方法'] = '平衡法'
            params['极板间距'] = '5.00 mm'
            
            # 从界面获取参数值（安全处理）
            if hasattr(self, 'balance_l_var'):
                l_val = self.balance_l_var.get()
                params['运动距离(l)'] = f"{l_val} mm" if l_val else "未设置"
            
            if hasattr(self, 'balance_T_var'):
                T_val = self.balance_T_var.get()
                params['室温(T)'] = f"{T_val} K" if T_val else "未设置"
            
            if hasattr(self, 'balance_P_var'):
                P_val = self.balance_P_var.get()
                params['大气压强(P)'] = f"{P_val} Pa" if P_val else "未设置"
            
            if hasattr(self, 'balance_g_var'):
                g_val = self.balance_g_var.get()
                params['重力加速度(g)'] = f"{g_val} m/s²" if g_val else "未设置"
            
            if hasattr(self, 'balance_rho1_var'):
                rho1_val = self.balance_rho1_var.get()
                params['油滴密度(ρ₁)'] = f"{rho1_val} kg/m³" if rho1_val else "未设置"
            
            if hasattr(self, 'balance_rho2_var'):
                rho2_val = self.balance_rho2_var.get()
                params['空气密度(ρ₂)'] = f"{rho2_val} kg/m³" if rho2_val else "未设置"
            
            if hasattr(self, 'balance_eta_var'):
                eta_val = self.balance_eta_var.get()
                params['空气粘滞系数(η)'] = f"{eta_val} kg/(m·s)" if eta_val else "未设置"
            
            # 分析结果（安全处理None值）
            if self.balance_e_experimental is not None:
                params['基本电荷实验值(e)'] = f"{self.balance_e_experimental:.3f} ×10⁻¹⁹ C"
            else:
                params['基本电荷实验值(e)'] = "未计算"
            
            if self.balance_e_error is not None:
                params['百分误差(δ)'] = f"{self.balance_e_error:.2f}%"
            else:
                params['百分误差(δ)'] = "未计算"
            
            params['基本电荷理论值(e₀)'] = "1.602 ×10⁻¹⁹ C"
            
        except Exception as e:
            print(f"获取平衡法参数时出错: {e}")
        
        return params

    def get_dynamic_params_dict(self):
        """获取动态法所有参数"""
        params = {}
        
        # 实验参数
        for key, value in self.dynamic_params_storage.items():
            params[f'动态法_{key}'] = value
        
        # 表格参数（从界面获取）
        params['动态法_极板间距'] = '5.00 mm'
        
        # 分析结果
        analysis = self.get_dynamic_analysis_results()
        params.update(analysis)
        
        return params

    def save_params_to_excel(self, writer, params_dict, sheet_name='实验参数'):
        """保存参数到Excel的单独sheet"""
        try:
            # 将参数转换为DataFrame
            param_df = pd.DataFrame(list(params_dict.items()), columns=['参数名称', '参数值'])
            
            # 保存到新的sheet
            param_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 调整列宽
            worksheet = writer.sheets[sheet_name]
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 20
            
            return True
        except Exception as e:
            print(f"保存参数时出错: {e}")
            return False

    def load_params_from_excel(self, file_path):
        """从Excel加载参数 - 修复版本"""
        try:
            # 首先检查文件是否存在
            if not os.path.exists(file_path):
                print(f"文件不存在: {file_path}")
                return False
            
            # 尝试读取Excel文件
            try:
                excel_file = pd.ExcelFile(file_path)
            except Exception as e:
                print(f"无法读取Excel文件 {file_path}: {e}")
                return False
            
            # 检查是否有实验参数sheet
            sheet_names = excel_file.sheet_names
            print(f"Excel文件中的sheet: {sheet_names}")
            
            param_data = []
            
            # 查找包含参数的sheet（可能是第一个sheet或者包含"参数"字样的sheet）
            target_sheet = None
            for sheet in sheet_names:
                if '参数' in sheet or 'param' in sheet.lower():
                    target_sheet = sheet
                    break
            
            if not target_sheet and sheet_names:
                target_sheet = sheet_names[0]  # 默认使用第一个sheet
            
            try:
                if target_sheet:
                    print(f"读取sheet: {target_sheet}")
                    df = pd.read_excel(file_path, sheet_name=target_sheet, header=None)
                    param_data = df.values.tolist()
            except Exception as e:
                print(f"读取sheet失败: {e}")
                return False
            
            if not param_data:
                return False
            
            # 处理参数数据 - 寻找参数区域
            param_section_start = -1
            param_section_end = -1
            
            for i, row in enumerate(param_data):
                if i >= 50:  # 只检查前50行
                    break
                if isinstance(row, list) and len(row) > 0:
                    first_cell = str(row[0]) if pd.notna(row[0]) else ""
                    if "实验参数" in first_cell or "=== 动态法实验参数 ===" in first_cell:
                        param_section_start = i
                    elif param_section_start != -1 and "=== 数据记录表 ===" in first_cell:
                        param_section_end = i
                        break
            
            # 如果没找到明确的结束标记，假设参数区域持续到数据记录表之前
            if param_section_start != -1 and param_section_end == -1:
                for i in range(param_section_start + 1, len(param_data)):
                    if i >= len(param_data):
                        break
                    row = param_data[i]
                    if isinstance(row, list) and len(row) > 0:
                        first_cell = str(row[0]) if pd.notna(row[0]) else ""
                        if "数据记录表" in first_cell or i - param_section_start > 20:
                            param_section_end = i
                            break
            
            # 提取参数
            if param_section_start != -1:
                param_dict = {}
                
                # 参数区域通常从第param_section_start+2行开始（跳过标题和表头）
                start_row = param_section_start + 2
                
                if param_section_end == -1:
                    param_section_end = min(start_row + 15, len(param_data))
                
                for i in range(start_row, param_section_end):
                    if i >= len(param_data):
                        break
                    
                    row = param_data[i]
                    if isinstance(row, list) and len(row) >= 2:
                        param_name = str(row[0]) if pd.notna(row[0]) else ""
                        param_value = str(row[1]) if pd.notna(row[1]) else ""
                        
                        if param_name and param_value:
                            param_dict[param_name.strip()] = param_value.strip()
                
                print(f"提取到的参数: {param_dict}")
                
                # 更新动态法参数
                self.update_dynamic_params_from_dict_extended(param_dict)
                return True
            
            return False
            
        except Exception as e:
            print(f"加载参数时出错: {e}")
            import traceback
            traceback.print_exc()
            return False

    def update_dynamic_params_from_dict_extended(self, param_dict):
        """从字典更新动态法参数到界面 - 简化版"""
        try:
            print(f"开始更新动态法参数，参数数量: {len(param_dict)}")
            
            # 简单的参数映射
            param_mapping = {
                '运动距离': 'l',
                '室温': 'T',
                '大气压强': 'P',
                '重力加速度': 'g',
                '油滴密度': 'rho1',
                '空气密度': 'rho2',
                '空气粘滞系数': 'eta'
            }
            
            updated_params = []
            
            for param_name, param_value in param_dict.items():
                # 查找匹配的参数
                for map_name, map_key in param_mapping.items():
                    if map_name in param_name:
                        if param_value and param_value.strip():
                            # 更新存储
                            self.dynamic_params_storage[map_key] = param_value.strip()
                            
                            # 更新界面
                            if map_key == 'l' and hasattr(self, 'dynamic_l_var'):
                                self.dynamic_l_var.set(param_value.strip())
                                updated_params.append(f"运动距离: {param_value}")
                            elif map_key == 'T' and hasattr(self, 'dynamic_T_var'):
                                self.dynamic_T_var.set(param_value.strip())
                                updated_params.append(f"室温: {param_value}")
                            elif map_key == 'P' and hasattr(self, 'dynamic_P_var'):
                                self.dynamic_P_var.set(param_value.strip())
                                updated_params.append(f"大气压强: {param_value}")
                            elif map_key == 'g' and hasattr(self, 'dynamic_g_var'):
                                self.dynamic_g_var.set(param_value.strip())
                                updated_params.append(f"重力加速度: {param_value}")
                            elif map_key == 'rho1' and hasattr(self, 'dynamic_rho1_var'):
                                self.dynamic_rho1_var.set(param_value.strip())
                                updated_params.append(f"油滴密度: {param_value}")
                            elif map_key == 'rho2' and hasattr(self, 'dynamic_rho2_var'):
                                self.dynamic_rho2_var.set(param_value.strip())
                                updated_params.append(f"空气密度: {param_value}")
                            elif map_key == 'eta' and hasattr(self, 'dynamic_eta_var'):
                                self.dynamic_eta_var.set(param_value.strip())
                                updated_params.append(f"空气粘滞系数: {param_value}")
                        break
            
            if updated_params:
                print(f"动态法参数更新完成，更新了 {len(updated_params)} 个参数")
            else:
                print("没有更新任何参数")
            
            print(f"最终参数存储: {self.dynamic_params_storage}")
                        
        except Exception as e:
            print(f"更新动态法参数时出错: {e}")


    def update_balance_params_from_dict(self, params_dict):
        """从字典更新平衡法参数到界面"""
        try:
            # 更新存储
            for key, value in params_dict.items():
                if key in self.balance_params_storage:
                    self.balance_params_storage[key] = str(value)
            
            # 更新界面显示
            if hasattr(self, 'balance_l_var'):
                self.balance_l_var.set(self.balance_params_storage.get('l', ''))
            if hasattr(self, 'balance_T_var'):
                self.balance_T_var.set(self.balance_params_storage.get('T', ''))
            if hasattr(self, 'balance_P_var'):
                self.balance_P_var.set(self.balance_params_storage.get('P', ''))
            if hasattr(self, 'balance_g_var'):
                self.balance_g_var.set(self.balance_params_storage.get('g', ''))
            if hasattr(self, 'balance_rho1_var'):
                self.balance_rho1_var.set(self.balance_params_storage.get('rho1', ''))
            if hasattr(self, 'balance_rho2_var'):
                self.balance_rho2_var.set(self.balance_params_storage.get('rho2', ''))
            if hasattr(self, 'balance_eta_var'):
                self.balance_eta_var.set(self.balance_params_storage.get('eta', ''))
                
        except Exception as e:
            print(f"更新平衡法参数时出错: {e}")

    def update_dynamic_params_from_dict(self, params_dict):
        """从字典更新动态法参数到界面"""
        try:
            # 更新存储
            for key, value in params_dict.items():
                if key in self.dynamic_params_storage:
                    self.dynamic_params_storage[key] = str(value)
            
            # 更新界面显示
            if hasattr(self, 'dynamic_l_var'):
                self.dynamic_l_var.set(self.dynamic_params_storage.get('l', ''))
            if hasattr(self, 'dynamic_T_var'):
                self.dynamic_T_var.set(self.dynamic_params_storage.get('T', ''))
            if hasattr(self, 'dynamic_P_var'):
                self.dynamic_P_var.set(self.dynamic_params_storage.get('P', ''))
            if hasattr(self, 'dynamic_g_var'):
                self.dynamic_g_var.set(self.dynamic_params_storage.get('g', ''))
            if hasattr(self, 'dynamic_rho1_var'):
                self.dynamic_rho1_var.set(self.dynamic_params_storage.get('rho1', ''))
            if hasattr(self, 'dynamic_rho2_var'):
                self.dynamic_rho2_var.set(self.dynamic_params_storage.get('rho2', ''))
            if hasattr(self, 'dynamic_eta_var'):
                self.dynamic_eta_var.set(self.dynamic_params_storage.get('eta', ''))
                
        except Exception as e:
            print(f"更新动态法参数时出错: {e}")

    def calculate_dynamic_data(self):
        """计算动态法数据"""
        # 清空图表
        self.draw_dynamic_chart_axes()
        # 1. 更新自动计算的参数
        try:
            T = float(self.dynamic_T_var.get())
            P = float(self.dynamic_P_var.get())
            g = float(self.dynamic_g_var.get())
            l_mm = float(self.dynamic_l_var.get())
            d=5.00e-3
            
            # 计算空气密度 ρ2
            R = 287.05
            rho2 = P / (R * T)
            self.dynamic_rho2_var.set(f"{rho2:.3f}")
            
            # 计算空气粘滞系数 η
            T0 = 273.15
            eta0 = 1.716e-5
            S = 110.4
            
            eta = eta0 * (T/T0)**1.5 * (T0 + S) / (T + S)
            eta_display = eta / 1e-5
            self.dynamic_eta_var.set(f"{eta_display:.3f}×10⁻⁵")
            
            # 油滴密度 ρ1 从输入框获取，不再自动计算
            rho1_str = self.dynamic_rho1_var.get()
            if not rho1_str.strip():
                rho1 = 981.0  # 默认值
                self.dynamic_rho1_var.set(f"{rho1}")
            else:
                rho1 = float(rho1_str)
            
            print(f"DEBUG 动态法: 油滴密度 rho1 = {rho1} (手动输入)")
            
            # 运动距离转换为米
            l = l_mm / 1000
            
        except ValueError as e:
            print(f"ERROR: 动态法参数数值转换错误: {e}")
            messagebox.showerror("错误", f"参数输入有误，请检查数字格式\n错误: {e}")
            return
        
        # 2. 计算表格中每一行的数据
        for item in self.dynamic_tree.get_children():
            values = list(self.dynamic_tree.item(item, 'values'))
            
            # 只计算有基本数据的行（需要电压和至少一个时间）
            if len(values) >= 9:
                try:
                    # 检查是否有足够的数据
                    has_rise_data = all(values[1:4])  # te1, te2, te3都有数据
                    has_fall_data = all(values[5:8])  # tg1, tg2, tg3都有数据
                    has_voltage = bool(values[0].strip())
                    
                    if not has_voltage:
                        continue
                    
                    # 提取电压 - 电压取整数
                    U = int(float(values[0]))  # 电压取整
                    
                    # 计算上升相关数据
                    if has_rise_data:
                        te1 = round(float(values[1]), 2)  # 时间保留2位小数
                        te2 = round(float(values[2]), 2)
                        te3 = round(float(values[3]), 2)
                        te_avg = round((te1 + te2 + te3) / 3, 2)  # 平均时间保留2位小数
                        values[4] = f"{te_avg:.2f}"
                        
                        # 计算上升速度 ve (×10⁻⁵m/s)
                        ve = l / te_avg  # m/s
                        ve_show = ve * 1e5
                        values[9] = f"{ve_show:.2f}"  # 2位小数
                    
                    # 计算下落相关数据
                    if has_fall_data:
                        tg1 = round(float(values[5]), 2)  # 时间保留2位小数
                        tg2 = round(float(values[6]), 2)
                        tg3 = round(float(values[7]), 2)
                        tg_avg = round((tg1 + tg2 + tg3) / 3, 2)  # 平均时间保留2位小数
                        values[8] = f"{tg_avg:.2f}"
                        
                        # 计算下落速度 vg (×10⁻⁵m/s)
                        vg = l / tg_avg  # m/s
                        vg_show = vg * 1e5
                        values[10] = f"{vg_show:.2f}"  # 2位小数
                    else:
                        values[8] = ""
                        values[10] = ""
                    
                    # 如果既有上升数据又有下落数据，计算油滴半径和电荷
                    if has_rise_data and has_fall_data:
                        # 修正：从显示值中正确提取eta的数值
                        eta_str = self.dynamic_eta_var.get()
                        # 移除×10⁻⁵部分并转换为浮点数
                        eta_val = float(eta_str.replace('×10⁻⁵', '')) * 1e-5
                        
                        # 计算油滴半径 a (动态法公式)
                        # 使用下落速度计算半径
                        a = math.sqrt((9 * eta_val * vg) / (2 * (rho1 - rho2) * g))
                        a_show = a * 1e7  # 转换为×10⁻⁷m
                        values[11] = f"{a_show:.2f}"
                        
                        b = 0.00823  # 默认修正常数 N/m
                        p = float(self.dynamic_P_var.get())  # 大气压强 Pa
                        
                        # 计算修正因子
                        correction = 1 + b / (p * a)
                        
                        # 计算上升时间和下落时间
                        te_avg = float(values[4]) if values[4] else 0
                        tg_avg = float(values[8]) if values[8] else 0
                        
                        if te_avg > 0 and tg_avg > 0:
                            # 标准动态法公式
                            part1 = 18 * math.pi / math.sqrt(2 * rho1 * g)
                            part2 = math.pow((eta_val * l) / correction, 1.5)
                            part3 = (1/te_avg + 1/tg_avg) * math.sqrt(1/tg_avg)
                            
                            q = part1 * d * part2 * part3 / U
                            q_show = q / 1e-19  # 转换为×10⁻¹⁹
                            values[12] = f"{q_show:.3f}"
                            
                            # 计算带电子数 n
                            e = 1.602e-19  # 元电荷
                            n = round(q / e)
                            values[13] = f"{n}"
                        else:
                            values[12] = ""
                            values[13] = ""
                    else:
                        # 缺少数据，清空计算结果
                        values[11:14] = ["", "", ""]
                    
                    # 更新行数据
                    self.dynamic_tree.item(item, values=values)
                    
                except (ValueError, ZeroDivisionError) as e:
                    print(f"动态法计算错误: {e}")
                    print(f"错误行数据: {values}")
                    # 如果计算错误，清空计算结果
                    if len(values) >= 14:
                        values[4] = ""  # 平均上升时间
                        values[8] = ""  # 平均下落时间
                        values[9:14] = ["", "", "", "", ""]  # 其他计算结果
                    self.dynamic_tree.item(item, values=values)
        
        # 在计算完成后保存数据
        self.save_dynamic_table_data()
        # 3. 绘制图表
        self.dynamic_chart_canvas.update_idletasks()
        self.plot_dynamic_chart()

    def draw_dynamic_chart_axes(self, n_min=None, n_max=None, q_min=None, q_max=None):
        """绘制动态法图表坐标轴 - 支持动态范围"""
        canvas = self.dynamic_chart_canvas
        canvas.delete("all")
        
        # 获取画布实际尺寸
        width = canvas.winfo_width()
        height = canvas.winfo_height()
        
        # 如果画布尺寸太小，使用更大的默认尺寸
        if width < 300 or height < 200:
            # 获取父容器的尺寸作为参考
            parent = canvas.master
            if parent:
                parent_width = parent.winfo_width()
                parent_height = parent.winfo_height()
                if parent_width > 100 and parent_height > 100:
                    width = parent_width - 40
                    height = parent_height - 40
                else:
                    width = 500
                    height = 300
            else:
                width = 500
                height = 300
            
            canvas.config(width=width, height=height)
        
        # 边距设置
        margin_top = 40
        margin_bottom = 70
        margin_left = 60
        margin_right = 20
        
        graph_width = width - margin_left - margin_right
        graph_height = height - margin_top - margin_bottom
        
        # 绘制坐标轴
        canvas.create_line(margin_left, margin_top, 
                        margin_left, margin_top + graph_height, 
                        width=3, fill='black')
        canvas.create_line(margin_left, margin_top + graph_height, 
                        margin_left + graph_width, margin_top + graph_height, 
                        width=3, fill='black')
        
        # 坐标轴标签
        font_size = 10
        
        canvas.create_text(margin_left - 40, margin_top + graph_height/2, 
                        text="q/×10⁻¹⁹C", 
                        angle=90, 
                        font=('Arial', font_size, 'bold'),
                        fill='darkblue')
        
        canvas.create_text(margin_left + graph_width/2, margin_top + graph_height + 40, 
                        text="电子数 n", 
                        font=('Arial', font_size, 'bold'),
                        fill='darkblue')
        
        canvas.create_text(width//2, 20, 
                        text="电荷-电子数关系图", 
                        font=('Arial', 12, 'bold'),
                        fill='darkred')
        
        # 设置坐标轴范围
        n_min_display = n_min if n_min is not None else 0
        n_max_display = n_max if n_max is not None else 10
        q_min_display = q_min if q_min is not None else 0
        q_max_display = q_max if q_max is not None else 20
        
        if n_max_display <= n_min_display:
            n_max_display = n_min_display + 1
        if q_max_display <= q_min_display:
            q_max_display = q_min_display + 1
        
        # 存储坐标信息
        self.dynamic_chart_info = {
            'margin_top': margin_top,
            'margin_left': margin_left,
            'graph_width': graph_width,
            'graph_height': graph_height,
            'n_min': n_min_display,
            'n_max': n_max_display,
            'q_min': q_min_display,
            'q_max': q_max_display
        }
        
        # 绘制Y轴刻度 - 固定5个刻度，保留小数点后一位
        y_ticks = 5
        for i in range(y_ticks + 1):
            y = margin_top + (graph_height / y_ticks) * (y_ticks - i)
            canvas.create_line(margin_left - 8, y, margin_left, y, width=2, fill='black')
            
            # 刻度标签 - 保留小数点后一位
            value = q_min_display + (q_max_display - q_min_display) * (i / y_ticks)
            label = f"{value:.1f}"  # 保留一位小数
                
            canvas.create_text(margin_left - 12, y, 
                            text=label, 
                            anchor='e', 
                            font=('Arial', font_size-1),
                            fill='darkgreen')
        
        # 绘制X轴刻度 - 显示合适的整数刻度
        tick_range = n_max_display - n_min_display
        
        # 确定刻度间隔
        if tick_range <= 2:
            tick_interval = 1
        elif tick_range <= 5:
            tick_interval = 1
        elif tick_range <= 10:
            tick_interval = 2
        elif tick_range <= 15:
            tick_interval = 3
        elif tick_range <= 20:
            tick_interval = 4
        elif tick_range <= 30:
            tick_interval = 5
        else:
            tick_interval = 10
        
        # 计算第一个刻度位置
        if n_min_display >= 0:
            first_tick = math.ceil(n_min_display / tick_interval) * tick_interval
        else:
            first_tick = math.floor(n_min_display / tick_interval) * tick_interval
        
        # 计算最后一个刻度位置
        last_tick = math.floor(n_max_display / tick_interval) * tick_interval
        
        # 生成刻度值
        tick_values = []
        current = first_tick
        while current <= last_tick:
            if n_min_display <= current <= n_max_display:
                tick_values.append(current)
            current += tick_interval
        
        # 如果刻度太少，确保至少有2个刻度
        if len(tick_values) < 2:
            if n_min_display >= 0:
                tick_values = [math.floor(n_min_display), math.ceil(n_max_display)]
            else:
                tick_values = [0, math.ceil(n_max_display)]
        
        # 绘制X轴刻度
        for value in tick_values:
            x = margin_left + ((value - n_min_display) / (n_max_display - n_min_display)) * graph_width
            
            canvas.create_line(x, margin_top + graph_height, 
                            x, margin_top + graph_height + 8, 
                            width=2, fill='black')
            
            canvas.create_text(x, margin_top + graph_height + 15, 
                            text=f"{int(round(value))}", 
                            anchor='n', 
                            font=('Arial', font_size-1),
                            fill='darkgreen')
        
        # 添加网格线
        grid_color = '#E0E0E0'
        grid_dash = (3, 2)
        
        # Y轴网格线
        for i in range(1, y_ticks):
            y = margin_top + (graph_height / y_ticks) * (y_ticks - i)
            canvas.create_line(margin_left, y, margin_left + graph_width, y, 
                            fill=grid_color, width=1, dash=grid_dash)
        
        # X轴网格线
        for value in tick_values:
            x = margin_left + ((value - n_min_display) / (n_max_display - n_min_display)) * graph_width
            canvas.create_line(x, margin_top, x, margin_top + graph_height, 
                            fill=grid_color, width=1, dash=grid_dash)
        
        # 添加坐标轴箭头
        arrow_size = 10
        canvas.create_polygon(
            margin_left - arrow_size//2, margin_top,
            margin_left + arrow_size//2, margin_top,
            margin_left, margin_top - arrow_size,
            fill='black'
        )
        
        canvas.create_polygon(
            margin_left + graph_width, margin_top + graph_height - arrow_size//2,
            margin_left + graph_width, margin_top + graph_height + arrow_size//2,
            margin_left + graph_width + arrow_size, margin_top + graph_height,
            fill='black'
        )

    # 修改 plot_dynamic_chart 方法
    def plot_dynamic_chart(self):
        """绘制动态法图表数据点"""
        canvas = self.dynamic_chart_canvas
        
        # 收集数据
        n_values = []
        q_values = []
        
        for item in self.dynamic_tree.get_children():
            values = self.dynamic_tree.item(item, 'values')
            if len(values) >= 14 and values[12] and values[13]:
                try:
                    q = float(values[12])
                    n = int(values[13])
                    n_values.append(n)
                    q_values.append(q)
                except (ValueError, TypeError):
                    continue
        
        if not n_values:
            self.draw_dynamic_chart_axes()
            canvas.create_text(canvas.winfo_width()//2, canvas.winfo_height()//2, 
                            text="暂无数据，请先计算数据", 
                            font=('Arial', 10), fill='gray')
            return
        
        # 计算数据范围
        n_min = min(n_values) if n_values else 0
        n_max = max(n_values) if n_values else 10
        q_min = min(q_values) if q_values else 0
        q_max = max(q_values) if q_values else 20
        
        n_range = max(1, n_max - n_min)
        q_range = max(1, q_max - q_min)
        
        n_min_display = max(0, n_min - n_range * 0.1)
        n_max_display = n_max + n_range * 0.1
        q_min_display = max(0, q_min - q_range * 0.1)
        q_max_display = q_max + q_range * 0.1
        
        n_min_display = max(0, n_min_display)
        q_min_display = max(0, q_min_display)
        
        if n_max_display - n_min_display < 2:
            n_max_display = n_min_display + 2
        
        if q_max_display - q_min_display < 2:
            q_max_display = q_min_display + 2
        
        # 绘制坐标轴
        self.draw_dynamic_chart_axes(n_min_display, n_max_display, q_min_display, q_max_display)
        
        # 获取坐标信息
        if not hasattr(self, 'dynamic_chart_info'):
            return
        
        info = self.dynamic_chart_info
        margin_top = info['margin_top']
        margin_left = info['margin_left']
        graph_width = info['graph_width']
        graph_height = info['graph_height']
        
        # 绘制数据点
        for n, q in zip(n_values, q_values):
            x = margin_left + ((n - n_min_display) / (n_max_display - n_min_display)) * graph_width
            y = margin_top + ((q_max_display - q) / (q_max_display - q_min_display)) * graph_height
            
            canvas.create_oval(x-4, y-4, x+4, y+4, 
                            fill='blue', outline='darkblue', width=2)
            
            label = f"({n}, {q:.3f})"
            canvas.create_text(x, y-15, text=label, 
                            font=('Arial', 8, 'bold'), fill='red')
        
        # 绘制拟合直线
        if len(n_values) >= 2:
            try:
                n_array = np.array(n_values)
                q_array = np.array(q_values)
                
                slope, intercept = np.polyfit(n_array, q_array, 1)
                
                x1 = margin_left
                x2 = margin_left + graph_width
                
                y1 = margin_top + ((q_max_display - (slope * n_min_display + intercept)) / 
                                (q_max_display - q_min_display)) * graph_height
                y2 = margin_top + ((q_max_display - (slope * n_max_display + intercept)) / 
                                (q_max_display - q_min_display)) * graph_height
                
                canvas.create_line(x1, y1, x2, y2, 
                                fill='red', width=2, dash=(4, 2))
                
                # 在图表左上角显示拟合公式和测量结果
                text_start_x = margin_left + 10
                text_start_y = margin_top + 10
                
                # 显示拟合公式
                formula_text = f"q = {slope:.3f}n + {intercept:.3f}"
                canvas.create_text(text_start_x, text_start_y, 
                                text=formula_text, font=('Arial', 9, 'bold'), 
                                fill='darkred', anchor='nw')
                
                e_standard = 1.602
                e_measured = abs(slope)
                
                if e_measured > 0:
                    error_percent = abs((e_measured - e_standard) / e_standard) * 100
                else:
                    error_percent = float('inf')
                
                # 显示标准值和测量值
                standard_text = f"e₀: {e_standard}"
                measure_text = f"e: {e_measured:.3f}"
                
                canvas.create_text(text_start_x, text_start_y + 15, 
                                text=standard_text, font=('Arial', 9), 
                                fill='darkgreen', anchor='nw')
                
                canvas.create_text(text_start_x, text_start_y + 30, 
                                text=measure_text, font=('Arial', 9, 'bold'), 
                                fill='blue', anchor='nw')
                
                # 显示误差
                error_color = 'red' if error_percent > 5 else 'orange' if error_percent > 2 else 'green'
                error_text = f"δ: {error_percent:.2f}%"
                canvas.create_text(text_start_x, text_start_y + 45, 
                                text=error_text, font=('Arial', 9, 'bold'), 
                                fill=error_color, anchor='nw')
                
                # 添加图例说明
                legend_x = text_start_x
                legend_y = text_start_y + 65
                canvas.create_text(legend_x, legend_y, 
                                text="(e₀:标准值, e:测量值, δ:误差)", 
                                font=('Arial', 8), 
                                fill='gray', anchor='nw')
                
                self.dynamic_e_experimental = e_measured
                self.dynamic_e_error = error_percent
                
            except Exception as e:
                print(f"绘制拟合直线时出错: {e}")

    def delete_dynamic_selected_row(self):
        """删除动态法选中行"""
        selection = self.dynamic_tree.selection()
        if not selection:
            messagebox.showinfo("提示", "请先选择要删除的行")
            return
        
        if messagebox.askyesno("确认", "确定要删除选中的行吗？"):
            for item in selection:
                self.dynamic_tree.delete(item)

        # 保存数据
        self.save_dynamic_table_data()

    def clear_dynamic_data(self):
        """清空动态法数据"""
        if messagebox.askyesno("确认", "确定要清空所有数据吗？"):
            # 清空表格数据
            for item in self.dynamic_tree.get_children():
                self.dynamic_tree.delete(item)
            
            # 添加初始的3行空行
            for _ in range(3):
                self.dynamic_tree.insert("", "end", values=[""] * len(self.dynamic_tree['columns']))
            
            # 清空图表
            self.draw_dynamic_chart_axes()
            
            # 保存数据
            self.save_dynamic_table_data()

    def export_dynamic_data(self):
        """导出动态法数据到Excel（参数和表格分开sheet）"""
        try:
            # 弹出保存文件对话框
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                title="保存动态法数据",
                initialfile="动态法实验数据.xlsx"
            )
            
            if not file_path:
                return
            
            # 收集参数数据
            params_data = []
            
            # 参数表头
            params_data.append(["=== 动态法实验参数 ==="])
            params_data.append(["参数名称", "参数值", "单位"])
            
            # 添加固定参数
            params_data.append(["实验方法", "动态法", ""])
            params_data.append(["极板间距", "5.00", "mm"])
            
            # 添加可调参数
            if hasattr(self, 'dynamic_l_var'):
                l_val = self.dynamic_l_var.get()
                params_data.append(["运动距离", l_val, "mm"])
            
            if hasattr(self, 'dynamic_T_var'):
                T_val = self.dynamic_T_var.get()
                params_data.append(["室温", T_val, "K"])
            
            if hasattr(self, 'dynamic_P_var'):
                P_val = self.dynamic_P_var.get()
                params_data.append(["大气压强", P_val, "Pa"])
            
            if hasattr(self, 'dynamic_g_var'):
                g_val = self.dynamic_g_var.get()
                params_data.append(["重力加速度", g_val, "m/s²"])
            
            if hasattr(self, 'dynamic_rho1_var'):
                rho1_val = self.dynamic_rho1_var.get()
                params_data.append(["油滴密度", rho1_val, "kg/m³"])
            
            if hasattr(self, 'dynamic_rho2_var'):
                rho2_val = self.dynamic_rho2_var.get()
                params_data.append(["空气密度", rho2_val, "kg/m³"])
            
            if hasattr(self, 'dynamic_eta_var'):
                eta_val = self.dynamic_eta_var.get()
                params_data.append(["空气粘滞系数", eta_val, "kg/(m·s)"])
            
            # 分析结果
            params_data.append([])  # 空行
            params_data.append(["=== 基本电荷分析结果 ==="])
            params_data.append(["项目", "数值", "单位"])
            params_data.append(["基本电荷理论值", "1.602", "×10⁻¹⁹ C"])
            
            if self.dynamic_e_experimental is not None:
                params_data.append(["基本电荷实验值", f"{self.dynamic_e_experimental:.3f}", "×10⁻¹⁹ C"])
            else:
                params_data.append(["基本电荷实验值", "未计算", ""])
            
            if self.dynamic_e_error is not None:
                params_data.append(["百分误差", f"{self.dynamic_e_error:.2f}", "%"])
            else:
                params_data.append(["百分误差", "未计算", ""])
            
            # 创建参数DataFrame
            params_df = pd.DataFrame(params_data)
            
            # 收集表格数据
            table_data = []
            
            # 表格标题
            table_data.append(["=== 动态法数据记录表 ==="])
            
            # 表头
            headers = [self.dynamic_tree.heading(col)['text'] for col in self.dynamic_tree['columns']]
            table_data.append(headers)
            
            # 表格数据
            has_data = False
            for item in self.dynamic_tree.get_children():
                values = self.dynamic_tree.item(item, 'values')
                row_values = []
                for v in values:
                    if v is None or str(v).strip() == "":
                        row_values.append("")
                    else:
                        row_values.append(str(v).strip())
                
                if any(row_values):  # 至少有一个非空值
                    table_data.append(row_values)
                    has_data = True
            
            if not has_data:
                table_data.append(["暂无数据"])
            
            # 创建表格DataFrame
            table_df = pd.DataFrame(table_data)
            
            # 保存到Excel（两个sheet）
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 保存参数到第一个sheet
                params_df.to_excel(writer, sheet_name='实验参数', index=False, header=False)
                
                # 调整参数sheet列宽
                worksheet_params = writer.sheets['实验参数']
                worksheet_params.column_dimensions['A'].width = 25
                worksheet_params.column_dimensions['B'].width = 20
                worksheet_params.column_dimensions['C'].width = 15
                
                # 保存表格到第二个sheet
                table_df.to_excel(writer, sheet_name='数据记录表', index=False, header=False)
                
                # 调整表格sheet列宽
                worksheet_table = writer.sheets['数据记录表']
                for i, col in enumerate(table_df.columns):
                    col_letter = get_column_letter(i + 1)
                    worksheet_table.column_dimensions[col_letter].width = 15
            
            # 显示导出成功消息
            e_exp_display = f"{self.dynamic_e_experimental:.3f}" if self.dynamic_e_experimental is not None else "未计算"
            e_err_display = f"{self.dynamic_e_error:.2f}%" if self.dynamic_e_error is not None else "未计算"
            
            messagebox.showinfo("导出成功", 
                            f"动态法数据已成功导出到：\n{file_path}\n\n"
                            f"导出内容：\n"
                            f"1. 实验参数（参数和计算结果）\n"
                            f"2. 数据记录表\n\n"
                            f"基本电荷分析结果：\n"
                            f"- e理论值: 1.602 ×10⁻¹⁹ C\n"
                            f"- e实验值: {e_exp_display} ×10⁻¹⁹ C\n"
                            f"- 百分误差: {e_err_display}\n\n"
                            f"数据记录表: {len(table_data)-2} 行数据")
            
            print(f"动态法数据已导出到: {file_path}")
            
        except Exception as e:
            messagebox.showerror("导出失败", f"导出数据时出错：\n{str(e)}")
            import traceback
            traceback.print_exc()


    def import_dynamic_data(self):
        """从Excel导入动态法数据（参数和表格分开sheet）"""
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
                title="选择要导入的动态法数据文件"
            )
            
            if not file_path:
                return
            
            print(f"开始导入文件: {file_path}")
            
            try:
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names
                print(f"Excel文件中包含的sheet: {sheet_names}")
            except Exception as e:
                messagebox.showerror("错误", f"无法读取Excel文件：\n{str(e)}")
                return
            
            # 1. 导入参数（从"实验参数"sheet）
            params_loaded = False
            if '实验参数' in sheet_names:
                print("找到'实验参数'sheet，开始导入参数...")
                params_loaded = self.load_params_from_sheet(file_path, '实验参数', 'dynamic')
            else:
                print("没有找到'实验参数'sheet，跳过参数导入")
            
            # 2. 导入表格数据（从"数据记录表"sheet）
            data_loaded = False
            imported_rows = 0
            
            if '数据记录表' in sheet_names:
                print("找到'数据记录表'sheet，开始导入表格数据...")
                imported_rows = self.load_table_data_from_sheet(file_path, '数据记录表', 'dynamic')
                if imported_rows > 0:
                    data_loaded = True
            else:
                print("没有找到'数据记录表'sheet，跳过表格数据导入")
            
            # 3. 显示导入结果
            if data_loaded:
                # 添加一个空行用于继续输入
                self.dynamic_tree.insert("", "end", values=[""] * len(self.dynamic_tree['columns']))
                
                # 重新计算数据
                print("重新计算数据...")
                self.root.after(500, self.calculate_dynamic_data)
            
            # 显示结果
            message_text = f"动态法数据导入完成：\n\n"
            message_text += f"参数导入: {'成功' if params_loaded else '失败'}\n"
            message_text += f"表格导入: {'成功' if data_loaded else '失败'}\n"
            if data_loaded:
                message_text += f"导入行数: {imported_rows}行\n"
            
            messagebox.showinfo("导入完成", message_text)
            
        except Exception as e:
            messagebox.showerror("导入失败", f"导入数据时出错：\n{str(e)}")
            import traceback
            traceback.print_exc()

    def load_params_from_sheet(self, file_path, sheet_name, method_type):
        """从指定sheet加载参数（通用方法）"""
        try:
            print(f"从sheet '{sheet_name}' 加载{method_type}参数...")
            
            # 读取sheet
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            data = df.values.tolist()
            
            if not data:
                print(f"sheet '{sheet_name}' 没有数据")
                return False
            
            # 提取参数
            param_dict = {}
            
            for i, row in enumerate(data):
                if i >= 50:  # 只检查前50行
                    break
                
                if isinstance(row, list) and len(row) >= 2:
                    param_name = str(row[0]) if row[0] is not None and not pd.isna(row[0]) else ""
                    param_value = str(row[1]) if len(row) > 1 and row[1] is not None and not pd.isna(row[1]) else ""
                    
                    # 跳过标题行
                    if any(marker in param_name for marker in ['===', '---', '###', '实验参数', '基本电荷']):
                        continue
                    
                    # 清理参数名并提取数值
                    if param_name.strip() and param_value.strip():
                        # 移除单位符号
                        clean_name = param_name.strip()
                        for char in ['=', ':', '(', ')', '[', ']', '{', '}', '单位', 'Unit']:
                            clean_name = clean_name.replace(char, "")
                        clean_name = clean_name.strip()
                        
                        if clean_name:
                            # 提取数值部分
                            value_str = param_value.strip()
                            # 移除单位
                            for unit in ['mm', 'K', 'Pa', 'm/s²', 'kg/m³', 'kg/(m·s)', '×10⁻⁹', '×10⁻¹⁹', 'C', '%']:
                                value_str = value_str.replace(unit, '')
                            
                            param_dict[clean_name] = value_str.strip()
                            print(f"提取参数: {clean_name} = {value_str}")
            
            print(f"从sheet '{sheet_name}' 提取到 {len(param_dict)} 个参数")
            
            if param_dict:
                if method_type == 'dynamic':
                    self.update_dynamic_params_from_dict_extended(param_dict)
                else:  # balance
                    self.update_balance_params_from_dict_extended(param_dict)
                return True
            else:
                print("没有提取到参数")
                return False
                
        except Exception as e:
            print(f"从sheet加载参数时出错: {e}")
            return False


    def load_table_data_from_sheet(self, file_path, sheet_name, method_type):
        """从指定sheet加载表格数据（通用方法）"""
        try:
            print(f"从sheet '{sheet_name}' 加载{method_type}表格数据...")
            
            # 读取sheet
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            data = df.values.tolist()
            
            if not data:
                print(f"sheet '{sheet_name}' 没有数据")
                return 0
            
            # 清空现有数据
            if method_type == 'dynamic':
                for item in self.dynamic_tree.get_children():
                    self.dynamic_tree.delete(item)
                column_count = len(self.dynamic_tree['columns'])
                # 动态法表头关键词
                header_keywords = ['电压U/V', '上升时间第一次te1/s', '下落时间第一次tg1/s']
            else:  # balance
                for item in self.balance_tree.get_children():
                    self.balance_tree.delete(item)
                column_count = len(self.balance_tree['columns'])
                # 平衡法表头关键词
                header_keywords = ['电压U/V', '下落时间第一次tg1/s', '平均下落时间tg/s']
            
            imported_rows = 0
            start_row = -1
            
            # 查找表格开始位置 - 更准确的表头识别
            print("查找表格开始位置...")
            for i, row in enumerate(data):
                if i >= 50:  # 最多检查50行
                    break
                
                if not isinstance(row, list):
                    continue
                
                # 检查是否是表头行
                is_header = False
                header_cells = []
                
                # 收集非空单元格
                for cell in row:
                    if cell is not None and not pd.isna(cell):
                        header_cells.append(str(cell).strip())
                
                # 检查是否包含表头关键词
                if len(header_cells) >= 3:  # 至少3个非空单元格
                    for keyword in header_keywords:
                        for cell in header_cells:
                            if keyword in cell:
                                is_header = True
                                print(f"找到表头行 {i}: {header_cells[:3]}...")
                                break
                        if is_header:
                            break
                
                if is_header:
                    start_row = i + 1  # 数据从表头下一行开始
                    break
            
            if start_row == -1:
                print("没有找到表头，尝试查找第一个数据行")
                # 如果没有找到表头，查找第一个包含数字的行
                for i, row in enumerate(data):
                    if i >= 30:
                        break
                    
                    if isinstance(row, list) and len(row) >= 3:
                        # 检查是否有数字数据
                        has_numbers = 0
                        for j in range(min(3, len(row))):
                            cell = row[j]
                            if cell is not None and not pd.isna(cell):
                                try:
                                    float(str(cell))
                                    has_numbers += 1
                                except:
                                    pass
                        
                        if has_numbers >= 2:  # 至少有2个数字
                            start_row = i
                            print(f"找到数据开始行 {i}")
                            break
            
            if start_row == -1:
                print("没有找到有效的数据行")
                return 0
            
            print(f"数据从第 {start_row} 行开始")
            
            # 读取数据行
            for i in range(start_row, len(data)):
                if i >= len(data):
                    break
                
                row = data[i]
                if not isinstance(row, list):
                    continue
                
                # 检查是否到达表格结束（空行或新标题）
                if len(row) > 0:
                    first_cell = str(row[0]) if row[0] is not None and not pd.isna(row[0]) else ""
                    # 如果遇到标题或说明行，停止
                    if any(marker in first_cell for marker in ['===', '---', '###', '实验参数', '基本电荷', '电荷-电子数']):
                        print(f"遇到非数据行，停止在第 {i} 行: {first_cell}")
                        break
                
                # 检查是否是空行（所有单元格都为空）
                is_empty = True
                for cell in row:
                    if cell is not None and not pd.isna(cell) and str(cell).strip():
                        is_empty = False
                        break
                
                if is_empty:
                    continue
                
                # 提取数据
                row_values = []
                has_data = False
                
                # 只读取前column_count列
                for j in range(min(column_count, len(row))):
                    cell = row[j]
                    if cell is None or pd.isna(cell):
                        row_values.append("")
                    else:
                        cell_str = str(cell).strip()
                        
                        # 处理数值
                        try:
                            # 尝试转换为数值
                            cell_float = float(cell_str)
                            # 根据数值大小决定小数位数
                            if cell_float == int(cell_float):
                                formatted = str(int(cell_float))
                            elif method_type == 'dynamic' and j >= 9 and j <= 12:  # 速度和半径列
                                formatted = f"{cell_float:.3f}"
                            elif method_type == 'balance' and j >= 5 and j <= 7:  # 速度和半径列
                                formatted = f"{cell_float:.3f}"
                            elif j == 0:  # 电压列
                                formatted = f"{cell_float:.0f}"
                            elif j in [1, 2, 3, 5, 6, 7]:  # 时间列（动态法和平衡法的时间列索引）
                                formatted = f"{cell_float:.2f}"
                            elif j in [4, 8]:  # 平均时间列
                                formatted = f"{cell_float:.2f}"
                            else:
                                formatted = f"{cell_float:.2f}"
                            row_values.append(formatted)
                        except ValueError:
                            # 如果不是数字，直接使用字符串
                            row_values.append(cell_str)
                        
                        if cell_str:
                            has_data = True
                
                # 补充缺失的列
                while len(row_values) < column_count:
                    row_values.append("")
                
                if has_data:
                    if method_type == 'dynamic':
                        self.dynamic_tree.insert("", "end", values=row_values)
                    else:  # balance
                        self.balance_tree.insert("", "end", values=row_values)
                    imported_rows += 1
                    
                    # 显示前几行的数据
                    if imported_rows <= 3:
                        print(f"导入第 {imported_rows} 行: {row_values[:3]}...")
            
            print(f"从sheet '{sheet_name}' 导入 {imported_rows} 行数据")
            return imported_rows
            
        except Exception as e:
            print(f"从sheet加载表格数据时出错: {e}")
            import traceback
            traceback.print_exc()
            return 0

def main():
    root = tk.Tk()
    app = ExperimentApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()  