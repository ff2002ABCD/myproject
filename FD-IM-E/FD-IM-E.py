import sys
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk
import math
import time

# === 强制导入，确保 PyInstaller 打包这些库 ===
# 这些是导出 Excel 功能需要的库
try:
    import openpyxl
    import pandas as pd
    import numpy as np
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils.dataframe import dataframe_to_rows
    print("✅ openpyxl, pandas, numpy 加载成功")
except ImportError as e:
    print(f"⚠️ 导入库失败: {e}")
# === 强制导入结束 ===

def get_resource_path(relative_path):
    """获取资源的绝对路径，支持打包后的环境"""
    try:
        # 打包后的临时文件夹路径
        base_path = sys._MEIPASS
    except Exception:
        # 开发环境的路径
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, relative_path)

class RigidBodyInertiaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("刚体转动惯量实验")
        self.root.geometry("1600x700")
        
        # 创建自定义按钮样式
        style = ttk.Style()
        style.configure("Accent.TButton", foreground="blue", font=("Arial", 9, "bold"))
        
        # 电源状态
        self.power_on = False  # 默认关机状态
    
        # 初始化数据存储
        self.table_data = {
            "测量空载物台的转动惯量": {
                "entries": [],  # 存储每行的item id
                "values": []    # 存储每行的数据值
            },
            "测量圆盘的转动惯量": {
                "entries": [],
                "values": []
            },
            "测量圆环的转动惯量": {
                "entries": [],
                "values": []
            },
            "验证平行轴定理": {  # 新增
            "entries": [],
            "values": []
            }
        }
        self.current_tab = "测量空载物台的转动惯量"
        
        self.max_omega = 2.0  # 最大角速度上限 (弧度/s)
        self.max_omega_deg = 0  # 最大角速度 (度/s)
        self.ramp_time = 0  # 加速到上限的时间
        self.ramp_angle = 0  # 加速阶段转过的角度
        self.is_accelerating = False  # 是否在加速阶段

        # self.cylinder_distance = 105  # 圆柱距离中心距离 (mm)
        # self.cylinder_positions = [45, 60, 75, 90, 105]  # 可选距离
        self.current_sample_type = "圆盘"  # 当前样品类型
        self.cylinder1_id = None  # 圆柱1的图形ID
        self.cylinder2_id = None  # 圆柱2的图形ID
        self.sample_label_id = None  # 样品标注矩形ID
        self.sample_text_id = None  # 样品标注文字ID

        # 保存计算结果的变量
        self.saved_j2_value = "待计算"  # 保存 J2 值
        self.saved_j3_value = "待计算"  # 保存 J3 值
        self.saved_error_value = "待计算"     # 保存相对误差
        self.saved_sample_type = None   # 保存当前计算的样品类型
        # 保存不同样品的计算结果（独立存储）
        self.saved_results = {
            "空载物台": {
                "j1": "待计算"
            },
            "圆盘": {
                "j2": "待计算",
                "j3": "待计算",
                "error": "待计算"
            },
            "圆环": {
                "j2": "待计算",
                "j3": "待计算",
                "error": "待计算"
            },
            "平行轴定理": {  # 新增
            "j2": "待计算",  # 圆柱+载物台转动惯量
            "j3": "待计算",  # 圆柱转动惯量
            "error": "待计算"      # 相对误差
            }
        }
        
        # 初始化标签引用（在 init_right_bottom_area 中创建）
        self.j2_label = None
        self.j2_value_label = None
        self.sample_param_label = None
        self.sample_param_value_label = None

        # 保存每个页面的参数设置（独立存储）
        self.saved_params = {
            "测量空载物台的转动惯量": {
                "weight": "5g",
                "radius": "15mm"
            },
            "测量圆盘的转动惯量": {
                "weight": "5g",
                "radius": "15mm"
            },
            "测量圆环的转动惯量": {
                "weight": "5g",
                "radius": "15mm"
            },
            "验证平行轴定理": {  # 新增
            "weight": "5g",
            "radius": "15mm",
            "distance": "45mm"  # 离载物台中心距离
            }
        }
        
        # 保存不同样品的计算结果（独立存储）
        # 实验状态变量
        self.is_running = False
        self.current_angle = 0  # 当前角度（度）
        self.angular_velocity = 0  # 当前角速度
        self.angular_acceleration = -0.0520  # 空载物台角加速度
        self.animation_id = None
        
        # 主机区域交互变量
        self.host_menu_items = ["次数", "开始", "查询", "清空"]
        self.host_selected_index = 0  # 当前高亮项索引
        self.host_is_setting = False  # 是否在次数设定模式
        self.host_setting_temp = 3  # 次数设定临时值
        self.host_target_count = 3  # 目标次数
        self.host_current_count = 0  # 当前已触发次数
        self.host_is_counting = False  # 是否在计时计数状态
        self.host_is_ready = False  # 是否在准备状态
        self.host_start_time = 0  # 开始计时的时间
        self.host_elapsed_time = 0  # 经过的时间
        self.host_query_data = []  # 查询数据列表 [{'count':次数, 'target':设定次数, 'time':
        self.host_query_index = 0  # 查询时显示的索引
        self.host_in_query = False  # 是否在查询界面
        self.host_trigger_count = 0  # 光电门触发计数
        self.host_last_trigger_angle = -1  # 上次触发角度
        
        # 数据曲线图相关变量
        self.chart_data = {}  # 存储所有测量数据
        # 结构: {
        #   "空载物台": {
        #       1: {"time": [t1, t2, ...], "angle": [a1, a2, ...], "count": n},
        #       2: {"time": [t1, t2, ...], "angle": [a1, a2, ...], "count": n},
        #       ...
        #   },
        #   "加砝码": {...},
        #   "加样品": {...},
        #   "加砝码和样品": {...}
        # }
        self.current_measurement_index = {}  # 记录每个实验条件的当前测量序号
        self.current_measurement_data = {}  # 当前正在进行的测量数据
        self.is_measuring = False  # 是否正在测量中
        
        # 滑轮和砝码相关变量
        self.pulley_img = None  # 滑轮图片
        self.weight_img = None  # 砝码图片
        self.pulley_id = None   # 滑轮画布ID
        self.pulley_line_id = None  # 连接线ID
        self.weight_line_id = None  # 砝码连接线ID（新增）
        self.weight_id = None   # 砝码画布ID
        self.weight_y_offset = 0  # 砝码垂直偏移量（动画用）
        self.base_pulley_y = 95  # 滑轮基础Y位置（距离顶部）
        self.base_weight_y = 0   # 砝码基础Y位置

        # 查询相关变量
        self.query_condition = tk.StringVar(value="空载物台")
        self.query_index = tk.IntVar(value=1)

        # 创建主框架
        main_frame = ttk.Frame(root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 左侧区域（实验操作区域）
        left_frame = ttk.LabelFrame(main_frame, text="实验操作区域")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        # 中间区域（包含主机和数据记录）
        middle_frame = ttk.Frame(main_frame)
        middle_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 右上区域（主机区域）
        right_top_frame = ttk.LabelFrame(middle_frame, text="实验仪主机区域")
        right_top_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        # 右下区域（数据记录区域）
        right_bottom_frame = ttk.LabelFrame(middle_frame, text="数据记录区域")
        right_bottom_frame.pack(fill=tk.BOTH, expand=True)

        # 右侧曲线图区域（放在最右边）
        chart_frame = ttk.LabelFrame(main_frame, text="数据曲线图")
        chart_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))

        # 初始化各个区域
        self.init_left_area(left_frame)
        self.init_right_top_area(right_top_frame)
        self.init_right_bottom_area(right_bottom_frame)
        self.init_chart_area(chart_frame)
        
        # 初始化参数显示
        self.update_params_display("测量空载物台的转动惯量")
    
    def init_chart_area(self, parent):
        """初始化数据曲线图区域"""
        # 查询控制区域
        query_frame = ttk.Frame(parent)
        query_frame.pack(fill=tk.X, pady=5, padx=5)
        
        ttk.Label(query_frame, text="实验条件:").pack(side=tk.LEFT, padx=2)
        condition_combo = ttk.Combobox(query_frame, textvariable=self.query_condition,
                                        values=["空载物台", "加砝码", "加样品", "加砝码和样品"],
                                        state="readonly", width=10)
        condition_combo.pack(side=tk.LEFT, padx=2)
        condition_combo.bind("<<ComboboxSelected>>", self.on_query_condition_change)
        
        ttk.Label(query_frame, text="序号:").pack(side=tk.LEFT, padx=2)
        self.query_index_spinbox = tk.Spinbox(query_frame, from_=1, to=99, width=4,
                                            textvariable=self.query_index,
                                            command=self.on_query_index_change)
        self.query_index_spinbox.pack(side=tk.LEFT, padx=2)
        
        ttk.Button(query_frame, text="查询", command=self.query_chart_data).pack(side=tk.LEFT, padx=5)
        
        # 曲线图显示区域（使用Canvas绘制）
        self.chart_canvas = tk.Canvas(parent, bg='white', highlightthickness=1, highlightbackground='#cccccc')
        self.chart_canvas.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 绑定鼠标事件
        self.chart_canvas.bind("<MouseWheel>", self.on_chart_scroll)
        self.chart_canvas.bind("<Motion>", self.on_chart_mouse_move)  # 鼠标移动
        self.chart_canvas.bind("<Leave>", self.on_chart_mouse_leave)  # 鼠标离开
        
        # 曲线图数据
        self.chart_plot_data = None  # 当前显示的曲线数据
        self.chart_x_scale = 1.0
        self.chart_y_scale = 1.0
        self.chart_x_offset = 0
        self.chart_y_offset = 0
        
        # 鼠标悬停提示相关
        self.tooltip_items = []  # 提示框所有元素的ID列表
        self.hover_point_index = -1  # 当前悬停的数据点索引

        # 按钮区域
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, pady=5, padx=5)
        
        ttk.Button(button_frame, text="导出数据", command=self.export_chart_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导入数据", command=self.import_chart_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="清空数据", command=self.clear_all_chart_data).pack(side=tk.LEFT, padx=5)

    def on_chart_mouse_move(self, event):
        """鼠标在曲线图上移动时显示对应点数值"""
        if not self.chart_plot_data:
            return
        
        # 获取画布尺寸
        width = self.chart_canvas.winfo_width()
        height = self.chart_canvas.winfo_height()
        if width < 100 or height < 100:
            return
        
        # 边距（与 draw_chart 保持一致）
        margin_left = 60
        margin_right = 30
        margin_top = 30
        margin_bottom = 40
        
        plot_width = width - margin_left - margin_right
        plot_height = height - margin_top - margin_bottom
        
        # 获取数据
        times = self.chart_plot_data["time"]
        angles = self.chart_plot_data["angle"]
        
        if len(times) < 2:
            return
        
        min_time = 0
        max_time = max(times)
        min_angle = 0
        max_angle = max(angles)
        
        time_range = max_time - min_time if max_time > min_time else 1
        angle_range = max_angle - min_angle if max_angle > min_angle else 1
        
        # 坐标转换函数
        def to_canvas_x(t):
            return margin_left + (t - min_time) / time_range * plot_width
        
        def to_canvas_y(a):
            return margin_top + plot_height - (a - min_angle) / angle_range * plot_height
        
        # 查找最近的数据点
        closest_index = -1
        min_dist = 20  # 像素阈值
        
        for i, (t, a) in enumerate(zip(times, angles)):
            cx = to_canvas_x(t)
            cy = to_canvas_y(a)
            dist = ((event.x - cx) ** 2 + (event.y - cy) ** 2) ** 0.5
            if dist < min_dist:
                min_dist = dist
                closest_index = i
        
        # 如果找到了最近的点且距离在阈值内
        if closest_index >= 0 and min_dist < 20:
            # 如果悬停的点变化了，更新提示
            if closest_index != self.hover_point_index:
                self.hover_point_index = closest_index
                self.show_tooltip(event.x, event.y, times[closest_index], angles[closest_index], closest_index + 1)
        else:
            # 没有悬停在点上，清除提示
            if self.hover_point_index != -1:
                self.hover_point_index = -1
                self.hide_tooltip()

    def on_chart_mouse_leave(self, event):
        """鼠标离开曲线图时隐藏提示"""
        self.hover_point_index = -1
        self.hide_tooltip()

    def show_tooltip(self, x, y, time_val, angle_val, count):
        """显示数据点提示框"""
        # 清除旧的提示框
        self.hide_tooltip()
        
        # 创建提示框背景
        text = f"第{count}次触发\n时间: {time_val:.3f}s\n角度: {angle_val:.3f}rad"
        
        # 创建提示框（使用矩形+文本）
        padding = 8
        font_size = 9
        
        # 估算文本尺寸
        lines = text.split('\n')
        max_len = max(len(line) for line in lines)
        text_width = max_len * 6 + padding * 2
        text_height = len(lines) * 16 + padding * 2
        
        # 调整位置，确保不超出画布
        canvas_width = self.chart_canvas.winfo_width()
        canvas_height = self.chart_canvas.winfo_height()
        
        tooltip_x = x + 15
        tooltip_y = y - 10
        
        if tooltip_x + text_width > canvas_width:
            tooltip_x = x - text_width - 15
        if tooltip_y + text_height > canvas_height:
            tooltip_y = canvas_height - text_height - 5
        if tooltip_y < 0:
            tooltip_y = 5
        
        # 使用列表存储提示框的所有元素
        self.tooltip_items = []
        
        # 绘制背景矩形（带半透明效果）
        rect_id = self.chart_canvas.create_rectangle(
            tooltip_x, tooltip_y, tooltip_x + text_width, tooltip_y + text_height,
            fill="#ffffcc", outline="#666666", width=1
        )
        self.tooltip_items.append(rect_id)
        
        # 绘制文本
        for i, line in enumerate(lines):
            text_id = self.chart_canvas.create_text(
                tooltip_x + padding, tooltip_y + padding + i * 16,
                text=line, anchor=tk.NW, font=("Arial", font_size), fill="#333333"
            )
            self.tooltip_items.append(text_id)
        
        # 高亮对应的数据点
        if self.chart_plot_data:
            times = self.chart_plot_data["time"]
            angles = self.chart_plot_data["angle"]
            if self.hover_point_index < len(times):
                # 获取画布尺寸
                width = self.chart_canvas.winfo_width()
                height = self.chart_canvas.winfo_height()
                if width < 100 or height < 100:
                    return
                
                margin_left = 60
                margin_right = 30
                margin_top = 30
                margin_bottom = 40
                
                plot_width = width - margin_left - margin_right
                plot_height = height - margin_top - margin_bottom
                
                min_time = 0
                max_time = max(times)
                min_angle = 0
                max_angle = max(angles)
                
                time_range = max_time - min_time if max_time > min_time else 1
                angle_range = max_angle - min_angle if max_angle > min_angle else 1
                
                def to_canvas_x(t):
                    return margin_left + (t - min_time) / time_range * plot_width
                
                def to_canvas_y(a):
                    return margin_top + plot_height - (a - min_angle) / angle_range * plot_height
                
                px = to_canvas_x(times[self.hover_point_index])
                py = to_canvas_y(angles[self.hover_point_index])
                
                # 绘制高亮圆点（覆盖在原有数据点上）
                highlight_id = self.chart_canvas.create_oval(
                    px-6, py-6, px+6, py+6,
                    fill="red", outline="red", width=2,
                    tags="highlight_point"
                )
                self.tooltip_items.append(highlight_id)

    def hide_tooltip(self):
        """隐藏提示框"""
        if hasattr(self, 'tooltip_items') and self.tooltip_items:
            for item in self.tooltip_items:
                self.chart_canvas.delete(item)
            self.tooltip_items = []
        
        # 删除高亮点（兼容旧方式）
        self.chart_canvas.delete("highlight_point")

   
    def on_query_condition_change(self, event=None):
        """查询条件变化时更新序号范围"""
        condition = self.query_condition.get()
        if condition in self.current_measurement_index:
            max_index = self.current_measurement_index[condition]
            self.query_index_spinbox.config(to=max(1, max_index))
            if self.query_index.get() > max_index:
                self.query_index.set(max(1, max_index))
        else:
            # 如果该条件没有数据，设置to为1
            self.query_index_spinbox.config(to=1)
            self.query_index.set(1)

    def on_query_index_change(self, event=None):
        """序号变化时自动查询"""
        # 确保序号在有效范围内
        condition = self.query_condition.get()
        if condition in self.current_measurement_index:
            max_index = self.current_measurement_index[condition]
            if self.query_index.get() > max_index:
                self.query_index.set(max(1, max_index))
        self.query_chart_data()

    def query_chart_data(self):
        """查询并显示曲线数据"""
        condition = self.query_condition.get()
        index = self.query_index.get()
        
        # 更新序号范围
        if condition in self.current_measurement_index:
            max_index = self.current_measurement_index[condition]
            self.query_index_spinbox.config(to=max(1, max_index))
        else:
            self.query_index_spinbox.config(to=1)
        
        if condition not in self.chart_data:
            self.chart_canvas.delete("all")
            self.chart_canvas.create_text(200, 150, text="暂无数据", font=("Arial", 16), fill="gray")
            return
        
        if index not in self.chart_data[condition]:
            self.chart_canvas.delete("all")
            self.chart_canvas.create_text(200, 150, text=f"第{index}次测量无数据", font=("Arial", 16), fill="gray")
            return
        
        data = self.chart_data[condition][index]
        self.chart_plot_data = data
        self.draw_chart(data)

    def draw_chart(self, data):
        """绘制曲线图"""
        self.chart_canvas.delete("all")
        
        if not data or len(data["time"]) < 1:
            self.chart_canvas.create_text(200, 150, text="数据点不足", font=("Arial", 16), fill="gray")
            return
        
        # 获取画布尺寸
        width = self.chart_canvas.winfo_width()
        height = self.chart_canvas.winfo_height()
        if width < 100 or height < 100:
            width = 400
            height = 300
        
        # 边距
        margin_left = 60
        margin_right = 30
        margin_top = 30
        margin_bottom = 40
        
        plot_width = width - margin_left - margin_right
        plot_height = height - margin_top - margin_bottom
        
        # 计算数据范围
        times = data["time"]
        angles = data["angle"]
        
        min_time = 0
        max_time = max(times) if times else 1
        min_angle = 0
        max_angle = max(angles) if angles else 1
        
        # 添加一些边距
        time_range = max_time - min_time
        angle_range = max_angle - min_angle
        if time_range == 0:
            time_range = 1
        if angle_range == 0:
            angle_range = 1
        
        # 坐标转换函数
        def to_canvas_x(t):
            return margin_left + (t - min_time) / time_range * plot_width
        
        def to_canvas_y(a):
            return margin_top + plot_height - (a - min_angle) / angle_range * plot_height
        
        # 绘制坐标轴
        self.chart_canvas.create_line(margin_left, margin_top, margin_left, margin_top + plot_height, width=2)
        self.chart_canvas.create_line(margin_left, margin_top + plot_height, margin_left + plot_width, margin_top + plot_height, width=2)
        
        # 绘制坐标轴标签
        self.chart_canvas.create_text(margin_left - 10, margin_top + plot_height // 2, text="角度(rad)", angle=90, font=("Arial", 9))
        self.chart_canvas.create_text(margin_left + plot_width // 2, margin_top + plot_height + 20, text="时间(s)", font=("Arial", 9))
        
        # 绘制网格和刻度
        for i in range(6):
            x = margin_left + i * plot_width / 5
            t = min_time + i * time_range / 5
            self.chart_canvas.create_line(x, margin_top, x, margin_top + plot_height, fill="#e0e0e0", width=1)
            self.chart_canvas.create_text(x, margin_top + plot_height + 15, text=f"{t:.1f}", font=("Arial", 8))
            
            y = margin_top + plot_height - i * plot_height / 5
            a = min_angle + i * angle_range / 5
            self.chart_canvas.create_line(margin_left, y, margin_left + plot_width, y, fill="#e0e0e0", width=1)
            self.chart_canvas.create_text(margin_left - 10, y, text=f"{a:.2f}", font=("Arial", 8))
        
        # 绘制数据点
        if len(times) >= 1:
            points = []
            for t, a in zip(times, angles):
                x = to_canvas_x(t)
                y = to_canvas_y(a)
                points.append((x, y))
            
            # 如果有多个点，绘制连线
            if len(points) > 1:
                for i in range(len(points) - 1):
                    self.chart_canvas.create_line(points[i][0], points[i][1], points[i+1][0], points[i+1][1], 
                                                fill="blue", width=2)
            
            # 绘制数据点
            for x, y in points:
                self.chart_canvas.create_oval(x-4, y-4, x+4, y+4, fill="red", outline="red", tags="data_point")
        
        # 显示数据信息（包括角加速度，如果有的话）
        info_lines = []
        info_lines.append(f"数据点数: {len(times)}")
        
        # 只有在数据中有角加速度时才显示（达到目标次数后才有）
        angular_accel = data.get("angular_accel")
        if angular_accel is not None:
            info_lines.append(f"角加速度: {angular_accel:.4f} rad/s²")
        
        info_text = "  ".join(info_lines)
        self.chart_canvas.create_text(margin_left + plot_width // 2, margin_top - 10, 
                                    text=info_text, font=("Arial", 9, "bold"), fill="blue")
        
        # 重置悬停状态
        self.hover_point_index = -1
        self.tooltip_items = []  # 清空提示框元素列表
        self.hide_tooltip()

    def on_chart_scroll(self, event):
        """鼠标滚轮缩放"""
        # 简单实现缩放功能
        if event.delta > 0:
            self.chart_x_scale *= 1.1
            self.chart_y_scale *= 1.1
        else:
            self.chart_x_scale *= 0.9
            self.chart_y_scale *= 0.9
        
        if self.chart_plot_data:
            self.draw_chart(self.chart_plot_data)

    def start_new_measurement(self):
        """开始新的测量"""
        condition = self.condition_var.get()
        
        # 初始化该条件的序号
        if condition not in self.current_measurement_index:
            self.current_measurement_index[condition] = 0
        
        # 序号加1
        self.current_measurement_index[condition] += 1
        index = self.current_measurement_index[condition]
        
        # 初始化该条件的数据结构（角加速度先设为None，完成后再计算）
        if condition not in self.chart_data:
            self.chart_data[condition] = {}
        
        self.chart_data[condition][index] = {
            "time": [],
            "angle": [],
            "count": 0,
            "angular_accel": None  # 初始为None，达到目标次数后才计算
        }
        
        self.current_measurement_data = {
            "condition": condition,
            "index": index,
            "data": self.chart_data[condition][index]
        }
        self.is_measuring = True

    def calculate_current_angular_acceleration(self):
        """计算当前实验条件下的角加速度"""
        condition = self.condition_var.get()
        g = 9.794
        J1 = 0.01033
        beta1 = -0.0520
        
        angular_accel = None
        
        if condition == "加砝码":
            try:
                weight_g = int(self.weight_var.get().replace('g', ''))
                total_weight_g = weight_g + 23.98
                total_weight_kg = total_weight_g / 1000.0
                radius_mm = int(self.radius_var.get().replace('mm', ''))
                radius_m = radius_mm / 1000.0
                numerator = total_weight_kg * radius_m * g + J1 * beta1
                denominator = J1 + total_weight_kg * radius_m * radius_m
                angular_accel = numerator / denominator
            except:
                angular_accel = None
                
        elif condition == "加样品":
            sample_type = self.sample_var.get() if hasattr(self, 'sample_var') else "圆盘"
            angular_accel = self.get_beta3(sample_type)
            
        elif condition == "加砝码和样品":
            try:
                sample_type = self.sample_var.get() if hasattr(self, 'sample_var') else "圆盘"
                beta3 = self.get_beta3(sample_type)
                J2 = self.get_J2(sample_type)
                weight_g = int(self.weight_var.get().replace('g', ''))
                total_weight_g = weight_g + 23.98
                total_weight_kg = total_weight_g / 1000.0
                radius_mm = int(self.radius_var.get().replace('mm', ''))
                radius_m = radius_mm / 1000.0
                numerator = total_weight_kg * radius_m * g + J2 * beta3
                denominator = J2 + total_weight_kg * radius_m * radius_m
                angular_accel = numerator / denominator
            except:
                angular_accel = None
                
        else:  # 空载物台
            angular_accel = -0.0520
        
        # --- 添加随机误差（最多±1%） ---
        if angular_accel is not None:
            import random
            error_factor = 1.0 + (random.random() - 0.5) * 0.02  # -5% 到 +5%
            angular_accel = angular_accel * error_factor
        # --- 误差添加结束 ---
        
        return angular_accel
    
    def add_chart_data_point(self, count):
        """添加数据点到曲线图"""
        if not self.is_measuring or not self.current_measurement_data:
            return
        
        # 计算当前时间
        current_time = self.host_elapsed_time
        # 角度 = 次数 * π
        angle = count * math.pi
        
        data = self.current_measurement_data["data"]
        data["time"].append(current_time)
        data["angle"].append(angle)
        data["count"] = count

    def end_measurement(self):
        """结束当前测量"""
        if self.is_measuring and self.current_measurement_data:
            condition = self.current_measurement_data["condition"]
            index = self.current_measurement_data["index"]
            
            # 如果数据点少于2个，删除该次测量
            data = self.current_measurement_data["data"]
            if len(data["time"]) < 2:
                del self.chart_data[condition][index]
                if not self.chart_data[condition]:
                    del self.chart_data[condition]
                if condition in self.current_measurement_index:
                    self.current_measurement_index[condition] -= 1
            else:
                # 数据有效，确保角加速度已计算并包含随机误差
                if data.get("angular_accel") is None:
                    # 如果角加速度未计算，使用当前条件计算并添加随机误差
                    accel = self.calculate_current_angular_acceleration()
                    if accel is not None:
                        data["angular_accel"] = accel  # calculate_current_angular_acceleration 已经包含误差
                
                if condition in self.current_measurement_index:
                    max_index = self.current_measurement_index[condition]
                    self.query_index_spinbox.config(to=max(1, max_index))
        
        self.is_measuring = False
        self.current_measurement_data = None
    
    def export_chart_data(self):
        """导出曲线图数据到Excel"""
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils.dataframe import dataframe_to_rows
            import pandas as pd
            import io
            from PIL import Image as PILImage
        except ImportError:
            messagebox.showerror("错误", "请安装必要的库: pip install openpyxl pandas")
            return
        
        if not self.chart_data:
            messagebox.showwarning("警告", "没有数据可导出")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="曲线图数据.xlsx"
        )
        if not file_path:
            return
        
        try:
            wb = openpyxl.Workbook()
            
            # 为每个实验条件和每次测量创建一个工作表
            for condition, indices in self.chart_data.items():
                for index, data in indices.items():
                    if len(data["time"]) < 2:
                        continue
                    
                    sheet_name = f"{condition}_{index}"
                    # 限制工作表名称长度
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:31]
                    ws = wb.create_sheet(sheet_name)
                    
                    # 写入数据
                    ws.cell(row=1, column=1, value="时间(s)")
                    ws.cell(row=1, column=2, value="角度(rad)")
                    ws.cell(row=1, column=3, value="触发次数")
                    
                    for i, (t, a) in enumerate(zip(data["time"], data["angle"])):
                        ws.cell(row=i+2, column=1, value=t)
                        ws.cell(row=i+2, column=2, value=a)
                        ws.cell(row=i+2, column=3, value=i+1)
                    
                    # 添加统计信息（包括角加速度）
                    row_offset = len(data["time"]) + 3
                    ws.cell(row=row_offset, column=1, value="总触发次数:")
                    ws.cell(row=row_offset, column=2, value=data["count"])
                    
                    angular_accel = data.get("angular_accel")
                    if angular_accel is not None:
                        ws.cell(row=row_offset + 1, column=1, value="角加速度:")
                        ws.cell(row=row_offset + 1, column=2, value=angular_accel)
            
            # 删除默认的Sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 保存
            wb.save(file_path)
            messagebox.showinfo("导出成功", f"数据已导出到: {file_path}")
            
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def import_chart_data(self):
        """导入曲线图数据"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("错误", "请安装必要的库: pip install openpyxl")
            return
        
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")],
            title="选择要导入的数据文件"
        )
        if not file_path:
            return
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            imported_count = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 解析工作表名称: "条件_序号"
                parts = sheet_name.rsplit('_', 1)
                if len(parts) != 2:
                    continue
                
                condition = parts[0]
                try:
                    index = int(parts[1])
                except ValueError:
                    continue
                
                # 检查条件是否有效
                valid_conditions = ["空载物台", "加砝码", "加样品", "加砝码和样品"]
                if condition not in valid_conditions:
                    continue
                
                # 读取数据
                times = []
                angles = []
                row = 2
                while True:
                    time_val = ws.cell(row=row, column=1).value
                    angle_val = ws.cell(row=row, column=2).value
                    if time_val is None or angle_val is None:
                        break
                    times.append(float(time_val))
                    angles.append(float(angle_val))
                    row += 1
                
                if len(times) < 2:
                    continue
                
                # 获取总触发次数
                count_row = len(times) + 3
                count_val = ws.cell(row=count_row, column=2).value
                count = int(count_val) if count_val else len(times)
                
                # 获取角加速度
                angular_accel = None
                accel_row = count_row + 1
                accel_label = ws.cell(row=accel_row, column=1).value
                if accel_label and "角加速度" in str(accel_label):
                    accel_val = ws.cell(row=accel_row, column=2).value
                    if accel_val is not None:
                        angular_accel = float(accel_val)
                
                # 保存数据
                if condition not in self.chart_data:
                    self.chart_data[condition] = {}
                
                self.chart_data[condition][index] = {
                    "time": times,
                    "angle": angles,
                    "count": count,
                    "angular_accel": angular_accel
                }
                
                # 更新序号
                if condition not in self.current_measurement_index:
                    self.current_measurement_index[condition] = 0
                if index > self.current_measurement_index[condition]:
                    self.current_measurement_index[condition] = index
                
                imported_count += 1
            
            if imported_count > 0:
                messagebox.showinfo("导入成功", f"成功导入 {imported_count} 组数据")
                self.query_chart_data()
            else:
                messagebox.showwarning("导入失败", "未找到有效数据")
                
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    def clear_all_chart_data(self):
        """清空所有曲线图数据"""
        if not self.chart_data:
            messagebox.showinfo("提示", "没有数据可清空")
            return
        
        result = messagebox.askyesno("确认清空", "确定要清空所有曲线图数据吗？\n此操作不可撤销！")
        if not result:
            return
        
        self.chart_data.clear()
        self.current_measurement_index.clear()
        self.is_measuring = False
        self.current_measurement_data = None
        
        self.chart_canvas.delete("all")
        self.chart_canvas.create_text(200, 150, text="数据已清空", font=("Arial", 16), fill="gray")
        
        messagebox.showinfo("清空完成", "所有曲线图数据已清空")

    def clear_chart(self):
        """清除当前显示的曲线"""
        self.chart_canvas.delete("all")
        self.chart_plot_data = None
        self.chart_canvas.create_text(200, 150, text="请查询数据", font=("Arial", 16), fill="gray")

    def init_left_area(self, parent):
        """初始化左侧实验操作区域"""
        # 导入装置图片
        img_path = get_resource_path(os.path.join("background", "装置.jpg"))
        self.device_img = None
        self.canvas_device = None
        self.screw_line = None
        
        # 新增：样品显示相关的变量
        self.sample_display_items = []  # 存储所有样品图形ID
        self.cylinder1_id = None
        self.cylinder2_id = None
        self.sample_rect_id = None  # 圆盘/圆环的ID
        
        # 加载滑轮和砝码图片
        try:
            pulley_path = get_resource_path(os.path.join("background", "滑轮.jpg"))
            pulley_img = Image.open(pulley_path)
            pulley_img = pulley_img.resize((60, 35), Image.Resampling.LANCZOS)
            self.pulley_img = ImageTk.PhotoImage(pulley_img)
        except Exception as e:
            print(f"无法加载滑轮图片: {e}")
            self.pulley_img = None
        
        try:
            weight_path = get_resource_path(os.path.join("background", "砝码.jpg"))
            weight_img = Image.open(weight_path)
            weight_img = weight_img.resize((40, 110), Image.Resampling.LANCZOS)
            self.weight_img = ImageTk.PhotoImage(weight_img)
        except Exception as e:
            print(f"无法加载砝码图片: {e}")
            self.weight_img = None
        
        try:
            img = Image.open(img_path)
            img = img.resize((400, 330), Image.Resampling.LANCZOS)
            self.device_img = ImageTk.PhotoImage(img)
            
            # 创建Canvas，宽度增加到470（左侧留70px给滑轮）
            self.canvas_device = tk.Canvas(parent, width=470, height=360, highlightthickness=0)
            self.canvas_device.pack(pady=10)
            
            # 添加分隔线（从x=70开始）
            self.canvas_device.create_line(70, 30, 470, 30, fill="#a0c4d8", width=1)
            
            # 在画布上放置装置图片，并保存其ID用于后续降低层级
            self.device_image_id = self.canvas_device.create_image(70, 30, anchor=tk.NW, image=self.device_img, tags=("device_image",))
            
            # 在图片上绘制竖线（螺钉）
            self.screw_line = self.canvas_device.create_line(
                70 + 200 - 150, 15 + 30, 70 + 200 - 150, 45 + 30,
                fill="red", width=3, tags="screw"
            )
            
        except Exception as e:
            tk.Label(parent, text=f"无法加载图片: 装置.jpg\n{str(e)}", fg="red").pack(pady=10)
        
        # 创建 condition_var
        condition_frame = ttk.Frame(parent)
        condition_frame.pack(pady=5)
        ttk.Label(condition_frame, text="实验条件:").pack(side=tk.LEFT, padx=5)
        self.condition_var = tk.StringVar()
        # 在 init_left_area 中
        condition_combo = ttk.Combobox(condition_frame, textvariable=self.condition_var, 
                                    values=["空载物台", "加砝码", "加样品", "加砝码和样品"], 
                                    state="readonly")  # 保持 readonly_
        condition_combo.current(0)
        condition_combo.pack(side=tk.LEFT, padx=5)
        condition_combo.bind("<<ComboboxSelected>>", self.on_condition_change)
        
        # 现在 condition_var 已存在，绘制滑轮和砝码
        self.draw_pulley_and_weight()
        
        # 现在 condition_var 已存在，更新样品显示
        self.update_sample_display()
        
        # === 关键：将所有非装置图片的元素提升到装置图片之上 ===
        # 方法：将装置图片降低到最底层，或者将其他元素提升到装置图片之上
        # 降低装置图片到最底层
        if hasattr(self, 'device_image_id'):
            self.canvas_device.tag_lower(self.device_image_id)
            # 或者使用 tag_lower 将 device_image 标签的所有元素降低
            self.canvas_device.tag_lower("device_image")
        
        # 然后重新提升螺钉、连接线等元素
        self.canvas_device.tag_raise("screw")
        self.canvas_device.tag_raise("connection_line")
        self.canvas_device.tag_raise("pulley_group")
        
        # 初始转速设定
        speed_frame = ttk.Frame(parent)
        speed_frame.pack(pady=5)
        ttk.Label(speed_frame, text="初始转速 (弧度/s²):").pack(side=tk.LEFT, padx=5)
        self.speed_var = tk.DoubleVar(value=0.5)
        self.speed_scale = ttk.Scale(speed_frame, from_=0, to=1, variable=self.speed_var, 
                                    orient=tk.HORIZONTAL, length=150)
        self.speed_scale.pack(side=tk.LEFT, padx=5)
        self.speed_label = ttk.Label(speed_frame, text="0.50", width=8)
        self.speed_label.pack(side=tk.LEFT, padx=5)
        self.speed_scale.bind("<Motion>", self.update_speed_label)
        
        # 砝码质量选项（初始隐藏）
        self.weight_frame = ttk.Frame(parent)
        ttk.Label(self.weight_frame, text="砝码质量:").pack(side=tk.LEFT, padx=5)
        self.weight_var = tk.StringVar(value="5g")
        self.weight_combo = ttk.Combobox(self.weight_frame, textvariable=self.weight_var,
                                        values=["5g", "10g", "15g", "20g", "25g", "30g", "35g"], 
                                        state="readonly", width=8)
        self.weight_combo.pack(side=tk.LEFT, padx=5)
        self.weight_frame.pack_forget()
        
        # 塔轮半径选项（初始隐藏）
        self.radius_frame = ttk.Frame(parent)
        ttk.Label(self.radius_frame, text="塔轮半径:").pack(side=tk.LEFT, padx=5)
        self.radius_var = tk.StringVar(value="15mm")
        self.radius_combo = ttk.Combobox(self.radius_frame, textvariable=self.radius_var,
                                        values=["15mm", "20mm", "25mm", "30mm", "35mm"],
                                        state="readonly", width=8)
        self.radius_combo.pack(side=tk.LEFT, padx=5)
        self.radius_combo.bind("<<ComboboxSelected>>", self.on_radius_change)
        self.radius_frame.pack_forget()
        
        # 样品选择（初始隐藏）
        self.sample_frame = ttk.Frame(parent)
        ttk.Label(self.sample_frame, text="样品:").pack(side=tk.LEFT, padx=5)
        self.sample_var = tk.StringVar(value="圆盘")
        self.sample_combo = ttk.Combobox(self.sample_frame, textvariable=self.sample_var,
                                        values=["圆盘", "圆环", "圆柱"],
                                        state="readonly", width=8)
        self.sample_combo.pack(side=tk.LEFT, padx=5)
        self.sample_combo.bind("<<ComboboxSelected>>", self.on_sample_change)
        self.sample_frame.pack_forget()
        
        # 圆柱距离选择（初始隐藏）
        self.distance_frame = ttk.Frame(parent)
        ttk.Label(self.distance_frame, text="离中心距离:").pack(side=tk.LEFT, padx=5)
        self.distance_var = tk.StringVar(value="105mm")
        self.distance_combo = ttk.Combobox(self.distance_frame, textvariable=self.distance_var,
                                        values=["45mm", "60mm", "75mm", "90mm", "105mm"],
                                        state="readonly", width=8)
        self.distance_combo.pack(side=tk.LEFT, padx=5)
        self.distance_combo.bind("<<ComboboxSelected>>", self.on_distance_change)
        self.distance_frame.pack_forget()

        # 开始和停止按钮
        button_frame = ttk.Frame(parent)
        button_frame.pack(pady=10)
        self.start_btn = ttk.Button(button_frame, text="开始", command=self.start_experiment)
        self.start_btn.pack(side=tk.LEFT, padx=10)
        self.stop_btn = ttk.Button(button_frame, text="停止", command=self.stop_experiment, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=10)
        
        # 实验状态标签
        self.status_label = ttk.Label(parent, text="状态: 就绪")
        self.status_label.pack(pady=5)

        # 在 init_left_area 中，创建 condition_combo 时保存引用
        self.condition_combo = condition_combo

        

    def raise_connection_lines(self):
        """将连接线提升到最上层"""
        if hasattr(self, 'canvas_device'):
            self.canvas_device.tag_raise("pulley_line")
            self.canvas_device.tag_raise("weight_line")

    def draw_pulley_and_weight(self):
        """绘制滑轮和砝码（在装置图片左侧）"""
        if self.canvas_device is None:
            return
        
        # 检查 condition_var 是否存在
        if not hasattr(self, 'condition_var'):
            return
        
        # 检查是否包含砝码
        condition = self.condition_var.get()
        if condition not in ["加砝码", "加砝码和样品"]:
            self.canvas_device.delete("pulley_group")
            self.canvas_device.delete("connection_line")
            self.pulley_id = None
            self.pulley_line_id = None
            self.weight_line_id = None
            self.weight_id = None
            return
        
        # 清除旧的滑轮和砝码
        self.canvas_device.delete("pulley_group")
        self.canvas_device.delete("connection_line")
        
        # 获取塔轮半径
        radius_str = self.radius_var.get()
        radius_mm = int(radius_str.replace('mm', ''))
        
        # 计算滑轮Y位置
        offset = ((radius_mm - 15) // 5) * 17
        pulley_y = self.base_pulley_y + 30 - offset
        self.base_weight_y = pulley_y + 80
        
        pulley_x = 5
        
        # 绘制滑轮
        if self.pulley_img:
            self.pulley_id = self.canvas_device.create_image(
                pulley_x + 30 + 25, pulley_y+10 + 30,
                anchor=tk.CENTER, image=self.pulley_img,
                tags=("pulley_group",)
            )
        else:
            self.pulley_id = self.canvas_device.create_oval(
                pulley_x + 25, pulley_y+10, pulley_x + 60, pulley_y + 60+10,
                fill="#888888", outline="#555555", width=2,
                tags=("pulley_group",)
            )
        
        # 绘制水平连接线
        self.pulley_line_id = self.canvas_device.create_line(
            pulley_x + 60 + 25 , pulley_y + 30, 230, pulley_y + 30,
            fill="#333333", width=3, dash=(),
            tags=("connection_line",)
        )
        
        # 绘制砝码连接线
        self.weight_line_id = self.canvas_device.create_line(
            pulley_x + 30, pulley_y + 47+10,
            pulley_x + 30, self.base_weight_y + self.weight_y_offset - 22+10,
            fill="#333333", width=3, dash=(),
            tags=("connection_line",)
        )
        
        # 绘制砝码
        weight_x = pulley_x + 10
        weight_y = self.base_weight_y + self.weight_y_offset+10
        
        if self.weight_img:
            self.weight_id = self.canvas_device.create_image(
                weight_x + 20, weight_y + 30,
                anchor=tk.CENTER, image=self.weight_img,
                tags=("pulley_group",)
            )
        else:
            self.weight_id = self.canvas_device.create_rectangle(
                weight_x, weight_y, weight_x + 40, weight_y + 60,
                fill="#CC6600", outline="#884400", width=2,
                tags=("pulley_group",)
            )
        
        # 保存砝码的初始位置
        self.weight_start_x = weight_x
        self.weight_start_y = weight_y
        
        # === 关键：将装置图片降低到最底层 ===
        if hasattr(self, 'device_image_id'):
            self.canvas_device.tag_lower(self.device_image_id)
            self.canvas_device.tag_lower("device_image")
        
        # 然后提升所有需要显示在装置图片之上的元素
        self.canvas_device.tag_raise("connection_line")
        self.canvas_device.tag_raise("pulley_group")
        self.canvas_device.tag_raise("screw")

    def update_pulley_position(self):
        """更新滑轮和砝码位置（当塔轮半径变化时）"""
        self.draw_pulley_and_weight()

    def update_sample_display(self):
        """更新样品正面视角显示（顶部30px区域）"""
        if not hasattr(self, 'condition_var'):
            return
        
        if self.canvas_device is None:
            return
        
        # 清除旧的样品图形
        for item in self.sample_display_items:
            self.canvas_device.delete(item)
        self.sample_display_items = []
        self.cylinder1_id = None
        self.cylinder2_id = None
        self.sample_rect_id = None
        
        # 获取当前实验条件
        condition = self.condition_var.get()
        
        # 只有在加样品或加砝码和样品时才显示样品
        if condition not in ["加样品", "加砝码和样品"]:
            return
        
        sample_type = self.sample_var.get()
        
        # === 修正：样品显示区域从 x=70 开始（装置图片左侧边缘） ===
        offset_x = 70  # 装置图片偏移量
        x_left = offset_x + 40
        x_right = offset_x + 360
        y_top = 0
        height = 30
        
        if sample_type == "圆盘":
            self.sample_rect_id = self.canvas_device.create_rectangle(
                x_left, y_top, x_right, y_top + height,
                fill="#4A90D9", outline="#2C5F8A", width=2
            )
            self.sample_display_items.append(self.sample_rect_id)
            
        elif sample_type == "圆环":
            self.sample_rect_id = self.canvas_device.create_rectangle(
                x_left, y_top, x_right, y_top + height,
                fill="#E67E22", outline="#D35400", width=2
            )
            self.sample_display_items.append(self.sample_rect_id)
            
        elif sample_type == "圆柱":
            distance_str = self.distance_var.get()
            distance_mm = int(distance_str.replace('mm', ''))
            px_per_mm = 150 / 105
            offset_px = distance_mm * px_per_mm
            
            center_x = offset_x + 200  # 装置图片中心
            width = 15
            
            x1 = center_x - offset_px - width/2
            self.cylinder1_id = self.canvas_device.create_rectangle(
                x1, y_top, x1 + width, y_top + height,
                fill="#27AE60", outline="#1A7A42", width=1
            )
            self.sample_display_items.append(self.cylinder1_id)
            
            x2 = center_x + offset_px - width/2
            self.cylinder2_id = self.canvas_device.create_rectangle(
                x2, y_top, x2 + width, y_top + height,
                fill="#27AE60", outline="#1A7A42", width=1
            )
            self.sample_display_items.append(self.cylinder2_id)

    def update_cylinder_positions(self):
        """更新圆柱位置（调用update_sample_display）"""
        self.update_sample_display()

    def hide_cylinders(self):
        """隐藏圆柱"""
        if self.cylinder1_id:
            self.canvas_device.delete(self.cylinder1_id)
            self.cylinder1_id = None
        if self.cylinder2_id:
            self.canvas_device.delete(self.cylinder2_id)
            self.cylinder2_id = None

    def update_speed_label(self, event=None):
        """更新转速标签"""
        self.speed_label.config(text=f"{self.speed_var.get():.2f}")
    
    def update_distance_label(self, event=None):
        """更新距离标签"""
        pass
    
    def on_condition_change(self, event=None):
        """实验条件变化时的处理"""
        condition = self.condition_var.get()
        
        # 保存当前页面的参数（在切换之前）
        self.save_current_params()
        
        # 隐藏所有额外选项
        self.weight_frame.pack_forget()
        self.radius_frame.pack_forget()
        self.sample_frame.pack_forget()
        self.distance_frame.pack_forget()
        
        # 重置角加速度
        self.angular_acceleration = -0.0520
        
        # 根据条件显示选项
        if condition == "加砝码" or condition == "加砝码和样品":
            # 包含砝码时，初始转速固定为0
            self.speed_scale.set(0)
            self.speed_scale.config(state="disabled")
            self.speed_label.config(text="0.00")
            self.weight_frame.pack(pady=2)
            self.radius_frame.pack(pady=2)
            # 角加速度根据砝码质量和塔轮半径变化
            self.update_angular_acceleration()
            
            # 同步参数显示到数据记录区域
            weight_g = int(self.weight_var.get().replace('g', '')) + 23.98
            self.m_value_label.config(text=f"{weight_g:.2f} g")
            self.r_value_label.config(text=self.radius_var.get())
            
            # 显示滑轮和砝码
            self.draw_pulley_and_weight()
        else:
            self.speed_scale.config(state="normal")
            # --- 隐藏滑轮、砝码和所有连接线 ---
            self.canvas_device.delete("pulley_group")
            self.canvas_device.delete("connection_line")
            self.pulley_id = None
            self.pulley_line_id = None
            self.weight_line_id = None
            self.weight_id = None
        
        if condition == "加样品" or condition == "加砝码和样品":
            self.sample_frame.pack(pady=2)
            if self.current_tab == "验证平行轴定理" and condition == "加砝码和样品":
                self.distance_frame.pack(pady=2)
            elif self.sample_var.get() == "圆柱":
                self.distance_frame.pack(pady=2)
            else:
                self.distance_frame.pack_forget()
        
        # 更新样品显示
        self.update_sample_display()
        
        # 更新状态
        if condition == "空载物台":
            self.status_label.config(text="状态: 空载物台 - 就绪")
        elif condition == "加砝码":
            self.status_label.config(text="状态: 加砝码 - 就绪")
        elif condition == "加样品":
            self.status_label.config(text="状态: 加样品 - 就绪")
        elif condition == "加砝码和样品":
            self.status_label.config(text="状态: 加砝码和样品 - 就绪")
            
    def on_sample_change(self, event=None):
        """样品选择变化时的处理"""
        sample = self.sample_var.get()
        if sample == "圆柱":
            self.distance_frame.pack(pady=2)
        else:
            self.distance_frame.pack_forget()
        
        # 更新样品显示
        self.update_sample_display()
        
        # 更新角加速度
        self.update_angular_acceleration()
    
    def save_current_params(self):
        """保存当前页面的参数设置"""
        if hasattr(self, 'current_tab'):
            params = {
                "weight": self.weight_var.get() if hasattr(self, 'weight_var') else "5g",
                "radius": self.radius_var.get() if hasattr(self, 'radius_var') else "15mm"
            }
            # 如果是平行轴定理，保存距离
            if self.current_tab == "验证平行轴定理" and hasattr(self, 'distance_var'):
                params["distance"] = self.distance_var.get()
            self.saved_params[self.current_tab] = params

    def restore_page_params(self, tab_name):
        """恢复指定页面的参数设置"""
        params = self.saved_params.get(tab_name, {"weight": "5g", "radius": "15mm"})
        
        # 恢复砝码质量
        if hasattr(self, 'weight_var'):
            self.weight_var.set(params["weight"])
        
        # 恢复塔轮半径
        if hasattr(self, 'radius_var'):
            self.radius_var.set(params["radius"])
        
        # 如果是平行轴定理，恢复距离
        if tab_name == "验证平行轴定理" and hasattr(self, 'distance_var'):
            distance = params.get("distance", "45mm")
            self.distance_var.set(distance)
        
        # 更新参数显示
        condition = self.condition_var.get()
        if condition in ["加砝码", "加砝码和样品"]:
            weight_g = int(self.weight_var.get().replace('g', '')) + 23.98
            self.m_value_label.config(text=f"{weight_g:.2f} g")
            self.r_value_label.config(text=self.radius_var.get())

    def update_angular_acceleration(self):
        """根据实验条件计算角加速度"""
        condition = self.condition_var.get()
        
        J1 = 0.01033
        g = 9.8
        beta1 = -0.0520
        
        if condition == "加砝码":
            # 原有的加砝码计算
            weight_g = int(self.weight_var.get().replace('g', ''))
            total_weight_g = weight_g + 24
            total_weight_kg = total_weight_g / 1000.0
            radius_mm = int(self.radius_var.get().replace('mm', ''))
            radius_m = radius_mm / 1000.0
            
            numerator = total_weight_kg * radius_m * g + J1 * beta1
            denominator = J1 + total_weight_kg * radius_m * radius_m
            beta2 = numerator / denominator
            self.angular_acceleration = beta2
            # self.status_label.config(text=f"状态: 加砝码 - β₂={beta2:.4f} rad/s²")
        
        elif condition == "加样品":
            # 仅加样品，使用β₃（负值，减速）
            sample_type = self.sample_var.get()
            beta3 = self.get_beta3(sample_type)
            self.angular_acceleration = beta3  # 负值
            # self.status_label.config(text=f"状态: 加样品 - β₃={beta3:.4f} rad/s²")
        
        elif condition == "加砝码和样品":
            # 加砝码和样品，计算β₄
            sample_type = self.sample_var.get()
            beta3 = self.get_beta3(sample_type)
            J2 = self.get_J2(sample_type)
            
            weight_g = int(self.weight_var.get().replace('g', ''))
            total_weight_g = weight_g + 24
            total_weight_kg = total_weight_g / 1000.0
            radius_mm = int(self.radius_var.get().replace('mm', ''))
            radius_m = radius_mm / 1000.0
            
            # β₄ = (m·R·g + J₂·β₃) / (J₂ + m·R²)
            numerator = total_weight_kg * radius_m * g + J2 * beta3
            denominator = J2 + total_weight_kg * radius_m * radius_m
            beta4 = numerator / denominator
            self.angular_acceleration = beta4
            # self.status_label.config(text=f"状态: 加砝码和样品 - β₄={beta4:.4f} rad/s²")
        
        else:
            # 空载物台
            self.angular_acceleration = -0.0520

    def get_beta3(self, sample_type):
        """获取样品的角加速度 β₃"""
        if sample_type == "圆盘":
            return -0.0220
        elif sample_type == "圆环":
            return -0.0192
        elif sample_type == "圆柱":
            # 根据距离获取对应的β₃值
            distance_str = self.distance_var.get()
            distance_mm = int(distance_str.replace('mm', ''))
            
            # 预定义的映射表
            beta3_map = {
                45: -0.0500,
                60: -0.0463,
                75: -0.0427,
                90: -0.0390,
                105: -0.0353
            }
            return beta3_map.get(distance_mm, -0.0353)
        return -0.0520

    def get_J2(self, sample_type):
        """获取样品的转动惯量 J₂"""
        if sample_type == "圆盘":
            return 0.0257
        elif sample_type == "圆环":
            return 0.0378
        elif sample_type == "圆柱":
            return 0.01713
        return 0.01033
    
    def init_right_top_area(self, parent):
        """初始化右上实验仪主机区域"""
        # 导入主机图片
        img_path = get_resource_path(os.path.join("background", "主机.jpg"))
        self.host_img = None
        
        try:
            img = Image.open(img_path)
            img = img.resize((660, 250), Image.Resampling.LANCZOS)
            self.host_img = ImageTk.PhotoImage(img)
            
            # 创建画布作为容器，按钮将叠加在画布上
            self.host_container = tk.Frame(parent, width=660, height=250)
            self.host_container.pack(pady=5)
            self.host_container.pack_propagate(False)  # 固定大小
            
            # 创建Canvas显示图片
            self.canvas = tk.Canvas(self.host_container, width=660, height=250, highlightthickness=0)
            self.canvas.place(x=0, y=0, relwidth=1, relheight=1)
            
            # 在画布上放置图片
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.host_img)
            
            # === 电源按钮（I 和 O 两个独立按钮） ===
            # I 按钮（开机）
            self.power_on_btn = tk.Button(
                self.host_container, 
                text="I", 
                font=("Arial", 12, "bold"),
                width=2, 
                height=1,
                bg="#DC6E8B",
                fg="black",
                relief="flat",
                bd=2,
                command=self.power_on_action
            )
            self.power_on_btn.place(x=588, y=128, width=24, height=20)
            
            # O 按钮（关机）
            self.power_off_btn = tk.Button(
                self.host_container, 
                text="O", 
                font=("Arial", 12, "bold"),
                width=2, 
                height=1,
                bg="#DC6E8B",
                fg="black",
                relief="flat",
                bd=2,
                command=self.power_off_action
            )
            self.power_off_btn.place(x=588, y=151, width=24, height=20)
            
            # 电源指示灯（显示在两个按钮旁边）
            # self.power_indicator = tk.Label(
            #     self.host_container,
            #     text="●",
            #     font=("Arial", 14),
            #     fg="#cc0000",  # 红色表示关机
            #     bg="#f0f0f0"
            # )
            # self.power_indicator.place(x=654, y=15)
            
            # 在画布上叠加放置文本（次数和时间）- 默认为隐藏（关机状态）
            self.count_text_id = self.canvas.create_text(60, 70, anchor=tk.NW, text="次数：3次", 
                                                        font=("Arial", 10, "bold"), fill="blue", tags="count", state='hidden')
            self.time_text_id = self.canvas.create_text(60, 95, anchor=tk.NW, text="时间：0s", 
                                                        font=("Arial", 10), fill="black", tags="time", state='hidden')
            self.accel_text_id = self.canvas.create_text(60, 115, anchor=tk.NW, text="", 
                                                        font=("Arial", 10), fill="red", tags="accel", state='hidden')
            
            # 创建三个文本框（开始、查询、清空）在图片上 - 默认为隐藏
            self.start_text_id = self.canvas.create_text(75, 140, text="开始", 
                                                        font=("Arial", 10, "bold"), fill="black", state='hidden')
            self.query_text_id = self.canvas.create_text(120, 140, text="查询", 
                                                        font=("Arial", 10), fill="black", state='hidden')
            self.clear_text_id = self.canvas.create_text(165, 140, text="清空", 
                                                        font=("Arial", 10), fill="black", state='hidden')
            
            # 存储文本ID列表用于高亮和隐藏
            self.host_text_ids = [self.count_text_id, self.start_text_id, 
                                self.query_text_id, self.clear_text_id]
            # 额外存储用于隐藏/显示
            self.host_text_objects = {
                'count': self.count_text_id,
                'time': self.time_text_id,
                'accel': self.accel_text_id,
                'start': self.start_text_id,
                'query': self.query_text_id,
                'clear': self.clear_text_id
            }

            # 初始高亮次数（电源关闭时不高亮）
            # self.highlight_host_item(0)
            
            # === 四个按钮叠加在图片上 ===
            btn_width = 20
            btn_height = 20
            
            # 长按相关变量
            self.long_press_id = None
            self.long_press_delay = 300  # 首次延迟 (ms)
            self.long_press_interval = 80  # 重复间隔 (ms)
            self.press_direction = None  # 'up' 或 'down'
            
            # 向上按钮 - 默认为禁用
            self.up_btn = tk.Button(self.host_container, text="", font=("Arial", 8),
                                width=6, height=1,
                                bg="#545454", fg="#545454", relief=tk.FLAT, bd=2,
                                state=tk.DISABLED)
            self.up_btn.place(x=258, y=98, width=btn_width, height=btn_height)
            # 绑定鼠标事件
            self.up_btn.bind("<ButtonPress-1>", lambda e: self.on_press_start('up'))
            self.up_btn.bind("<ButtonRelease-1>", self.on_press_end)
            self.up_btn.bind("<Leave>", self.on_press_end)
            
            # 向下按钮 - 默认为禁用
            self.down_btn = tk.Button(self.host_container, text="", font=("Arial", 8),
                                    width=6, height=1, 
                                    bg="#545454", fg="#545454", relief=tk.FLAT, bd=2,
                                    state=tk.DISABLED)
            self.down_btn.place(x=259, y=158, width=btn_width, height=btn_height)
            self.down_btn.bind("<ButtonPress-1>", lambda e: self.on_press_start('down'))
            self.down_btn.bind("<ButtonRelease-1>", self.on_press_end)
            self.down_btn.bind("<Leave>", self.on_press_end)
            
            # 确定按钮 - 默认为禁用
            self.confirm_btn = tk.Button(self.host_container, text="", font=("Arial", 8),
                                        width=6, height=1, command=self.host_confirm,
                                        bg="#545454", fg="#545454",relief=tk.FLAT, bd=2,
                                        state=tk.DISABLED)
            self.confirm_btn.place(x=325, y=98, width=btn_width, height=btn_height)
            
            # 返回按钮 - 默认为禁用
            self.return_btn = tk.Button(self.host_container, text="", font=("Arial", 8),
                                    width=6, height=1, command=self.host_return,
                                    bg="#545454", fg="#545454", relief=tk.FLAT, bd=2,
                                    state=tk.DISABLED)
            self.return_btn.place(x=325, y=158, width=btn_width, height=btn_height)
            
        except Exception as e:
            tk.Label(parent, text=f"无法加载图片: 主机.jpg\n{str(e)}", fg="red").pack(pady=10)
            # 如果没有图片，仍然显示文本
            text_frame = ttk.Frame(parent)
            text_frame.pack(pady=10)
            
            start_label = tk.Label(text_frame, text="开始", font=("Arial", 10, "bold"), bg="orange", fg="white", padx=10, pady=2)
            start_label.pack(side=tk.LEFT, padx=5)
            query_label = tk.Label(text_frame, text="查询", font=("Arial", 10), bg="lightgray", fg="black", padx=10, pady=2)
            query_label.pack(side=tk.LEFT, padx=5)
            clear_label = tk.Label(text_frame, text="清空", font=("Arial", 10), bg="lightgray", fg="black", padx=10, pady=2)
            clear_label.pack(side=tk.LEFT, padx=5)
            
            # 即使图片加载失败也创建电源按钮
            self.power_on_btn = tk.Button(
                text_frame, 
                text="I", 
                font=("Arial", 12, "bold"),
                width=2, 
                height=1,
                bg="#DC6E8B",
                fg="black",
                relief=tk.RAISED,
                bd=2,
                command=self.power_on_action
            )
            self.power_on_btn.pack(side=tk.RIGHT, padx=2)
            
            self.power_off_btn = tk.Button(
                text_frame, 
                text="O", 
                font=("Arial", 12, "bold"),
                width=2, 
                height=1,
                bg="#DC6E8B",
                fg="black",
                relief=tk.RAISED,
                bd=2,
                command=self.power_off_action
            )
            self.power_off_btn.pack(side=tk.RIGHT, padx=2)
    
    def power_on_action(self):
        """开机操作"""
        if self.power_on:
            return
        self.power_on = True
        
        # 开机
        # self.power_indicator.config(fg="#00cc00")  # 绿色表示开机
        self.show_host_display()
        # 启用操作按钮
        self.up_btn.config(state=tk.NORMAL)
        self.down_btn.config(state=tk.NORMAL)
        self.confirm_btn.config(state=tk.NORMAL)
        self.return_btn.config(state=tk.NORMAL)
        # 高亮默认项
        self.highlight_host_item(0)
        self.status_label.config(text="状态: 主机已开机")
        # messagebox.showinfo("电源", "主机已开机")

    def power_off_action(self):
        """关机操作"""
        if not self.power_on:
            return
        self.power_on = False
        
        # 关机 - 清空数据并隐藏
        # self.power_indicator.config(fg="#cc0000")  # 红色表示关机
        self.clear_host_data()
        self.hide_host_display()
        # 禁用操作按钮
        self.up_btn.config(state=tk.DISABLED)
        self.down_btn.config(state=tk.DISABLED)
        self.confirm_btn.config(state=tk.DISABLED)
        self.return_btn.config(state=tk.DISABLED)
        self.status_label.config(text="状态: 主机已关机")
        # messagebox.showinfo("电源", "主机已关机，数据已清空")

    def show_host_display(self):
        """显示主机信息"""
        # 显示次数、时间、角加速度
        self.canvas.itemconfig(self.count_text_id, state='normal')
        self.canvas.itemconfig(self.time_text_id, state='normal')
        self.canvas.itemconfig(self.accel_text_id, state='normal')
        # 显示开始、查询、清空
        self.canvas.itemconfig(self.start_text_id, state='normal')
        self.canvas.itemconfig(self.query_text_id, state='normal')
        self.canvas.itemconfig(self.clear_text_id, state='normal')

    def hide_host_display(self):
        """隐藏主机信息"""
        # 隐藏次数、时间、角加速度
        self.canvas.itemconfig(self.count_text_id, state='hidden')
        self.canvas.itemconfig(self.time_text_id, state='hidden')
        self.canvas.itemconfig(self.accel_text_id, state='hidden')
        # 隐藏开始、查询、清空
        self.canvas.itemconfig(self.start_text_id, state='hidden')
        self.canvas.itemconfig(self.query_text_id, state='hidden')
        self.canvas.itemconfig(self.clear_text_id, state='hidden')

    def clear_host_data(self):
        """清空主机数据（关机时调用）"""
        # 重置主机状态
        self.host_is_counting = False
        self.host_is_ready = False
        self.host_current_count = 0
        self.host_trigger_count = 0
        self.host_last_trigger_angle = -1
        self.host_elapsed_time = 0
        self.host_target_count = 3
        self.host_query_data = []
        self.host_in_query = False
        self.host_is_setting = False
        self.host_selected_index = 0
        
        # 重置显示
        self.update_host_display("count", "次数：3次")
        self.update_host_display("time", "时间：0s")
        self.update_host_display("accel", "")
        self.canvas.itemconfig(self.start_text_id, text="开始")
        
        # 取消高亮
        for text_id in self.host_text_ids:
            self.canvas.itemconfig(text_id, fill="black", font=("Arial", 10))
        
        # 如果实验正在运行，停止实验
        if self.is_running:
            self.is_running = False
            if self.animation_id:
                self.root.after_cancel(self.animation_id)
                self.animation_id = None
            self.current_angle = 0
            self.update_screw_position(0)

    def on_press_start(self, direction):
        """鼠标按下按钮时触发"""
        self.press_direction = direction
        # 立即执行一次
        self.do_press_action(direction)
        # 启动长按定时器
        self.start_long_press()

    def on_press_end(self, event=None):
        """鼠标释放按钮时触发"""
        self.press_direction = None
        self.stop_long_press()

    def start_long_press(self):
        """启动长按定时器"""
        if self.long_press_id:
            self.root.after_cancel(self.long_press_id)
            self.long_press_id = None
        # 延迟后开始重复
        self.long_press_id = self.root.after(self.long_press_delay, self.repeat_press)

    def stop_long_press(self):
        """停止长按定时器"""
        if self.long_press_id:
            self.root.after_cancel(self.long_press_id)
            self.long_press_id = None

    def repeat_press(self):
        """长按重复执行"""
        if self.press_direction is None:
            return
        self.do_press_action(self.press_direction)
        # 继续下一次重复
        self.long_press_id = self.root.after(self.long_press_interval, self.repeat_press)

    def do_press_action(self, direction):
        """执行按钮动作"""
        if direction == 'up':
            self.host_up()
        elif direction == 'down':
            self.host_down()

    def highlight_host_item(self, index):
        """高亮指定的主机菜单项"""
        # 重置所有文本颜色
        for i, text_id in enumerate(self.host_text_ids):
            if i == 0:  # 次数
                color = "blue" if i == index else "black"
                font = ("Arial", 10, "bold") if i == index else ("Arial", 10)
                self.canvas.itemconfig(text_id, fill=color, font=font)
            else:
                color = "blue" if i == index else "black"
                font = ("Arial", 10, "bold") if i == index else ("Arial", 10)
                self.canvas.itemconfig(text_id, fill=color, font=font)
        
        self.host_selected_index = index
    
    def host_up(self):
        """向上按钮 - 上移高亮"""
        if self.host_is_setting:
            # 次数设定模式：+1
            self.host_setting_temp = min(99, self.host_setting_temp + 1)
            self.update_host_display("count", f"次数：{self.host_setting_temp}次")
        elif self.host_in_query:
            # 查询模式：上一组
            if self.host_query_index > 0:
                self.host_query_index -= 1
                # 先删除旧的设定次数显示
                self.canvas.delete("setting_count")
                self.display_query_data(self.host_query_index)
        else:
            # 正常模式：向上循环
            self.host_selected_index = (self.host_selected_index - 1) % 4
            self.highlight_host_item(self.host_selected_index)

    def host_down(self):
        """向下按钮 - 下移高亮"""
        if self.host_is_setting:
            # 次数设定模式：-1
            self.host_setting_temp = max(3, self.host_setting_temp - 1)
            self.update_host_display("count", f"次数：{self.host_setting_temp}次")
        elif self.host_in_query:
            # 查询模式：下一组
            if self.host_query_index < len(self.host_query_data) - 1:
                self.host_query_index += 1
                # 先删除旧的设定次数显示
                self.canvas.delete("setting_count")
                self.display_query_data(self.host_query_index)
        else:
            # 正常模式：向下循环
            self.host_selected_index = (self.host_selected_index + 1) % 4
            self.highlight_host_item(self.host_selected_index)
    
    def host_confirm(self):
        """确定按钮"""
        if self.host_is_setting:
            # 退出次数设定
            self.host_target_count = self.host_setting_temp
            self.host_is_setting = False
            self.update_host_display("count", f"次数：{self.host_target_count}次")
            self.highlight_host_item(0)
            return
        
        if self.host_in_query:
            # 查询模式按确定无操作
            return
        
        selected_item = self.host_menu_items[self.host_selected_index]
        
        if selected_item == "次数":
            # 进入次数设定
            self.host_is_setting = True
            self.host_setting_temp = self.host_target_count
            self.update_host_display("count", f"次数：{self.host_setting_temp}次")
            # 高亮显示设定状态
            self.canvas.itemconfig(self.count_text_id, fill="red")
        
        elif selected_item == "开始":
            if not self.host_is_ready and not self.host_is_counting and not self.is_running:
                # 进入计时计数准备状态
                self.host_is_ready = True
                self.host_is_counting = False
                self.host_current_count = 0
                self.host_trigger_count = 0
                self.host_last_trigger_angle = -1
                self.host_start_time = None
                self.host_elapsed_time = 0
                self.update_host_display("count", f"次数：{self.host_target_count}次")
                self.update_host_display("time", "时间：0s")
                self.update_host_display("accel", "")
                self.canvas.itemconfig(self.start_text_id, text="复位")
                self.status_label.config(text="状态: 准备就绪，请点击左侧'开始'启动运动")
                messagebox.showinfo("准备就绪", "主机已就绪，请点击左侧'开始'按钮启动运动")
            elif self.host_is_ready or self.host_is_counting:
                # 复位操作（在准备状态或计数状态都可以复位）
                self.host_is_ready = False
                self.host_is_counting = False
                self.is_running = False
                if self.animation_id:
                    self.root.after_cancel(self.animation_id)
                    self.animation_id = None
                self.host_current_count = 0
                self.host_trigger_count = 0
                self.host_last_trigger_angle = -1
                self.host_elapsed_time = 0
                self.host_start_time = None
                self.current_angle = 0
                self.is_accelerating = False  # 重置加速状态
                self.update_screw_position(0)
                self.update_host_display("count", f"次数：{self.host_target_count}次")
                self.update_host_display("time", "时间：0s")
                self.update_host_display("accel", "")
                self.canvas.itemconfig(self.start_text_id, text="开始")
                self.status_label.config(text="状态: 已复位")
                # 复位操作 - 结束当前测量
                self.end_measurement()
                    
        elif selected_item == "查询":
            if len(self.host_query_data) > 0:
                self.host_in_query = True
                self.host_query_index = 0
                # 先删除之前的设定次数显示
                self.canvas.delete("setting_count")
                self.display_query_data(0)
                self.status_label.config(text="状态: 查询模式")
            else:
                messagebox.showinfo("提示", "暂无查询数据")
        
        elif selected_item == "清空":
            # 重置所有
            self.host_is_counting = False
            self.host_is_ready = False
            self.host_current_count = 0
            self.host_trigger_count = 0
            self.host_last_trigger_angle = -1
            self.host_elapsed_time = 0
            self.host_target_count = 3
            self.host_query_data = []
            self.host_in_query = False
            self.host_is_setting = False
            self.update_host_display("count", "次数：3次")
            self.update_host_display("time", "时间：0s")
            self.update_host_display("accel", "")
            self.canvas.itemconfig(self.start_text_id, text="开始")
            self.highlight_host_item(0)
            self.status_label.config(text="状态: 已清空 - 就绪")
    
    def host_return(self):
        """返回按钮"""
        if self.host_is_setting:
            # 退出次数设定，不保存
            self.host_is_setting = False
            self.update_host_display("count", f"次数：{self.host_target_count}次")
            self.highlight_host_item(0)
        elif self.host_in_query:
            # 退出查询
            self.host_in_query = False
            # 恢复显示开始、查询、清空文本
            self.canvas.itemconfig(self.start_text_id, state='normal')
            self.canvas.itemconfig(self.query_text_id, state='normal')
            self.canvas.itemconfig(self.clear_text_id, state='normal')
            # 删除设定次数显示
            self.canvas.delete("setting_count")
            # 恢复正常显示
            self.update_host_display("count", f"次数：{self.host_target_count}次")
            self.update_host_display("time", f"时间：{self.host_elapsed_time:.2f}s")
            self.update_host_display("accel", "")
            self.highlight_host_item(0)
            self.status_label.config(text="状态: 就绪")
        else:
            # 正常模式返回无操作
            pass
    
    def display_query_data(self, index):
        """显示查询数据"""
        if index < len(self.host_query_data):
            data = self.host_query_data[index]
            
            # 隐藏开始、查询、清空文本
            self.canvas.itemconfig(self.start_text_id, state='hidden')
            self.canvas.itemconfig(self.query_text_id, state='hidden')
            self.canvas.itemconfig(self.clear_text_id, state='hidden')
            
            # 先删除旧的设定次数显示
            self.canvas.delete("setting_count")
            
            # 合并显示：第X组数据 和 当时的设定次数
            self.update_host_display("count", f"第{index}组：{data['target']}次")
            
            # 更新时间显示
            self.update_host_display("time", f"时间：{data['time']:.2f}s")
            
            # 更新角加速度显示
            self.update_host_display("accel", f"角加速度：{data['accel']:.4f} rad/s²")
    
    def check_photogate(self, angle_deg):
        """检查是否触发光电门"""
        if not self.host_is_counting:
            return
        
        # 获取当前角度在0-360度范围内的值
        angle_mod = angle_deg % 360
        
        # 首次检测，初始化上次角度
        if self.host_last_trigger_angle < 0:
            self.host_last_trigger_angle = angle_mod
            return
        
        # 检测是否跨越了90度或270度
        # 使用角度跨越检测：从<90到>=90，或从<270到>=270
        crossed_90 = self.host_last_trigger_angle < 90 and angle_mod >= 90
        crossed_270 = self.host_last_trigger_angle < 270 and angle_mod >= 270
        
        # 处理角度从接近360回到0的情况（一圈结束）
        # 如果角度从>350到<10，说明经过0度，此时270度到90度之间的跨越需要特殊处理
        if self.host_last_trigger_angle > 350 and angle_mod < 10:
            # 经过了0度，检查是否跨越了270度（在一圈结束时）
            # 这种情况不需要额外触发，因为270度的触发已经在前面处理了
            pass
        
        if crossed_90:
            self.trigger_photogate()
        elif crossed_270:
            self.trigger_photogate()
        
        self.host_last_trigger_angle = angle_mod
    
    def trigger_photogate(self):
        """触发光电门"""
        if not self.host_is_counting:
            return
            
        self.host_trigger_count += 1
        
        if self.host_start_time is None:
            # 第一次触发，开始计时
            self.host_start_time = time.time()
            self.host_current_count = 0
            self.update_host_display("count", "次数：0次")
            self.update_host_display("time", "时间：0.00s")
            self.update_host_display("accel", "")
            self.status_label.config(text="状态: 计时中...")
            
            # 开始新的测量
            self.start_new_measurement()
            return
        
        # 计算时间
        self.host_elapsed_time = time.time() - self.host_start_time
        self.host_current_count += 1
        
        # 添加数据点到曲线图（每次触发都添加）
        self.add_chart_data_point(self.host_current_count)
        
        # --- 实时更新曲线图 ---
        if self.is_measuring and self.current_measurement_data:
            condition = self.current_measurement_data["condition"]
            index = self.current_measurement_data["index"]
            self.query_condition.set(condition)
            self.query_index.set(index)
            data = self.chart_data[condition][index]
            self.chart_plot_data = data
            self.draw_chart(data)
        
        # 更新显示
        self.update_host_display("count", f"次数：{self.host_current_count}次")
        self.update_host_display("time", f"时间：{self.host_elapsed_time:.2f}s")
        
        # 检查是否达到目标次数
        if self.host_current_count >= self.host_target_count:
            # --- 达到目标次数后计算角加速度 ---
            condition = self.condition_var.get()
            g = 9.794
            
            if condition == "加砝码":
                J1 = 0.01033
                beta1 = -0.0520
                weight_g = int(self.weight_var.get().replace('g', ''))
                total_weight_g = weight_g + 23.98
                total_weight_kg = total_weight_g / 1000.0
                radius_mm = int(self.radius_var.get().replace('mm', ''))
                radius_m = radius_mm / 1000.0
                numerator = total_weight_kg * radius_m * g + J1 * beta1
                denominator = J1 + total_weight_kg * radius_m * radius_m
                angular_accel = numerator / denominator
                
            elif condition == "加样品":
                sample_type = self.sample_var.get()
                angular_accel = self.get_beta3(sample_type)
                
            elif condition == "加砝码和样品":
                sample_type = self.sample_var.get()
                beta3 = self.get_beta3(sample_type)
                J2 = self.get_J2(sample_type)
                weight_g = int(self.weight_var.get().replace('g', ''))
                total_weight_g = weight_g + 23.98
                total_weight_kg = total_weight_g / 1000.0
                radius_mm = int(self.radius_var.get().replace('mm', ''))
                radius_m = radius_mm / 1000.0
                numerator = total_weight_kg * radius_m * g + J2 * beta3
                denominator = J2 + total_weight_kg * radius_m * radius_m
                angular_accel = numerator / denominator
                
            else:
                angular_accel = self.angular_acceleration
            
            # 添加随机误差
            import random
            error_factor = 1.0 + (random.random() - 0.5) * 0.02
            angular_accel = angular_accel * error_factor

            # 更新数据中的角加速度
            if self.is_measuring and self.current_measurement_data:
                self.current_measurement_data["data"]["angular_accel"] = angular_accel
                data = self.current_measurement_data["data"]
                self.chart_plot_data = data
                self.draw_chart(data)
            
            # 显示角加速度
            self.update_host_display("accel", f"角加速度：{angular_accel:.4f} rad/s²")
            
            # 记录查询数据
            self.host_query_data.append({
                'count': self.host_current_count,
                'target': self.host_target_count,
                'time': self.host_elapsed_time,
                'accel': angular_accel
            })
            
            # 停止计数，退出准备状态
            self.host_is_counting = False
            self.host_is_ready = False
            self.is_running = False
            if self.animation_id:
                self.root.after_cancel(self.animation_id)
                self.animation_id = None
            
            # === 达到目标次数，启用控件 ===
            self.set_controls_enabled(True)
            
            # 保持当前位置（不复位）
            self.canvas.itemconfig(self.start_text_id, text="开始")
            self.status_label.config(text=f"状态: 计时完成 (共{self.host_current_count}次)")
            messagebox.showinfo("计时完成", f"达到设定次数{self.host_target_count}次")

            # 结束测量
            self.end_measurement()
            
            # 自动查询显示刚刚完成的测量
            condition = self.condition_var.get()
            if condition in self.current_measurement_index:
                index = self.current_measurement_index[condition]
                self.query_condition.set(condition)
                self.query_index.set(index)
                self.query_index_spinbox.config(to=max(1, index))
                self.query_chart_data()
        else:
            # 未达到目标次数，不显示角加速度
            self.update_host_display("accel", "")
        
    def init_right_bottom_area(self, parent):
        """初始化右下数据记录区域"""
        # 选项卡 - 改用按钮组
        tab_frame = ttk.Frame(parent)
        tab_frame.pack(pady=5)
        
        # 创建四个按钮
        self.btn_empty = ttk.Button(tab_frame, text="测量空载物台的转动惯量", 
                                    command=lambda: self.switch_tab("测量空载物台的转动惯量"))
        self.btn_empty.pack(side=tk.LEFT, padx=5)
        
        self.btn_disc = ttk.Button(tab_frame, text="测量圆盘的转动惯量",
                                command=lambda: self.switch_tab("测量圆盘的转动惯量"))
        self.btn_disc.pack(side=tk.LEFT, padx=5)
        
        self.btn_ring = ttk.Button(tab_frame, text="测量圆环的转动惯量",
                                command=lambda: self.switch_tab("测量圆环的转动惯量"))
        self.btn_ring.pack(side=tk.LEFT, padx=5)
        
        self.btn_axis = ttk.Button(tab_frame, text="验证平行轴定理",  # 新增
                                command=lambda: self.switch_tab("验证平行轴定理"))
        self.btn_axis.pack(side=tk.LEFT, padx=5)
        
        # 设置当前选中的按钮样式（初始状态）
        self.current_tab = "测量空载物台的转动惯量"
        self.update_button_style()
        
        # 创建表格和数据显示的容器
        self.table_frame = ttk.Frame(parent)
        self.table_frame.pack(fill=tk.BOTH, pady=5)
        
        # 初始化表格（传入当前选项卡名称）
        self.create_empty_table(self.current_tab)
        
        # 参数显示区域 - 拆分成多个独立栏目
        params_container = ttk.LabelFrame(parent, text="实验参数")
        params_container.pack(fill=tk.X, pady=5, padx=5)
        
        # 第一行：砝码质量 + 塔轮半径 + 重力加速度
        row1_frame = ttk.Frame(params_container)
        row1_frame.pack(fill=tk.X, pady=2)
        
        self.m_label = ttk.Label(row1_frame, text="砝码和砝码架总质量 m:", font=("Arial", 9))
        self.m_label.pack(side=tk.LEFT, padx=5)
        self.m_value_label = ttk.Label(row1_frame, text="      g", font=("Arial", 9, "bold"), foreground="blue")
        self.m_value_label.pack(side=tk.LEFT, padx=5)
        
        self.r_label = ttk.Label(row1_frame, text="塔轮半径 R:", font=("Arial", 9))
        self.r_label.pack(side=tk.LEFT, padx=20)
        self.r_value_label = ttk.Label(row1_frame, text="   mm", font=("Arial", 9, "bold"), foreground="blue")
        self.r_value_label.pack(side=tk.LEFT, padx=5)
        
        self.g_label = ttk.Label(row1_frame, text="重力加速度 g:", font=("Arial", 9))
        self.g_label.pack(side=tk.LEFT, padx=20)
        self.g_value_label = ttk.Label(row1_frame, text="9.794 m/s²", font=("Arial", 9, "bold"), foreground="blue")
        self.g_value_label.pack(side=tk.LEFT, padx=5)
        
        # 第二行：J₁ + 样品参数 + J₂
        self.row2_frame = ttk.Frame(params_container)
        self.row2_frame.pack(fill=tk.X, pady=2)
        
        # J₁ (空载物台转动惯量) - 默认显示，在空载物台页面使用
        self.j1_label = ttk.Label(self.row2_frame, text="空载物台转动惯量 J₁:", font=("Arial", 9))
        self.j1_label.pack(side=tk.LEFT, padx=5)
        self.j1_value_label = ttk.Label(self.row2_frame, text="待计算", font=("Arial", 9, "bold"), foreground="blue")
        self.j1_value_label.pack(side=tk.LEFT, padx=5)
        
        # 样品参数标签（用于显示圆盘/圆环的参数）- 默认隐藏
        self.sample_param_label = ttk.Label(self.row2_frame, text="", font=("Arial", 9))
        self.sample_param_value_label = ttk.Label(self.row2_frame, text="", font=("Arial", 9, "bold"), foreground="blue")
        # 先不pack，在update_params_display中根据需要显示
        
        # J2 标签（用于显示转动惯量结果）- 默认隐藏
        self.j2_label = ttk.Label(self.row2_frame, text="", font=("Arial", 9))
        self.j2_value_label = ttk.Label(self.row2_frame, text="", font=("Arial", 9, "bold"), foreground="blue")
        # 先不pack，在update_params_display中根据需要显示
        
        # 第三行 - 用于显示转动惯量和相对误差（动态创建）
        self.row3_frame = ttk.Frame(params_container)
        self.row3_frame.pack(fill=tk.X, pady=2)

        # 初始化额外参数的标签引用
        self.j3_label = None
        self.j3_value_label = None
        self.error_label = None
        self.error_value_label = None

        # 按钮区域
        button_frame = ttk.Frame(parent)
        button_frame.pack(pady=5)
        ttk.Button(button_frame, text="计算", command=self.calculate).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="清空数据", command=self.clear_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导出数据", command=self.export_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导入数据", command=self.import_data).pack(side=tk.LEFT, padx=5)
    
    def switch_tab(self, new_tab):
        """切换选项卡"""
        if new_tab == self.current_tab:
            return
        
        # 保存当前选项卡的数据和参数
        self.save_current_table_data()
        self.save_current_params()
        
        # 更新当前选项卡
        self.current_tab = new_tab
        
        # 更新按钮样式
        self.update_button_style()
        
        # 恢复新选项卡的参数
        self.restore_page_params(new_tab)
        
        # 更新参数显示（先清除额外参数）
        self.clear_extra_params()
        self.update_params_display(new_tab)
        
        # 创建新表格
        self.create_empty_table(new_tab)
        
        # if new_tab != "测量空载物台的转动惯量":
        #     messagebox.showinfo("提示", f"'{new_tab}' 功能已启用，请确保左侧实验条件为'加砝码和样品'")

    def clear_extra_params(self):
        """清除额外参数显示（J3和相对误差）"""
        if hasattr(self, 'row3_frame'):
            for widget in self.row3_frame.winfo_children():
                widget.destroy()
    
    def update_button_style(self):
        """更新按钮样式，高亮当前选中的按钮"""
        for btn in [self.btn_empty, self.btn_disc, self.btn_ring, self.btn_axis]:
            btn.configure(style="TButton")
        
        if self.current_tab == "测量空载物台的转动惯量":
            self.btn_empty.configure(style="Accent.TButton")
        elif self.current_tab == "测量圆盘的转动惯量":
            self.btn_disc.configure(style="Accent.TButton")
        elif self.current_tab == "测量圆环的转动惯量":
            self.btn_ring.configure(style="Accent.TButton")
        elif self.current_tab == "验证平行轴定理":
            self.btn_axis.configure(style="Accent.TButton")

    def update_params_display(self, tab_selection):
        """根据选项卡更新参数显示"""
        
        # 获取该页面的保存参数
        params = self.saved_params.get(tab_selection, {"weight": "5g", "radius": "15mm"})
        weight_g = int(params["weight"].replace('g', '')) + 23.98
        
        # 先隐藏所有可能显示的标签
        self.j1_label.pack_forget()
        self.j1_value_label.pack_forget()
        self.sample_param_label.pack_forget()
        self.sample_param_value_label.pack_forget()
        self.j2_label.pack_forget()
        self.j2_value_label.pack_forget()
        
        if tab_selection == "测量空载物台的转动惯量":
            # 第一行：砝码质量 + 塔轮半径 + 重力加速度
            self.m_label.config(text="砝码和砝码架总质量 m:")
            self.m_value_label.config(text=f"{weight_g:.2f} g")
            self.r_label.config(text="塔轮半径 R:")
            self.r_value_label.config(text=params["radius"])
            self.g_label.config(text="重力加速度 g:")
            self.g_value_label.config(text="9.794 m/s²")
            
            # 第二行：显示J₁
            self.j1_label.config(text="空载物台转动惯量 J₁:")
            self.j1_label.pack(side=tk.LEFT, padx=5)
            self.j1_value_label.config(text=self.saved_results["空载物台"]["j1"])
            self.j1_value_label.pack(side=tk.LEFT, padx=5)
            
            # 清空第三行
            self.clear_extra_params()
            
        elif tab_selection == "测量圆盘的转动惯量":
            # 第一行：砝码质量 + 塔轮半径 + 重力加速度
            self.m_label.config(text="砝码和砝码架总质量 m:")
            self.m_value_label.config(text=f"{weight_g:.2f} g")
            self.r_label.config(text="塔轮半径 R:")
            self.r_value_label.config(text=params["radius"])
            self.g_label.config(text="重力加速度 g:")
            self.g_value_label.config(text="9.794 m/s²")
            
            # 第二行：样品参数 + J₂
            # 显示样品参数
            self.sample_param_label.config(text="圆盘样品:")
            self.sample_param_label.pack(side=tk.LEFT, padx=5)
            self.sample_param_value_label.config(text="M=2166g, d₁=240mm")
            self.sample_param_value_label.pack(side=tk.LEFT, padx=5)
            
            # 显示J2值
            self.j2_label.config(text="载物台加样品转动惯量J₂:")
            self.j2_label.pack(side=tk.LEFT, padx=20)
            self.j2_value_label.config(text=self.saved_results["圆盘"]["j2"])
            self.j2_value_label.pack(side=tk.LEFT, padx=5)
            
            # 清空第三行并重新创建
            self.clear_extra_params()
            
            # 在第三行显示J3和误差
            j3_frame = ttk.Frame(self.row3_frame)
            j3_frame.pack(fill=tk.X, pady=1)
            
            j3_label = ttk.Label(j3_frame, text="圆盘转动惯量 J₃ (理论值0.0156 kg·m²):", font=("Arial", 9))
            j3_label.pack(side=tk.LEFT, padx=5)
            self.j3_value_label = ttk.Label(j3_frame, text=self.saved_results["圆盘"]["j3"], 
                                            font=("Arial", 9, "bold"), foreground="blue")
            self.j3_value_label.pack(side=tk.LEFT, padx=5)
            
            error_label = ttk.Label(j3_frame, text="相对误差:", font=("Arial", 9))
            error_label.pack(side=tk.LEFT, padx=20)
            self.error_value_label = ttk.Label(j3_frame, text=self.saved_results["圆盘"]["error"], 
                                            font=("Arial", 9, "bold"), foreground="red")
            self.error_value_label.pack(side=tk.LEFT, padx=5)
            
        elif tab_selection == "测量圆环的转动惯量":
            # 第一行：砝码质量 + 塔轮半径 + 重力加速度
            self.m_label.config(text="砝码和砝码架总质量 m:")
            self.m_value_label.config(text=f"{weight_g:.2f} g")
            self.r_label.config(text="塔轮半径 R:")
            self.r_value_label.config(text=params["radius"])
            self.g_label.config(text="重力加速度 g:")
            self.g_value_label.config(text="9.794 m/s²")
            
            # 第二行：样品参数 + J₂
            # 显示样品参数
            self.sample_param_label.config(text="圆环样品:")
            self.sample_param_label.pack(side=tk.LEFT, padx=5)
            self.sample_param_value_label.config(text="M=2172g, d₁=200mm, d₂=240mm")
            self.sample_param_value_label.pack(side=tk.LEFT, padx=5)
            
            # 显示J2值
            self.j2_label.config(text="载物台加样品转动惯量J₂:")
            self.j2_label.pack(side=tk.LEFT, padx=20)
            self.j2_value_label.config(text=self.saved_results["圆环"]["j2"])
            self.j2_value_label.pack(side=tk.LEFT, padx=5)
            
            # 清空第三行并重新创建
            self.clear_extra_params()
            
            # 在第三行显示J3和误差
            j3_frame = ttk.Frame(self.row3_frame)
            j3_frame.pack(fill=tk.X, pady=1)
            
            j3_label = ttk.Label(j3_frame, text="圆环转动惯量 J₃ (理论值0.0264 kg·m²):", font=("Arial", 9))
            j3_label.pack(side=tk.LEFT, padx=5)
            self.j3_value_label = ttk.Label(j3_frame, text=self.saved_results["圆环"]["j3"], 
                                            font=("Arial", 9, "bold"), foreground="blue")
            self.j3_value_label.pack(side=tk.LEFT, padx=5)
            
            error_label = ttk.Label(j3_frame, text="相对误差:", font=("Arial", 9))
            error_label.pack(side=tk.LEFT, padx=20)
            self.error_value_label = ttk.Label(j3_frame, text=self.saved_results["圆环"]["error"], 
                                            font=("Arial", 9, "bold"), foreground="red")
            self.error_value_label.pack(side=tk.LEFT, padx=5)
            
        elif tab_selection == "验证平行轴定理":
            # 第一行：砝码质量 + 塔轮半径 + 重力加速度
            self.m_label.config(text="砝码和砝码架总质量 m:")
            self.m_value_label.config(text=f"{weight_g:.2f} g")
            self.r_label.config(text="塔轮半径 R:")
            self.r_value_label.config(text=params["radius"])
            self.g_label.config(text="重力加速度 g:")
            self.g_value_label.config(text="9.794 m/s²")
            
            # 获取距离参数
            distance = params.get("distance", "45mm")
            
            # 第二行：样品参数 + J₂
            # 显示样品参数
            self.sample_param_label.config(text="圆柱样品:")
            self.sample_param_label.pack(side=tk.LEFT, padx=5)
            self.sample_param_value_label.config(text=f"M₁=300.9g, d₁=29.86mm  M₂=300.6g, d₂=29.85mm  x={distance}")
            self.sample_param_value_label.pack(side=tk.LEFT, padx=5)
            
            # 显示J2值
            self.j2_label.config(text="载物台加样品转动惯量J₂:")
            self.j2_label.pack(side=tk.LEFT, padx=20)
            self.j2_value_label.config(text=self.saved_results["平行轴定理"]["j2"])
            self.j2_value_label.pack(side=tk.LEFT, padx=5)
            
            # 清空第三行并重新创建
            self.clear_extra_params()
            
            # 在第三行显示J3和误差
            j3_frame = ttk.Frame(self.row3_frame)
            j3_frame.pack(fill=tk.X, pady=1)
            
            j3_label = ttk.Label(j3_frame, text="圆柱转动惯量 J₃ (理论值0.00670 kg·m²):", font=("Arial", 9))
            j3_label.pack(side=tk.LEFT, padx=5)
            self.j3_value_label = ttk.Label(j3_frame, text=self.saved_results["平行轴定理"]["j3"], 
                                            font=("Arial", 9, "bold"), foreground="blue")
            self.j3_value_label.pack(side=tk.LEFT, padx=5)
            
            error_label = ttk.Label(j3_frame, text="相对误差:", font=("Arial", 9))
            error_label.pack(side=tk.LEFT, padx=20)
            self.error_value_label = ttk.Label(j3_frame, text=self.saved_results["平行轴定理"]["error"], 
                                            font=("Arial", 9, "bold"), foreground="red")
            self.error_value_label.pack(side=tk.LEFT, padx=5)
            

    def show_extra_params(self, sample_type):
        """显示额外参数（J3和相对误差）"""
        # 清空第三行
        for widget in self.row3_frame.winfo_children():
            widget.destroy()
        
        # 获取该样品类型的保存结果
        result = self.saved_results.get(sample_type, {"j2": "待计算", "j3": "待计算", "error": ""})
        
        if sample_type == "圆盘":
            self.j3_label = ttk.Label(self.row3_frame, text="圆盘转动惯量 J₃ (理论值0.0156 kg·m²):", font=("Arial", 9))
            self.j3_label.pack(side=tk.LEFT, padx=5)
            self.j3_value_label = ttk.Label(self.row3_frame, text=result["j3"], font=("Arial", 9, "bold"), foreground="blue")
            self.j3_value_label.pack(side=tk.LEFT, padx=5)
            
            self.error_label = ttk.Label(self.row3_frame, text="相对误差:", font=("Arial", 9))
            self.error_label.pack(side=tk.LEFT, padx=20)
            self.error_value_label = ttk.Label(self.row3_frame, text=result["error"], font=("Arial", 9, "bold"), foreground="red")
            self.error_value_label.pack(side=tk.LEFT, padx=5)
            
        elif sample_type == "圆环":
            self.j3_label = ttk.Label(self.row3_frame, text="圆环转动惯量 J₃ (理论值0.0264 kg·m²):", font=("Arial", 9))
            self.j3_label.pack(side=tk.LEFT, padx=5)
            self.j3_value_label = ttk.Label(self.row3_frame, text=result["j3"], font=("Arial", 9, "bold"), foreground="blue")
            self.j3_value_label.pack(side=tk.LEFT, padx=5)
            
            self.error_label = ttk.Label(self.row3_frame, text="相对误差:", font=("Arial", 9))
            self.error_label.pack(side=tk.LEFT, padx=20)
            self.error_value_label = ttk.Label(self.row3_frame, text=result["error"], font=("Arial", 9, "bold"), foreground="red")
            self.error_value_label.pack(side=tk.LEFT, padx=5)

    def create_empty_table(self, tab_name=None):
        """创建表格"""
        if tab_name is None:
            tab_name = self.current_tab
        
        for widget in self.table_frame.winfo_children():
            widget.destroy()
        
        # 根据选项卡确定列名
        if tab_name == "测量空载物台的转动惯量":
            columns = ("次数", "beta1", "beta2")
            col_labels = {
                "次数": "测量次数",
                "beta1": "β₁/(弧度·s⁻²)",
                "beta2": "β₂/(弧度·s⁻²)"
            }
        elif tab_name == "测量圆盘的转动惯量":
            columns = ("次数", "beta3", "beta4")
            col_labels = {
                "次数": "测量次数",
                "beta3": "β₃/(弧度·s⁻²)",
                "beta4": "β₄/(弧度·s⁻²)"
            }
        elif tab_name == "测量圆环的转动惯量":
            columns = ("次数", "beta3", "beta4")
            col_labels = {
                "次数": "测量次数",
                "beta3": "β₃/(弧度·s⁻²)",
                "beta4": "β₄/(弧度·s⁻²)"
            }
        elif tab_name == "验证平行轴定理":  # 新增
            columns = ("次数", "beta3", "beta4")
            col_labels = {
                "次数": "测量次数",
                "beta3": "β₃/(弧度·s⁻²)",
                "beta4": "β₄/(弧度·s⁻²)"
            }
        else:
            columns = ("次数", "beta1", "beta2")
            col_labels = {
                "次数": "测量次数",
                "beta1": "β₁/(弧度·s⁻²)",
                "beta2": "β₂/(弧度·s⁻²)"
            }
    
    # ... 其余代码保持不变 ...
        
        self.tree = ttk.Treeview(self.table_frame, columns=columns, show="headings", height=7)
        
        for col in columns:
            self.tree.heading(col, text=col_labels[col])
            if col == "次数":
                self.tree.column(col, width=100, anchor=tk.CENTER)
            else:
                self.tree.column(col, width=150, anchor=tk.CENTER)
        
        # 创建5行数据行
        entries = []
        for i in range(1, 6):
            item_id = self.tree.insert("", tk.END, values=(f"第{i}次", "", ""))
            entries.append(item_id)
        
        # 创建平均值行
        avg_item_id = self.tree.insert("", tk.END, values=("平均值", "", ""), tags=("avg",))
        self.tree.tag_configure("avg", background="#f0f0f0")
        
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        self.tree.pack(fill=tk.X, pady=5)
        
        self.table_data[tab_name]["entries"] = entries
        
        # 恢复保存的数据
        saved_values = self.table_data[tab_name]["values"]
        if saved_values:
            # 恢复前5行数据
            for i, item_id in enumerate(entries):
                if i < len(saved_values):
                    values_tuple = tuple(saved_values[i])
                    if len(values_tuple) == len(columns):
                        self.tree.item(item_id, values=values_tuple)
            
            # 恢复平均值行（如果保存的数据中包含平均值）
            # 平均值行在 saved_values 中的索引为 len(entries)（即第6行）
            if len(saved_values) > len(entries):
                avg_values = tuple(saved_values[len(entries)])
                if len(avg_values) == len(columns):
                    self.tree.item(avg_item_id, values=avg_values)
    
    def save_current_table_data(self):
        """保存当前表格数据到 table_data"""
        if hasattr(self, 'tree') and self.tree:
            current_values = []
            for item in self.tree.get_children():
                try:
                    values = self.tree.item(item, "values")
                    # 保存所有行（包括平均值行）
                    current_values.append(list(values))
                except tk.TclError:
                    pass
            self.table_data[self.current_tab]["values"] = current_values
        
    def on_tree_double_click(self, event):
        """双击表格单元格进行编辑"""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        
        column = self.tree.identify_column(event.x)
        item = self.tree.identify_row(event.y)
        
        current_values = list(self.tree.item(item, "values"))
        
        if current_values[0] == "平均值":
            return
        
        col_idx = int(column[1:]) - 1
        current_value = current_values[col_idx] if col_idx < len(current_values) else ""
        
        x, y, width, height = self.tree.bbox(item, column)
        entry = tk.Entry(self.tree, justify="center")
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, current_value)
        entry.focus()
        entry.select_range(0, tk.END)  # 选中所有文本方便重新输入
        
        def save_edit(event=None):
            """保存编辑"""
            new_value = entry.get().strip()
            current_values[col_idx] = new_value
            self.tree.item(item, values=tuple(current_values))
            entry.destroy()
            # 自动保存到 table_data
            self.save_current_table_data()
        
        def cancel_edit(event=None):
            """取消编辑"""
            entry.destroy()
        
        # 按回车保存
        entry.bind("<Return>", save_edit)
        
        # 按 Escape 取消编辑
        entry.bind("<Escape>", cancel_edit)
        
        # 点击其他地方保存（FocusOut）
        entry.bind("<FocusOut>", save_edit)
    
    def start_experiment(self):
        if self.is_running:
            return
        
        # 检查主机是否已进入准备状态
        if not self.host_is_ready:
            messagebox.showwarning("提示", "请先在主机界面点击'开始'")
            return
        
        # === 禁用所有实验控制控件 ===
        self.set_controls_enabled(False)
        
        condition = self.condition_var.get()
        self.is_running = True
        self.status_label.config(text=f"状态: 运行中 - {condition}")
        # 在 start_experiment 方法中，开始动画前
        if condition in ["加样品", "加砝码和样品"] and self.sample_var.get() == "圆柱":
            self.update_cylinder_positions()
        # 设置主机为计数状态
        self.host_is_counting = True
        self.host_is_ready = False
        self.host_current_count = 0
        self.host_trigger_count = 0
        self.host_last_trigger_angle = -1
        self.host_elapsed_time = 0
        self.host_start_time = None
        self.update_host_display("count", f"次数：{self.host_target_count}次")
        self.update_host_display("time", "时间：0s")
        self.update_host_display("accel", "")
        self.canvas.itemconfig(self.start_text_id, text="复位")
        self.status_label.config(text="状态: 计时中...")
        
        # ... 其余代码保持不变 ...
        
        # 速度倍率
        speed_multiplier = 10
        rad_to_deg = 180 / math.pi
        
        initial_omega = self.speed_var.get()
        alpha = abs(self.angular_acceleration)
        
        # 角速度上限 (弧度/s)
        self.max_omega = 2.0
        
        if condition in ["加砝码", "加砝码和样品"]:
            # 砝码下落是加速运动，初始角速度为0
            initial_omega = 0
            self.speed_scale.set(0)
            self.speed_label.config(text="0.00")
            alpha = self.angular_acceleration
            
            if alpha > 0:
                self.ramp_time = self.max_omega / alpha
            else:
                self.ramp_time = 0
            
            ramp_angle = 0.5 * alpha * self.ramp_time * self.ramp_time if alpha > 0 else 0
            
            # --- 修改：根据目标次数计算需要的总角度 ---
            # 每次触发需要经过180度（从90度到270度），每圈触发2次
            # 需要的圈数 = ceil(目标次数 / 2)
            import math as math_module
            needed_cycles = math_module.ceil(self.host_target_count / 2)
            # 每圈360度，加上额外的半圈确保最后一次触发完成
            min_angle_needed = needed_cycles * 360 + 180
            
            # 至少转4圈（1440度），但根据目标次数调整
            max_angle = max(1440, min_angle_needed)
            self.total_angle_deg = max_angle
            self.ramp_angle = ramp_angle
            self.is_accelerating = True
        else:
            # 空载物台：匀减速运动
            # 计算总转过的角度
            self.total_angle_rad = (initial_omega ** 2) / (2 * alpha) if alpha > 0 else 0
            self.total_angle_deg = self.total_angle_rad * rad_to_deg
            
            # 确保至少能触发目标次数
            import math as math_module
            needed_cycles = math_module.ceil(self.host_target_count / 2)
            min_angle_needed = needed_cycles * 360 + 180
            
            if self.total_angle_deg < min_angle_needed and initial_omega > 0:
                needed_omega = math.sqrt(2 * alpha * min_angle_needed / rad_to_deg)
                if needed_omega <= 1.0:
                    initial_omega = needed_omega
                    self.speed_scale.set(initial_omega)
                    self.speed_label.config(text=f"{initial_omega:.2f}")
                    self.total_angle_rad = (initial_omega ** 2) / (2 * alpha) if alpha > 0 else 0
                    self.total_angle_deg = self.total_angle_rad * rad_to_deg
                else:
                    self.total_angle_deg = max(self.total_angle_deg, min_angle_needed)
            
            self.stop_time = initial_omega / alpha if alpha > 0 else 0
            self.is_accelerating = False
        
        self.current_angle = 0
        if condition in ["加砝码", "加砝码和样品"]:
            self.angular_velocity = 0
            self.effective_acceleration = alpha * rad_to_deg * speed_multiplier
            self.max_omega_deg = self.max_omega * rad_to_deg * speed_multiplier
            # 计算到达上限后的匀速阶段时间
            if self.max_omega > 0 and alpha > 0:
                # 加速阶段时间
                ramp_time = self.max_omega / alpha
                # 加速阶段角度
                ramp_angle_deg = 0.5 * alpha * ramp_time * ramp_time * rad_to_deg
                # 剩余匀速阶段角度
                remaining_angle = self.total_angle_deg - ramp_angle_deg
                # 匀速阶段时间
                cruise_time = remaining_angle / self.max_omega_deg if self.max_omega_deg > 0 else 0
                self.stop_time = ramp_time + cruise_time
            else:
                self.stop_time = 0
        else:
            self.angular_velocity = initial_omega * rad_to_deg * speed_multiplier
            self.effective_acceleration = -alpha * rad_to_deg * speed_multiplier
            self.max_omega_deg = 0
        
        self.start_time = 0
        self.elapsed_time = 0
        # 更新样品显示（重置位置）
        self.update_sample_display()
        self.update_screw_position(0)
        
        self.animate()

    def animate(self):
        if not self.is_running:
            return

        dt = 0.05
        self.elapsed_time += dt

        condition = self.condition_var.get()

        # --- 更新砝码位置（包含砝码时） ---
        if condition in ["加砝码", "加砝码和样品"] and self.weight_id is not None:
            # 砝码下落距离与转过的角度成正比
            radius_mm = int(self.radius_var.get().replace('mm', ''))
            radius_m = radius_mm / 1000.0
            circumference = 2 * math.pi * radius_m * 1000
            
            drop_distance = (self.current_angle / 360.0) * circumference
            max_drop = 400
            drop_distance = min(drop_distance, max_drop)
            
            new_y = self.base_weight_y + drop_distance
            
            if self.weight_img:
                self.canvas_device.coords(self.weight_id, 
                                            self.weight_start_x + 20, 
                                            new_y + 30 +10)
            else:
                self.canvas_device.coords(self.weight_id,
                                            self.weight_start_x,
                                            new_y,
                                            self.weight_start_x + 40,
                                            new_y + 60 +10)
            
            if self.weight_line_id:
                pulley_x = 5
                radius_str = self.radius_var.get()
                radius_mm = int(radius_str.replace('mm', ''))
                offset = ((radius_mm - 15) // 5) * 17
                pulley_y = self.base_pulley_y + 30 - offset
                
                self.canvas_device.coords(
                    self.weight_line_id,
                    pulley_x + 30, pulley_y + 47 +10,
                    pulley_x + 30, new_y - 22 +10
                )

        # 检查是否到达停止时间
        if self.elapsed_time >= self.stop_time:
            self.update_screw_position(self.current_angle)
            self.is_running = False
            self.status_label.config(text="状态: 已停止")
            
            # === 动画自然停止，启用控件 ===
            self.set_controls_enabled(True)
            
            if self.host_is_counting:
                self.host_is_counting = False
                self.host_is_ready = False
                self.canvas.itemconfig(self.start_text_id, text="开始")
                self.status_label.config(text="状态: 已停止")
            return

        # --- 更新角速度 ---
        if condition in ["加砝码", "加砝码和样品"]:
            if self.is_accelerating:
                self.angular_velocity += self.effective_acceleration * dt
                if self.angular_velocity >= self.max_omega_deg:
                    self.angular_velocity = self.max_omega_deg
                    self.is_accelerating = False
            self.current_angle += self.angular_velocity * dt
        else:
            self.angular_velocity += self.effective_acceleration * dt
            
            if self.angular_velocity <= 0:
                self.angular_velocity = 0
                self.update_screw_position(self.current_angle)
                self.is_running = False
                if self.animation_id:
                    self.root.after_cancel(self.animation_id)
                    self.animation_id = None
                self.status_label.config(text="状态: 已停止 (速度为零)")
                
                # === 速度为零停止，启用控件 ===
                self.set_controls_enabled(True)
                
                if self.host_is_counting:
                    self.host_is_counting = False
                    self.host_is_ready = False
                    self.canvas.itemconfig(self.start_text_id, text="开始")
                    self.status_label.config(text="状态: 已停止 (速度为零)")
                return
            
            self.current_angle += self.angular_velocity * dt
        
        # 更新螺钉位置
        self.update_screw_position(self.current_angle)
        
        # 更新圆柱位置（如果是圆柱样品）
        if condition in ["加样品", "加砝码和样品"] and self.sample_var.get() == "圆柱":
            self.update_cylinder_animation(self.current_angle)
        
        # 检查光电门
        if self.host_is_counting:
            self.check_photogate(self.current_angle)
        
        self.animation_id = self.root.after(int(dt * 1000), self.animate)
    
    def update_cylinder_animation(self, angle_deg):
        """根据角度更新圆柱位置（在顶部样品显示区域摆动）"""
        if self.canvas_device is None:
            return
        
        condition = self.condition_var.get()
        if condition not in ["加样品", "加砝码和样品"]:
            return
        
        if self.sample_var.get() != "圆柱":
            return
        
        if self.cylinder1_id is None or self.cylinder2_id is None:
            return
        
        # 获取距离 (mm)
        distance_str = self.distance_var.get()
        distance_mm = int(distance_str.replace('mm', ''))
        px_per_mm = 150 / 105
        offset_px = distance_mm * px_per_mm
        
        # === 修正：装置图片偏移70px ===
        offset_x = 70
        center_x = offset_x + 200
        y_top = 0
        width = 15
        height = 30
        
        # 角度转弧度
        angle_rad = math.radians(angle_deg)
        
        # 左右摆动：使用余弦函数
        swing_offset = offset_px * math.cos(angle_rad)
        
        # 左侧圆柱
        x1 = center_x - swing_offset - width/2
        self.canvas_device.coords(
            self.cylinder1_id,
            x1, y_top, x1 + width, y_top + height
        )
        
        # 右侧圆柱
        x2 = center_x + swing_offset - width/2
        self.canvas_device.coords(
            self.cylinder2_id,
            x2, y_top, x2 + width, y_top + height
        )

    def on_radius_change(self, event=None):
        """塔轮半径变化时的处理"""
        # 更新滑轮位置
        self.draw_pulley_and_weight()
        # 更新角加速度
        self.update_angular_acceleration()

    def on_distance_change(self, event=None):
        """圆柱距离变化时的处理"""
        # 更新圆柱位置
        if self.sample_var.get() == "圆柱":
            self.update_sample_display()
        # 重新计算角加速度
        self.update_angular_acceleration()

    def update_screw_position(self, angle_deg):
        """更新螺钉位置 - 左右水平运动"""
        if self.canvas_device is None or self.screw_line is None:
            return
        
        # 角度转弧度
        angle_rad = math.radians(angle_deg)
        
        # 装置图片偏移70px，中心位置在 (70+200, 30+165)
        offset_x = 70
        offset_y = 30
        center_x = offset_x + 200
        center_y = offset_y + 30  # 螺钉在图片中的位置
        radius = 150  # 运动半径
        
        x = center_x - radius * math.cos(angle_rad)
        y = center_y
        
        # 更新竖线位置
        self.canvas_device.coords(self.screw_line, x, y - 15, x, y + 15)
    
    def stop_experiment(self):
        """停止实验按钮回调"""
        self.is_running = False
        if self.animation_id:
            self.root.after_cancel(self.animation_id)
            self.animation_id = None
        
        # 复位到初始位置
        self.current_angle = 0
        self.update_screw_position(0)
        # 复位圆柱位置
        if self.sample_var.get() == "圆柱":
            self.update_cylinder_animation(0)
        
        # --- 复位砝码位置和连接线 ---
        if self.weight_id is not None:
            if self.weight_img:
                self.canvas_device.coords(self.weight_id,
                                        self.weight_start_x + 20,
                                        self.base_weight_y + 30 +10)
            else:
                self.canvas_device.coords(self.weight_id,
                                        self.weight_start_x,
                                        self.base_weight_y+10,
                                        self.weight_start_x + 40,
                                        self.base_weight_y +10+ 60)
            
            # 复位砝码连接线
            if self.weight_line_id:
                pulley_x = 5  # 与 draw_pulley_and_weight 保持一致
                radius_str = self.radius_var.get()
                radius_mm = int(radius_str.replace('mm', ''))
                offset = ((radius_mm - 15) // 5) * 17
                pulley_y = self.base_pulley_y + 30 - offset
                
                self.canvas_device.coords(
                    self.weight_line_id,
                    pulley_x + 30, pulley_y + 47+10,
                    pulley_x + 30, self.base_weight_y - 22 +10
                )
        
        # 重置加速状态
        self.is_accelerating = False
        
        # 重置主机状态，退出准备状态
        self.host_is_ready = False
        self.host_is_counting = False
        self.host_current_count = 0
        self.host_trigger_count = 0
        self.host_last_trigger_angle = -1
        self.host_elapsed_time = 0
        self.host_start_time = None
        self.update_host_display("count", f"次数：{self.host_target_count}次")
        self.update_host_display("time", "时间：0s")
        self.update_host_display("accel", "")
        self.canvas.itemconfig(self.start_text_id, text="开始")
        
        # === 启用所有实验控制控件 ===
        self.set_controls_enabled(True)
        
        self.status_label.config(text="状态: 已停止 (复位)")
        messagebox.showinfo("实验停止", "实验已停止，螺钉已复位")
    
    def set_controls_enabled(self, enabled):
        """启用或禁用所有实验控制控件"""
        # ttk.Combobox 使用 "readonly"(只读) 或 "disabled"(禁用)
        # 注意：不是 "normal"，"normal" 会让 Combobox 变成可编辑的
        combo_state = "readonly" if enabled else "disabled"
        # 普通按钮使用 tk.NORMAL 或 tk.DISABLED
        button_state = tk.NORMAL if enabled else tk.DISABLED
        
        # 实验条件下拉菜单
        if hasattr(self, 'condition_combo'):
            self.condition_combo.config(state=combo_state)
        
        # 砝码质量下拉菜单
        if hasattr(self, 'weight_combo'):
            self.weight_combo.config(state=combo_state)
        
        # 塔轮半径下拉菜单
        if hasattr(self, 'radius_combo'):
            self.radius_combo.config(state=combo_state)
        
        # 样品下拉菜单
        if hasattr(self, 'sample_combo'):
            self.sample_combo.config(state=combo_state)
        
        # 圆柱距离下拉菜单
        if hasattr(self, 'distance_combo'):
            self.distance_combo.config(state=combo_state)
        
        # 初始转速滑块
        if hasattr(self, 'speed_scale'):
            if not enabled:
                self.speed_scale.config(state="disabled")
            else:
                # 恢复之前的禁用状态（如果是加砝码条件，仍然禁用）
                condition = self.condition_var.get() if hasattr(self, 'condition_var') else "空载物台"
                if condition in ["加砝码", "加砝码和样品"]:
                    self.speed_scale.config(state="disabled")
                else:
                    self.speed_scale.config(state="normal")
        
        # 开始按钮
        if hasattr(self, 'start_btn'):
            self.start_btn.config(state=button_state)
        
        # 停止按钮（动画运行时启用，停止后禁用）
        if hasattr(self, 'stop_btn'):
            self.stop_btn.config(state=tk.NORMAL if not enabled else tk.DISABLED)

    def update_host_display(self, tag, text):
        """更新右上主机区域的显示"""
        if hasattr(self, 'canvas'):
            if tag == "count":
                self.canvas.itemconfig(self.count_text_id, text=text)
            elif tag == "time":
                self.canvas.itemconfig(self.time_text_id, text=text)
            elif tag == "accel":
                self.canvas.itemconfig(self.accel_text_id, text=text)
    
    def calculate(self):
        """计算平均值和转动惯量"""
        # 先保存当前数据
        self.save_current_table_data()
        
        if self.current_tab == "测量空载物台的转动惯量":
            self.calculate_empty()
        elif self.current_tab == "测量圆盘的转动惯量":
            self.calculate_sample("圆盘")
        elif self.current_tab == "测量圆环的转动惯量":
            self.calculate_sample("圆环")
        elif self.current_tab == "验证平行轴定理":  # 新增
            self.calculate_parallel_axis()
        else:
            messagebox.showinfo("提示", "当前功能开发中")

    def calculate_empty(self):
        """计算空载物台转动惯量"""
        # 保存当前页面的参数
        self.save_current_params()
        
        # 1. 同步实验参数
        condition = self.condition_var.get()
        
        if condition in ["加砝码", "加砝码和样品"]:
            weight_str = self.weight_var.get()
            weight_g = int(weight_str.replace('g', '')) + 23.98
            self.m_value_label.config(text=f"{weight_g:.2f} g")
        else:
            self.m_value_label.config(text="53.98 g")
        
        if condition in ["加砝码", "加砝码和样品"]:
            radius_str = self.radius_var.get()
            self.r_value_label.config(text=radius_str)
        else:
            self.r_value_label.config(text="25 mm")
        
        self.g_value_label.config(text="9.794 m/s²")
        
        # 2. 从表格中读取数据
        beta1_values = []
        beta2_values = []
        
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            if values[0] == "平均值":
                continue
            try:
                beta1_str = values[1].strip()
                beta2_str = values[2].strip()
                
                if beta1_str:
                    beta1 = float(beta1_str)
                    beta1_values.append(beta1)
                if beta2_str:
                    beta2 = float(beta2_str)
                    beta2_values.append(beta2)
            except (ValueError, IndexError):
                continue
        
        if len(beta1_values) == 0 or len(beta2_values) == 0:
            messagebox.showwarning("数据不足", "请先在表格中输入数据！")
            return
        
        avg_beta1 = sum(beta1_values) / len(beta1_values)
        avg_beta2 = sum(beta2_values) / len(beta2_values)
        
        # 更新平均值行
        for item in self.tree.get_children():
            if self.tree.item(item, "values")[0] == "平均值":
                self.tree.item(item, values=("平均值", f"{avg_beta1:.4f}", f"{avg_beta2:.4f}"))
                break
        
        self.save_current_table_data()
        
        # 3. 计算J1
        if avg_beta1 >= 0:
            self.j1_value_label.config(text="数据无效: β₁应为负数")
            self.saved_results["空载物台"]["j1"] = "数据无效"
            messagebox.showwarning("计算失败", "β₁应为负数（表示减速运动）")
            return
        
        if avg_beta2 <= 0:
            self.j1_value_label.config(text="数据无效: β₂应为正数")
            self.saved_results["空载物台"]["j1"] = "数据无效"
            messagebox.showwarning("计算失败", "β₂应为正数（表示加速运动）")
            return
        
        if avg_beta2 <= avg_beta1:
            self.j1_value_label.config(text="数据无效: β₂应大于β₁")
            self.saved_results["空载物台"]["j1"] = "数据无效"
            messagebox.showwarning("计算失败", f"β₂({avg_beta2:.4f})应大于β₁({avg_beta1:.4f})")
            return
        
        try:
            weight_str = self.weight_var.get()
            weight_g = int(weight_str.replace('g', '')) + 23.98
            m = weight_g / 1000.0
            
            radius_str = self.radius_var.get()
            radius_mm = int(radius_str.replace('mm', ''))
            R = radius_mm / 1000.0
            
            g = 9.794
            
            numerator = m * R * (g - R * avg_beta2)
            denominator = avg_beta2 - avg_beta1
            
            if abs(denominator) > 1e-10:
                J1_calculated = numerator / denominator
                # 保存计算结果
                self.saved_results["空载物台"]["j1"] = f"{J1_calculated:.6f} kg·m²"
                self.j1_value_label.config(text=self.saved_results["空载物台"]["j1"])
                
                messagebox.showinfo("计算完成", f"J₁ = {J1_calculated:.6f} kg·m²")
            else:
                self.j1_value_label.config(text="计算错误: 分母为零")
                self.saved_results["空载物台"]["j1"] = "计算错误"
                messagebox.showwarning("计算失败", "β₂ - β₁ = 0")
                
        except Exception as e:
            self.j1_value_label.config(text="计算错误")
            self.saved_results["空载物台"]["j1"] = "计算错误"
            messagebox.showerror("计算失败", str(e))

    def calculate_sample(self, sample_type):
        """计算样品转动惯量（圆盘或圆环）"""
        # 保存当前页面的参数
        self.save_current_params()
        # 1. 同步实验参数
        condition = self.condition_var.get()
        
        if condition in ["加砝码", "加砝码和样品"]:
            weight_str = self.weight_var.get()
            weight_g = int(weight_str.replace('g', '')) + 23.98
            self.m_value_label.config(text=f"{weight_g:.2f} g")
        else:
            self.m_value_label.config(text="53.98 g")
        
        if condition in ["加砝码", "加砝码和样品"]:
            radius_str = self.radius_var.get()
            self.r_value_label.config(text=radius_str)
        else:
            self.r_value_label.config(text="25 mm")
        
        self.g_value_label.config(text="9.794 m/s²")
        
        # 2. 从表格中读取数据 (beta3, beta4)
        beta3_values = []
        beta4_values = []
        
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            if values[0] == "平均值":
                continue
            try:
                beta3_str = values[1].strip()
                beta4_str = values[2].strip()
                
                if beta3_str:
                    beta3 = float(beta3_str)
                    beta3_values.append(beta3)
                if beta4_str:
                    beta4 = float(beta4_str)
                    beta4_values.append(beta4)
            except (ValueError, IndexError):
                continue
        
        if len(beta3_values) == 0 or len(beta4_values) == 0:
            messagebox.showwarning("数据不足", "请先在表格中输入数据！")
            return
        
        avg_beta3 = sum(beta3_values) / len(beta3_values)
        avg_beta4 = sum(beta4_values) / len(beta4_values)
        
        # 更新平均值行
        for item in self.tree.get_children():
            if self.tree.item(item, "values")[0] == "平均值":
                self.tree.item(item, values=("平均值", f"{avg_beta3:.4f}", f"{avg_beta4:.4f}"))
                break
        
        self.save_current_table_data()
        
        # 3. 计算J2 (样品+载物台转动惯量)
        if avg_beta3 >= 0:
            # 使用 j2_value_label 替代 j1_value_label
            if hasattr(self, 'j2_value_label'):
                self.j2_value_label.config(text="数据无效: β₃应为负数")
            # 保存错误状态到对应的样品类型
            self.saved_results[sample_type] = {
                "j2": "数据无效",
                "j3": "数据无效",
                "error": ""
            }
            messagebox.showwarning("计算失败", "β₃应为负数（表示减速运动）")
            return
        
        if avg_beta4 <= 0:
            if hasattr(self, 'j2_value_label'):
                self.j2_value_label.config(text="数据无效: β₄应为正数")
            self.saved_results[sample_type] = {
                "j2": "数据无效",
                "j3": "数据无效",
                "error": ""
            }
            messagebox.showwarning("计算失败", "β₄应为正数（表示加速运动）")
            return
        
        if avg_beta4 <= avg_beta3:
            if hasattr(self, 'j2_value_label'):
                self.j2_value_label.config(text="数据无效: β₄应大于β₃")
            self.saved_results[sample_type] = {
                "j2": "数据无效",
                "j3": "数据无效",
                "error": ""
            }
            messagebox.showwarning("计算失败", f"β₄({avg_beta4:.4f})应大于β₃({avg_beta3:.4f})")
            return
        
        try:
            weight_str = self.weight_var.get()
            weight_g = int(weight_str.replace('g', '')) + 23.98
            m = weight_g / 1000.0
            
            radius_str = self.radius_var.get()
            radius_mm = int(radius_str.replace('mm', ''))
            R = radius_mm / 1000.0
            
            g = 9.794
            
            # J₂ = mR(g - R·β₄) / (β₄ - β₃)
            numerator = m * R * (g - R * avg_beta4)
            denominator = avg_beta4 - avg_beta3
            
            if abs(denominator) > 1e-10:
                J2_calculated = numerator / denominator
                # 使用 j2_value_label 替代 j1_value_label
                if hasattr(self, 'j2_value_label'):
                    self.j2_value_label.config(text=f"{J2_calculated:.6f} kg·m²")
                
                # 计算J3 = J2 - J1
                J1 = 0.01033  # 已知空载物台转动惯量
                J3_calculated = J2_calculated - J1
                
                # 理论值
                if sample_type == "圆盘":
                    J3_theory = 0.0156
                else:  # 圆环
                    J3_theory = 0.0264
                
                # 计算相对误差
                relative_error = abs(J3_calculated - J3_theory) / J3_theory * 100
                
                # 保存计算结果到对应的样品类型
                self.saved_results[sample_type] = {
                    "j2": f"{J2_calculated:.6f} kg·m²",
                    "j3": f"{J3_calculated:.6f} kg·m²",
                    "error": f"{relative_error:.2f}%"
                }
                
                # 更新额外参数显示
                if hasattr(self, 'j3_value_label'):
                    self.j3_value_label.config(text=self.saved_results[sample_type]["j3"])
                if hasattr(self, 'error_value_label'):
                    self.error_value_label.config(text=self.saved_results[sample_type]["error"])
                if hasattr(self, 'j2_value_label'):
                    self.j2_value_label.config(text=self.saved_results[sample_type]["j2"])
                
                messagebox.showinfo("计算完成", 
                                f"β₃平均值: {avg_beta3:.4f} rad/s²\n"
                                f"β₄平均值: {avg_beta4:.4f} rad/s²\n"
                                f"\n计算结果:\n"
                                f"J₂ (样品+载物台) = {J2_calculated:.6f} kg·m²\n"
                                f"J₃ (样品) = {J3_calculated:.6f} kg·m²\n"
                                f"理论值: {J3_theory:.4f} kg·m²\n"
                                f"相对误差: {relative_error:.2f}%")
            else:
                if hasattr(self, 'j2_value_label'):
                    self.j2_value_label.config(text="计算错误: 分母为零")
                self.saved_results[sample_type] = {
                    "j2": "计算错误",
                    "j3": "计算错误",
                    "error": ""
                }
                messagebox.showwarning("计算失败", "β₄ - β₃ = 0")
                
        except Exception as e:
            if hasattr(self, 'j2_value_label'):
                self.j2_value_label.config(text="计算错误")
            self.saved_results[sample_type] = {
                "j2": "计算错误",
                "j3": "计算错误",
                "error": ""
            }
            messagebox.showerror("计算失败", str(e))
    
    def calculate_parallel_axis(self):
        """计算平行轴定理验证"""
        # 保存当前页面的参数
        self.save_current_params()
        
        # 1. 同步实验参数
        condition = self.condition_var.get()
        
        if condition in ["加砝码", "加砝码和样品"]:
            weight_str = self.weight_var.get()
            weight_g = int(weight_str.replace('g', '')) + 23.98
            self.m_value_label.config(text=f"{weight_g:.2f} g")
        else:
            self.m_value_label.config(text="53.98 g")
        
        if condition in ["加砝码", "加砝码和样品"]:
            radius_str = self.radius_var.get()
            self.r_value_label.config(text=radius_str)
        else:
            self.r_value_label.config(text="25 mm")
        
        self.g_value_label.config(text="9.794 m/s²")
        
        # 2. 从表格中读取数据 (beta3, beta4)
        beta3_values = []
        beta4_values = []
        
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            if values[0] == "平均值":
                continue
            try:
                beta3_str = values[1].strip()
                beta4_str = values[2].strip()
                
                if beta3_str:
                    beta3 = float(beta3_str)
                    beta3_values.append(beta3)
                if beta4_str:
                    beta4 = float(beta4_str)
                    beta4_values.append(beta4)
            except (ValueError, IndexError):
                continue
        
        if len(beta3_values) == 0 or len(beta4_values) == 0:
            messagebox.showwarning("数据不足", "请先在表格中输入数据！")
            return
        
        avg_beta3 = sum(beta3_values) / len(beta3_values)
        avg_beta4 = sum(beta4_values) / len(beta4_values)
        
        # 更新平均值行
        for item in self.tree.get_children():
            if self.tree.item(item, "values")[0] == "平均值":
                self.tree.item(item, values=("平均值", f"{avg_beta3:.4f}", f"{avg_beta4:.4f}"))
                break
        
        self.save_current_table_data()
        
        # 3. 计算J2 (圆柱+载物台转动惯量)
        if avg_beta3 >= 0:
            if hasattr(self, 'j2_value_label'):
                self.j2_value_label.config(text="数据无效: β₃应为负数")
            self.saved_results["平行轴定理"] = {
                "j2": "数据无效",
                "j3": "数据无效",
                "error": ""
            }
            messagebox.showwarning("计算失败", "β₃应为负数（表示减速运动）")
            return
        
        if avg_beta4 <= 0:
            if hasattr(self, 'j2_value_label'):
                self.j2_value_label.config(text="数据无效: β₄应为正数")
            self.saved_results["平行轴定理"] = {
                "j2": "数据无效",
                "j3": "数据无效",
                "error": ""
            }
            messagebox.showwarning("计算失败", "β₄应为正数（表示加速运动）")
            return
        
        if avg_beta4 <= avg_beta3:
            if hasattr(self, 'j2_value_label'):
                self.j2_value_label.config(text="数据无效: β₄应大于β₃")
            self.saved_results["平行轴定理"] = {
                "j2": "数据无效",
                "j3": "数据无效",
                "error": ""
            }
            messagebox.showwarning("计算失败", f"β₄({avg_beta4:.4f})应大于β₃({avg_beta3:.4f})")
            return
        
        try:
            weight_str = self.weight_var.get()
            weight_g = int(weight_str.replace('g', '')) + 23.98
            m = weight_g / 1000.0
            
            radius_str = self.radius_var.get()
            radius_mm = int(radius_str.replace('mm', ''))
            R = radius_mm / 1000.0
            
            g = 9.794
            
            # J₂ = mR(g - R·β₄) / (β₄ - β₃)
            numerator = m * R * (g - R * avg_beta4)
            denominator = avg_beta4 - avg_beta3
            
            if abs(denominator) > 1e-10:
                J2_calculated = numerator / denominator
                if hasattr(self, 'j2_value_label'):
                    self.j2_value_label.config(text=f"{J2_calculated:.6f} kg·m²")
                
                # 计算J3 = J2 - J1
                J1 = 0.01033  # 已知空载物台转动惯量
                J3_calculated = J2_calculated - J1
                
                # 理论值
                J3_theory = 0.00670
                
                # 计算相对误差
                relative_error = abs(J3_calculated - J3_theory) / J3_theory * 100
                
                # 保存计算结果
                self.saved_results["平行轴定理"] = {
                    "j2": f"{J2_calculated:.6f} kg·m²",
                    "j3": f"{J3_calculated:.6f} kg·m²",
                    "error": f"{relative_error:.2f}%"
                }
                
                # 更新额外参数显示
                if hasattr(self, 'j3_value_label'):
                    self.j3_value_label.config(text=self.saved_results["平行轴定理"]["j3"])
                if hasattr(self, 'error_value_label'):
                    self.error_value_label.config(text=self.saved_results["平行轴定理"]["error"])
                if hasattr(self, 'j2_value_label'):
                    self.j2_value_label.config(text=self.saved_results["平行轴定理"]["j2"])
                
                messagebox.showinfo("计算完成", 
                                f"β₃平均值: {avg_beta3:.4f} rad/s²\n"
                                f"β₄平均值: {avg_beta4:.4f} rad/s²\n"
                                f"\n计算结果:\n"
                                f"J₂ (圆柱+载物台) = {J2_calculated:.6f} kg·m²\n"
                                f"J₃ (圆柱) = {J3_calculated:.6f} kg·m²\n"
                                f"理论值: {J3_theory:.4f} kg·m²\n"
                                f"相对误差: {relative_error:.2f}%")
            else:
                if hasattr(self, 'j2_value_label'):
                    self.j2_value_label.config(text="计算错误: 分母为零")
                self.saved_results["平行轴定理"] = {
                    "j2": "计算错误",
                    "j3": "计算错误",
                    "error": ""
                }
                messagebox.showwarning("计算失败", "β₄ - β₃ = 0")
                
        except Exception as e:
            if hasattr(self, 'j2_value_label'):
                self.j2_value_label.config(text="计算错误")
            self.saved_results["平行轴定理"] = {
                "j2": "计算错误",
                "j3": "计算错误",
                "error": ""
            }
            messagebox.showerror("计算失败", str(e))

    def clear_data(self):
        """清空数据"""
        # 弹出确认对话框
        result = messagebox.askyesno(
            "确认清空", 
            f"确定要清空当前选项卡 '{self.current_tab}' 的所有数据吗？\n\n此操作不可撤销！",
            icon='warning'
        )
        
        if not result:
            return
        
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            if values[0] == "平均值":
                self.tree.item(item, values=("平均值", "", ""))
            else:
                self.tree.item(item, values=(values[0], "", ""))
        
        # 清空保存的数据
        self.table_data[self.current_tab]["values"] = []
        
        # 根据当前标签重置显示
        if self.current_tab == "测量空载物台的转动惯量":
            self.saved_results["空载物台"]["j1"] = "待计算"
            self.j1_value_label.config(text="待计算")
            # 重置参数显示
            params = self.saved_params.get(self.current_tab, {"weight": "5g", "radius": "15mm"})
            weight_g = int(params["weight"].replace('g', '')) + 23.98
            self.m_value_label.config(text=f"{weight_g:.2f} g")
            self.r_value_label.config(text=params["radius"])
            
        elif self.current_tab == "测量圆盘的转动惯量":
            self.saved_results["圆盘"] = {"j2": "待计算", "j3": "待计算", "error": "待计算"}
            if hasattr(self, 'j2_value_label'):
                self.j2_value_label.config(text="待计算")
            if hasattr(self, 'j3_value_label'):
                self.j3_value_label.config(text="待计算")
            if hasattr(self, 'error_value_label'):
                self.error_value_label.config(text="待计算")
            # 恢复参数显示
            params = self.saved_params.get(self.current_tab, {"weight": "5g", "radius": "15mm"})
            weight_g = int(params["weight"].replace('g', '')) + 23.98
            self.m_value_label.config(text=f"{weight_g:.2f} g")
            self.r_value_label.config(text=params["radius"])
            # 恢复样品参数
            self.sample_param_label.config(text="圆盘样品:")
            self.sample_param_value_label.config(text="M=2166g, d₁=240mm")
            self.j2_label.config(text="载物台加样品转动惯量J₂:")
            
        elif self.current_tab == "测量圆环的转动惯量":
            self.saved_results["圆环"] = {"j2": "待计算", "j3": "待计算", "error": "待计算"}
            if hasattr(self, 'j2_value_label'):
                self.j2_value_label.config(text="待计算")
            if hasattr(self, 'j3_value_label'):
                self.j3_value_label.config(text="待计算")
            if hasattr(self, 'error_value_label'):
                self.error_value_label.config(text="待计算")
            # 恢复参数显示
            params = self.saved_params.get(self.current_tab, {"weight": "5g", "radius": "15mm"})
            weight_g = int(params["weight"].replace('g', '')) + 23.98
            self.m_value_label.config(text=f"{weight_g:.2f} g")
            self.r_value_label.config(text=params["radius"])
            # 恢复样品参数
            self.sample_param_label.config(text="圆环样品:")
            self.sample_param_value_label.config(text="M=2172g, d₁=200mm, d₂=240mm")
            self.j2_label.config(text="载物台加样品转动惯量J₂:")
            
        elif self.current_tab == "验证平行轴定理":
            self.saved_results["平行轴定理"] = {"j2": "待计算", "j3": "待计算", "error": "待计算"}
            if hasattr(self, 'j2_value_label'):
                self.j2_value_label.config(text="待计算")
            if hasattr(self, 'j3_value_label'):
                self.j3_value_label.config(text="待计算")
            if hasattr(self, 'error_value_label'):
                self.error_value_label.config(text="待计算")
            # 恢复参数显示
            params = self.saved_params.get(self.current_tab, {"weight": "5g", "radius": "15mm", "distance": "45mm"})
            weight_g = int(params["weight"].replace('g', '')) + 23.98
            self.m_value_label.config(text=f"{weight_g:.2f} g")
            self.r_value_label.config(text=params["radius"])
            # 恢复样品参数
            distance = params.get("distance", "45mm")
            self.sample_param_label.config(text="圆柱样品:")
            self.sample_param_value_label.config(text=f"M₁=300.9g, d₁=29.86mm  M₂=300.6g, d₂=29.85mm  x={distance}")
            self.j2_label.config(text="载物台加样品转动惯量J₂:")
        
        messagebox.showinfo("清空完成", f"当前选项卡 '{self.current_tab}' 的数据已清空")

    def export_data(self):
        """导出数据到文件"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("XLSX files", "*.xlsx")],
            initialfile=f"{self.current_tab}.xlsx"
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8-sig') as f:
                    # 写入选项卡名称
                    f.write(f"# 选项卡:{self.current_tab}\n")
                    f.write(f"# 导出时间:{time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write("\n")
                    
                    # 写入表格数据
                    headers = []
                    if self.current_tab == "测量空载物台的转动惯量":
                        headers = ["测量次数", "β1（弧度*s^-2）", "β2（弧度*s^-2）"]
                    elif self.current_tab == "测量圆盘的转动惯量":
                        headers = ["测量次数", "β3（弧度*s^-2）", "β4（弧度*s^-2）"]
                    elif self.current_tab == "测量圆环的转动惯量":
                        headers = ["测量次数", "β3（弧度*s^-2）", "β4（弧度*s^-2）"]
                    elif self.current_tab == "验证平行轴定理":
                        headers = ["测量次数", "β3（弧度*s^-2）", "β4（弧度*s^-2）"]
                    else:
                        headers = ["测量次数", "β1（弧度*s^-2）", "β2（弧度*s^-2）"]
                    
                    f.write(",".join(headers) + "\n")
                    
                    # 写入5行数据
                    for item in self.tree.get_children():
                        values = self.tree.item(item, "values")
                        if values[0] == "平均值":
                            continue
                        row = []
                        for v in values:
                            v_str = str(v).strip()
                            if ',' in v_str or '\n' in v_str:
                                v_str = f'"{v_str}"'
                            row.append(v_str)
                        f.write(",".join(row) + "\n")
                    
                    # 写入平均值行
                    for item in self.tree.get_children():
                        if self.tree.item(item, "values")[0] == "平均值":
                            values = self.tree.item(item, "values")
                            row = []
                            for v in values:
                                v_str = str(v).strip()
                                if ',' in v_str or '\n' in v_str:
                                    v_str = f'"{v_str}"'
                                row.append(v_str)
                            f.write(",".join(row) + "\n")
                            break
                    
                    f.write("\n")
                    f.write("# 实验参数:\n")
                    
                    # 获取当前页面的保存参数
                    params = self.saved_params.get(self.current_tab, {"weight": "5g", "radius": "15mm"})
                    
                    # 写入通用参数
                    f.write(f"砝码质量,{params.get('weight', '5g')}\n")
                    f.write(f"塔轮半径,{params.get('radius', '15mm')}\n")
                    f.write(f"砝码和砝码架总质量 m,{self.m_value_label.cget('text')}\n")
                    f.write(f"塔轮半径 R,{self.r_value_label.cget('text')}\n")
                    f.write(f"重力加速度 g,{self.g_value_label.cget('text')}\n")
                    
                    # 写入当前选项卡特定的参数
                    if self.current_tab == "测量空载物台的转动惯量":
                        f.write(f"空载物台转动惯量 J₁,{self.saved_results['空载物台']['j1']}\n")
                    elif self.current_tab == "测量圆盘的转动惯量":
                        f.write(f"圆盘+载物台转动惯量 J₂,{self.saved_results['圆盘']['j2']}\n")
                        f.write(f"圆盘转动惯量 J₃,{self.saved_results['圆盘']['j3']}\n")
                        f.write(f"相对误差,{self.saved_results['圆盘']['error']}\n")
                        f.write(f"圆盘质量 M,2166g\n")
                        f.write(f"圆盘直径 d₁,240mm\n")
                    elif self.current_tab == "测量圆环的转动惯量":
                        f.write(f"圆环+载物台转动惯量 J₂,{self.saved_results['圆环']['j2']}\n")
                        f.write(f"圆环转动惯量 J₃,{self.saved_results['圆环']['j3']}\n")
                        f.write(f"相对误差,{self.saved_results['圆环']['error']}\n")
                        f.write(f"圆环质量 M,2172g\n")
                        f.write(f"圆环内径 d₁,200mm\n")
                        f.write(f"圆环外径 d₂,240mm\n")
                    elif self.current_tab == "验证平行轴定理":
                        f.write(f"圆柱+载物台转动惯量 J₂,{self.saved_results['平行轴定理']['j2']}\n")
                        f.write(f"圆柱转动惯量 J₃,{self.saved_results['平行轴定理']['j3']}\n")
                        f.write(f"相对误差,{self.saved_results['平行轴定理']['error']}\n")
                        f.write(f"圆柱1质量 M₁,300.9g\n")
                        f.write(f"圆柱1直径 d₁,29.86mm\n")
                        f.write(f"圆柱2质量 M₂,300.6g\n")
                        f.write(f"圆柱2直径 d₂,29.85mm\n")
                        f.write(f"离中心距离 x,{params.get('distance', '45mm')}\n")
                        f.write(f"理论值 J,0.00670 kg·m²\n")
                    
                    f.write(f"当前实验条件,{self.condition_var.get()}\n")
                    f.write(f"初始转速,{self.speed_var.get():.2f} 弧度/s²\n")
                    
                    messagebox.showinfo("导出成功", f"数据已导出到: {file_path}")
            except Exception as e:
                messagebox.showerror("导出失败", f"导出时出错: {str(e)}")
    
    def import_data(self):
        """从文件导入数据"""
        file_path = filedialog.askopenfilename(filetypes=[("XLSX files", "*.xlsx")])
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    lines = f.readlines()
                
                # 解析文件头，获取选项卡信息
                tab_name = None
                for line in lines:
                    if line.startswith('# 选项卡:'):
                        tab_name = line.strip().replace('# 选项卡:', '').strip()
                        break
                
                # 如果文件中的选项卡与当前选项卡不同，询问是否切换
                if tab_name and tab_name != self.current_tab:
                    result = messagebox.askyesno(
                        "选项卡不匹配", 
                        f"文件中的选项卡为 '{tab_name}'，当前选项卡为 '{self.current_tab}'\n"
                        f"是否切换到 '{tab_name}' 并导入数据？"
                    )
                    if result:
                        # 切换到文件中的选项卡
                        self.switch_tab(tab_name)
                    else:
                        messagebox.showinfo("取消导入", "导入已取消")
                        return
                
                # 查找数据起始行
                data_start_line = 0
                for i, line in enumerate(lines):
                    if "测量次数" in line or ("β1" in line and "β2" in line) or ("β3" in line and "β4" in line):
                        data_start_line = i + 1
                        break
                
                # 导入表格数据
                imported_count = 0
                for i in range(5):  # 最多5行数据
                    line_idx = data_start_line + i
                    if line_idx >= len(lines):
                        break
                    
                    line = lines[line_idx].strip()
                    if not line:
                        continue
                    
                    if line.startswith('#') or "实验参数" in line or "砝码" in line:
                        break
                    
                    parts = []
                    in_quotes = False
                    current_part = ""
                    
                    for char in line:
                        if char == '"':
                            in_quotes = not in_quotes
                        elif char == ',' and not in_quotes:
                            parts.append(current_part.strip())
                            current_part = ""
                        else:
                            current_part += char
                    parts.append(current_part.strip())
                    parts = [p.strip('"') for p in parts]
                    
                    if i < len(self.table_data[self.current_tab]["entries"]):
                        current_values = list(self.tree.item(self.table_data[self.current_tab]["entries"][i], "values"))
                        if len(parts) >= 2:
                            current_values[1] = parts[1]
                        if len(parts) >= 3:
                            current_values[2] = parts[2]
                        self.tree.item(self.table_data[self.current_tab]["entries"][i], values=tuple(current_values))
                        imported_count += 1
                
                # 查找并导入平均值行
                for i, line in enumerate(lines):
                    if line.strip().startswith("平均值"):
                        parts = []
                        in_quotes = False
                        current_part = ""
                        for char in line.strip():
                            if char == '"':
                                in_quotes = not in_quotes
                            elif char == ',' and not in_quotes:
                                parts.append(current_part.strip())
                                current_part = ""
                            else:
                                current_part += char
                        parts.append(current_part.strip())
                        parts = [p.strip('"') for p in parts]
                        
                        for item in self.tree.get_children():
                            if self.tree.item(item, "values")[0] == "平均值":
                                if len(parts) >= 2:
                                    current_values = list(self.tree.item(item, "values"))
                                    current_values[1] = parts[1]
                                if len(parts) >= 3:
                                    current_values[2] = parts[2]
                                self.tree.item(item, values=tuple(current_values))
                                break
                        break
                
                # 导入实验参数
                params = {}
                for line in lines:
                    line = line.strip()
                    if not line or line.startswith('#') or "测量次数" in line:
                        continue
                    if ',' in line and not line.startswith('"'):
                        parts = line.split(',')
                        if len(parts) >= 2:
                            key = parts[0].strip()
                            value = ','.join(parts[1:]).strip()
                            params[key] = value
                
                # --- 关键修改：先保存参数到 saved_params ---
                # 更新 saved_params
                if self.current_tab in self.saved_params:
                    if '砝码质量' in params:
                        self.saved_params[self.current_tab]["weight"] = params['砝码质量']
                    if '塔轮半径' in params:
                        self.saved_params[self.current_tab]["radius"] = params['塔轮半径']
                    if '离中心距离 x' in params and self.current_tab == "验证平行轴定理":
                        self.saved_params[self.current_tab]["distance"] = params['离中心距离 x']
                
                # 更新左侧实验操作区域的参数
                if '砝码质量' in params and hasattr(self, 'weight_var'):
                    self.weight_var.set(params['砝码质量'])
                if '塔轮半径' in params and hasattr(self, 'radius_var'):
                    self.radius_var.set(params['塔轮半径'])
                if '离中心距离 x' in params and hasattr(self, 'distance_var'):
                    self.distance_var.set(params['离中心距离 x'])
                
                # 恢复计算结果
                if self.current_tab == "测量空载物台的转动惯量":
                    if '空载物台转动惯量 J₁' in params:
                        self.saved_results["空载物台"]["j1"] = params['空载物台转动惯量 J₁']
                elif self.current_tab == "测量圆盘的转动惯量":
                    if '圆盘+载物台转动惯量 J₂' in params:
                        self.saved_results["圆盘"]["j2"] = params['圆盘+载物台转动惯量 J₂']
                    if '圆盘转动惯量 J₃' in params:
                        self.saved_results["圆盘"]["j3"] = params['圆盘转动惯量 J₃']
                    if '相对误差' in params:
                        self.saved_results["圆盘"]["error"] = params['相对误差']
                elif self.current_tab == "测量圆环的转动惯量":
                    if '圆环+载物台转动惯量 J₂' in params:
                        self.saved_results["圆环"]["j2"] = params['圆环+载物台转动惯量 J₂']
                    if '圆环转动惯量 J₃' in params:
                        self.saved_results["圆环"]["j3"] = params['圆环转动惯量 J₃']
                    if '相对误差' in params:
                        self.saved_results["圆环"]["error"] = params['相对误差']
                elif self.current_tab == "验证平行轴定理":
                    if '圆柱+载物台转动惯量 J₂' in params:
                        self.saved_results["平行轴定理"]["j2"] = params['圆柱+载物台转动惯量 J₂']
                    if '圆柱转动惯量 J₃' in params:
                        self.saved_results["平行轴定理"]["j3"] = params['圆柱转动惯量 J₃']
                    if '相对误差' in params:
                        self.saved_results["平行轴定理"]["error"] = params['相对误差']
                
                # 保存导入的数据
                self.save_current_table_data()
                
                # --- 关键修改：调用 update_params_display 刷新显示 ---
                # 清除第三行并重新显示
                self.clear_extra_params()
                self.update_params_display(self.current_tab)
                
                # 额外更新：确保塔轮半径和砝码质量显示正确
                condition = self.condition_var.get()
                if condition in ["加砝码", "加砝码和样品"]:
                    if '砝码质量' in params:
                        weight_g = int(params['砝码质量'].replace('g', '')) + 23.98
                        self.m_value_label.config(text=f"{weight_g:.2f} g")
                    if '塔轮半径' in params:
                        self.r_value_label.config(text=params['塔轮半径'])
                
                if imported_count > 0:
                    messagebox.showinfo("导入成功", f"已导入 {imported_count} 行数据和所有实验参数")
                else:
                    messagebox.showwarning("导入警告", "未找到有效数据，但参数可能已导入")
                        
            except Exception as e:
                messagebox.showerror("导入失败", f"导入时出错: {str(e)}")

def main():
    root = tk.Tk()
    app = RigidBodyInertiaApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()